"""
Multi-API Academic Paper Search Tool
Searches Crossref, OpenAlex, and PubMed APIs, filters, downloads, and searches full-text.

FLOW:
Step 1: Initial menu (New Search / Continue Downloading / Search Full-texts)
Step 2: Master Query (4 API search options)
Step 3: In-memory filtering (remove no-abstract, deduplicate, assign study numbers)
Step 4: Option (Filter OA now OR continue to subquery)
Step 5: Subquery (local AND/OR filtering on abstracts)
Step 6: Filter by OA/non-OA, save CSV
Step 7: Download OA papers
Step 8: Search full-text of downloaded papers
Step 9: Full-text search results

Usage:
    python multi_api_search.py
"""

import requests
import csv
import re
import html
import time
import os
import sys
import threading
from datetime import datetime
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
import xml.etree.ElementTree as ET

# Add scihub_downloader to path for non-OA paper downloads
SCIHUB_PATH = Path(__file__).parent / "scihub_downloader" / "scihub"
if SCIHUB_PATH.exists():
    sys.path.insert(0, str(SCIHUB_PATH))

# =============================================================================
# TERMINAL COLORS (ANSI escape codes)
# =============================================================================
GREEN = "[32m"
DIM_RED = "[2;31m"
RESET = "[0m"

# =============================================================================
# CONFIGURATION
# =============================================================================

# Memory-efficient mode: streams writes instead of holding all in RAM
MEMORY_EFFICIENT_MODE = False

# =============================================================================
# TEXTILE/MATERIALS SCIENCE JOURNAL LIST (528 entries)
# =============================================================================
# Textile/Materials Science Journals
# Format: "Full Name": "ACRONYM"
TEXTILE_JOURNALS = {
    "2015 International Conference on Material Science and Applications (icmsa-15)": "ICMSA",
    "2016 International Conference on Mechanics and Materials Science (MMS2016)": "ICMMS",
    "2017 2nd International Conference on Materials Science, Machinery and Energy Engineering (MSMEE 2017)": "ICMSMEE",
    "59th Appita Annual Conference and Exhibition: Incorporating the 13th ISWFPC (International Symposium on Wood, Fibre and Pulping Chemistry), Auckland, New Zealand, 16-19 May 2005: Proceedings": "AACEII",
    "AATCC Journal of Research": "AJR",
    "AATCC Review": "AR",
    "AATCC Review: the magazine of the textile dyeing, printing, and finishing industry": "ARMTDPFI",
    "ADVANCED MATERIAL SCIENCE AND INNOVATIONS": "AMSI",
    "AIMS Materials Science": "AMS",
    "ATA Journal for Asia on Textile and Apparel": "AJATA",
    "AUTEX Research Journal (AUTEXRJ / Autex Research Journal)": "ARJ",
    "Academia Materials Science": "AMS",
    "Advance Research in Textile Engineering": "ARTE",
    "Advanced Fiber Materials": "AFM",
    "Advanced Material Science": "AMS",
    "Advanced Materials Science and Technology": "AMST",
    "Advanced Textile Technology": "ATT",
    "Advances in Material Science": "AMS",
    "Advances in Materials Science": "AMS",
    "Advances in Materials Science and Applications": "AMSA",
    "Advances in Materials Science and Engineering": "AMSE",
    "Advances in Materials Science and Engineering An International Journal (MSEJ)": "AMSEIJ",
    "Advances in computer and materials science research.": "ACMSR",
    "Advances in materials science and engineering": "AMSE",
    "Advances in materials science research": "AMSR",
    "Africa and Middle East Textiles": "AMET",
    "Agriculture Association of Textile Chemical and Critical Reviews Journal": "AATCCRJ",
    "Al-Fiziya Journal of Materials Science Geophysics Instrumentation and Theoretical Physics": "AJMSGITP",
    "American Journal of Materials Science": "AJMS",
    "American journal of materials science": "AJMS",
    "American journal of materials science and engineering": "AJMSE",
    "American journal of materials science and technology": "AJMST",
    "Annals of Material Science": "AMS",
    "Annals of Materials Science & Engineering": "AMSE",
    "Annals of computational physics and material science.": "ACPMS",
    "Annals of the University of Oradea. Fascicle of Textils-Leatherwork/Annals of the University of Oradea. Fascicle of Textile, Leatherwork": "AUOFT",
    "Annual Review of Materials Science": "ARMS",
    "Apparel Views": "AV",
    "Applied Chemistry and Materials Science": "ACMS",
    "Arashi Journal of Metals & Material Sciences": "AJMMS",
    "Archives of Materials Science and Engineering": "AMSE",
    "Asian Journal of Material Science and Engineering": "AJMSE",
    "Asian Journal of Materials Science": "AJMS",
    "Asian Journal of Textile": "AJT",
    "Asian Technical Textile (ATT)": "ATT",
    "Asian Textile Journal (ATJ)": "ATJ",
    "Atlantis Highlights in Materials Science and Technology/Atlantis highlights in materials science and technology": "AHMST",
    "Atlas Journal of Materials Science": "AJMS",
    "BAJU Journal of Fashion & Textile Design Unesa": "BJFTDU",
    "BSU- Journal of Chemistry and Material Sciences": "BJCMS",
    "Bangladesh Textile Journal": "BTJ",
    "Basic Sciences Journal of Textile Universities": "BSJTU",
    "Beijing Textile Journal": "BTJ",
    "Bioactive Carbohydrates and Dietary Fibre": "BCDF",
    "Biology of the Cotton Stainer Dysdercus cingulatus Fabricius (Pyrrhocoridae Hemiptera)": "BCSDCF",
    "Bulletin of Materials Science": "BMS",
    "Bulletin of the United States Institute for Textile Research Inc": "BUSITRI",
    "CRASHWORTHINESS AND ENERGY ABSORPTION ANALYSIS OF MESOCARP COIR-FIBRE/EPOXY RESIN-REINFORCED GLASS FIBER (GFRP) HYBRID COMPOSITE LAMINATE": "CEAAMC",
    "Canadian Textile Journal (CTJ)": "CTJ",
    "Cellulose": "",
    "Chemical Fiber & Textile Technology": "CFTT",
    "Chemistry & Material Sciences Research Journal": "CMSRJ",
    "China Synthetic Fiber Industry": "CSFI",
    "China's Fiber and Products": "CFP",
    "Clothing & Textiles Research Journal": "CTRJ",
    "Clothing Cultures": "CC",
    "Clothing and Textiles Research Journal": "CTRJ",
    "Clotho": "",
    "Coloration Technology": "CT",
    "Colour Research & Application": "CRA",
    "Colourage": "",
    "Communications in development and assembling of textile products": "CDATP",
    "Compendium of Materials Science 材料學輯要": "CMS材",
    "Composite Science and Technology": "CST",
    "Composite materials science": "CMS",
    "Composites à renfort textile": "CÀRT",
    "Computational Materials Science": "CMS",
    "Computer Methods in Material Science": "CMMS",
    "Computer Methods in Materials Science": "CMMS",
    "Computer Methods in Materials Science.": "CMMS",
    "Conference Papers in Materials Science": "CPMS",
    "Construction materials science mechanical engineering": "CMSME",
    "Construction, materials science, mechanical engineering": "CMSME",
    "Costume: The Journal of the Costume Society": "CJCS",
    "Current Materials Science": "CMS",
    "Current Opinion in Solid State and Materials Science": "COSSMS",
    "Current Trends in Fashion Technology & Textile Engineering": "CTFTTE",
    "DEStech Transactions on Materials Science and Engineering": "DTMSE",
    "DORA Empa (Swiss Federal Laboratories for Materials Science and Technology (Empa))": "DE",
    "Dandong Textile College": "DTC",
    "Dataset Papers in Materials Science": "DPMS",
    "Digital Textile": "DT",
    "Dress": "",
    "Duck: Journal for Research in Textile and Textile Design": "DJRTTD",
    "Dyes and Pigments": "DP",
    "Ecotextile News": "EN",
    "Euro courses. Mechanical and materials science": "ECMMS",
    "European Journal of Material Sciences": "EJMS",
    "European Journal of Materials Science and Engineering": "EJMSE",
    "Exploratory Materials Science Research": "EMSR",
    "FIBER Journal – Online Journal of the International Fashion and Apparel Industry": "FJOJIFAI",
    "FIBRES & TEXTILES in Eastern Europe": "FTEE",
    "Family and Consumer Sciences Research Journal": "FCSRJ",
    "Fashion & Textile Research Journal": "FTRJ",
    "Fashion Theory: The Journal of Dress, Body and Culture": "FTJDBC",
    "Fashion and Textiles": "FT",
    "Fashion and Textiles Review": "FTR",
    "Fiber Composites": "FC",
    "Fiber Glass": "FG",
    "Fiber Reinforced Plastics/Composite": "FRP",
    "Fiber Reinforced Plastics/Composites": "FRP",
    "Fibers": "",
    "Fibers and Polymers": "FP",
    "Fibre Chemistry": "FC",
    "Fibre Diffraction Review": "FDR",
    "Fibre Science and Technology": "FST",
    "Fibre eBooks": "FE",
    "Fibres and Textiles": "FT",
    "Fibres and Textiles in Eastern Europe": "FTEE",
    "Foundations of materials science and engineering": "FMSE",
    "Frontiers of Materials Science": "FMS",
    "Frontiers of Materials Science in China": "FMSC",
    "GQ": "",
    "Geotextiles and Geomembranes": "GG",
    "Global Journal of Materials Scienceand Engineering": "GJMSE",
    "Guangdong Chemical Fiber": "GCF",
    "Handwoven": "",
    "Heilongjiang Textile": "HT",
    "Hi-tech Fiber & Application": "HFA",
    "High temperature and materials science": "HTMS",
    "Home Textiles Today": "HTT",
    "IOP Conference Series Materials Science and Engineering": "ICSMSE",
    "IOSR Journal of Polymer and Textile Engineering": "IJPTE",
    "IRESPUB Journal of Environmental & Material Sciences": "IJEMS",
    "ISRN Materials Science": "IMS",
    "ISRN Textiles": "IT",
    "Indian Journal of Fiber and Textile Research (IJFTR)": "IJFTR",
    "Indian Journal of Fibre & Textile Research": "IJFTR",
    "Indian Journal of Fibre & Textile Research (IJFTR)": "IJFTR",
    "Indian Journal of Fibre and Textile Engineering": "IJFTE",
    "Indian Journal of Materials Science": "IJMS",
    "Indian Textile Journal (ITJ)": "ITJ",
    "Indonesian Journal of Materials Science": "IJMS",
    "Industria Textila": "IT",
    "Industrie textile": "IT",
    "Innovations in Corrosion and Materials Science (Formerly Recent Patents on Corrosion Science)": "ICMS",
    "Insight - Material Science": "IMS",
    "Institute for Problems in Materials Science eBooks": "IPMSE",
    "International Conference Inventive Material Science Applications": "ICIMSA",
    "International Conference Material Science and Engineering Technology": "ICMSET",
    "International Conference Trends Material Science and Inventive Materials": "ICTMSIM",
    "International Conference on Materials Science ICMS2018": "ICMSI",
    "International Fiber Journal (IFJ)": "IFJ",
    "International Forum “Microelectronics – 2020”. Joung Scientists Scholarship “Microelectronics – 2020”. XIII International conference «Silicon – 2020». XII young scientists scholarship for silicon nanostructures and devices physics, material science, process and analysis": "IFJSSXICXYSSSNDPMSPA",
    "International Journal of Advanced Materials Science": "IJAMS",
    "International Journal of Advances in Materials Science and Engineering": "IJAMSE",
    "International Journal of Chemical & Material Sciences": "IJCMS",
    "International Journal of Clothing Science and Technology": "IJCST",
    "International Journal of Computational Materials Science and Engineering": "IJCMSE",
    "International Journal of Computational Materials Science and Surface Engineering": "IJCMSSE",
    "International Journal of Design Management and Professional Practice": "IJDMPP",
    "International Journal of Fashion Design, Technology and Education": "IJFDTE",
    "International Journal of Fashion Studies": "IJFS",
    "International Journal of Fashion Technology & Textile Engineering": "IJFTTE",
    "International Journal of Fiber and Textile Research": "IJFTR",
    "International Journal of Green Nanotechnology Materials Science & Engineering": "IJGNMSE",
    "International Journal of Green Nanotechnology: Materials Science & Engineering": "IJGNMSE",
    "International Journal of Material Science": "IJMS",
    "International Journal of Material Science and Engineering": "IJMSE",
    "International Journal of Material Science and Research": "IJMSR",
    "International Journal of Material Sciences and Technology": "IJMST",
    "International Journal of Materials Science": "IJMS",
    "International Journal of Materials Science and Applications": "IJMSA",
    "International Journal of Materials Science and Engineering": "IJMSE",
    "International Journal of Materials Science and Technology Studies": "IJMSTS",
    "International Journal of Materials Sciences": "IJMS",
    "International Journal of Metallurgical & Materials Science and Engineering": "IJMMSE",
    "International Journal of Polymer and Textile Engineering": "IJPTE",
    "International Journal of Surface Engineering and Interdisciplinary Materials Science": "IJSEIMS",
    "International Journal of Sustainable Fashion & Textiles": "IJSFT",
    "International Journal of Textile Engineering and Management": "IJTEM",
    "International Journal of Textile Science": "IJTS",
    "International Journal of Textile Science & Engineering": "IJTSE",
    "International Journal of Textile Science Research": "IJTSR",
    "International Journal of Textile and Fashion Technology": "IJTFT",
    "International Journal of Textile and Fashion Technology (IJTFT)": "IJTFT",
    "International Textile and Apparel Association Annual Conference Proceedings": "ITAAACP",
    "International conference Material Science,Smart Structures and Applications": "ICMSSA",
    "International fiber science and technology series": "IFSTS",
    "International journal of chemical & material sciences": "IJCMS",
    "International journal of chemistry and materials science.": "IJCMS",
    "Iranian Journal of Textile Nano-bio Modification (IJTNBM)": "IJTNM",
    "Iranian journal materials science and engineering": "IJMSE",
    "JOURNAL OF INNOVATIVE RESEARCH IN TEXTILE AND LIGHT INDUSTRY": "JIRTLI",
    "Journal for Dyeing and Finishing": "JDF",
    "Journal for Weavers, Spinners and Dyers": "JWSD",
    "Journal of Advanced Materials Science and Engineering": "JAMSE",
    "Journal of Advanced Research in Manufacturing Material Science & Metallurgical Engineering": "JARMMSME",
    "Journal of Advanced Research in Materials Science": "JARMS",
    "Journal of Anhui Textile Vocational Technical Institute": "JATVTI",
    "Journal of Applied Material Science & Engineering Research": "JAMSER",
    "Journal of Beijing Institute of Clothing": "JBIC",
    "Journal of Beijing Institute of Clothing Technology": "JBICT",
    "Journal of Building Material Science": "JBMS",
    "Journal of Chemical Engineering and Materials Science": "JCEMS",
    "Journal of Chemistry and Material Sciences (JCMS)": "JCMS",
    "Journal of Chengdu Textile College": "JCTC",
    "Journal of Clothing Science": "JCS",
    "Journal of Coated Fabrics": "JCF",
    "Journal of Composite Materials": "JCM",
    "Journal of Computational Intelligence in Materials Science": "JCIMS",
    "Journal of Cotton Science": "JCS",
    "Journal of Design Textiles Engineering and Technology": "JDTET",
    "Journal of Design and Textiles": "JDT",
    "Journal of Engineered Fibers and Fabrics": "JEFF",
    "Journal of Fashion Marketing and Management": "JFMM",
    "Journal of Fashion Technology & Textile Engineering": "JFTTE",
    "Journal of Fashion Technology and Textile Engineering (JFTTE)": "JFTTE",
    "Journal of Fiber Bioengineering and Informatics": "JFBI",
    "Journal of Fiber Science and Technology": "JFST",
    "Journal of Fibers and Polymer Composites": "JFPC",
    "Journal of Henan Textile College": "JHTC",
    "Journal of Industrial Textile": "JIT",
    "Journal of Industrial Textiles": "JIT",
    "Journal of Japanese Association for Dietary Fiber Research": "JJADFR",
    "Journal of Material Science & Engineering": "JMSE",
    "Journal of Material Science and Engineering": "JMSE",
    "Journal of Material Science and Radiation": "JMSR",
    "Journal of Material Science and Technology": "JMST",
    "Journal of Material Science and Technology Research": "JMSTR",
    "Journal of Material Sciences & Applied Engineering": "JMSAE",
    "Journal of Material Sciences & Manfacturing Research": "JMSMR",
    "Journal of Material Sciences & Manufacturing Research": "JMSMR",
    "Journal of Material Sciences and Engineering Technology": "JMSET",
    "Journal of Materials Science": "JMS",
    "Journal of Materials Science Composites": "JMSC",
    "Journal of Materials Science Letters": "JMSL",
    "Journal of Materials Science Materials Theory": "JMSMT",
    "Journal of Materials Science Materials in Electronics": "JMSME",
    "Journal of Materials Science Materials in Energy": "JMSME",
    "Journal of Materials Science Materials in Engineering": "JMSME",
    "Journal of Materials Science Materials in Medicine": "JMSMM",
    "Journal of Materials Science Metallurgy": "JMSM",
    "Journal of Materials Science Polymers": "JMSP",
    "Journal of Materials Science Research": "JMSR",
    "Journal of Materials Science Research and Reviews": "JMSRR",
    "Journal of Materials Science and Applied Energy": "JMSAE",
    "Journal of Materials Science and Chemical Engineering": "JMSCE",
    "Journal of Materials Science and Engineering A": "JMSE",
    "Journal of Materials Science and Engineering B": "JMSEB",
    "Journal of Materials Science and Engineering with Advanced Technology": "JMSEAT",
    "Journal of Materials Science and Nanotechnology": "JMSN",
    "Journal of Materials Science and Surface Engineering": "JMSSE",
    "Journal of Modern Textile Science and Engineering": "JMTSE",
    "Journal of Nanotechnology and Materials Science": "JNMS",
    "Journal of Nantong Textile Vocational Technology College": "JNTVTC",
    "Journal of Natural Fibers": "JNF",
    "Journal of Northwest Institute of Textile Science and Technology": "JNITST",
    "Journal of Sustainable Materials Science and Engineering": "JSMSE",
    "Journal of Suzhou Institute of Silk Textile Technology": "JSISTT",
    "Journal of Textile": "JT",
    "Journal of Textile Association (JTA)": "JTA",
    "Journal of Textile Design Research and Practice": "JTDRP",
    "Journal of Textile Design, Research and Practice": "JTDRP",
    "Journal of Textile Engineering": "JTE",
    "Journal of Textile Engineering & Fashion Technology": "JTEFT",
    "Journal of Textile Engineering (JTE)": "JTE",
    "Journal of Textile Engineering and Fashion Technology": "JTEFT",
    "Journal of Textile Institute": "JTI",
    "Journal of Textile Research": "JTR",
    "Journal of Textile Science & Engineering": "JTSE",
    "Journal of Textile Science & Fashion Technology": "JTSFT",
    "Journal of Textile Science and Engineering": "JTSE",
    "Journal of Textile Science and Technology": "JTST",
    "Journal of Textile and Apparel, Technology and Management (JTATM)": "JTATM",
    "Journal of Textiles": "JT",
    "Journal of Textiles Coloration and Polymer Science": "JTCPS",
    "Journal of Textiles and Engineer": "JTE",
    "Journal of Textiles and Fibrous Materials": "JTFM",
    "Journal of Textiles, Coloration and Polymer Science": "JTCPS",
    "Journal of Thermoplastic Composite Materials": "JTCM",
    "Journal of Vinyl & Additive Technology": "JVAT",
    "Journal of Wuhan Textile S.h.t. Institute": "JWTSI",
    "Journal of Wuhan Textile S.n.T. Institute": "JWTSI",
    "Journal of Wuhan Textile University": "JWTU",
    "Journal of Wuhan University of Technology-materials Science Edition": "JWUTSE",
    "Journal of Zhejiang Textile & Fashion College": "JZTFC",
    "Journal of Zhejiang Textile & Fashion Vocational College": "JZTFVC",
    "Journal of Zhengzhou Textile Institute": "JZTI",
    "Journal of metallurgy and materials science": "JMMS",
    "Journal of textile and apparel technology and management": "JTATM",
    "Journal of the China Textile Institute": "JCTI",
    "Journal of the Institution of Engineers (India). Part TX: Textile Engineering Division": "JIE",
    "Journal of the Japan Research Association for Textile End-Uses": "JJRATE",
    "Journal of the Korean Society of Clothing and Textiles": "JKSCT",
    "Journal of the Society of Materials Science Japan": "JSMSJ",
    "Journal of the Textile Association": "JTA",
    "Journal of the Textile Institute": "JTI",
    "Journal of the Textile Institute Proceedings": "JTIP",
    "Journal of the Textile Institute Transactions": "JTIT",
    "Journal of the Textile Institute. Proceedings": "JTIP",
    "Journal of the Textile Machinary Society": "JTMS",
    "Journal of the Textile Machinery Society of Japan": "JTMSJ",
    "Journal of the Textile Machinery Society of Japan - Transactions": "JTMSJT",
    "Journal of the Textile Machinery of Japan - Proceedings -": "JTMJP",
    "Juniper Online Journal Material Science": "JOJMS",
    "Jusami | Indonesian Journal of Materials Science": "JIJMS",
    "KnE Materials Science": "KMS",
    "Knitting International": "KI",
    "Knitting Trade Journal": "KTJ",
    "Latest Trends in Textile and Fashion Designing": "LTTFD",
    "MRS Bulletin": "MB",
    "Machine intelligence for materials science": "MIMS",
    "Man-Made Textiles in India": "MTI",
    "Maofang ke-ji": "MK",
    "Material Science": "MS",
    "Material Science & Engineering International Journal": "MSEIJ",
    "Material Science Engineering and Applications": "MSEA",
    "Material Science Research India": "MSRI",
    "Material Science and Applied Chemistry": "MSAC",
    "Material Science and Engineering": "MSE",
    "Material Science and Engineering with Advanced Research": "MSEAR",
    "Material Sciences": "MS",
    "Materials Australia": "MA",
    "Materials Science": "MS",
    "Materials Science & Engineering R-reports": "MSER",
    "Materials Science & Technology Conference and Exhibition": "MSTCE",
    "Materials Science Advanced Composite Materials": "MSACM",
    "Materials Science Forum": "MSF",
    "Materials Science Foundations": "MSF",
    "Materials Science Materials Review": "MSMR",
    "Materials Science Monographs": "MSM",
    "Materials Science Reports": "MSR",
    "Materials Science Textile and Clothing Technology": "MSTCT",
    "Materials Science and Condensed Matter Physics": "MSCMP",
    "Materials Science and Engineering": "MSE",
    "Materials Science and Engineering A": "MSE",
    "Materials Science and Engineering B": "MSEB",
    "Materials Science and Engineering B-advanced Functional Solid-state Materials": "MSEBFSM",
    "Materials Science and Engineering C": "MSEC",
    "Materials Science and Engineering R Reports": "MSERR",
    "Materials Science and Engineering of Powder Metallargy": "MSEPM",
    "Materials Science and Engineering of Powder Metallurgy": "MSEPM",
    "Materials Science and Engineering: R Reports": "MSERR",
    "Materials Science and Metallurgy Engineering": "MSME",
    "Materials Science and Nanotechnology": "MSN",
    "Materials Science and Surface Engineering": "MSSE",
    "Materials Science and Technology": "MST",
    "Materials Science and Welding Technologies": "MSWT",
    "Materials Science for Energy Technologies": "MSET",
    "Materials Science in Additive Manufacturing": "MSAM",
    "Materials Science in Semiconductor Processing": "MSSP",
    "Materials Science-Poland": "MS",
    "Materials Science-medziagotyra": "MS",
    "Materials Science. Non-Equilibrium Phase Transformations.": "MSNPT",
    "Materials Sciences and Applications": "MSA",
    "Materials Today Sustainability": "MTS",
    "Materials science : an Indian journal": "MSIJ",
    "Materials science and metallurgy engineering": "MSME",
    "Materials science and technology/Materials science & technology": "MST",
    "Materials science forum": "MSF",
    "Materials science foundations": "MSF",
    "Materials science monographs": "MSM",
    "Materials science research international": "MSRI",
    "Mechanics & materials science/The mechanics and materials science series": "MMS",
    "Metal Working and Material Science": "MWMS",
    "Metallurgical and Materials Transactions A-physical Metallurgy and Materials Science": "MMTAMMS",
    "Mexican journal of materials science and engineering": "MJMSE",
    "Modeling and Numerical Simulation of Material Science": "MNSMS",
    "Modelling and Simulation in Materials Science and Engineering": "MSMSE",
    "Modern Approaches on Material Science": "MAMS",
    "Modern Concepts in Material Science": "MCMS",
    "Nano Materials Science": "NMS",
    "Nanotechnology and Advanced Material Science": "NAMS",
    "Nature Materials": "NM",
    "New Cloth Market (NCM) Magazine": "NCM",
    "Non-Metallic Material Science": "NMS",
    "OAJRC Material Science": "OMS",
    "Online Journal of Materials Science": "OJMS",
    "Open Material Sciences": "OMS",
    "Opérations unitaires et process de fabrication de produits alimentaires": "OUEPDFDPA",
    "Ordnance Material Science and Engineering": "OMSE",
    "Oxford Open Materials Science": "OOMS",
    "PNIPU Bulletin The mechanical engineering materials science": "PBMEMS",
    "PROCEEDING OF INTERNATIONAL CONFERENCE ON RECENT TRENDS IN APPLIED PHYSICS AND MATERIAL SCIENCE: RAM 2013": "PICRTAPMSR",
    "PROCEEDINGS OF THE III INTERNATIONAL CONFERENCE ON ADVANCED TECHNOLOGIES IN MATERIALS SCIENCE, MECHANICAL AND AUTOMATION ENGINEERING: MIP: Engineering-III – 2021": "PIICATMSMAEME",
    "Pakistan Textile Journal (PTJ)": "PTJ",
    "Particle and Fibre Toxicology": "PFT",
    "PeerJ Materials Science": "PMS",
    "Plant Fiber Sciences in China": "PFSC",
    "Plasturgie fabrications de corps creux de films et de fils": "PFDCCDFEDF",
    "Polymeric Materials Science & Cngineering": "PMSC",
    "Polymeric Materials Science and Engineering, Proceedings of the ACS Division of Polymeric Materials Science and Engineering": "PMSEPADPMSE",
    "Polymeric materials science and engineering": "PMSE",
    "Prabha Materials Science Letters": "PMSL",
    "Practice of Fashion and Textile Education Journal": "PFTEJ",
    "Procedia Materials Science": "PMS",
    "Proceeding of the ... annual International Conference on Materials Science, Metal & Manufacturing/Proceedings of the ... annual International Conference  on Materials Science, Metals & Manufacturing": "PAICMSMM",
    "Proceedings of Higher Education Institutions Textile Industry Technology": "PHEITIT",
    "Proceedings of Higher Education Institutions. Textile Industry Technology": "PHEITIT",
    "Proceedings of National Polytechnic University of Armenia METALLURGY MATERIAL SCIENCE MINING ENGINEERING": "PNPUAMMSME",
    "Proceedings of the 19th World Textile Conference - Autex 2019": "PWTCA",
    "Proceedings of the 1st International Conference on Mechanical Engineering and Material Science": "PICMEMS",
    "Proceedings of the Korean Fiber Society Conference": "PKFSC",
    "Procédés textiles": "PT",
    "Progress in Chinese Materials Sciences": "PCMS",
    "Progress in Color, Colorants and Coatings": "PCCC",
    "Progress in Materials Science": "PMS",
    "Progress in Textile Science & Technology": "PTST",
    "QUARKS Brazilian Electronic Journal of Physics Chemistry and Materials Science": "QBEJPCMS",
    "REVIEWS ON ADVANCED MATERIALS SCIENCE": "RAMS",
    "Recent Patents on Materials Science": "RPMS",
    "Reference Module in Materials Science and Materials Engineering#R##N#Encyclopedia of Materials: Science and Technology (Second Edition)": "RMMSMEMST",
    "Reports in development and assembly of textile products": "RDATP",
    "Research & Development in Material Science": "RDMS",
    "Research & Reviews Journal of Material Sciences": "RRJMS",
    "Research Journal of Fashion Technology (RJFT)": "RJFT",
    "Research Journal of Textile and Apparel": "RJTA",
    "Research Journal of Textile and Apparel (RJTA)": "RJTA",
    "Research Journal of Textile and Leather": "RJTL",
    "Research Letters in Materials Science": "RLMS",
    "Research Papers Faculty of Materials Science and Technology Slovak University of Technology": "RPFMSTSUT",
    "Research and Application of Materials Science": "RAMS",
    "Research in Materials Science": "RMS",
    "Russian Science and Technology Conference \"Fundamental and applied problems of materials creation and phases of technologies for textile industry\"": "RSTCAPMCPTTI",
    "SCIREA Journal of Textile Science and Technology": "SJTST",
    "SDRP Journal of Nanotechnology & Material Science": "SJNMS",
    "SOJ Materials Science & Engineering": "SMSE",
    "SRX Materials Science": "SMS",
    "SVOA Materials Science & Technology": "SMST",
    "SVOA materials science & technology.": "SMST",
    "Scandinavian Journal of Materials Science": "SJMS",
    "ScienceOpen Material Science": "SMS",
    "Sen i Gakkaishi": "SIG",
    "Sen i Kikai Gakkaishi (Journal of the Textile Machinery Society of Japan)": "SIKG",
    "Sen'i Kikai Gakkai ronbunshū/Sen'i Kikai Gakkai ronbunshu": "SKGR",
    "Sen'i Kikai Gakkaishi": "SKG",
    "Sen'i Kikai Gakkaishi (journal of The Textile Machinery Society of Japan)": "SKG",
    "Sen'i kogaku/Sen'i kōgaku": "SK",
    "Sen'i to kougyou/Sen'i to kōgyō": "SK",
    "Sen-i Gakkaishi": "SG",
    "Series in materials science and engineering": "SMSE",
    "Series in microscopy in materials science": "SMMS",
    "Shandong Textile Science & Technology": "STST",
    "Shanghai Textile Science & Technology": "STST",
    "Shuttle, Spindle & Dyepot": "SSD",
    "Sichuan Textile Technology": "STT",
    "Society of Materials Science, Japan eBooks": "SMSJE",
    "Soviet Materials Science": "SMS",
    "Springer series in materials science": "SSMS",
    "SunText Review of Material Science": "SRMS",
    "Sustainable Materials and Technologies": "SMT",
    "Synthetic Fiber in China": "SFC",
    "TEXT Journal": "TJ",
    "TEXTIL Plus": "TP",
    "TEXTILE": "",
    "THE JOURNAL OF THE SOCIETY OF TEXTILE INDUSTRY JAPAN": "JSTIJ",
    "TMI (Textile Month International)": "",
    "Tekstil ve Konfeksiyon": "TVK",
    "Texere": "",
    "Textile": "",
    "Textile & Leather Review": "TLR",
    "Textile Accessories": "TA",
    "Textile Asia": "TA",
    "Textile Auxiliaries": "TA",
    "Textile Chemist and Colorist & American Dyestuff Reporter": "TCCADR",
    "Textile Coloration and Finishing": "TCF",
    "Textile Dyeing and Finishing Journal": "TDFJ",
    "Textile History": "TH",
    "Textile Industry Trade Journal": "TITJ",
    "Textile Institute Journal Proceedings and Abstracts": "TIJPA",
    "Textile Institute eBooks": "TIE",
    "Textile Magazine": "TM",
    "Textile Museum journal.": "TMJ",
    "Textile Outlook International": "TOI",
    "Textile Progress": "TP",
    "Textile Research": "TR",
    "Textile Research Journal": "TRJ",
    "Textile Review": "TR",
    "Textile Science & Research Journal": "TSRJ",
    "Textile Science and Engineering": "TSE",
    "Textile Society of America Symposium Proceedings": "TSASP",
    "Textile View Magazine": "TVM",
    "Textile World": "TW",
    "Textile chemist and colorist": "TCC",
    "Textile month": "TM",
    "Textile rental": "TR",
    "Textile science and clothing technology": "TSCT",
    "Textile science and technology": "TST",
    "Textiles": "",
    "Textiles Trends": "TT",
    "Textiles and Clothing Sustainability": "TCS",
    "Textiles and Light Industrial Science and Technology": "TLIST",
    "Textiles and Materials Research Journal": "TMRJ",
    "Textiles industriels": "TI",
    "Textiles techniques": "TT",
    "The 2nd Annual International Workshop on Materials Science and Engineering (IWMSE 2016)": "AIWMSE",
    "The 3rd International Conference on Machinery, Materials Science and Energy Engineering (ICMMSEE 2015)": "ICMMSEE",
    "The Annals of “Dunarea de Jos” University of Galati Fascicle IX Metallurgy and Materials Science": "ADJUGFIMMS",
    "The Fibreculture Journal": "FJ",
    "The Journal of Cloth and Culture (JCC)": "JCC",
    "The Journal of Design, Creative Process & the Fashion Industry": "JDCPFI",
    "The Journal of Textile Design Research and Practice": "JTDRP",
    "The Journal of the Textile Institute": "JTI",
    "The Open Materials Science Journal": "OMSJ",
    "The Open Textile Journal": "OTJ",
    "The Textile Institute eBooks": "TIE",
    "The Textile Journal": "TJ",
    "The Textile Museum Journal": "TMJ",
    "Time to knit": "TK",
    "Treatise on Materials Science and Technology": "TMST",
    "Treatise on materials science and technology": "TMST",
    "Trends in Textile Engineering & Fashion Technology": "TTEFT",
    "Ukrainian Journal of Mechanical Engineering and Materials Science": "UJMEMS",
    "Universal Journal of Materials Science": "UJMS",
    "Usak University Journal of Material Sciences": "UUJMS",
    "Utilization of personal protective equipment and associated factors among Kombolcha Textile Factory workers Kombolcha Ethiopia A cross-sectional study": "UPPEAFAKTFWKECS",
    "Wood Material Science and Engineering": "WMSE",
    "Wood and Fiber Science": "WFS",
    "Wood and Fiber Science (Society of Wood Science and Technology)": "WFS",
    "World Journal of Materials Science": "WJMS",
    "XII Conference on Solid State Crystals: Materials Science and      Applications": "XCSSCMSA",
    "i-manager s Journal on Material Science": "ISJMS",
    "繊維製品消費科学 = Journal of the Japan Reseach Association for textile end-uses": "繊JJRATE",
    "繊維製品消費科学 = Journal of the Japan Research Association for Textile End-uses": "繊JJRATE",
    "韓國染色加工學會誌 = Textile coloration and finishing": "韓TCF",
    "韓國纖維工學會誌 = Textile science and engineering": "韓TSE",
    "한국의류산업학회지 = Journal of the Korean Society for Clothing Industry": "한JKSCI",
}

# Total: 528 unique journals

# Build reverse lookup: acronym -> list of full names
ACRONYM_TO_JOURNALS = {}
for name, acronym in TEXTILE_JOURNALS.items():
    if acronym:
        if acronym not in ACRONYM_TO_JOURNALS:
            ACRONYM_TO_JOURNALS[acronym] = []
        ACRONYM_TO_JOURNALS[acronym].append(name)

BATCH_SIZE = 500  # For Crossref/OpenAlex
PUBMED_BATCH_SIZE = 200

# =============================================================================
# PRESET MODE - Set to True to run non-interactively with preset parameters
# =============================================================================
PRESET_MODE = False  # Set to False for interactive mode

PRESET = {
    "initial_option": "1",  # 1=New Search, 2=Continue downloading, 3=Search full-texts
    "search_mode": "4",     # 1-4, see menu options
    "max_results": 0,       # 0 = unlimited

    # CrossRef keywords (comma-separated)
    "crossref_keywords": "textile, textiles, fabric, fabrics, garment, garments, wicking, moisture, drying time, drying times, drying rate, drying rates, drying curve, drying curves, drying test, drying tests",

    # PubMed/OpenAlex abstract search (comma-separated, multi-word = exact phrase)
    "abstract_search": "moisture wicking, thermophysiological, moisture management, wicking, textile, textiles, fabric, fabrics, garment, garments",

    # OpenAlex full-text search (comma-separated)
    "fulltext_search": "AATCC TM199, AATCC 199, AATCC 200, AATCC 201, ISO 13029, ISO 17617, AATCC TM200, AATCC TM201, GB/T 38473, JIS L 1096, ASTM D2654, moisture management, moisture wicking, drying rates, drying rate, drying times, drying time, drying curve, drying curves, drying test, drying tests",

    # Step 4: 1=Filter OA now, 2=Continue to subquery
    "step4_option": "1",  # Skip subquery, go straight to OA filter

    # Subquery (only used if step4_option="2")
    "subquery": "",

    # Download settings
    "oa_download_count": 0,      # 0 = skip downloads
    "non_oa_download_count": 0,  # 0 = skip downloads
}

def get_input(prompt: str, preset_key: str = None, default: str = "") -> str:
    """Get input from user or preset config."""
    if PRESET_MODE and preset_key and preset_key in PRESET:
        value = str(PRESET[preset_key])
        print(f"{prompt} [PRESET: {value[:50]}{'...' if len(value) > 50 else ''}]")
        return value
    return input(prompt).strip() or default

HEADERS = {
    "User-Agent": "MultiAPISearchTool/1.0 (mailto:textile.research.query@gmail.com)"
}

BASE_DIR = Path(r"C:\Users\azt12\OneDrive\Documents\Wrestling Robe\Materials Science - Wickability\Studies for Analysis by LLM AI\ScrapedResearch")

# =============================================================================
# JOURNAL FILTERING FUNCTIONS
# =============================================================================

def resolve_journal_input(user_input: str) -> list[str]:
    """Resolve user input to list of full journal names.
    
    Accepts:
    - Full journal name (exact match)
    - Acronym (e.g., 'TRJ' for Textile Research Journal)
    - Partial name (fuzzy match)
    
    Returns list of matching full journal names.
    """
    user_input = user_input.strip()
    matches = []
    
    # Check if it's an exact full name
    if user_input in TEXTILE_JOURNALS:
        return [user_input]
    
    # Check if it's an acronym
    upper_input = user_input.upper()
    if upper_input in ACRONYM_TO_JOURNALS:
        return ACRONYM_TO_JOURNALS[upper_input]
    
    # Fuzzy match on journal names (case-insensitive contains)
    lower_input = user_input.lower()
    for name in TEXTILE_JOURNALS:
        if lower_input in name.lower():
            matches.append(name)
    
    return matches

def display_journal_matches(matches: list[str]) -> list[str]:
    """Display matching journals and let user select which ones to use."""
    if not matches:
        return []
    
    if len(matches) == 1:
        print(f"{GREEN}  Found: {matches[0]}{RESET}")
        return matches
    
    print(f"{GREEN}  Found {len(matches)} matching journals:{RESET}")
    for i, name in enumerate(matches[:20], 1):  # Show max 20
        acronym = TEXTILE_JOURNALS.get(name, "")
        print(f"    {i}: {name}" + (f" [{acronym}]" if acronym else ""))
    
    if len(matches) > 20:
        print(f"    ... and {len(matches) - 20} more")
    
    print("\nEnter numbers to select (comma-separated), 'all', or press Enter for all:")


    selection = input("  > ").strip()
    
    if not selection or selection.lower() == 'all':
        return matches
    
    try:
        indices = [int(x.strip()) - 1 for x in selection.split(",")]
        return [matches[i] for i in indices if 0 <= i < len(matches)]
    except:
        return matches

def build_crossref_journal_filter(journal_names: list[str]) -> str:
    """Build CrossRef filter string for journal names.
    
    CrossRef uses: filter=container-title:Name1,container-title:Name2
    """
    if not journal_names:
        return ""
    return ",".join(f"container-title:{name}" for name in journal_names)

def build_openalex_journal_filter(journal_names: list[str]) -> str:
    """Build OpenAlex filter string for journal names.
    
    OpenAlex requires source IDs, but also accepts display_name search.
    Uses: filter=primary_location.source.display_name:Name
    For multiple: pipe-separated
    """
    if not journal_names:
        return ""
    # OpenAlex can filter by source display name with search
    return "|".join(journal_names)

def filter_papers_by_journal(papers: list[dict], journal_names: list[str]) -> list[dict]:
    """Client-side filter papers by journal name (for PubMed).
    
    Matches against paper's 'journal' or 'container-title' field.
    Uses fuzzy matching (case-insensitive contains).
    """
    if not journal_names:
        return papers
    
    lower_journals = [j.lower() for j in journal_names]
    filtered = []
    
    for paper in papers:
        paper_journal = paper.get("journal", "") or paper.get("container-title", "") or ""
        paper_journal_lower = paper_journal.lower()
        
        for journal in lower_journals:
            if journal in paper_journal_lower or paper_journal_lower in journal:
                filtered.append(paper)
                break
    
    return filtered

def interactive_journal_selection() -> list[str]:
    """Interactive prompt for journal selection."""
    print(f"\n{GREEN}JOURNAL FILTER{RESET}")

    print("  Enter journal name(s) or acronym(s), comma-separated")
    print("  Examples: TRJ, Textile Research Journal, JEFF")
    print(f"{DIM_RED}  (Press Enter to skip journal filtering){RESET}")
    
    user_input = input("  > ").strip()
    
    if not user_input:
        return []
    
    # Process each term
    terms = [t.strip() for t in user_input.split(",") if t.strip()]
    all_selected = []
    
    for term in terms:
        matches = resolve_journal_input(term)
        if matches:
            selected = display_journal_matches(matches)
            all_selected.extend(selected)
        else:
            print(f"{DIM_RED}  No match found for: {term}{RESET}")
    
    # Remove duplicates while preserving order
    seen = set()
    unique = []
    for j in all_selected:
        if j not in seen:
            seen.add(j)
            unique.append(j)
    
    if unique:
        print(f"\n{GREEN}  Selected {len(unique)} journal(s) for filtering{RESET}")

    
    return unique

# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def clean_text(text: str) -> str:
    """Clean HTML/XML tags and normalize whitespace."""
    if not text:
        return ""
    text = re.sub(r'<[^>]+>', ' ', str(text))
    text = html.unescape(text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def normalize_doi(doi: str) -> str:
    """Normalize DOI for comparison."""
    if not doi:
        return ""
    doi = doi.lower().strip()
    for prefix in ["https://doi.org/", "http://doi.org/", "doi:"]:
        if doi.startswith(prefix):
            doi = doi[len(prefix):]
    return doi

def normalize_title(title: str) -> str:
    """Normalize title for fuzzy matching."""
    if not title:
        return ""
    title = clean_text(title).lower()
    title = re.sub(r'[^\w\s]', '', title)
    title = re.sub(r'\s+', ' ', title).strip()
    return title

def get_timestamp_str(dt: datetime = None) -> str:
    """Get timestamp string in MM-DD-YY-HHmm format (e.g., 01-01-26-0439)."""
    if dt is None:
        dt = datetime.now()
    return dt.strftime("%m-%d-%y-%H%M")

def setup_query_folders(timestamp_str: str) -> dict:
    """Create query folder structure and return paths dict.

    Structure:
        ScrapedResearch/
        └── MM-DD-YY-HHmm Query/
            ├── Query Parameters/
            │   └── MM-DD-YY-HHmm Query Parameters.txt
            ├── Downloaded Papers/
            │   ├── OA Papers/
            │   └── Non-OA Papers/
            ├── MM-DD-YY-HHmm Paper API Query.xlsx
            ├── MM-DD-YY-HHmm BACKUP.json
            └── MM-DD-YY-HHmm Full-Text Search.xlsx
    """
    query_folder = BASE_DIR / f"{timestamp_str} Query"
    params_folder = query_folder / "Query Parameters"
    download_folder = query_folder / "Downloaded Papers"
    oa_folder = download_folder / "OA Papers"
    non_oa_folder = download_folder / "Non-OA Papers"

    # Create all directories
    for folder in [query_folder, params_folder, oa_folder, non_oa_folder]:
        folder.mkdir(parents=True, exist_ok=True)

    return {
        "query_folder": query_folder,
        "params_folder": params_folder,
        "download_folder": download_folder,
        "oa_folder": oa_folder,
        "non_oa_folder": non_oa_folder,
        "params_file": params_folder / f"{timestamp_str} Query Parameters.txt",
        "excel_file": query_folder / f"{timestamp_str} Paper API Query.xlsx",
        "backup_file": query_folder / f"{timestamp_str} BACKUP.json",
        "fulltext_file": query_folder / f"{timestamp_str} Full-Text Search.xlsx",
    }

# Global to hold current query paths (set when query starts)
QUERY_PATHS = {}

# =============================================================================
# CROSSREF API
# =============================================================================

def crossref_search(query: str, max_results: int = None, journal_filter: str = "") -> list[dict]:
    """Search Crossref API with cursor-based pagination."""
    print(f"\n{'='*60}")
    print("CROSSREF API")
    print(f"{'='*60}")

    API_URL = "https://api.crossref.org/works"
    all_papers = []
    cursor = "*"
    batch_num = 0

    while True:
        batch_num += 1
        params = {"query": query, "rows": BATCH_SIZE, "cursor": cursor}
        if journal_filter:
            params["filter"] = journal_filter

        try:
            response = requests.get(API_URL, params=params, headers=HEADERS, timeout=120)
            response.raise_for_status()
            data = response.json()

            if data.get("status") != "ok":
                print(f"  API Error: {data}")
                break

            items = data["message"].get("items", [])
            next_cursor = data["message"].get("next-cursor")
            total = data["message"].get("total-results", 0)

            if batch_num == 1:
                print(f"  Total available: {total:,} papers")

            if not items:
                print(f"  No more results")
                break

            for item in items:
                authors = item.get("author", [])
                if authors:
                    first = authors[0]
                    author = f"{first.get('family', '')}, {first.get('given', '')}"
                else:
                    author = ""

                titles = item.get("title", [])
                title = clean_text(titles[0]) if titles else ""

                year = None
                for df in ["issued", "published", "published-print"]:
                    dp = item.get(df, {}).get("date-parts", [[]])
                    if dp and dp[0] and dp[0][0]:
                        year = dp[0][0]
                        break

                doi = item.get("DOI", "")
                abstract = clean_text(item.get("abstract", ""))

                all_papers.append({
                    "source": "Crossref",
                    "author": author,
                    "title": title,
                    "year": year or "",
                    "doi": doi,
                    "doi_url": f"https://doi.org/{doi}" if doi else "",
                    "abstract": abstract,
                    "is_oa": None,
                    "pdf_url": None,
                })

            print(f"  Batch {batch_num}: +{len(items)} | Total: {len(all_papers):,}")

            if max_results and len(all_papers) >= max_results:
                print(f"  Reached max results ({max_results})")
                break

            if not next_cursor:
                break

            cursor = next_cursor
            time.sleep(0.3)

        except Exception as e:
            print(f"  Error in batch {batch_num}: {e}")
            break

    print(f"  TOTAL from Crossref: {len(all_papers):,}")
    return all_papers

# =============================================================================
# OPENALEX API
# =============================================================================

def openalex_reconstruct_abstract(inverted_index: dict) -> str:
    """Reconstruct abstract from OpenAlex inverted index format."""
    if not inverted_index:
        return ""
    words = []
    for word, positions in inverted_index.items():
        for pos in positions:
            words.append((pos, word))
    words.sort(key=lambda x: x[0])
    return " ".join(w[1] for w in words)

def openalex_search(query: str, max_results: int = None, search_type: str = "default", journal_filter: str = "") -> list[dict]:
    """Search OpenAlex API with cursor-based pagination."""
    search_label = f"OPENALEX API ({search_type})"
    print(f"\n{'='*60}")
    print(search_label)
    print(f"{'='*60}")

    API_URL = "https://api.openalex.org/works"
    all_papers = []
    cursor = "*"
    batch_num = 0

    while True:
        batch_num += 1
        params = {"per_page": min(BATCH_SIZE, 200), "cursor": cursor}

        if search_type == "abstract":
            params["filter"] = f"title_and_abstract.search:{query}"
        elif search_type == "fulltext":
            params["filter"] = f"fulltext.search:{query}"
        else:
            params["search"] = query
        
        # Add journal filter if specified
        if journal_filter:
            existing_filter = params.get("filter", "")
            journal_filter_str = f"primary_location.source.display_name.search:{journal_filter}"
            if existing_filter:
                params["filter"] = f"{existing_filter},{journal_filter_str}"
            else:
                params["filter"] = journal_filter_str

        try:
            response = requests.get(API_URL, params=params, headers=HEADERS, timeout=120)
            response.raise_for_status()
            data = response.json()

            items = data.get("results", [])
            meta = data.get("meta", {})
            next_cursor = meta.get("next_cursor")
            total = meta.get("count", 0)

            if batch_num == 1:
                print(f"  Total available: {total:,} papers")

            if not items:
                print(f"  No more results")
                break

            for item in items:
                authorships = item.get("authorships", [])
                if authorships:
                    author_info = authorships[0].get("author", {})
                    author = author_info.get("display_name", "")
                else:
                    author = ""

                title = clean_text(item.get("title", ""))
                year = item.get("publication_year", "")
                doi = item.get("doi", "")
                if doi and doi.startswith("https://doi.org/"):
                    doi = doi[16:]

                abstract_inv = item.get("abstract_inverted_index", {})
                abstract = openalex_reconstruct_abstract(abstract_inv)

                oa_info = item.get("open_access", {})
                is_oa = oa_info.get("is_oa", False)
                pdf_url = oa_info.get("oa_url", "")

                all_papers.append({
                    "source": "OpenAlex",
                    "author": author,
                    "title": title,
                    "year": year,
                    "doi": doi,
                    "doi_url": f"https://doi.org/{doi}" if doi else "",
                    "abstract": abstract,
                    "is_oa": is_oa,
                    "pdf_url": pdf_url,
                })

            print(f"  Batch {batch_num}: +{len(items)} | Total: {len(all_papers):,}")

            if max_results and len(all_papers) >= max_results:
                print(f"  Reached max results ({max_results})")
                break

            if not next_cursor:
                break

            cursor = next_cursor
            time.sleep(0.1)

        except Exception as e:
            print(f"  Error in batch {batch_num}: {e}")
            break

    print(f"  TOTAL from OpenAlex: {len(all_papers):,}")
    return all_papers

# =============================================================================
# PUBMED API
# =============================================================================

def pubmed_search(query: str, max_results: int = None) -> list[dict]:
    """Search PubMed using E-utilities."""
    print(f"\n{'='*60}")
    print("PUBMED E-UTILITIES API")
    print(f"{'='*60}")

    ESEARCH_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
    EFETCH_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"

    print("  Searching for PMIDs...")
    search_params = {
        "db": "pubmed",
        "term": query,
        "retmax": max_results or 100000,
        "retmode": "json",
        "usehistory": "y",
    }

    try:
        response = requests.get(ESEARCH_URL, params=search_params, headers=HEADERS, timeout=120)
        response.raise_for_status()
        data = response.json()

        result = data.get("esearchresult", {})
        pmids = result.get("idlist", [])
        total = int(result.get("count", 0))

        print(f"  Total available: {total:,} papers")
        print(f"  Retrieved {len(pmids):,} PMIDs")

        if not pmids:
            return []

    except Exception as e:
        print(f"  Error searching PubMed: {e}")
        return []

    all_papers = []
    batch_num = 0

    for start in range(0, len(pmids), PUBMED_BATCH_SIZE):
        batch_num += 1
        batch_pmids = pmids[start:start + PUBMED_BATCH_SIZE]

        fetch_params = {
            "db": "pubmed",
            "id": ",".join(batch_pmids),
            "retmode": "xml",
            "rettype": "abstract",
        }

        try:
            response = requests.get(EFETCH_URL, params=fetch_params, headers=HEADERS, timeout=120)
            response.raise_for_status()
            root = ET.fromstring(response.content)

            for article in root.findall(".//PubmedArticle"):
                medline = article.find(".//MedlineCitation")
                if medline is None:
                    continue

                pmid_elem = medline.find(".//PMID")
                pmid = pmid_elem.text if pmid_elem is not None else ""

                title_elem = medline.find(".//ArticleTitle")
                title = clean_text(title_elem.text) if title_elem is not None else ""

                author_list = medline.findall(".//Author")
                if author_list:
                    first_author = author_list[0]
                    lastname = first_author.find("LastName")
                    firstname = first_author.find("ForeName")
                    author = f"{lastname.text if lastname is not None else ''}, {firstname.text if firstname is not None else ''}"
                else:
                    author = ""

                pub_date = medline.find(".//PubDate")
                year = ""
                if pub_date is not None:
                    year_elem = pub_date.find("Year")
                    if year_elem is not None:
                        year = year_elem.text

                abstract_elem = medline.find(".//Abstract/AbstractText")
                abstract = clean_text(abstract_elem.text) if abstract_elem is not None else ""

                doi = ""
                for id_elem in article.findall(".//ArticleId"):
                    if id_elem.get("IdType") == "doi":
                        doi = id_elem.text
                        break

                # Get journal name
                journal_elem = medline.find(".//MedlineJournalInfo/MedlineTA")
                journal = journal_elem.text if journal_elem is not None else ""
                
                all_papers.append({
                    "source": "PubMed",
                    "author": author,
                    "title": title,
                    "year": year,
                    "doi": doi,
                    "doi_url": f"https://doi.org/{doi}" if doi else "",
                    "pmid": pmid,
                    "abstract": abstract,
                    "journal": journal,
                    "is_oa": None,
                    "pdf_url": None,
                })

            print(f"  Batch {batch_num}: +{len(batch_pmids)} | Total: {len(all_papers):,}")
            time.sleep(0.35)

        except Exception as e:
            print(f"  Error fetching batch {batch_num}: {e}")
            continue

    print(f"  TOTAL from PubMed: {len(all_papers):,}")
    return all_papers

# =============================================================================
# OPEN ACCESS CHECK (Unpaywall)
# =============================================================================

def check_unpaywall_oa(doi: str) -> dict:
    """Check if a paper is Open Access using Unpaywall API."""
    if not doi:
        return {"is_oa": None, "pdf_url": None}

    doi_clean = normalize_doi(doi)
    url = f"https://api.unpaywall.org/v2/{doi_clean}"
    params = {"email": "textile.research.query@gmail.com"}

    try:
        response = requests.get(url, params=params, timeout=10)
        if response.status_code == 200:
            data = response.json()
            is_oa = data.get("is_oa", False)
            best_oa = data.get("best_oa_location", {}) or {}
            pdf_url = best_oa.get("url_for_pdf", "") or best_oa.get("url", "")
            return {"is_oa": is_oa, "pdf_url": pdf_url}
        else:
            return {"is_oa": None, "pdf_url": None}
    except Exception:
        return {"is_oa": None, "pdf_url": None}

def check_oa_status(papers: list[dict]) -> list[dict]:
    """Check OA status for papers using Unpaywall where needed."""
    needs_check = [p for p in papers if p.get("is_oa") is None and p.get("doi")]
    already_known = len(papers) - len(needs_check)

    print(f"  Papers with OA status (from OpenAlex): {already_known:,}")
    print(f"  Papers needing Unpaywall check: {len(needs_check):,}")

    if not needs_check:
        return papers

    print(f"  Checking via Unpaywall API...")

    checked = 0
    with ThreadPoolExecutor(max_workers=10) as executor:
        future_to_paper = {
            executor.submit(check_unpaywall_oa, p["doi"]): p
            for p in needs_check
        }

        for future in as_completed(future_to_paper):
            paper = future_to_paper[future]
            try:
                result = future.result()
                paper["is_oa"] = result["is_oa"]
                if result["pdf_url"] and not paper.get("pdf_url"):
                    paper["pdf_url"] = result["pdf_url"]
            except Exception:
                pass

            checked += 1
            if checked % 100 == 0:
                print(f"    Checked {checked:,}/{len(needs_check):,}")

    print(f"  Unpaywall check complete")
    return papers

# =============================================================================
# DEDUPLICATION
# =============================================================================

def deduplicate_papers(papers: list[dict]) -> list[dict]:
    """Deduplicate papers by DOI, then by title."""
    print(f"\n{'='*60}")
    print("DEDUPLICATION")
    print(f"{'='*60}")

    seen_dois = {}
    seen_titles = {}
    unique = []
    dup_doi = 0
    dup_title = 0

    for paper in papers:
        doi = normalize_doi(paper.get("doi", ""))
        title = normalize_title(paper.get("title", ""))

        if doi and doi in seen_dois:
            dup_doi += 1
            existing = seen_dois[doi]
            if not existing.get("abstract") and paper.get("abstract"):
                existing["abstract"] = paper["abstract"]
            if not existing.get("is_oa") and paper.get("is_oa"):
                existing["is_oa"] = paper["is_oa"]
            if not existing.get("pdf_url") and paper.get("pdf_url"):
                existing["pdf_url"] = paper["pdf_url"]
            if paper["source"] not in existing.get("sources", existing["source"]):
                existing["sources"] = existing.get("sources", existing["source"]) + f", {paper['source']}"
            continue

        if title and len(title) > 20 and title in seen_titles:
            dup_title += 1
            existing = seen_titles[title]
            if not existing.get("abstract") and paper.get("abstract"):
                existing["abstract"] = paper["abstract"]
            if not existing.get("doi") and paper.get("doi"):
                existing["doi"] = paper["doi"]
                existing["doi_url"] = paper["doi_url"]
            if not existing.get("is_oa") and paper.get("is_oa"):
                existing["is_oa"] = paper["is_oa"]
            if not existing.get("pdf_url") and paper.get("pdf_url"):
                existing["pdf_url"] = paper["pdf_url"]
            if paper["source"] not in existing.get("sources", existing["source"]):
                existing["sources"] = existing.get("sources", existing["source"]) + f", {paper['source']}"
            continue

        paper["sources"] = paper["source"]
        unique.append(paper)

        if doi:
            seen_dois[doi] = paper
        if title and len(title) > 20:
            seen_titles[title] = paper

    print(f"  Input papers: {len(papers):,}")
    print(f"  Duplicates by DOI: {dup_doi:,}")
    print(f"  Duplicates by title: {dup_title:,}")
    print(f"  UNIQUE papers: {len(unique):,}")

    return unique

# =============================================================================
# SUBQUERY (Local Abstract Filtering)
# =============================================================================

def parse_subquery(subquery: str) -> list[list[str]]:
    """Parse subquery with AND/OR operators."""
    and_groups = [g.strip() for g in re.split(r'\s+AND\s+', subquery) if g.strip()]
    parsed_groups = []
    for group in and_groups:
        or_terms = [t.strip() for t in re.split(r'\s+OR\s+', group) if t.strip()]
        parsed_groups.append(or_terms)
    return parsed_groups

def matches_subquery(abstract: str, parsed_groups: list[list[str]]) -> bool:
    """Check if abstract matches the subquery."""
    if not abstract:
        return False

    abstract_lower = abstract.lower()

    for or_terms in parsed_groups:
        group_matched = False
        for term in or_terms:
            term_lower = term.lower()
            if ' ' in term:
                if term_lower in abstract_lower:
                    group_matched = True
                    break
            else:
                if re.search(r'\b' + re.escape(term_lower) + r'\b', abstract_lower):
                    group_matched = True
                    break

        if not group_matched:
            return False

    return True

def filter_by_subquery(papers: list[dict], subquery: str) -> list[dict]:
    """Filter papers by subquery on abstracts."""
    parsed = parse_subquery(subquery)

    print(f"\n  Subquery parsed as {len(parsed)} AND-group(s):")
    for i, grp in enumerate(parsed, 1):
        print(f"    Group {i}: {grp}")

    filtered = []
    for paper in papers:
        if matches_subquery(paper.get("abstract", ""), parsed):
            filtered.append(paper)

    return filtered

# =============================================================================
# QUERY BUILDING
# =============================================================================

def build_keywords_query(keywords: str, api: str) -> str:
    """Build API query from comma-separated keywords."""
    terms = [t.strip() for t in keywords.split(",") if t.strip()]
    if len(terms) == 1:
        return terms[0]
    if api == "pubmed":
        return " OR ".join(f"({t})" for t in terms)
    else:
        return " OR ".join(terms)

def parse_exact_and_keywords(input_str: str) -> list[dict]:
    """Parse comma-separated input (multi-word = exact phrase)."""
    terms = [t.strip() for t in input_str.split(",") if t.strip()]
    parsed = []
    for term in terms:
        if ' ' in term:
            parsed.append({"term": term, "type": "phrase"})
        else:
            parsed.append({"term": term, "type": "keyword"})
    return parsed

def build_exact_keywords_query(input_str: str, api: str) -> str:
    """Build query with exact phrases and keywords."""
    parsed = parse_exact_and_keywords(input_str)
    if not parsed:
        return ""

    terms_formatted = []
    for item in parsed:
        term = item["term"]
        if item["type"] == "phrase":
            terms_formatted.append(f'"{term}"')
        else:
            terms_formatted.append(term)

    if api == "pubmed":
        field = "[Title/Abstract]"
        return " OR ".join(f"{t}{field}" for t in terms_formatted)
    else:
        return " OR ".join(terms_formatted)

# =============================================================================
# CSV/EXCEL HANDLING WITH ROW COLORS
# =============================================================================

def save_papers_csv(papers: list[dict], filepath: Path, with_status: bool = True):
    """Save papers to CSV with optional status column."""
    fieldnames = ["study_number", "status", "sources", "author", "title", "year", "doi", "doi_url", "is_oa", "pdf_url", "abstract"]
    if not with_status:
        fieldnames.remove("status")

    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(papers)

    print(f"  Saved {len(papers):,} papers to: {filepath}")

def sanitize_for_excel(text):
    """Remove all characters illegal in XML 1.0 (which openpyxl/Excel uses).

    XML 1.0 legal characters are:
    - #x9 (tab), #xA (newline), #xD (carriage return)
    - #x20-#xD7FF (includes all standard printable chars and unicode)
    - #xE000-#xFFFD
    - #x10000-#x10FFFF

    Illegal characters that must be removed:
    - 0x00-0x08 (C0 control chars)
    - 0x0B-0x0C (vertical tab, form feed)
    - 0x0E-0x1F (C0 control chars)
    - 0x7F (DEL)
    - 0x80-0x9F (C1 control chars - common in Windows text!)
    - 0xD800-0xDFFF (surrogate pairs)
    - 0xFFFE-0xFFFF (non-characters)
    """
    if not isinstance(text, str):
        return text

    # Build pattern for all illegal XML 1.0 characters
    # C0 controls (except tab, LF, CR), DEL, C1 controls, surrogates, non-chars
    illegal_pattern = (
        r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F'  # C0/C1 control chars + DEL
        r'\uD800-\uDFFF'  # Surrogate pairs
        r'\uFFFE\uFFFF]'  # Non-characters
    )
    return re.sub(illegal_pattern, '', text)

def save_oa_split_excel(oa_papers: list[dict], no_oa_papers: list[dict], filepath: Path):
    """Save papers to Excel with OA and No OA worksheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        headers = ["Study #", "Status", "Sources", "Author", "Title", "Year", "DOI", "DOI URL", "PDF URL", "Abstract"]

        def write_sheet(ws, papers, sheet_name):
            ws.title = sheet_name
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header).font = Font(bold=True)

            for row, paper in enumerate(papers, 2):
                ws.cell(row=row, column=1, value=paper.get("study_number", ""))
                ws.cell(row=row, column=2, value=paper.get("status", "pending"))
                ws.cell(row=row, column=3, value=sanitize_for_excel(paper.get("sources", "")))
                ws.cell(row=row, column=4, value=sanitize_for_excel(paper.get("author", "")))
                ws.cell(row=row, column=5, value=sanitize_for_excel(paper.get("title", "")))
                ws.cell(row=row, column=6, value=paper.get("year", ""))
                ws.cell(row=row, column=7, value=paper.get("doi", ""))
                ws.cell(row=row, column=8, value=paper.get("doi_url", ""))
                ws.cell(row=row, column=9, value=paper.get("pdf_url", "") or "")
                ws.cell(row=row, column=10, value=sanitize_for_excel((paper.get("abstract", "") or "")[:32000]))

                # Apply color based on status
                status = paper.get("status", "pending")
                if status == "downloaded":
                    fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Green
                    for c in range(1, 11):
                        ws.cell(row=row, column=c).fill = fill
                elif status == "searched":
                    fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Red
                    for c in range(1, 11):
                        ws.cell(row=row, column=c).fill = fill

            # Column widths
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 50
            ws.column_dimensions['F'].width = 8
            ws.column_dimensions['G'].width = 25
            ws.column_dimensions['H'].width = 35
            ws.column_dimensions['I'].width = 40
            ws.column_dimensions['J'].width = 80

        ws1 = wb.active
        write_sheet(ws1, oa_papers, "OA")

        ws2 = wb.create_sheet()
        write_sheet(ws2, no_oa_papers, "No OA")

        wb.save(filepath)
        print(f"  Saved to: {filepath}")
        print(f"    OA worksheet: {len(oa_papers):,} papers")
        print(f"    No OA worksheet: {len(no_oa_papers):,} papers")

    except ImportError:
        print("  openpyxl not installed. Saving as CSV.")
        save_papers_csv(oa_papers, filepath.with_suffix('.csv').with_stem(filepath.stem + '_OA'))
        save_papers_csv(no_oa_papers, filepath.with_suffix('.csv').with_stem(filepath.stem + '_NoOA'))

def load_papers_from_excel(filepath: Path) -> tuple[list[dict], list[dict]]:
    """Load papers from Excel file with OA/No OA sheets."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(filepath)

        def read_sheet(ws):
            papers = []
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                paper = {}
                for i, header in enumerate(headers):
                    if header and i < len(row):
                        key = header.lower().replace(" ", "_").replace("#", "number")
                        paper[key] = row[i]
                papers.append(paper)
            return papers

        oa_papers = read_sheet(wb["OA"]) if "OA" in wb.sheetnames else []
        no_oa_papers = read_sheet(wb["No OA"]) if "No OA" in wb.sheetnames else []

        return oa_papers, no_oa_papers

    except Exception as e:
        print(f"  Error loading Excel: {e}")
        return [], []

def update_paper_status_in_excel(filepath: Path, study_numbers: list[int], new_status: str):
    """Update status and color of specific papers in Excel."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill

        wb = load_workbook(filepath)

        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

        fill = green_fill if new_status == "downloaded" else red_fill if new_status == "searched" else None

        for sheet_name in ["OA", "No OA"]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]

            for row in range(2, ws.max_row + 1):
                study_num = ws.cell(row=row, column=1).value
                if study_num in study_numbers:
                    ws.cell(row=row, column=2).value = new_status
                    if fill:
                        for col in range(1, 11):
                            ws.cell(row=row, column=col).fill = fill

        wb.save(filepath)

    except Exception as e:
        print(f"  Error updating Excel: {e}")

# =============================================================================
# PDF DOWNLOAD
# =============================================================================

def download_pdf(url: str, filepath: Path) -> bool:
    """Download a PDF from URL."""
    try:
        response = requests.get(url, headers=HEADERS, timeout=60, allow_redirects=True)
        if response.status_code == 200 and 'pdf' in response.headers.get('Content-Type', '').lower():
            with open(filepath, 'wb') as f:
                f.write(response.content)
            return True
        return False
    except Exception:
        return False

def download_oa_papers(papers: list[dict], download_dir: Path, csv_filepath: Path, count: int) -> int:
    """Download OA papers and update CSV status."""
    pending = [p for p in papers if p.get("status", "pending") == "pending" and p.get("pdf_url")]

    print(f"\n  OA papers with PDF URLs pending: {len(pending):,}")

    if not pending:
        print("  No papers to download.")
        return 0

    to_download = pending[:count]
    print(f"  Downloading {len(to_download)} papers...")

    download_dir.mkdir(parents=True, exist_ok=True)
    downloaded = 0

    for i, paper in enumerate(to_download, 1):
        study_num = paper.get("study_number", i)
        title_slug = re.sub(r'[^\w\s-]', '', paper.get("title", "")[:50]).strip().replace(' ', '_')
        filename = f"{study_num:04d}_{title_slug}.pdf"
        filepath = download_dir / filename

        print(f"  [{i}/{len(to_download)}] Downloading study #{study_num}...", end=" ")

        if download_pdf(paper.get("pdf_url", ""), filepath):
            paper["status"] = "downloaded"
            update_paper_status_in_excel(csv_filepath, [study_num], "downloaded")
            downloaded += 1
            print("OK")
        else:
            print("FAILED")

    print(f"\n  Downloaded: {downloaded}/{len(to_download)}")
    return downloaded

def download_scihub_paper(doi: str, filepath: Path) -> bool:
    """Download a paper from SciHub using DOI."""
    try:
        from scihub import SciHub
        sh = SciHub()
        result = sh.fetch(doi)
        if result and 'pdf' in result and not result.get('err'):
            with open(filepath, 'wb') as f:
                f.write(result['pdf'])
            return True
        return False
    except Exception as e:
        return False

def download_non_oa_papers(papers: list[dict], download_dir: Path, csv_filepath: Path, count: int, progress_lock: threading.Lock = None) -> int:
    """Download non-OA papers via SciHub and update CSV status."""
    # Filter papers with DOI but no OA URL
    pending = [p for p in papers if p.get("status", "pending") == "pending" and p.get("doi") and not p.get("pdf_url")]

    print(f"\n  [SciHub] Non-OA papers with DOI pending: {len(pending):,}")

    if not pending:
        print("  [SciHub] No papers to download.")
        return 0

    to_download = pending[:count]
    print(f"  [SciHub] Attempting {len(to_download)} papers via Sci-Hub...")

    download_dir.mkdir(parents=True, exist_ok=True)
    downloaded = 0

    for i, paper in enumerate(to_download, 1):
        study_num = paper.get("study_number", i)
        doi = paper.get("doi", "")
        title_slug = re.sub(r'[^\w\s-]', '', paper.get("title", "")[:50]).strip().replace(' ', '_')
        filename = f"{study_num:04d}_{title_slug}.pdf"
        filepath = download_dir / filename

        print(f"  [SciHub] [{i}/{len(to_download)}] Study #{study_num} (DOI: {doi[:30]}...)...", end=" ")

        if download_scihub_paper(doi, filepath):
            paper["status"] = "downloaded"
            if progress_lock:
                with progress_lock:
                    update_paper_status_in_excel(csv_filepath, [study_num], "downloaded")
            else:
                update_paper_status_in_excel(csv_filepath, [study_num], "downloaded")
            downloaded += 1
            print("OK")
        else:
            print("FAILED")

        # Rate limit to avoid captchas
        time.sleep(1)

    print(f"\n  [SciHub] Downloaded: {downloaded}/{len(to_download)}")
    return downloaded

def parallel_download_all(oa_papers: list[dict], non_oa_papers: list[dict],
                          oa_folder: Path, non_oa_folder: Path, csv_filepath: Path,
                          oa_count: int, non_oa_count: int) -> tuple[int, int]:
    """Download OA and non-OA papers in parallel threads to separate folders."""
    progress_lock = threading.Lock()
    oa_result = [0]
    non_oa_result = [0]

    def download_oa_thread():
        oa_result[0] = download_oa_papers(oa_papers, oa_folder, csv_filepath, oa_count)

    def download_non_oa_thread():
        non_oa_result[0] = download_non_oa_papers(non_oa_papers, non_oa_folder, csv_filepath, non_oa_count, progress_lock)

    # Start both threads
    oa_thread = threading.Thread(target=download_oa_thread, name="OA-Downloader")
    non_oa_thread = threading.Thread(target=download_non_oa_thread, name="SciHub-Downloader")

    print("\n  Starting parallel download (OA + SciHub)...")
    oa_thread.start()
    non_oa_thread.start()

    # Wait for both to complete
    oa_thread.join()
    non_oa_thread.join()

    return oa_result[0], non_oa_result[0]

# =============================================================================
# FULL-TEXT SEARCH
# =============================================================================

def extract_pdf_text(filepath: Path) -> str:
    """Extract text from PDF using pdfplumber."""
    try:
        import pdfplumber
        text = ""
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        return text
    except Exception as e:
        print(f"    Error extracting {filepath.name}: {e}")
        return ""

def count_term_occurrences(text: str, term: str) -> int:
    """Count occurrences of a term in text."""
    if not text or not term:
        return 0
    text_lower = text.lower()
    term_lower = term.lower()

    if ' ' in term:
        # Exact phrase
        return text_lower.count(term_lower)
    else:
        # Keyword with word boundaries
        return len(re.findall(r'\b' + re.escape(term_lower) + r'\b', text_lower))

def search_fulltext_papers(papers: list[dict], pdf_dir: Path, search_terms: list[str], output_filepath: Path, csv_filepath: Path):
    """Search full-text of downloaded papers for terms."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Full-Text Search"

        # Headers
        headers = ["Study #", "Author", "Title", "DOI URL", "Abstract"] + search_terms
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)

        downloaded_papers = [p for p in papers if p.get("status") == "downloaded"]
        print(f"\n  Searching {len(downloaded_papers)} downloaded papers...")

        red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

        for row_idx, paper in enumerate(downloaded_papers, 2):
            study_num = paper.get("study_number", row_idx - 1)
            title_slug = re.sub(r'[^\w\s-]', '', paper.get("title", "")[:50]).strip().replace(' ', '_')
            pdf_filename = f"{study_num:04d}_{title_slug}.pdf"
            pdf_path = pdf_dir / pdf_filename

            print(f"  [{row_idx-1}/{len(downloaded_papers)}] Searching study #{study_num}...", end=" ")

            # Write basic info
            ws.cell(row=row_idx, column=1, value=study_num)
            ws.cell(row=row_idx, column=2, value=paper.get("author", ""))
            ws.cell(row=row_idx, column=3, value=paper.get("title", ""))
            ws.cell(row=row_idx, column=4, value=paper.get("doi_url", ""))
            ws.cell(row=row_idx, column=5, value=(paper.get("abstract", "") or "")[:32000])

            # Extract and search text
            if pdf_path.exists():
                text = extract_pdf_text(pdf_path)
                for term_idx, term in enumerate(search_terms, 6):
                    count = count_term_occurrences(text, term)
                    ws.cell(row=row_idx, column=term_idx, value=count)
                print("OK")
            else:
                print("PDF not found")

            # Mark as searched (red)
            paper["status"] = "searched"
            update_paper_status_in_excel(csv_filepath, [study_num], "searched")
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = red_fill

        # Column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 60

        wb.save(output_filepath)
        print(f"\n  Saved search results to: {output_filepath}")

    except Exception as e:
        print(f"  Error in full-text search: {e}")

# =============================================================================
# MAIN
# =============================================================================

def main():
    global QUERY_PATHS  # Declare global at start of function

    print("=" * 70)
    print("MULTI-API ACADEMIC PAPER SEARCH TOOL")
    print("Crossref + OpenAlex + PubMed")
    print("=" * 70)

    # STEP 1: Initial Menu
    print("\n" + "=" * 60)
    print("STEP 1: SELECT OPTION")
    print("=" * 60)
    print("\n  1: New Search")
    print("  2: Continue downloading from existing query")
    print("  3: Search full-texts of downloaded papers")

    initial_option = get_input("\nSelect option (1-3): ", "initial_option")

    if initial_option == "2":
        # Skip to Step 7 - Continue downloading
        print("\n" + "-" * 40)
        print("Enter the query folder name (e.g., '01-01-26-0439 Query'):")
        folder_name = input("> ").strip()
        query_folder = BASE_DIR / folder_name

        if not query_folder.exists():
            print(f"Folder not found: {query_folder}")
            return

        # Extract timestamp and set up paths
        timestamp_str = folder_name.replace(" Query", "")
        QUERY_PATHS = setup_query_folders(timestamp_str)

        csv_path = QUERY_PATHS["excel_file"]
        if not csv_path.exists():
            print(f"Excel file not found: {csv_path}")
            return

        oa_papers, no_oa_papers = load_papers_from_excel(csv_path)

        # Go to Step 7
        step7_download(oa_papers, no_oa_papers, csv_path, timestamp_str)
        return

    elif initial_option == "3":
        # Skip to Step 8 - Search full-texts
        print("\n" + "-" * 40)
        print("Enter the query folder name (e.g., '01-01-26-0439 Query'):")
        folder_name = input("> ").strip()
        query_folder = BASE_DIR / folder_name

        if not query_folder.exists():
            print(f"Folder not found: {query_folder}")
            return

        # Extract timestamp and set up paths
        timestamp_str = folder_name.replace(" Query", "")
        QUERY_PATHS = setup_query_folders(timestamp_str)

        csv_path = QUERY_PATHS["excel_file"]
        pdf_dir = QUERY_PATHS["oa_folder"]  # Search in OA papers folder

        if not csv_path.exists():
            print(f"Excel file not found: {csv_path}")
            return

        oa_papers, no_oa_papers = load_papers_from_excel(csv_path)

        # Go to Step 8
        step8_fulltext_search(oa_papers, pdf_dir, csv_path, timestamp_str)
        return

    # Continue with new search (option 1)
    query_start_time = datetime.now()
    timestamp_str = get_timestamp_str(query_start_time)

    # Set up query folder structure
    QUERY_PATHS = setup_query_folders(timestamp_str)
    print(f"\n  Query folder: {QUERY_PATHS['query_folder']}")

    # STEP 2: Master Query
    print("\n" + "=" * 60)
    print("STEP 2: MASTER QUERY (API Search)")
    print("=" * 60)
    print("\nSelect search mode:\n")
    print("  1: All APIs (keywords, best match, abstract)")
    print("  2: CrossRef (keywords, best match, abstract) +")
    print("     PubMed and OpenAlex (exact phrase and keywords, abstract)")
    print("  3: CrossRef (keywords, best match, abstract) +")
    print("     PubMed (exact phrase and keywords, abstract) +")
    print("     OpenAlex (exact phrase and keywords, full-text)")
    print("  4: CrossRef (keywords, best match, abstract) +")
    print("     PubMed and OpenAlex (exact phrase and keywords, abstract) +")
    print("     OpenAlex (exact phrase and keywords, full-text)\n")

    option = get_input("Select option (1-4): ", "search_mode")
    if option not in ["1", "2", "3", "4"]:
        print("Invalid option. Defaulting to 1.")
        option = "1"

    # Get max results
    print("\nMax results per API? (default: 10000, 0 for unlimited)")
    try:
        max_input = get_input("> ", "max_results")
        max_results = int(max_input) if max_input else 10000
        if max_results == 0:
            max_results = None
    except ValueError:
        max_results = 10000

    # Journal filter selection
    selected_journals = interactive_journal_selection()
    journal_filter_crossref = build_crossref_journal_filter(selected_journals) if selected_journals else ""
    journal_filter_openalex = build_openalex_journal_filter(selected_journals) if selected_journals else ""
    
    # Execute queries based on option
    all_papers = []
    query_for_filename = ""

    if option == "1":
        print("\n" + "-" * 40)
        print("All APIs (keywords, best match, abstract):")
        print("Enter your search terms as a comma-separated list")
        keywords = input("> ").strip()
        if not keywords:
            print("No query provided. Exiting.")
            return
        query_for_filename = keywords

        all_papers.extend(crossref_search(build_keywords_query(keywords, "crossref"), max_results, journal_filter_crossref))
        all_papers.extend(openalex_search(build_keywords_query(keywords, "openalex"), max_results, "default", journal_filter_openalex))
        pubmed_results = pubmed_search(build_keywords_query(keywords, "pubmed"), max_results)
        if selected_journals:
            pubmed_results = filter_papers_by_journal(pubmed_results, selected_journals)
            print(f"  After journal filter: {len(pubmed_results):,} papers")
        all_papers.extend(pubmed_results)

    elif option == "2":
        print("\n" + "-" * 40)
        print("CrossRef (keywords, best match, abstract):")
        print("Enter your search terms as a comma-separated list")
        crossref_kw = input("> ").strip()

        print("\n" + "-" * 40)
        print("PubMed and OpenAlex (exact phrase and keywords, abstract):")
        print("Enter your search terms as a comma-separated list with more than one term meaning exact phrase")
        pm_oa_input = input("> ").strip()

        query_for_filename = crossref_kw or pm_oa_input

        if crossref_kw:
            all_papers.extend(crossref_search(build_keywords_query(crossref_kw, "crossref"), max_results, journal_filter_crossref))
        if pm_oa_input:
            pubmed_results = pubmed_search(build_exact_keywords_query(pm_oa_input, "pubmed"), max_results)
            if selected_journals:
                pubmed_results = filter_papers_by_journal(pubmed_results, selected_journals)
                print(f"  After journal filter: {len(pubmed_results):,} papers")
            all_papers.extend(pubmed_results)
            all_papers.extend(openalex_search(build_exact_keywords_query(pm_oa_input, "openalex"), max_results, "abstract", journal_filter_openalex))

    elif option == "3":
        print("\n" + "-" * 40)
        print("CrossRef (keywords, best match, abstract):")
        print("Enter your search terms as a comma-separated list")
        crossref_kw = input("> ").strip()

        print("\n" + "-" * 40)
        print("PubMed (exact phrase and keywords, abstract):")
        print("Enter your search terms as a comma-separated list with more than one term meaning exact phrase")
        pm_input = input("> ").strip()

        print("\n" + "-" * 40)
        print("OpenAlex (exact phrase and keywords, full-text):")
        print("Enter your search terms as a comma-separated list with more than one term meaning exact phrase")
        oa_ft_input = input("> ").strip()

        query_for_filename = crossref_kw or pm_input or oa_ft_input

        if crossref_kw:
            all_papers.extend(crossref_search(build_keywords_query(crossref_kw, "crossref"), max_results, journal_filter_crossref))
        if pm_input:
            pubmed_results = pubmed_search(build_exact_keywords_query(pm_input, "pubmed"), max_results)
            if selected_journals:
                pubmed_results = filter_papers_by_journal(pubmed_results, selected_journals)
                print(f"  After journal filter: {len(pubmed_results):,} papers")
            all_papers.extend(pubmed_results)
        if oa_ft_input:
            all_papers.extend(openalex_search(build_exact_keywords_query(oa_ft_input, "openalex"), max_results, "fulltext", journal_filter_openalex))

    elif option == "4":
        print("\n" + "-" * 40)
        print("CrossRef (keywords, best match, abstract):")
        print("Enter your search terms as a comma-separated list")
        crossref_kw = get_input("> ", "crossref_keywords")

        print("\n" + "-" * 40)
        print("PubMed and OpenAlex (exact phrase and keywords, abstract):")
        print("Enter your search terms as a comma-separated list with more than one term meaning exact phrase")
        abstract_input = get_input("> ", "abstract_search")

        print("\n" + "-" * 40)
        print("OpenAlex (exact phrase and keywords, full-text):")
        print("Enter your search terms as a comma-separated list with more than one term meaning exact phrase")
        oa_ft_input = get_input("> ", "fulltext_search")

        query_for_filename = crossref_kw or abstract_input or oa_ft_input

        if crossref_kw:
            all_papers.extend(crossref_search(build_keywords_query(crossref_kw, "crossref"), max_results, journal_filter_crossref))
        if abstract_input:
            pubmed_results = pubmed_search(build_exact_keywords_query(abstract_input, "pubmed"), max_results)
            if selected_journals:
                pubmed_results = filter_papers_by_journal(pubmed_results, selected_journals)
                print(f"  After journal filter: {len(pubmed_results):,} papers")
            all_papers.extend(pubmed_results)
            all_papers.extend(openalex_search(build_exact_keywords_query(abstract_input, "openalex"), max_results, "abstract", journal_filter_openalex))
        if oa_ft_input:
            all_papers.extend(openalex_search(build_exact_keywords_query(oa_ft_input, "openalex"), max_results, "fulltext", journal_filter_openalex))

    # STEP 3: Filter and assign study numbers
    print(f"\n{'='*60}")
    print("STEP 3: FILTER AND ASSIGN STUDY NUMBERS")
    print(f"{'='*60}")

    # Remove papers without abstracts
    papers_with_abstracts = [p for p in all_papers if p.get("abstract")]
    print(f"  Total from APIs: {len(all_papers):,}")
    print(f"  Without abstracts (removed): {len(all_papers) - len(papers_with_abstracts):,}")
    print(f"  With abstracts: {len(papers_with_abstracts):,}")

    # Deduplicate
    unique_papers = deduplicate_papers(papers_with_abstracts)

    # Assign study numbers
    for i, paper in enumerate(unique_papers, 1):
        paper["study_number"] = i
        paper["status"] = "pending"

    print(f"\n  Assigned study numbers 1 to {len(unique_papers):,}")

    # STEP 4: Option to filter OA now or continue
    print(f"\n{'='*60}")
    print("STEP 4: CHOOSE NEXT ACTION")
    print(f"{'='*60}")
    print("\n  1: Filter by OA/non-OA now and save CSV (then download)")
    print("  2: Continue to subquery first")

    step4_option = get_input("\nSelect option (1-2): ", "step4_option")

    if step4_option == "1":
        # Go directly to OA filtering and save
        papers_to_filter = unique_papers
    else:
        # STEP 5: Subquery
        print(f"\n{'='*60}")
        print("STEP 5: SUBQUERY (Local Abstract Filtering)")
        print(f"{'='*60}")
        print("Enter AND/OR query to filter abstracts:")
        print("  - AND splits into groups (ALL must match)")
        print("  - OR within groups (ANY can match)")
        print("  - Multi-word = exact phrase")
        print("\nExample: moisture OR wicking AND transport OR flux AND AATCC TM199")

        subquery = get_input("\nSubquery: ", "subquery")

        if subquery:
            filtered = filter_by_subquery(unique_papers, subquery)
            print(f"\n  Before: {len(unique_papers):,}")
            print(f"  After: {len(filtered):,}")

            # Re-assign study numbers
            for i, paper in enumerate(filtered, 1):
                paper["study_number"] = i

            papers_to_filter = filtered
        else:
            papers_to_filter = unique_papers

    # STEP 6: Check OA status and save
    print(f"\n{'='*60}")
    print("STEP 6: CHECK OPEN ACCESS STATUS")
    print(f"{'='*60}")

    papers_to_filter = check_oa_status(papers_to_filter)

    # Split by OA
    oa_papers = [p for p in papers_to_filter if p.get("is_oa") == True]
    no_oa_papers = [p for p in papers_to_filter if p.get("is_oa") == False]
    unknown = [p for p in papers_to_filter if p.get("is_oa") is None]

    # Add unknown to no_oa for now
    no_oa_papers.extend(unknown)

    print(f"\n  Open Access: {len(oa_papers):,}")
    print(f"  Not Open Access: {len(no_oa_papers):,}")

    # Use QUERY_PATHS for file locations
    csv_path = QUERY_PATHS["excel_file"]
    backup_path = QUERY_PATHS["backup_file"]

    print(f"\n{'='*60}")
    print("SAVING RESULTS")
    print(f"{'='*60}")

    # BACKUP: Save JSON first to prevent data loss
    import json
    try:
        with open(backup_path, 'w', encoding='utf-8') as f:
            json.dump({"oa": oa_papers, "no_oa": no_oa_papers}, f, ensure_ascii=False, indent=2)
        print(f"  JSON backup saved: {backup_path}")
    except Exception as e:
        print(f"  WARNING: JSON backup failed: {e}")

    # Now save Excel (with fallback to CSV if it fails)
    try:
        save_oa_split_excel(oa_papers, no_oa_papers, csv_path)
    except Exception as e:
        print(f"  ERROR saving Excel: {e}")
        print("  Falling back to CSV...")
        save_papers_csv(oa_papers, csv_path.with_suffix('.csv').with_stem(csv_path.stem + '_OA'))
        save_papers_csv(no_oa_papers, csv_path.with_suffix('.csv').with_stem(csv_path.stem + '_NoOA'))

    # STEP 7: Download (OA + non-OA via SciHub in parallel)
    step7_download(oa_papers, no_oa_papers, csv_path, timestamp_str)

def step7_download(oa_papers: list[dict], non_oa_papers: list[dict], csv_path: Path, timestamp_str: str):
    """Step 7: Download OA papers (direct) and non-OA papers (via SciHub) in parallel."""
    print(f"\n{'='*60}")
    print("STEP 7: DOWNLOAD PAPERS (OA + SciHub parallel)")
    print(f"{'='*60}")

    # Count pending for both types
    oa_pending = [p for p in oa_papers if p.get("status", "pending") == "pending" and p.get("pdf_url")]
    non_oa_pending = [p for p in non_oa_papers if p.get("status", "pending") == "pending" and p.get("doi") and not p.get("pdf_url")]

    print(f"\n  OA papers (direct download): {len(oa_pending):,}")
    print(f"  Non-OA papers (via SciHub):  {len(non_oa_pending):,}")
    print(f"  Total pending:               {len(oa_pending) + len(non_oa_pending):,}")

    if not oa_pending and not non_oa_pending:
        print("  No papers to download.")
        return

    first_iteration = True
    while True:
        # Ask for OA count
        if oa_pending:
            print(f"\n  How many OA papers to download? (max {len(oa_pending)}, 0 to skip)")
            try:
                if first_iteration:
                    oa_count = int(get_input("  OA > ", "oa_download_count"))
                else:
                    oa_count = int(input("  OA > ").strip())
            except ValueError:
                oa_count = 10
        else:
            oa_count = 0

        # Ask for non-OA count
        if non_oa_pending:
            print(f"\n  How many non-OA papers to download via SciHub? (max {len(non_oa_pending)}, 0 to skip)")
            try:
                if first_iteration:
                    non_oa_count = int(get_input("  SciHub > ", "non_oa_download_count"))
                else:
                    non_oa_count = int(input("  SciHub > ").strip())
            except ValueError:
                non_oa_count = 10
        else:
            non_oa_count = 0

        first_iteration = False

        if oa_count == 0 and non_oa_count == 0:
            print("  Skipping downloads.")
            return

        # Run parallel downloads (separate folders for OA and Non-OA)
        oa_downloaded, non_oa_downloaded = parallel_download_all(
            oa_papers, non_oa_papers,
            QUERY_PATHS["oa_folder"], QUERY_PATHS["non_oa_folder"],
            csv_path, oa_count, non_oa_count
        )

        print(f"\n  === BATCH COMPLETE ===")
        print(f"  OA downloaded:     {oa_downloaded}")
        print(f"  SciHub downloaded: {non_oa_downloaded}")

        # Update pending counts
        oa_pending = [p for p in oa_papers if p.get("status", "pending") == "pending" and p.get("pdf_url")]
        non_oa_pending = [p for p in non_oa_papers if p.get("status", "pending") == "pending" and p.get("doi") and not p.get("pdf_url")]

        print(f"\n  Remaining OA:     {len(oa_pending):,}")
        print(f"  Remaining SciHub: {len(non_oa_pending):,}")

        if not oa_pending and not non_oa_pending:
            print("  All papers downloaded!")
            break

        cont = input("\n  Continue downloading? (Y/N): ").strip().upper()
        if cont != "Y":
            print("  Returning to main menu...")
            return

def step8_fulltext_search(oa_papers: list[dict], pdf_dir: Path, csv_path: Path, timestamp_str: str):
    """Step 8: Search full-text of downloaded papers."""
    print(f"\n{'='*60}")
    print("STEP 8: FULL-TEXT SEARCH")
    print(f"{'='*60}")

    # Check remaining downloads
    pending = [p for p in oa_papers if p.get("status", "pending") == "pending" and p.get("pdf_url")]
    downloaded = [p for p in oa_papers if p.get("status") == "downloaded"]

    print(f"\n  Downloaded papers: {len(downloaded):,}")
    print(f"  Remaining to download: {len(pending):,}")

    if pending:
        print("\n  1: Begin search of all downloaded papers")
        print("  2: Continue downloading remaining papers")

        choice = input("\nSelect option (1-2): ").strip()

        if choice == "2":
            step7_download(oa_papers, csv_path, timestamp_str)
            return

    if not downloaded:
        print("  No downloaded papers to search.")
        return

    # Get search terms
    print("\n  Enter comma-separated exact phrases and keywords to search:")
    search_input = input("> ").strip()

    if not search_input:
        print("  No search terms provided.")
        return

    search_terms = [t.strip() for t in search_input.split(",") if t.strip()]
    print(f"\n  Searching for: {search_terms}")

    # Create output file
    output_path = QUERY_PATHS["fulltext_file"]

    search_fulltext_papers(oa_papers, pdf_dir, search_terms, output_path, csv_path)

    print(f"\n{'='*60}")
    print("DONE!")
    print(f"{'='*60}")

if __name__ == "__main__":
    main()
