"""
LLM Filter - Relevance Checking with SambaNova

Filters extracted text to determine if papers are relevant to fabric research.

PROCESS:
1. Read metadata Excel file (contains PDF Number and empty Breakpoint column)
2. For each PDF, read the full text extraction file
3. Send to SambaNova LLM with filtering prompt
4. Update Breakpoint column with response
5. Save updated Excel file

INPUTS:
- Metadata Excel file (.xlsx)
- Full text extraction files (.txt)

OUTPUTS:
- Updated Excel file with Breakpoint column filled

BREAKPOINT VALUES:
- "Passes" - Paper performs experiment with knit/woven fabrics or yarns
- "No experiment" - Paper does not perform an experiment
- "No knit/woven/yarn" - Paper performs experiment but not with relevant materials
- "Error: <message>" - Processing error occurred
"""

import os
import sys
from pathlib import Path
from typing import Dict, List, Tuple
import time

# Import timing utilities
sys.path.insert(0, str(Path(__file__).parent.parent))
from Timers.timing_utils import save_timer

# Load environment variables
from dotenv import load_dotenv
load_dotenv()

# Check for required libraries
try:
    import pandas as pd
except ImportError:
    print("ERROR: Missing required library 'pandas'")
    print("Install it with: pip install pandas openpyxl")
    sys.exit(1)

try:
    import requests
except ImportError:
    print("ERROR: Missing required library 'requests'")
    print("Install it with: pip install requests")
    sys.exit(1)

# SambaNova API configuration
SAMBANOVA_API_KEY = os.getenv("SAMBANOVA_API_KEY")
SAMBANOVA_CHAT_ENDPOINT = os.getenv("SAMBANOVA_CHAT_ENDPOINT")

if not SAMBANOVA_API_KEY or not SAMBANOVA_CHAT_ENDPOINT:
    print("ERROR: SambaNova API credentials not found in .env file")
    print("Required: SAMBANOVA_API_KEY, SAMBANOVA_CHAT_ENDPOINT")
    sys.exit(1)


def load_filter_prompt() -> str:
    """Load the filtering prompt from Config/LLMRelevancePrompt.txt."""
    prompt_file = Path(__file__).parent.parent / "Config" / "LLMRelevancePrompt.txt"

    if not prompt_file.exists():
        print(f"ERROR: Filter prompt file not found: {prompt_file}")
        sys.exit(1)

    with open(prompt_file, 'r', encoding='utf-8') as f:
        return f.read()


# Load filtering prompt from config
FILTER_PROMPT = load_filter_prompt()


def call_sambanova(text: str, max_retries: int = 3) -> str:
    """
    Call SambaNova API with filtering prompt.

    Args:
        text: Full text of paper to filter
        max_retries: Number of retry attempts on failure

    Returns:
        Filter decision: "Passes", "No experiment", "No knit/woven/yarn", or "Error: <message>"
    """
    prompt = FILTER_PROMPT.format(text=text)

    headers = {
        "Authorization": f"Bearer {SAMBANOVA_API_KEY}",
        "Content-Type": "application/json"
    }

    payload = {
        "model": "Meta-Llama-3.3-70B-Instruct",
        "messages": [
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.0,  # Deterministic output
        "max_tokens": 50  # Only need a few words
    }

    for attempt in range(max_retries):
        try:
            response = requests.post(
                SAMBANOVA_CHAT_ENDPOINT,
                headers=headers,
                json=payload,
                timeout=30
            )

            if response.status_code == 200:
                result = response.json()
                decision = result["choices"][0]["message"]["content"].strip()

                # Validate response
                valid_responses = ["Passes", "No experiment", "No knit/woven/yarn"]
                if decision in valid_responses:
                    return decision
                else:
                    # LLM gave unexpected response, try to map it
                    decision_lower = decision.lower()
                    if "no experiment" in decision_lower:
                        return "No experiment"
                    elif "no knit" in decision_lower or "no woven" in decision_lower or "no yarn" in decision_lower:
                        return "No knit/woven/yarn"
                    elif "pass" in decision_lower:
                        return "Passes"
                    else:
                        return f"Error: Unexpected response '{decision}'"

            elif response.status_code == 429:
                # Rate limit - wait and retry
                wait_time = 2 ** attempt  # Exponential backoff
                print(f"  [Rate limit] Waiting {wait_time}s before retry {attempt+1}/{max_retries}...")
                time.sleep(wait_time)

            else:
                return f"Error: API returned {response.status_code}"

        except requests.exceptions.Timeout:
            if attempt < max_retries - 1:
                print(f"  [Timeout] Retry {attempt+1}/{max_retries}...")
                time.sleep(1)
            else:
                return "Error: API timeout"

        except Exception as e:
            return f"Error: {str(e)}"

    return "Error: Max retries exceeded"


def filter_papers(pdf_folder: Path, excel_path: Path) -> Path:
    """
    Filter papers using SambaNova LLM and update Breakpoint column.
    Deletes irrelevant papers and their files.

    Args:
        pdf_folder: Folder containing PDFs and text extraction files
        excel_path: Path to metadata Excel file

    Returns:
        Path to updated Excel file
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    print("[DEBUG] Loading metadata Excel file...")
    wb = load_workbook(excel_path)

    # Find the sheet with data (usually "OA" or first sheet)
    if "OA" in wb.sheetnames:
        ws = wb["OA"]
    else:
        ws = wb.active

    # Get headers
    headers = [cell.value for cell in ws[1]]

    if "Breakpoint" not in headers:
        raise ValueError("ERROR: 'Breakpoint' column not found in Excel file")

    breakpoint_col = headers.index("Breakpoint") + 1
    pdf_num_col = headers.index("PDF Number") + 1

    total_papers = ws.max_row - 1  # Exclude header
    print(f"[DEBUG] Found {total_papers} papers to filter")
    print(f"[DEBUG] Excel columns: {headers}\n")

    # Process each paper
    total_start = time.time()
    deleted_count = 0
    first_llm_call_time = None

    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

    for row_idx in range(2, ws.max_row + 1):
        pdf_num = ws.cell(row=row_idx, column=pdf_num_col).value

        print("=" * 80)
        print(f"Processing PDF {row_idx-1}/{total_papers} (PDF #{pdf_num})")
        print("=" * 80)

        # Find text file (in Extracted Text folder, not Downloaded Papers)
        text_file = pdf_folder.parent / "Extracted Text" / f"{pdf_num}.txt"

        if not text_file.exists():
            print(f"  [WARNING] Text file not found: {text_file}")
            ws.cell(row=row_idx, column=breakpoint_col).value = "Error: Text file not found"
            # Turn row red
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = red_fill
            continue

        # Read full text
        try:
            with open(text_file, 'r', encoding='utf-8') as f:
                text = f.read()

            if not text.strip():
                print(f"  [WARNING] Text file is empty")
                ws.cell(row=row_idx, column=breakpoint_col).value = "Error: Empty text file"
                # Turn row red
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = red_fill
                continue

            print(f"  [1/3] Loaded text: {len(text):,} chars, {len(text.split()):,} words")

        except Exception as e:
            print(f"  [ERROR] Failed to read text file: {e}")
            ws.cell(row=row_idx, column=breakpoint_col).value = f"Error: {str(e)}"
            # Turn row red
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = red_fill
            continue

        # Call SambaNova LLM (Timer 8: First call)
        print(f"  [2/3] Calling SambaNova LLM...")
        start_time = time.time()

        decision = call_sambanova(text)

        elapsed = time.time() - start_time

        # Save first LLM call time
        if first_llm_call_time is None:
            first_llm_call_time = elapsed
            # Get query folder from pdf_folder path
            query_folder = pdf_folder.parent
            save_timer(query_folder, "8_llm_filter_first", first_llm_call_time)

        print(f"  [RESULT] {decision} ({elapsed:.1f}s)")

        # Update Excel with decision
        ws.cell(row=row_idx, column=breakpoint_col).value = decision

        # If irrelevant, turn row red and DELETE files
        if decision in ["No experiment", "No knit/woven/yarn"]:
            print(f"  [3/3] Marking as irrelevant and deleting files...")

            # Turn row red
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = red_fill

            # Delete files
            extracted_text_folder = pdf_folder.parent / "Extracted Text"
            files_to_delete = [
                pdf_folder / "OA Papers" / f"{pdf_num}.pdf",
                extracted_text_folder / f"{pdf_num}.txt",
                extracted_text_folder / f"{pdf_num}_filtered.txt",
                extracted_text_folder / f"{pdf_num}_filtered_NEW.txt",
            ]

            for file_path in files_to_delete:
                if file_path.exists():
                    try:
                        file_path.unlink()
                        print(f"    [DELETED] {file_path.name}")
                    except Exception as e:
                        print(f"    [ERROR] Failed to delete {file_path.name}: {e}")

            deleted_count += 1

        elif decision.startswith("Error:"):
            print(f"  [3/3] Error occurred, keeping files for retry")
            # Turn row red but keep files
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = red_fill

        else:
            # Passes - leave green, keep files
            print(f"  [3/3] Paper passes filter, keeping all files")

        # Small delay to avoid rate limits
        time.sleep(0.5)

    # Save updated Excel
    total_time = time.time() - total_start

    print("\n" + "=" * 80)
    print("SAVING UPDATED EXCEL FILE")
    print("=" * 80)

    wb.save(excel_path)
    print(f"Updated Excel saved to: {excel_path.name}")

    # Summary
    print("\n" + "=" * 80)
    print("FILTERING SUMMARY")
    print("=" * 80)

    # Count decisions
    decisions = {}
    for row_idx in range(2, ws.max_row + 1):
        decision = ws.cell(row=row_idx, column=breakpoint_col).value or "Passes"
        decisions[decision] = decisions.get(decision, 0) + 1

    for value, count in sorted(decisions.items()):
        print(f"  {value}: {count}")

    print(f"\nFiles deleted: {deleted_count} papers")
    print(f"Total time: {total_time:.1f}s")
    print(f"Avg per paper: {total_time/total_papers:.1f}s")
    print("=" * 80)

    return excel_path


def find_latest_query_folder():
    """Find the latest Query folder in Output directory."""
    output_base = Path(__file__).parent.parent / "Output"
    query_folders = [d for d in output_base.iterdir() if d.is_dir() and "Query" in d.name]
    if not query_folders:
        return None
    return max(query_folders, key=lambda d: d.stat().st_mtime)


def main():
    """CLI entry point for LLM filtering."""
    import argparse

    parser = argparse.ArgumentParser(description="Filter papers using SambaNova LLM")
    parser.add_argument("pdf_folder", nargs='?', type=Path, help="Folder containing PDFs and text files")
    parser.add_argument("excel_file", nargs='?', type=Path, help="Metadata Excel file path")

    args = parser.parse_args()

    # Auto-detect if no arguments provided
    if args.pdf_folder is None:
        query_folder = find_latest_query_folder()
        if query_folder is None:
            print("ERROR: No Query folder found in Output directory")
            sys.exit(1)

        args.pdf_folder = query_folder / "Downloaded Papers"
        args.excel_file = query_folder / "metadata.xlsx"

        print(f"Auto-detected Query folder: {query_folder.name}")

    if not args.pdf_folder.exists():
        print(f"ERROR: PDF folder not found: {args.pdf_folder}")
        sys.exit(1)

    if not args.excel_file.exists():
        print(f"ERROR: Excel file not found: {args.excel_file}")
        sys.exit(1)

    print("=" * 80)
    print("LLM FILTER - SambaNova Paper Relevance Checking")
    print("=" * 80)
    print(f"PDF Folder: {args.pdf_folder}")
    print(f"Excel File: {args.excel_file}")
    print("=" * 80 + "\n")

    filter_papers(args.pdf_folder, args.excel_file)


if __name__ == "__main__":
    main()
