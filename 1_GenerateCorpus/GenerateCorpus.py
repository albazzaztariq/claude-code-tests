"""
GenerateCorpus - Research Paper Collection Module
API search → Deduplicate → OA check → Excel → Mirror Testing → Download

FLOW:
1. Read config file (terms, journals, fields)
2. Build OpenAlex query with AND/OR/NOT logic
3. Call API and paginate through results
4. Deduplicate by DOI
5. Check OA status via Unpaywall (get all mirrors)
6. Save to Excel (2 sheets: OA / No OA)
7. Test all unique mirrors in parallel (50KB sample, 32 concurrent)
8. Download PDFs using fastest mirrors (sorted by measured speed)

MIRROR TESTING:
- Extracts unique mirrors from all papers
- Tests in batches of 50, max 32 concurrent connections
- Downloads 50KB sample from each to measure bandwidth
- Sorts mirrors by actual speed (not ping or static rankings)
- Downloads use fastest available mirror for each paper

Usage:
    python GenerateCorpus.py [config_file.txt]

TEMPORARY: Limited to first 250 PDFs for testing
"""

import sys
import re
import time
import json
import asyncio
import threading

# Check for required third-party libraries
try:
    import requests
except ImportError:
    print("ERROR: Missing required library 'requests'")
    print("Install it with: pip install requests")
    sys.exit(1)

try:
    import aiohttp
except ImportError:
    print("ERROR: Missing required library 'aiohttp'")
    print("Install it with: pip install aiohttp")
    sys.exit(1)

try:
    from dotenv import load_dotenv
except ImportError:
    print("ERROR: Missing required library 'python-dotenv'")
    print("Install it with: pip install python-dotenv")
    sys.exit(1)

from pathlib import Path
from datetime import datetime
from urllib.parse import urlencode, urlparse
from urllib.request import urlretrieve
import os

# Import timing utilities and corpus utilities
sys.path.insert(0, str(Path(__file__).parent.parent))
from Timers.timing_utils import save_timer
from Utils.corpus_utils import save_corpus_json

# =============================================================================
# CONFIGURATION
# =============================================================================

# Load .env from FabricETL base directory
load_dotenv(Path(__file__).parent.parent / ".env")

OPENALEX_API_URL = "https://api.openalex.org/works"
OPENALEX_EMAIL = os.getenv("OPENALEX_EMAIL")  # Polite pool access (10 req/sec)
UNPAYWALL_EMAIL = os.getenv("UNPAYWALL_EMAIL")

# Default config file location
BASE_DIR = Path(__file__).parent
DEFAULT_CONFIG = BASE_DIR.parent / "Config" / "query.txt"  # DEV/Config or SBX/Config or PROD/Config
OUTPUT_DIR = Path(__file__).parent.parent / "Output"  # DEV/Output or SBX/Output or PROD/Output

# Journal list import (optional)
try:
    sys.path.insert(0, str(BASE_DIR.parent.parent / "v02" / "v02-2" / "src"))
    from JournalList import TEXTILE_JOURNALS
except ImportError:
    TEXTILE_JOURNALS = {}

# =============================================================================
# CONFIG FILE PARSER
# =============================================================================

def parse_config(filepath: Path) -> dict:
    """Parse config file and return search parameters.

    Returns:
        dict with keys: terms, journals, fields, max_results
    """
    if not filepath.exists():
        raise FileNotFoundError(f"Config file not found: {filepath}")

    with open(filepath, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    # Parse sections
    config = {
        "terms": "",
        "journals": [],
        "fields": ["title", "abstract"],
        "max_results": None
    }

    current_section = None

    for line in lines:
        line = line.strip()

        # Skip comments and blanks
        if not line or line.startswith("#"):
            continue

        # Section headers
        if line.startswith("[") and line.endswith("]"):
            current_section = line[1:-1].lower()
            continue

        # Parse based on section
        if current_section == "openalex":
            if line.startswith("field:"):
                field_str = line.split(":", 1)[1].strip()
                config["fields"] = [f.strip() for f in field_str.split(",")]
            elif line.startswith("terms:"):
                config["terms"] = line.split(":", 1)[1].strip()

        elif current_section == "journals":
            # Expand acronyms to full names
            journal = line.strip()
            if journal in TEXTILE_JOURNALS.values():
                # Find full name from acronym
                for full_name, acronym in TEXTILE_JOURNALS.items():
                    if acronym == journal:
                        config["journals"].append(full_name)
                        break
            else:
                config["journals"].append(journal)

        elif current_section == "max results":
            try:
                config["max_results"] = int(line.replace(",", ""))
            except ValueError:
                pass

    if not config["terms"]:
        raise ValueError("No search terms found in config file")

    return config

# =============================================================================
# OPENALEX QUERY BUILDER
# =============================================================================

def build_openalex_query(config: dict) -> str:
    """Build OpenAlex API URL from config.

    Handles AND/OR/NOT logic in search terms.

    CRITICAL: OpenAlex requires comma-separated filters for AND logic.
    Example: (woven OR knit) AND moisture-wicking becomes:
      filter=fulltext.search:woven OR knit,fulltext.search:moisture-wicking

    Args:
        config: Dict with terms, journals, fields, max_results

    Returns:
        Full API URL with query parameters
    """
    terms = config["terms"]
    fields = config["fields"]

    # Parse terms: /AND/ splits into separate filter parts (comma-separated)
    # Within each AND group, /// = OR
    and_groups = [g.strip() for g in terms.split("/AND/") if g.strip()]

    filter_parts = []

    for and_group in and_groups:
        # Split by /// for OR terms within this AND group
        or_terms = [t.strip() for t in and_group.split("///") if t.strip()]

        # Format each term (quote multi-word phrases)
        formatted_terms = []
        for term in or_terms:
            term = term.strip()
            if ' ' in term:
                # Multi-word phrase: needs quotes
                formatted_terms.append(f'"{term}"')
            else:
                # Single word or hyphenated: no quotes
                formatted_terms.append(term)

        # Join OR terms with " OR "
        or_query = " OR ".join(formatted_terms)

        # Create filter part for this AND group
        if "fulltext" in fields:
            filter_parts.append(f"fulltext.search:{or_query}")
        elif "title" in fields and "abstract" in fields:
            filter_parts.append(f"title_and_abstract.search:{or_query}")
        elif "title" in fields:
            filter_parts.append(f"title.search:{or_query}")
        else:
            # Fallback to search param (not recommended)
            filter_parts.append(f"search:{or_query}")

    # Add journal filter if specified
    if config["journals"]:
        journal_filter = "|".join(config["journals"])
        filter_parts.append(f"primary_location.source.display_name.search:{journal_filter}")

    # Build params
    params = {
        "per_page": 200,
        "cursor": "*",
        "mailto": OPENALEX_EMAIL
    }

    if filter_parts:
        params["filter"] = ",".join(filter_parts)

    return f"{OPENALEX_API_URL}?{urlencode(params, safe=':,|')}"

# =============================================================================
# OPENALEX API CALLER
# =============================================================================

def call_openalex(url: str, max_results: int = None, query_folder: Path = None) -> list[dict]:
    """Call OpenAlex API and paginate through all results.

    Args:
        url: OpenAlex API URL with cursor=*
        max_results: Maximum papers to retrieve (None = unlimited)
        query_folder: Query folder for timer saving (optional)

    Returns:
        List of paper metadata dicts
    """
    papers = []
    current_url = url

    # Timing variables for all calls
    all_call_times = []

    print("Calling OpenAlex API...")

    while current_url:
        batch_num = len(papers)//200 + 1

        # Timer 2: Track all call times
        call_start = time.time()

        try:
            response = requests.get(current_url, timeout=30)
            response.raise_for_status()
            data = response.json()
        except Exception as e:
            break

        call_time = time.time() - call_start
        all_call_times.append(call_time)

        results = data.get("results", [])

        # Parse results
        for item in results:
            paper = {
                "title": item.get("title", ""),
                "author": (item.get("authorships", [{}])[0].get("author", {}).get("display_name", "") if item.get("authorships") else ""),
                "year": item.get("publication_year", ""),
                "doi": item.get("doi", "").replace("https://doi.org/", "") if item.get("doi") else "",
                "abstract": (item.get("abstract_inverted_index", None) and "Abstract available" or ""),
                "pdf_url": "",
                "sources": "OpenAlex"
            }

            # Get DOI URL
            if paper["doi"]:
                paper["doi_url"] = f"https://doi.org/{paper['doi']}"
            else:
                paper["doi_url"] = ""

            # Get PDF URL if available
            if item.get("open_access", {}).get("oa_url"):
                paper["pdf_url"] = item["open_access"]["oa_url"]

            papers.append(paper)

            # Check max results
            if max_results and len(papers) >= max_results:
                print(f"  Reached max results limit ({max_results})")
                # Save timer before early return
                if all_call_times and query_folder:
                    total_openalex_time = sum(all_call_times)
                    save_timer(query_folder, "1_openalex_total", total_openalex_time)
                return papers

        # Get next cursor
        meta = data.get("meta", {})
        next_cursor = meta.get("next_cursor")

        if next_cursor:
            current_url = url.replace("cursor=*", f"cursor={next_cursor}")
            time.sleep(0.1)  # Rate limiting
        else:
            current_url = None

    print(f"Results: {len(papers):,}")

    # Save total OpenAlex API time (Timer 2)
    if all_call_times and query_folder:
        total_openalex_time = sum(all_call_times)
        save_timer(query_folder, "1_openalex_total", total_openalex_time)

    return papers

# =============================================================================
# DEDUPLICATION
# =============================================================================

def deduplicate(papers: list[dict]) -> list[dict]:
    """Deduplicate papers by DOI and title.

    Args:
        papers: List of paper metadata dicts

    Returns:
        Deduplicated list
    """
    before_count = len(papers)

    seen_dois = set()
    seen_titles = set()
    unique = []

    for paper in papers:
        doi = paper.get("doi", "")
        title = paper.get("title", None)

        # Handle None titles
        if title:
            title = title.lower().strip()
        else:
            title = ""

        # Skip if DOI or title already seen
        if doi and doi in seen_dois:
            continue
        if title and title in seen_titles:
            continue

        # Add to unique list
        unique.append(paper)

        if doi:
            seen_dois.add(doi)
        if title:
            seen_titles.add(title)

    print(f"Duplicates Removed: {before_count - len(unique):,}")

    # DON'T assign study numbers yet - only after filtering out paywalls
    # Use temporary index for download tracking
    for i, paper in enumerate(unique, 1):
        paper["temp_index"] = i  # Temporary identifier for downloading
        paper["status"] = "pending"

    return unique

# =============================================================================
# DYNAMIC MIRROR SPEED TESTING
# =============================================================================

def extract_domain(url: str) -> str:
    """Extract domain from URL for mirror identification."""
    try:
        parsed = urlparse(url)
        return parsed.netloc.replace("www.", "")
    except:
        return ""

async def test_mirror_speed(session, url: str, semaphore, timeout=2):
    """Test download speed from a mirror by downloading 50KB sample.

    Args:
        session: aiohttp ClientSession
        url: Mirror URL to test
        semaphore: Asyncio semaphore for concurrency limiting
        timeout: Timeout in seconds

    Returns:
        Tuple of (domain, speed_bytes_per_sec)
    """
    domain = extract_domain(url)

    async with semaphore:
        try:
            start = time.time()
            async with session.get(url, timeout=timeout) as resp:
                if resp.status != 200:
                    return domain, 0

                # Download 50KB sample
                data = await resp.content.read(51200)
                elapsed = time.time() - start

                if elapsed > 0:
                    speed = len(data) / elapsed  # bytes/sec
                    return domain, speed
                else:
                    return domain, 0
        except:
            return domain, 0

async def test_mirrors_batched(mirrors: list[str], batch_size=50, concurrency=32):
    """Test all mirrors in batches with concurrency limits.

    Args:
        mirrors: List of mirror URLs to test
        batch_size: Number of mirrors per batch
        concurrency: Max concurrent connections

    Returns:
        Dict mapping domain -> speed (bytes/sec)
    """
    semaphore = asyncio.Semaphore(concurrency)
    speeds = {}

    async with aiohttp.ClientSession(
        connector=aiohttp.TCPConnector(limit=concurrency),
        timeout=aiohttp.ClientTimeout(total=3)
    ) as session:
        # Process in batches
        for i in range(0, len(mirrors), batch_size):
            batch = mirrors[i:i+batch_size]

            # Test batch in parallel (up to concurrency limit)
            tasks = [test_mirror_speed(session, url, semaphore) for url in batch]
            batch_results = await asyncio.gather(*tasks, return_exceptions=True)

            # Collect results
            for result in batch_results:
                if isinstance(result, tuple):
                    domain, speed = result
                    if domain and speed > 0:
                        # Keep best speed if domain tested multiple times
                        speeds[domain] = max(speeds.get(domain, 0), speed)

            # Rate limit courtesy between batches
            if i + batch_size < len(mirrors):
                await asyncio.sleep(1.5)

    return speeds

def rank_mirrors_by_speed(oa_papers: list[dict]) -> dict:
    """Extract unique mirrors from all papers and test speeds.

    Args:
        oa_papers: List of OA papers with oa_locations

    Returns:
        Dict mapping domain -> speed (bytes/sec)
    """
    # Extract all unique mirror URLs
    unique_mirrors = set()
    for paper in oa_papers:
        for loc in paper.get("oa_locations", []):
            url = loc.get("url_for_pdf") or loc.get("url")
            if url:
                unique_mirrors.add(url)

    if not unique_mirrors:
        return {}

    # Test mirrors asynchronously
    speeds = asyncio.run(test_mirrors_batched(list(unique_mirrors)))

    return speeds

# =============================================================================
# UNPAYWALL OA CHECKER
# =============================================================================

async def check_single_unpaywall(session, paper: dict, semaphore):
    """Check single paper's OA status via Unpaywall API.

    Args:
        session: aiohttp ClientSession
        paper: Paper dict with DOI
        semaphore: Asyncio semaphore for concurrency limiting

    Returns:
        Tuple of (paper, call_time, is_oa)
    """
    doi = paper.get("doi", "")

    if not doi:
        paper["access"] = "Non-OA"
        return paper, 0, False

    async with semaphore:
        try:
            url = f"https://api.unpaywall.org/v2/{doi}?email={UNPAYWALL_EMAIL}"

            call_start = time.time()
            async with session.get(url, timeout=aiohttp.ClientTimeout(total=5)) as resp:
                call_time = time.time() - call_start

                if resp.status == 200:
                    data = await resp.json()

                    # Get ALL OA locations
                    oa_locations = data.get("oa_locations", [])

                    if oa_locations:
                        # Add OpenAlex URL to locations if not already present
                        openalex_url = paper.get("pdf_url", "")
                        if openalex_url:
                            unpaywall_urls = [loc.get("url_for_pdf") or loc.get("url") for loc in oa_locations]
                            if openalex_url not in unpaywall_urls:
                                oa_locations.insert(0, {"url_for_pdf": openalex_url, "host_type": "publisher"})

                        # Save ALL locations
                        paper["oa_locations"] = oa_locations

                        # Set pdf_url to first location
                        first_loc = oa_locations[0]
                        if first_loc:
                            paper["pdf_url"] = first_loc.get("url_for_pdf") or first_loc.get("url")

                        paper["access"] = "OA"
                        return paper, call_time, True
                    else:
                        paper["access"] = "Non-OA"
                        return paper, call_time, False
                else:
                    paper["access"] = "Non-OA"
                    return paper, call_time, False

        except Exception:
            paper["access"] = "Non-OA"
            return paper, 0, False

async def check_unpaywall_parallel(papers: list[dict], concurrency=15):
    """Check all papers' OA status in parallel.

    Args:
        papers: List of paper dicts
        concurrency: Max concurrent requests (default 15)

    Returns:
        Tuple of (all_papers, all_call_times)
    """
    semaphore = asyncio.Semaphore(concurrency)
    all_call_times = []

    async with aiohttp.ClientSession(
        connector=aiohttp.TCPConnector(limit=concurrency),
        timeout=aiohttp.ClientTimeout(total=10)
    ) as session:
        # Create tasks for all papers
        tasks = [check_single_unpaywall(session, paper, semaphore) for paper in papers]

        # Execute all in parallel
        results = await asyncio.gather(*tasks, return_exceptions=True)

        # Process results
        for result in results:
            if isinstance(result, tuple):
                paper, call_time, is_oa = result
                if call_time > 0:
                    all_call_times.append(call_time)

    return papers, all_call_times

def check_oa_status(papers: list[dict], query_folder: Path = None) -> tuple[list[dict], list[dict]]:
    """Check OA status via Unpaywall and split into OA/non-OA.

    Args:
        papers: List of paper metadata dicts
        query_folder: Query folder for timer saving (optional)

    Returns:
        Tuple of (oa_papers, non_oa_papers)
    """
    # Run parallel checks
    papers, all_call_times = asyncio.run(check_unpaywall_parallel(papers, concurrency=15))

    # Split into OA and Non-OA
    oa_papers = [p for p in papers if p.get("access") == "OA"]
    non_oa_papers = [p for p in papers if p.get("access") == "Non-OA"]

    # Save Timer 3: Total Unpaywall API time
    if all_call_times and query_folder:
        total_unpaywall_time = sum(all_call_times)
        save_timer(query_folder, "2_unpaywall_total", total_unpaywall_time)

    return oa_papers, non_oa_papers

# =============================================================================
# EXCEL SAVER
# =============================================================================

def save_to_excel(all_papers: list[dict], filepath: Path):
    """Save papers to Excel with single worksheet.

    Args:
        all_papers: List of all paper metadata (OA + Non-OA combined)
        filepath: Path to save Excel file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Corpus"

        # Updated headers with "Access" after Breakpoint
        headers = ["PDF Number", "FileType", "Breakpoint", "Access", "Download Source", "Sources", "Author", "Title", "Year", "DOI", "DOI URL", "PDF URL", "Abstract"]

        # Headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)

        # Data
        for row, paper in enumerate(all_papers, 2):
            status = paper.get("status", "pending")

            ws.cell(row=row, column=1, value=paper.get("study_number", ""))

            # FileType column: PDF, HTML, XML, JSON, Unknown, Non-text/Unknown, or empty
            if status == "failed":
                filetype_value = ""  # No file downloaded
            elif status == "unreadable":
                filetype_value = "Non-text/Unknown"
            elif status == "unknown":
                filetype_value = "Unknown"
            elif status in ["html", "xml", "json"]:
                filetype_value = status.upper()  # "HTML", "XML", or "JSON"
            else:
                filetype_value = "PDF"  # Valid PDF
            ws.cell(row=row, column=2, value=filetype_value)

            # Breakpoint column: Download errors or unreadable files
            if status == "failed":
                # Check if timeout occurred
                if paper.get("timeout_occurred", False):
                    breakpoint_value = "Download Timeout"
                else:
                    breakpoint_value = "Download Failed"
            elif status == "unreadable":
                breakpoint_value = "Not Readable"
            else:
                breakpoint_value = ""  # No error
            ws.cell(row=row, column=3, value=breakpoint_value)

            # Access column: OA or Non-OA
            ws.cell(row=row, column=4, value=paper.get("access", ""))

            # Remaining columns
            ws.cell(row=row, column=5, value=paper.get("download_source", ""))  # Which mirror succeeded
            ws.cell(row=row, column=6, value=paper.get("sources", ""))
            ws.cell(row=row, column=7, value=paper.get("author", ""))
            ws.cell(row=row, column=8, value=paper.get("title", ""))
            ws.cell(row=row, column=9, value=paper.get("year", ""))
            ws.cell(row=row, column=10, value=paper.get("doi", ""))
            ws.cell(row=row, column=11, value=paper.get("doi_url", ""))
            ws.cell(row=row, column=12, value=paper.get("pdf_url", ""))
            ws.cell(row=row, column=13, value=paper.get("abstract", "")[:32000])

            # Apply color based on status
            if status == "downloaded":
                fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Green (valid PDF)
                for c in range(1, 14):
                    ws.cell(row=row, column=c).fill = fill
            elif status in ["html", "xml", "json", "unknown"]:
                fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange (non-PDF format or unknown)
                for c in range(1, 14):
                    ws.cell(row=row, column=c).fill = fill
            elif status == "failed" or status == "unreadable":
                fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Red (download failed or unreadable)
                for c in range(1, 14):
                    ws.cell(row=row, column=c).fill = fill

        # Column widths
        ws.column_dimensions['A'].width = 10  # PDF Number
        ws.column_dimensions['B'].width = 16  # FileType
        ws.column_dimensions['C'].width = 15  # Breakpoint
        ws.column_dimensions['D'].width = 10  # Access
        ws.column_dimensions['E'].width = 20  # Download Source
        ws.column_dimensions['F'].width = 15  # Sources
        ws.column_dimensions['G'].width = 25  # Author
        ws.column_dimensions['H'].width = 50  # Title
        ws.column_dimensions['I'].width = 8   # Year
        ws.column_dimensions['J'].width = 25  # DOI
        ws.column_dimensions['K'].width = 35  # DOI URL
        ws.column_dimensions['L'].width = 40  # PDF URL
        ws.column_dimensions['M'].width = 80  # Abstract

        # Save
        wb.save(filepath)

    except ImportError:
        print("  ERROR: openpyxl not installed")
        print("  Install with: pip install openpyxl")

# =============================================================================
# PDF DOWNLOADER WITH PROGRESS BAR
# =============================================================================

def format_time(seconds):
    """Format seconds into HH:MM:SS."""
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    secs = int(seconds % 60)
    if hours > 0:
        return f"{hours:02d}:{minutes:02d}:{secs:02d}"
    else:
        return f"{minutes:02d}:{secs:02d}"

def format_size(bytes_val):
    """Format bytes into MB."""
    return f"{bytes_val / 1024 / 1024:.2f} MB"

def download_with_progress(url, filepath, study_num, paper_type, max_retries=2):
    """Download file with progress tracking and retry logic.

    Args:
        url: URL to download from
        filepath: Path to save file
        study_num: Study number for tracking
        paper_type: "OA" or "Non-OA"
        max_retries: Maximum number of download attempts (default: 2)

    Returns:
        tuple: (success: bool, size: int, file_format: str or None, source_name: str or None, timeout: bool)
               file_format is "HTML", "XML", "JSON", or None (for PDF/unknown)
               source_name is extracted from URL (e.g., "nature.com", "doaj.org")
               timeout is True if download timed out
    """
    for attempt in range(max_retries):
        try:
            response = requests.get(url, stream=True, timeout=15)
            if response.status_code != 200:
                continue  # Try again

            total_size = int(response.headers.get('content-length', 0))
            downloaded = 0
            first_chunk = None

            with open(filepath, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        if first_chunk is None:
                            first_chunk = chunk
                        f.write(chunk)
                        downloaded += len(chunk)

            # Detect file format (PDF, HTML, XML, JSON)
            file_format = None
            if first_chunk:
                header = first_chunk[:100].lower()
                stripped_header = header.strip()

                # Detect specific formats
                if header.startswith(b'<!doctype html') or header.startswith(b'<html'):
                    file_format = "HTML"
                elif header.startswith(b'<?xml') or header.startswith(b'<article'):
                    file_format = "XML"
                elif stripped_header.startswith(b'{') or stripped_header.startswith(b'['):
                    file_format = "JSON"
                elif header.startswith(b'%pdf'):
                    file_format = None  # Valid PDF - no special status needed
                else:
                    file_format = "UNKNOWN"  # Unknown format

            # Extract source name from URL
            source_name = None
            try:
                from urllib.parse import urlparse
                parsed = urlparse(url)
                source_name = parsed.netloc.replace('www.', '')
            except:
                source_name = "unknown"

            return True, downloaded, file_format, source_name, False
        except requests.exceptions.Timeout:
            # Timeout - return timeout flag
            if attempt == max_retries - 1:
                return False, 0, None, None, True
            time.sleep(1)  # Wait 1 second before retry
        except Exception as e:
            if attempt == max_retries - 1:
                return False, 0, None, None, False
            time.sleep(1)  # Wait 1 second before retry

    return False, 0, None, None, False


def download_with_multiple_sources(locations, filepath, study_num, paper_type, mirror_speeds):
    """Download file trying multiple OA sources sorted by measured speed.

    NEW STRATEGY: Try ALL mirrors and prefer PDF over HTML/XML.
    - Try all mirrors (sorted by speed)
    - Collect all successful downloads
    - Return PDF if any mirror provides it
    - Only return HTML/XML if NO PDF available

    Args:
        locations: List of OA location dicts from UnPaywall
        filepath: Path to save file
        study_num: Study number for tracking
        paper_type: "OA" or "Non-OA"
        mirror_speeds: Dict mapping domain -> speed (bytes/sec)

    Returns:
        tuple: (success: bool, size: int, file_format: str or None, source_name: str or None, timeout: bool)
    """
    if not locations:
        return False, 0, None, None, False

    # Sort locations by measured speed (fastest first)
    def get_speed(loc):
        url = loc.get("url_for_pdf") or loc.get("url") or ""
        domain = extract_domain(url)
        return mirror_speeds.get(domain, 0)

    sorted_locations = sorted(locations, key=get_speed, reverse=True)

    # Try mirrors one at a time until we find a PDF
    best_non_pdf = None  # Fallback if no PDF found

    for loc in sorted_locations:
        url = loc.get("url_for_pdf") or loc.get("url")
        if not url:
            continue

        success, size, file_format, source_name, timeout = download_with_progress(url, filepath, study_num, paper_type, max_retries=1)

        if success:
            # Check if it's a PDF
            if file_format in ["HTML", "XML", "JSON"]:
                # Not a PDF - delete and save as fallback
                if filepath.exists():
                    filepath.unlink()
                if best_non_pdf is None:
                    best_non_pdf = (success, size, file_format, source_name, timeout)
                # Try next mirror
                continue
            elif file_format == "UNKNOWN":
                # Unknown format - delete and save as fallback
                if filepath.exists():
                    filepath.unlink()
                if best_non_pdf is None:
                    best_non_pdf = (success, size, file_format, source_name, timeout)
                # Try next mirror
                continue
            else:
                # PDF found! Stop trying other mirrors
                return success, size, file_format, source_name, timeout

    # No PDF found from any mirror - return best non-PDF if we have one
    if best_non_pdf:
        # Re-download the fallback (we deleted it earlier)
        for loc in sorted_locations:
            url = loc.get("url_for_pdf") or loc.get("url")
            if not url:
                continue
            success, size, file_format, source_name, timeout = download_with_progress(url, filepath, study_num, paper_type, max_retries=1)
            if success and file_format == best_non_pdf[2]:
                return success, size, file_format, source_name, timeout

    return False, 0, None, None, False

def download_pdfs(all_papers: list[dict], download_folder: Path,
                  api_start_time: float, mirror_speeds: dict, query_folder: Path = None,
                  on_pdf_validated=None):
    """Download PDFs to staging folder, validate, then move to Corpus PDFs.

    Args:
        all_papers: List of all papers (OA + Non-OA combined)
        download_folder: Final destination directory for validated PDFs
        api_start_time: Time when API calls started (for runtime tracking)
        mirror_speeds: Dict mapping domain -> speed (bytes/sec) from testing
        query_folder: Query folder for timer saving (optional)
        on_pdf_validated: Optional callback(pdf_num, pdf_path) called when each PDF is validated
    """
    # Create staging folder for temporary downloads
    staging_folder = download_folder.parent / "Staging Downloads"
    staging_folder.mkdir(exist_ok=True)

    # Count total downloads needed (OA and Non-OA)
    downloadable = [p for p in all_papers if p.get("status") == "pending" and (p.get("pdf_url") or p.get("doi"))]

    total_downloads = len(downloadable)

    if total_downloads == 0:
        return

    print(f"Unique papers for download: {total_downloads}")

    # Download tracking
    downloaded_count = 0
    failed_count = 0
    total_bytes = 0
    download_start_time = time.time()

    # Create stdout lock for thread-safe progress bar (Windows console requires this)
    stdout_lock = threading.Lock()

    def update_progress():
        """Update progress bar display."""
        elapsed = time.time() - download_start_time
        api_elapsed = time.time() - api_start_time

        # Calculate progress
        progress = downloaded_count / total_downloads if total_downloads > 0 else 0
        percent = progress * 100

        # Calculate speed
        if elapsed > 0:
            mbps = (total_bytes * 8 / 1024 / 1024) / elapsed  # Megabits per second
            avg_time_per_file = elapsed / downloaded_count if downloaded_count > 0 else 0
            remaining_files = total_downloads - downloaded_count
            eta = avg_time_per_file * remaining_files if downloaded_count > 0 else 0
        else:
            mbps = 0
            eta = 0

        # Progress bar (50 chars) - ASCII for Windows compatibility
        bar_length = 50
        filled = int(bar_length * progress)
        bar = '=' * filled + '-' * (bar_length - filled)

        # Print on same line
        runtime_str = format_time(api_elapsed)
        eta_str = format_time(eta) if eta > 0 else "--:--"

        print(f"\r[{bar}] {percent:5.1f}% | {downloaded_count}/{total_downloads} | "
              f"{mbps:6.2f} Mbps | ETA: {eta_str} | Runtime: {runtime_str}", end='')

    # Import queue manager
    import sys
    from pathlib import Path
    sys.path.insert(0, str(Path(__file__).parent.parent / "QueueManager"))
    from QueueManager import DownloadValidationManager

    # Initialize queue manager
    manager = DownloadValidationManager(
        staging_folder=staging_folder,
        corpus_folder=download_folder,
        mirror_speeds=mirror_speeds,
        max_download_workers=30,  # Increased from 20 to reduce timeout overhead
        max_validation_workers=10,
        on_validated_callback=on_pdf_validated  # Pass callback for inter-pipeline streaming
    )

    # Create download wrapper function
    def download_wrapper(url, filepath, temp_idx):
        """Wrapper to match expected signature for queue manager."""
        return download_with_progress(url, filepath, temp_idx, "OA")

    # Progress callback (track last values to avoid redundant prints)
    last_printed = {"downloaded": -1, "failed": -1}

    def progress_callback(downloaded, failed, validated):
        nonlocal downloaded_count, failed_count

        # Only print if values changed
        if downloaded == last_printed["downloaded"] and failed == last_printed["failed"]:
            return

        last_printed["downloaded"] = downloaded
        last_printed["failed"] = failed
        downloaded_count = downloaded
        failed_count = failed

        elapsed = time.time() - download_start_time
        api_elapsed = time.time() - api_start_time

        progress = (downloaded + failed) / total_downloads if total_downloads > 0 else 0
        percent = progress * 100

        if elapsed > 0:
            mbps = (manager.total_bytes * 8 / 1024 / 1024) / elapsed
            avg_time_per_file = elapsed / (downloaded + failed) if (downloaded + failed) > 0 else 0
            remaining_files = total_downloads - (downloaded + failed)
            eta = avg_time_per_file * remaining_files if (downloaded + failed) > 0 else 0
        else:
            mbps = 0
            eta = 0

        bar_length = 50
        filled = int(bar_length * progress)
        bar = '=' * filled + '-' * (bar_length - filled)

        runtime_str = format_time(api_elapsed)
        eta_str = format_time(eta) if eta > 0 else "--:--"

        # Thread-safe progress update (simple newline logging for Claude Code TUI compatibility)
        with stdout_lock:
            # Print progress as a new line (no \r) to avoid TUI rendering issues
            print(f"[{bar}] {percent:5.1f}% | Downloaded: {downloaded}/{total_downloads} | Failed: {failed} | "
                  f"{mbps:6.2f} Mbps | ETA: {eta_str} | Runtime: {runtime_str}")

    # Process downloads with rolling window of workers
    all_papers = manager.process_papers(downloadable, download_wrapper, progress_callback)

    # Get final stats
    stats = manager.get_stats()
    downloaded_count = stats["valid_pdfs"]
    failed_count = stats["failed_downloads"] + stats["paywall_files"] + stats["unreadable_files"]
    total_bytes = stats["total_bytes"]

    # OLD SEQUENTIAL LOOP (commented out for reference)
    """
    for paper in downloadable:
        temp_idx = paper.get("temp_index", "unknown")
        filename = f"temp_{temp_idx}.pdf"
        staging_path = staging_folder / filename  # Download to staging first
        final_path = download_folder / filename   # Move here after validation
        access_type = paper.get("access", "")

        # Download based on access type
        if access_type == "OA":
            # Use multi-source download with all OA locations
            locations = paper.get("oa_locations", [])
            if not locations:
                # Fallback to single URL if no locations saved
                pdf_url = paper.get("pdf_url", "")
                if pdf_url:
                    locations = [{"url_for_pdf": pdf_url}]

            success, size, file_format, source_name = download_with_multiple_sources(locations, staging_path, temp_idx, "OA", mirror_speeds)
        else:
            # Non-OA: Try SciHub mirrors
            doi = paper.get("doi", "")
            if not doi:
                paper["status"] = "failed"
                failed_count += 1
                update_progress()
                continue

            scihub_urls = [
                f"https://sci-hub.se/{doi}",
                f"https://sci-hub.st/{doi}",
                f"https://sci-hub.ru/{doi}"
            ]

            success = False
            size = 0
            file_format = None
            source_name = None
            for scihub_url in scihub_urls:
                success, size, file_format, source_name = download_with_progress(scihub_url, staging_path, temp_idx, "Non-OA")
                if success:
                    break

        # Handle download result
        if success:
            # Validate file in staging folder
            try:
                import fitz
                doc = fitz.open(staging_path)
                page_count = len(doc)

                # Check if file is readable (has pages and can extract text)
                # Try to extract text from first page to verify readability
                has_text = False
                try:
                    if page_count > 0:
                        first_page_text = doc[0].get_text()
                        has_text = len(first_page_text.strip()) > 0
                except:
                    has_text = False
                finally:
                    doc.close()

                if page_count == 0 or (file_format == "UNKNOWN" and not has_text) or file_format in ["HTML", "XML", "JSON"]:
                    # No pages OR unknown format with no text OR paywall HTML/XML/JSON - unreadable
                    if file_format in ["HTML", "XML", "JSON"]:
                        paper["status"] = "paywall"  # HTML/XML/JSON paywall pages
                    else:
                        paper["status"] = "unreadable"
                    staging_path.unlink()  # Delete from staging
                    failed_count += 1
                else:
                    # File has pages/text - MOVE to Corpus PDFs
                    staging_path.rename(final_path)

                    if file_format:
                        if file_format == "UNKNOWN":
                            paper["status"] = "unknown"  # Unknown format but has text
                        else:
                            paper["status"] = file_format.lower()  # "html", "xml", or "json"
                    else:
                        paper["status"] = "downloaded"  # Valid PDF

                    paper["download_source"] = source_name
                    downloaded_count += 1
                    total_bytes += size

            except Exception:
                # PyMuPDF can't open file - unreadable
                paper["status"] = "unreadable"
                if staging_path.exists():
                    staging_path.unlink()  # Delete from staging
                failed_count += 1
        else:
            paper["status"] = "failed"  # Download failed
            paper["download_source"] = None
            failed_count += 1

        update_progress()
    """
    # END OLD SEQUENTIAL LOOP

    # Final progress update
    print()  # New line after progress bar

    # Summary
    elapsed = time.time() - download_start_time

    # Timer 4: Total download time
    if query_folder:
        save_timer(query_folder, "3_corpus_download_total", elapsed)

    # NOTE: study_number assignment and file renaming now handled by queue manager

    # Update JSON with status changes (after downloads complete)
    if downloaded_count > 0 or failed_count > 0:
        save_corpus_json(all_papers, query_folder)

    # Clean up staging folder (Windows may have file locks, so retry with delay)
    import shutil
    import gc
    if staging_folder.exists():
        # Use PowerShell Remove-Item -Force for reliable Windows deletion
        import subprocess
        result = subprocess.run(
            ['powershell', '-Command', f'Remove-Item -Path "{staging_folder}" -Force -Recurse -ErrorAction SilentlyContinue'],
            capture_output=True,
            text=True,
            timeout=5
        )

# =============================================================================
# MAIN ORCHESTRATOR
# =============================================================================

def main(on_pdf_validated=None, on_query_folder_created=None):
    """Main orchestrator.

    Args:
        on_pdf_validated: Optional callback(pdf_num, pdf_path) called when each PDF is validated
        on_query_folder_created: Optional callback(query_folder) called immediately after query folder is created
    """
    # Track API start time
    api_start_time = time.time()

    # Get config file path
    if len(sys.argv) > 1:
        config_path = Path(sys.argv[1])
    else:
        config_path = DEFAULT_CONFIG

    # Parse config
    config = parse_config(config_path)

    # Create output folders
    timestamp = datetime.now().strftime("%m-%d-%y-%H%M")
    query_folder = OUTPUT_DIR / f"{timestamp} Query"
    download_folder = query_folder / "Corpus PDFs"

    download_folder.mkdir(parents=True, exist_ok=True)

    # Notify other pipelines that query folder is ready (allows Pipeline 2 to start immediately)
    if on_query_folder_created:
        on_query_folder_created(query_folder)

    # Build query
    url = build_openalex_query(config)

    # Call API - pass query_folder for timer saving
    max_results = config.get('max_results', None)
    if max_results:
        print(f"Max results limit: {max_results}")
    else:
        print("Max results limit: Unlimited")

    papers = call_openalex(url, max_results=max_results, query_folder=query_folder)

    if not papers:
        return

    # Deduplicate
    papers = deduplicate(papers)
    print(f"Unique papers for download: {len(papers)}")

    # Check OA status (pass query_folder for timer saving)
    oa_papers, non_oa_papers = check_oa_status(papers, query_folder=query_folder)

    # Combine all papers into single list (sorted by temp_index for now)
    all_papers = sorted(oa_papers + non_oa_papers, key=lambda p: p.get("temp_index", 0))

    # Save to JSON (will be converted to Excel in Pipeline 5)
    save_corpus_json(all_papers, query_folder)

    # Test mirror speeds (before downloading) - only for OA papers
    mirror_speeds = rank_mirrors_by_speed(oa_papers)

    # Download PDFs (with progress tracking, using measured mirror speeds)
    download_pdfs(all_papers, download_folder, api_start_time, mirror_speeds, query_folder=query_folder, on_pdf_validated=on_pdf_validated)

    # Return query folder path for Pipeline orchestrator
    return query_folder

if __name__ == "__main__":
    main()
