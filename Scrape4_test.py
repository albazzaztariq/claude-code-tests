import os
import re
import sys
from pathlib import Path
import fitz  # PyMuPDF
import requests
import openpyxl
from openpyxl import Workbook
import pdfplumber
import camelot
import pytest
from azure.ai.formrecognizer import DocumentAnalysisClient
from azure.core.credentials import AzureKeyCredential
from PIL import Image
import io

# Nougat OCR for table extraction
# Fix for albumentations v1.4.0+ compatibility with nougat-ocr
try:
    import albumentations as alb

    # Monkeypatch ImageCompression to accept old API (quality as first positional arg)
    _original_ImageCompression = alb.ImageCompression
    class ImageCompression(_original_ImageCompression):
        def __init__(self, quality_or_type=95, p=0.5, **kwargs):
            # If first arg is int, treat as quality bounds and default to 'jpeg'
            if isinstance(quality_or_type, int):
                # New API: quality_lower, quality_upper, compression_type
                super().__init__(quality_lower=quality_or_type, quality_upper=quality_or_type,
                               compression_type='jpeg', p=p, **kwargs)
            else:
                # If first arg is string, treat as compression_type (new API call)
                super().__init__(compression_type=quality_or_type, p=p, **kwargs)
    alb.ImageCompression = ImageCompression

    # Monkeypatch GaussNoise to accept old API (std as single int)
    _original_GaussNoise = alb.GaussNoise
    class GaussNoise(_original_GaussNoise):
        def __init__(self, std_or_limit=20, p=0.5, **kwargs):
            # If first arg is int/float, convert to std_range tuple
            if isinstance(std_or_limit, (int, float)):
                # New API expects std_range as tuple (min, max)
                super().__init__(std_range=(0, std_or_limit), p=p, **kwargs)
            else:
                # If already a tuple, pass as std_range
                super().__init__(std_range=std_or_limit, p=p, **kwargs)
    alb.GaussNoise = GaussNoise

except ImportError:
    pass  # albumentations not installed, skip patch

try:
    from nougat import NougatModel
    from nougat.utils.checkpoint import get_checkpoint
    NOUGAT_AVAILABLE = True
except ImportError:
    NOUGAT_AVAILABLE = False
    print("Warning: Nougat OCR not available. Install with: pip install nougat-ocr")

# ================== CONFIGURATION ==================
OLLAMA_URL = "http://localhost:11434/api/generate"
OLLAMA_MODEL = "gemma2:2b"
AZURE_VISION_KEY = os.getenv("AZURE_VISION_KEY", "YOUR_AZURE_VISION_KEY_HERE")
AZURE_VISION_ENDPOINT = os.getenv("AZURE_VISION_ENDPOINT", "YOUR_AZURE_ENDPOINT_HERE")
BASE_DIR = Path(
    r"C:\Users\azt12\OneDrive\Documents\Wrestling Robe\Materials Science - Wickability\Studies for Analysis by LLM AI\Datafiles & Python Scripts"
)
INPUT_EXCEL = BASE_DIR / "Simplified Table Format.xlsx"
OUTPUT_EXCEL = BASE_DIR / "Sample-Level-Metadata.xlsx"
PDF_FOLDER = Path(
    r"C:\Users\azt12\OneDrive\Documents\Wrestling Robe\Materials Science - Wickability\Studies for Analysis by LLM AI"
)

# ================== HELPER FUNCTIONS (not tests) ==================
def call_local_llm(prompt: str) -> str:
    """Call local Ollama LLM."""
    payload = {
        "model": OLLAMA_MODEL,
        "prompt": prompt,
        "stream": False,
    }
    resp = requests.post(OLLAMA_URL, json=payload, timeout=240)
    resp.raise_for_status()
    data = resp.json()
    return data.get("response", "")

def extract_title_with_formatting(pdf_path: str) -> str:
    """
    Extract title by finding the largest font size in first page.
    """
    doc = fitz.open(pdf_path)
    page = doc[0]
    blocks = page.get_text("dict")["blocks"]
    text_items = []
    for block in blocks:
        if "lines" in block:
            for line in block["lines"]:
                for span in line["spans"]:
                    text = span["text"].strip()
                    size = span["size"]
                    if text and len(text) > 3:
                        text_items.append(
                            {
                                "text": text,
                                "size": size,
                                "y": span["bbox"][1],
                                "x": span["bbox"][0],
                            }
                        )
    if not text_items:
        doc.close()
        return None
    text_items.sort(key=lambda x: (x["y"], x["x"]))
    skip_phrases = [
        "preprint",
        "not peer reviewed",
        "copyright",
        "license",
        "doi:",
        "arxiv",
        "biorxiv",
        "medrxiv",
        "journal of",
        "page ",
        "volume ",
        "issue ",
    ]
    candidates = []
    for item in text_items[:30]:
        if any(phrase in item["text"].lower() for phrase in skip_phrases):
            continue
        if len(item["text"]) < 10 and item["size"] < 15:
            continue
        candidates.append(item)
    if not candidates:
        doc.close()
        return None
    max_size = max(c["size"] for c in candidates)
    title_parts = []
    last_y = -999
    for item in candidates:
        if item["size"] >= max_size - 2:
            if abs(item["y"] - last_y) > 5:
                if title_parts and abs(item["y"] - last_y) > 20:
                    break
            title_parts.append(item["text"])
            last_y = item["y"]
        elif title_parts:
            break
    doc.close()
    if title_parts:
        full_title = " ".join(title_parts)
        full_title = re.sub(r"\s+", " ", full_title)
        return full_title.strip()
    return None

def extract_first_author_with_formatting(pdf_path: str) -> str:
    """
    Extract first author by finding text after the title.
    """
    doc = fitz.open(pdf_path)
    page = doc[0]
    blocks = page.get_text("dict")["blocks"]
    text_items = []
    for block in blocks:
        if "lines" in block:
            for line in block["lines"]:
                for span in line["spans"]:
                    text = span["text"].strip()
                    if text:
                        text_items.append(
                            {
                                "text": text,
                                "size": span["size"],
                                "y": span["bbox"][1],
                            }
                        )
    if not text_items:
        doc.close()
        return None
    text_items.sort(key=lambda x: x["y"])
    # Find max font size (title)
    max_size = max(item["size"] for item in text_items[:30])
    # Find where title ends
    title_end_idx = 0
    for i, item in enumerate(text_items[:30]):
        if item["size"] >= max_size - 2:
            title_end_idx = i + 1
    # Look for author in next 10 items
    for i in range(title_end_idx, min(title_end_idx + 10, len(text_items))):
        text = text_items[i]["text"]
        # Skip institutional lines
        if any(
            word in text.lower()
            for word in ["university", "school", "department", "@", "http"]
        ):
            continue
        # Look for any capitalized word that's 3+ letters (last name)
        words = text.split()
        for j, word in enumerate(words):
            # Clean the word
            clean_word = re.sub(r"[^A-Za-z]", "", word)
            # If it's capitalized and 3+ letters, it's probably a name
            if clean_word and clean_word[0].isupper() and len(clean_word) >= 3:
                # If next word is also capitalized, this is first name, take next word (last name)
                if j + 1 < len(words):
                    next_word = re.sub(r"[^A-Za-z]", "", words[j + 1])
                    if next_word and next_word[0].isupper() and len(next_word) >= 3:
                        doc.close()
                        return next_word  # Return last name
                # Otherwise just return this word
                doc.close()
                return clean_word
    doc.close()
    return None

def extract_year_from_text(text: str) -> int:
    """
    Extract publication year.
    """
    year_pattern = r"\b(19[9]\d|20[0-2]\d)\b"
    first_part = text[:6000]
    keywords_patterns = [
        (r"published[:\s]+.*?(\d{4})", 1),
        (r"received[:\s]+.*?(\d{4})", 1),
        (r"accepted[:\s]+.*?(\d{4})", 1),
        (r"copyright[:\s]+.*?(\d{4})", 1),
        (r"©[:\s]*(\d{4})", 1),
        (r"(\d{4})\s+published", 1),
        (r"online:?\s+.*?(\d{4})", 1),
        (r"\b(\d{4})\s+elsevier", 1),
        (r"\b(\d{4})\s+wiley", 1),
        (r"\b(\d{4})\s+springer", 1),
    ]
    for pattern, group in keywords_patterns:
        matches = re.finditer(pattern, first_part, re.IGNORECASE)
        for match in matches:
            year_str = match.group(group)
            year_int = int(year_str)
            if 1990 <= year_int <= 2025:
                return year_int
    lines = first_part.split("\n")
    for i, line in enumerate(lines[:50]):
        if re.match(r"^\s*\d{4}\s*$", line):
            year_int = int(line.strip())
            if 2000 <= year_int <= 2025:
                return year_int
    all_years = re.findall(year_pattern, first_part)
    if all_years:
        valid_years = [int(y) for y in all_years if 1990 <= int(y) <= 2025]
        if valid_years:
            from collections import Counter
            year_counts = Counter(valid_years)
            multiple_years = [y for y, count in year_counts.items() if count > 1]
            if multiple_years:
                return max(multiple_years)
            return max(valid_years)
    return None

def extract_table_row_count(pdf_path: str, table_number: int) -> int:
    """
    Extract row count from a specific table number in the PDF.
    Returns the number of data rows (excluding header) if found, None otherwise.
    Tries: Camelot, Azure Computer Vision, Nougat OCR (in that order).
    """
    print(f"\n    --- Extracting Table {table_number} ---")

    best_count = None
    best_method = None

    # Try CAMELOT
    print(f"\n    Trying CAMELOT...")
    try:
        tables_lattice = camelot.read_pdf(pdf_path, pages="all", flavor="lattice")
        tables_stream = camelot.read_pdf(pdf_path, pages="all", flavor="stream")
        all_camelot_tables = list(tables_lattice) + list(tables_stream)

        if table_number <= len(all_camelot_tables):
            table = all_camelot_tables[table_number - 1]
            df = table.df
            rows = len(df) - 1  # Exclude header
            print(f"    Camelot: Found Table {table_number} with {rows} data rows")

            # Show FULL table with ACTUAL CELL DATA
            print(f"    Extracted table data (ALL ROWS):")
            for idx in range(len(df)):
                row_data = df.iloc[idx].tolist()
                row_str = " | ".join([str(cell)[:50] for cell in row_data])
                print(f"      Row {idx}: {row_str}")

            if rows > 0 and not best_count:
                best_count = rows
                best_method = "Camelot"
        else:
            print(f"    Camelot: Table {table_number} not found (only {len(all_camelot_tables)} tables)")
    except Exception as e:
        print(f"    Camelot ERROR: {e}")

    # Try AZURE COMPUTER VISION (Document Intelligence)
    print(f"\n    Trying AZURE COMPUTER VISION...")
    try:
        azure_result = extract_table_with_azure(pdf_path, table_number)
        if azure_result:
            rows, table_data = azure_result
            print(f"    Azure: Found Table {table_number} with {rows} data rows")

            # Show FULL table with ACTUAL CELL DATA
            print(f"    Extracted table data (ALL ROWS):")
            for idx, row in enumerate(table_data):
                row_str = " | ".join([str(cell)[:50] for cell in row])
                print(f"      Row {idx}: {row_str}")

            if rows > 0 and not best_count:
                best_count = rows
                best_method = "Azure Computer Vision"
        else:
            print(f"    Azure: Table {table_number} not found or no rows detected")
    except Exception as e:
        print(f"    Azure ERROR: {e}")

    # Try NOUGAT OCR
    if NOUGAT_AVAILABLE:
        print(f"\n    Trying NOUGAT OCR...")
        try:
            nougat_result = extract_table_with_nougat(pdf_path, table_number)
            if nougat_result:
                rows, table_data = nougat_result
                print(f"    Nougat: Found Table {table_number} with {rows} data rows")

                # Show FULL table with ACTUAL CELL DATA
                print(f"    Extracted table data (ALL ROWS):")
                for idx, row in enumerate(table_data):
                    row_str = " | ".join([str(cell)[:50] for cell in row])
                    print(f"      Row {idx}: {row_str}")

                if rows > 0 and not best_count:
                    best_count = rows
                    best_method = "Nougat OCR"
            else:
                print(f"    Nougat: Table {table_number} not found or no rows detected")
        except Exception as e:
            print(f"    Nougat ERROR: {e}")

    if best_count:
        print(f"\n    ✓ Best result: {best_count} rows from {best_method}")
        return best_count
    else:
        print(f"\n    ✗ Could not extract Table {table_number} from any method")
        return None

def extract_table_with_azure(pdf_path: str, table_number: int):
    """
    Use Azure Document Intelligence to extract tables from PDF.
    Extracts table area as compressed image to avoid file size limits.
    Returns tuple of (row_count, table_data) if found, None otherwise.
    """
    try:
        # First, use Camelot to find where the table is located
        print(f"    Finding table location with Camelot...")
        tables_lattice = camelot.read_pdf(pdf_path, pages="all", flavor="lattice")
        tables_stream = camelot.read_pdf(pdf_path, pages="all", flavor="stream")
        all_tables = list(tables_lattice) + list(tables_stream)

        if table_number > len(all_tables):
            print(f"    Table {table_number} not found by Camelot")
            return None

        # Get the table and its location
        camelot_table = all_tables[table_number - 1]

        # VERIFY this is actually a table with structure
        # Check Camelot's accuracy and ensure it has multiple rows/columns
        df = camelot_table.df
        accuracy = camelot_table.parsing_report.get('accuracy', 0) if hasattr(camelot_table, 'parsing_report') else 0

        print(f"    Camelot accuracy: {accuracy:.1f}%, Rows: {len(df)}, Cols: {len(df.columns)}")

        camelot_failed = False

        # Check if Camelot failed to properly detect the table
        if accuracy < 60 or len(df) < 3 or len(df.columns) < 3:
            print(f"    ⚠ Camelot has low accuracy or insufficient table structure")
            camelot_failed = True
        else:
            # Check if table has actual data (at least some cells with numbers)
            # Real tables have numeric data, not just text or empty cells
            all_values = df.values.flatten()
            numeric_cells = sum(1 for val in all_values if str(val).strip() and any(c.isdigit() for c in str(val)))
            total_cells = len(all_values)
            numeric_ratio = numeric_cells / total_cells if total_cells > 0 else 0

            print(f"    Table has {numeric_cells}/{total_cells} cells with numbers ({numeric_ratio*100:.1f}%)")

            # Require at least 20% of cells to have numeric content
            if numeric_ratio < 0.2:
                print(f"    ⚠ Too few numeric cells")
                camelot_failed = True

        # If Camelot failed, try searching for "Table X" in the PDF and extract that area
        if camelot_failed:
            print(f"    → Falling back to text search for 'Table {table_number}'")
            doc = fitz.open(pdf_path)

            # Search all pages for "Table {table_number}"
            table_found = False
            for page_idx in range(len(doc)):
                page = doc[page_idx]
                text = page.get_text()

                # Look for "Table X" in the text
                if re.search(rf'\bTable\s+{table_number}\b', text, re.IGNORECASE):
                    print(f"    ✓ Found 'Table {table_number}' on page {page_idx + 1}")

                    # Search for the text location to get approximate position
                    search_results = page.search_for(f"Table {table_number}")

                    if search_results:
                        # Get the position of "Table X" text
                        table_caption_rect = search_results[0]

                        # Extract a large area below the caption (likely contains the table)
                        # Take from caption down to 60% of page height
                        page_rect = page.rect
                        extraction_rect = fitz.Rect(
                            page_rect.x0,  # Left edge of page
                            table_caption_rect.y0,  # Start at table caption
                            page_rect.x1,  # Right edge of page
                            min(table_caption_rect.y0 + 400, page_rect.y1)  # 400 points below or bottom of page
                        )

                        rect = extraction_rect
                        page_num = page_idx
                        table_found = True
                        print(f"    → Extracting area around 'Table {table_number}' caption")
                        break

            doc.close()

            if not table_found:
                print(f"    ✗ Could not find 'Table {table_number}' text in PDF")
                return None

            # Reopen for extraction
            doc = fitz.open(pdf_path)
            page = doc[page_num]
        else:
            print(f"    ✓ Verified as real data table")
            page_num = camelot_table.page - 1  # Camelot uses 1-indexed, fitz uses 0-indexed

            # Get table bounding box from Camelot
            # Camelot bbox format: (x1, y1, x2, y2) in PDF coordinates
            table_bbox = camelot_table._bbox if hasattr(camelot_table, '_bbox') else None

            # Open PDF with PyMuPDF to extract table area as image
            doc = fitz.open(pdf_path)
            page = doc[page_num]

            # If we have bbox, use it; otherwise use full page
            if table_bbox:
                # Convert Camelot bbox to fitz rect
                rect = fitz.Rect(table_bbox)
            else:
                # Use full page
                rect = page.rect

        # Verify the extracted area contains "Table {table_number}" text
        text_in_area = page.get_text("text", clip=rect)
        if not re.search(rf'\bTable\s+{table_number}\b', text_in_area, re.IGNORECASE):
            print(f"    ✗ Extracted area doesn't contain 'Table {table_number}' - wrong region")
            doc.close()
            return None

        print(f"    ✓ Verified area contains 'Table {table_number}' text")

        # Render the table area as image with zoom for quality
        zoom = 2.0  # 2x resolution
        mat = fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=mat, clip=rect)

        # Convert to PIL Image
        img_data = pix.tobytes("png")
        img = Image.open(io.BytesIO(img_data))

        # Compress as JPEG to reduce size
        img_buffer = io.BytesIO()
        img.convert('RGB').save(img_buffer, format='JPEG', quality=85, optimize=True)
        img_buffer.seek(0)
        compressed_size = len(img_buffer.getvalue())

        # Save image to disk for debugging
        # Extract study number from PDF filename (e.g., "Study3.pdf" → 3)
        pdf_dir = Path(pdf_path).parent
        pdf_name = Path(pdf_path).stem
        study_match = re.search(r'(\d+)', pdf_name)
        study_number = study_match.group(1) if study_match else "unknown"

        # Format: 1_study_{study_number}_table{table_number}.jpeg
        debug_image_path = pdf_dir / f"1_study_{study_number}_table{table_number}.jpeg"
        with open(debug_image_path, 'wb') as f:
            f.write(img_buffer.getvalue())
        img_buffer.seek(0)  # Reset buffer position after writing

        print(f"    Table image extracted: {compressed_size / 1024:.1f} KB")
        print(f"    Saved to: {debug_image_path}")
        doc.close()

        # Send compressed image to Azure
        credential = AzureKeyCredential(AZURE_VISION_KEY)
        client = DocumentAnalysisClient(
            endpoint=AZURE_VISION_ENDPOINT,
            credential=credential
        )

        poller = client.begin_analyze_document("prebuilt-layout", document=img_buffer)
        result = poller.result()

        print(f"    Azure found {len(result.tables) if result.tables else 0} tables in extracted image")

        # Azure should find 1 table (the one we extracted)
        if result.tables and len(result.tables) > 0:
            table = result.tables[0]  # Use first table found in the image

            print(f"    Table structure: {table.row_count} rows x {table.column_count} columns, {len(table.cells)} cells")

            # Build table data structure
            rows_dict = {}
            for cell in table.cells:
                row_idx = cell.row_index
                if row_idx not in rows_dict:
                    rows_dict[row_idx] = {}
                rows_dict[row_idx][cell.column_index] = cell.content

            # Convert to list of lists
            table_data = []
            for row_idx in sorted(rows_dict.keys()):
                row = []
                for col_idx in sorted(rows_dict[row_idx].keys()):
                    row.append(rows_dict[row_idx][col_idx])
                table_data.append(row)

            if table_data:
                first_row_preview = " | ".join([str(cell)[:20] for cell in table_data[0]])
                print(f"    First row: {first_row_preview}")

            row_count = len(table_data) - 1 if len(table_data) > 0 else 0
            return (row_count, table_data)
        else:
            return None

    except Exception as e:
        raise Exception(f"Azure extraction failed: {str(e)}")

def extract_table_with_nougat(pdf_path: str, table_number: int):
    """
    Use Nougat OCR to extract tables from PDF.
    Nougat outputs markdown, which we parse for table data.
    Returns tuple of (row_count, table_data) if found, None otherwise.
    """
    if not NOUGAT_AVAILABLE:
        print("    ⚠ Nougat OCR not available, skipping")
        return None

    try:
        print(f"    Extracting with Nougat OCR...")

        # Initialize Nougat model (cached after first use)
        if not hasattr(extract_table_with_nougat, 'model'):
            print("    Loading Nougat model (first time only)...")
            checkpoint = get_checkpoint("facebook/nougat-base")
            extract_table_with_nougat.model = NougatModel.from_pretrained(checkpoint)
            print("    Model loaded successfully")

        model = extract_table_with_nougat.model

        # Process PDF with Nougat
        predictions = model.inference(pdf_path=str(pdf_path), batch_size=1)

        if not predictions or len(predictions) == 0:
            print("    ✗ Nougat returned no predictions")
            return None

        markdown_text = predictions[0]

        # Parse markdown for tables
        # Markdown tables look like:
        # | Header 1 | Header 2 |
        # |----------|----------|
        # | Cell 1   | Cell 2   |

        # Find all markdown tables
        table_pattern = r'\|[^\n]+\|\n\|[-:\s|]+\|\n(?:\|[^\n]+\|\n)+'
        tables = re.findall(table_pattern, markdown_text)

        if table_number > len(tables):
            print(f"    ✗ Table {table_number} not found (Nougat found {len(tables)} tables)")
            return None

        # Get the requested table
        table_md = tables[table_number - 1]

        # Parse markdown table into rows
        lines = [line.strip() for line in table_md.split('\n') if line.strip()]

        # Remove the separator line (contains --- and |)
        data_lines = [line for line in lines if not all(c in '-:|' for c in line.replace(' ', ''))]

        # Parse each row
        table_data = []
        for line in data_lines:
            # Split by | and clean up
            cells = [cell.strip() for cell in line.split('|')]
            # Remove empty first/last cells (from leading/trailing |)
            cells = [c for c in cells if c]
            if cells:
                table_data.append(cells)

        if len(table_data) > 0:
            row_count = len(table_data) - 1  # Subtract header row
            print(f"    ✓ Nougat extracted {row_count} data rows from Table {table_number}")
            return (row_count, table_data)
        else:
            print("    ✗ No data rows found in table")
            return None

    except Exception as e:
        print(f"    Nougat ERROR: {e}")
        return None

def extract_sample_count_from_table(pdf_path: str, full_text: str) -> int:
    """
    Extract number of samples with SIMPLE, CLEAR, VISIBLE logic.
    """
    print("\n    === SAMPLE DETECTION DIAGNOSTICS ===")
    search_text = full_text
    print(f"    Searching through {len(search_text)} characters of text")
    
    # ==================================================================
    # PRIORITY 1: EXPLICIT COUNT - GROUP LOGIC
    # ==================================================================
    print("\n    PRIORITY 1: Looking for EXPLICIT COUNT statements...")
    print("    GROUP 1a: 'total of' + number immediately following")
    print("    GROUP 1b: number that immediately precedes a Group 2 term (with optional words)")
    print(
        "    GROUP 2: fabrics/materials/variants/garments/samples/textiles/specimens (singular + plural)"
    )
    print(
        "    GROUP 3: tested/produced/used/analyzed/evaluated/studied/prepared/examined"
    )
    print("    Numbers: Arabic (1-100), Roman (I-L), Words (one-fifty)")
    print("    ")
    print("    Valid combinations:")
    print("    - Group 1a OR 1b + Group 2 + Group 3 (BEST)")
    print("    - Group 2 + Group 3 (COMMON)")
    print("    - Group 1a OR 1b + Group 2 (COMMON)")
    print("    - Group 1a OR 1b + Group 3 (RARE)")
    
    explicit_count = None
    # Split on period/question/exclamation followed by space and capital letter
    # This handles both "word. Next" and "word.\nNext" patterns
    sentences = re.split(r'[.!?](?=\s+[A-Z])', search_text)
    print(f"    Split text into {len(sentences)} sentences")

    # ==================================================================
    # PRIORITY 0: SAMPLE NUMBER COLUMN - DISABLED (too many false positives)
    # ==================================================================
    # DISABLED - Camelot extracts tables incorrectly, causing false matches
    # Using other detection methods instead
    print("\n    PRIORITY 0: Sample number column detection DISABLED")
    print("    (Too many false positives with Camelot table extraction)")


    # Number word to digit mapping - EVERY NUMBER 1-50 (lowercase only)
    # We convert sentences to lowercase before searching, so only lowercase keys needed
    word_to_num = {
        "one": 1,
        "two": 2,
        "three": 3,
        "four": 4,
        "five": 5,
        "six": 6,
        "seven": 7,
        "eight": 8,
        "nine": 9,
        "ten": 10,
        "eleven": 11,
        "twelve": 12,
        "thirteen": 13,
        "fourteen": 14,
        "fifteen": 15,
        "sixteen": 16,
        "seventeen": 17,
        "eighteen": 18,
        "nineteen": 19,
        "twenty": 20,
        "twenty-one": 21,
        "twenty-two": 22,
        "twenty-three": 23,
        "twenty-four": 24,
        "twenty-five": 25,
        "twenty-six": 26,
        "twenty-seven": 27,
        "twenty-eight": 28,
        "twenty-nine": 29,
        "thirty": 30,
        "thirty-one": 31,
        "thirty-two": 32,
        "thirty-three": 33,
        "thirty-four": 34,
        "thirty-five": 35,
        "thirty-six": 36,
        "thirty-seven": 37,
        "thirty-eight": 38,
        "thirty-nine": 39,
        "forty": 40,
        "forty-one": 41,
        "forty-two": 42,
        "forty-three": 43,
        "forty-four": 44,
        "forty-five": 45,
        "forty-six": 46,
        "forty-seven": 47,
        "forty-eight": 48,
        "forty-nine": 49,
        "fifty": 50,
    }
    
    # Roman numerals - EVERY NUMBER 1-50
    roman_to_num = {
        "i": 1,
        "ii": 2,
        "iii": 3,
        "iv": 4,
        "v": 5,
        "vi": 6,
        "vii": 7,
        "viii": 8,
        "ix": 9,
        "x": 10,
        "xi": 11,
        "xii": 12,
        "xiii": 13,
        "xiv": 14,
        "xv": 15,
        "xvi": 16,
        "xvii": 17,
        "xviii": 18,
        "xix": 19,
        "xx": 20,
        "xxi": 21,
        "xxii": 22,
        "xxiii": 23,
        "xxiv": 24,
        "xxv": 25,
        "xxvi": 26,
        "xxvii": 27,
        "xxviii": 28,
        "xxix": 29,
        "xxx": 30,
        "xxxi": 31,
        "xxxii": 32,
        "xxxiii": 33,
        "xxxiv": 34,
        "xxxv": 35,
        "xxxvi": 36,
        "xxxvii": 37,
        "xxxviii": 38,
        "xxxix": 39,
        "xl": 40,
        "xli": 41,
        "xlii": 42,
        "xliii": 43,
        "xliv": 44,
        "xlv": 45,
        "xlvi": 46,
        "xlvii": 47,
        "xlviii": 48,
        "xlix": 49,
        "l": 50,
    }

    # Build word number pattern for regex (used in multiple places)
    word_pattern_check = "|".join(word_to_num.keys())

    # Track if we find the BEST combination (Groups 1+2+3 together)
    found_best_combination = False

    # ===== SUPERSEDING CASE 2: "total of" + NUMBER immediately followed by Group 2 =====
    # This takes absolute priority - if found, use it and stop searching
    print("\n    SUPERSEDING CASE 2: Checking for 'total of' + number + Group 2 term...")

    for i, sentence in enumerate(sentences):
        sentence_lower = sentence.lower()
        
        # Check for "total of" + arabic number + group2
        # (Less likely to match sample IDs due to "total of", but check for consistency)
        total_of_arabic = re.search(
            r"total\s+of\s+(?<![a-zA-Z0-9.])(\d+)(?![a-zA-Z0-9.])\s+(?:different\s+)?(?:types?\s+of\s+)?(?:fabrics?|materials?|samples?|variants?|garments?|textiles?|specimens?|jerseys?|structures?)\b",
            sentence_lower
        )
        if total_of_arabic:
            num = int(total_of_arabic.group(1))
            if 1 <= num <= 100:
                print(f"\n    ✓✓✓ SUPERSEDING CASE FOUND - 'total of' + {num} ✓✓✓")
                print(f"    Sentence {i}: '{sentence.strip()[:200]}...'")
                print("\n    ═══════════════════════════════════════════════")
                print(f"    DECISION: EXPLICIT COUNT = {num}")
                print("    SUPERSEDES ALL OTHER METHODS - RETURNING NOW")
                print("    ═══════════════════════════════════════════════\n")
                return num
        
        # Check for "total of" + word number + group2 (CASE-INSENSITIVE)
        total_of_word = re.search(
            rf"total\s+of\s+({word_pattern_check})\s+(?:different\s+)?(?:types?\s+of\s+)?(?:fabrics?|materials?|samples?|variants?|garments?|textiles?|specimens?|jerseys?|structures?)\b",
            sentence_lower
        )
        if total_of_word:
            word_num = total_of_word.group(1)
            if word_num in word_to_num:
                num = word_to_num[word_num]
                print(f"\n    ✓✓✓ SUPERSEDING CASE FOUND - 'total of' + '{word_num}' → {num} ✓✓✓")
                print(f"    Sentence {i}: '{sentence.strip()[:200]}...'")
                print("\n    ═══════════════════════════════════════════════")
                print(f"    DECISION: EXPLICIT COUNT = {num}")
                print("    SUPERSEDES ALL OTHER METHODS - RETURNING NOW")
                print("    ═══════════════════════════════════════════════\n")
                return num
    
    print("    Result: NO 'total of' + number + Group 2 found")
    
    # ===== MAIN GROUP LOGIC =====
    print("\n    MAIN GROUP LOGIC: Checking all sentences for valid combinations...")

    for i, sentence in enumerate(sentences):
        sentence_lower = sentence.lower()

        # OPTIMIZATION: Only process sentences that contain numbers
        # Check for: arabic numerals (0-9), roman numerals, or word numbers
        has_any_number = (
            re.search(r'\d', sentence_lower) or  # Arabic numerals
            re.search(rf'\b({word_pattern_check})\b', sentence_lower) or  # Word numbers
            re.search(rf'\b({"|".join(roman_to_num.keys())})\b', sentence_lower)  # Roman numerals
        )
        if not has_any_number:
            continue  # Skip sentences without any numbers

        # CHECK GROUP 1: "total of" (1a) OR number immediately before group 2 (1b)
        # These are TWO SEPARATE cases
        has_total_of = "total of" in sentence_lower

        # Check if WHOLE, POSITIVE number IMMEDIATELY precedes group 2 words
        # IMMEDIATELY means the very next word, with NO words in between
        # Pattern: number + optional whitespace + Group 2 word (NO words in between)
        has_number_before_group2_check = False

        # STRICT pattern - number must be followed by Group 2 term with NO intervening words
        # Only whitespace allowed between number and Group 2 term
        # Exclude sample IDs like "S18", "F3" (number must be standalone)
        if re.search(
            r"(?<![a-zA-Z0-9.])(\d+)(?![a-zA-Z0-9.])\s+(?:fabrics?|materials?|samples?|variants?|garments?|textiles?|specimens?|jerseys?|structures?)\b",
            sentence_lower
        ):
            has_number_before_group2_check = True

        # Check for word numbers IMMEDIATELY before Group 2 (CASE-INSENSITIVE)
        if re.search(
            rf"\b({word_pattern_check})\s+(?:fabrics?|materials?|samples?|variants?|garments?|textiles?|specimens?|jerseys?|structures?)\b",
            sentence_lower
        ):
            has_number_before_group2_check = True

        # SPECIAL PATTERN: "X different types of [modifiers] [Group2]"
        # Allows: "nine different types of single jersey weft knitted fabrics"
        # Allows: "four types of tri-layer fabrics" (handles hyphens)
        # Modifiers between "types of" and Group 2 are allowed for this pattern ONLY
        if re.search(
            rf"\b({word_pattern_check})\s+(?:different\s+)?types?\s+of\s+(?:[\w-]+\s+){{0,6}}(?:fabrics?|materials?|samples?|variants?|garments?|textiles?|specimens?|jerseys?|structures?)\b",
            sentence_lower
        ):
            has_number_before_group2_check = True

        # SPECIAL PATTERN: digit + "different types of [modifiers] [Group2]"
        # Allows: "9 different types of single jersey weft knitted fabrics"
        # Allows: "4 types of tri-layer fabrics" (handles hyphens)
        # Modifiers between "types of" and Group 2 are allowed for this pattern ONLY
        # Exclude sample IDs like "S9", "F4" (number must be standalone)
        if re.search(
            r"(?<![a-zA-Z0-9.])(\d+)(?![a-zA-Z0-9.])\s+(?:different\s+)?types?\s+of\s+(?:[\w-]+\s+){{0,6}}(?:fabrics?|materials?|samples?|variants?|garments?|textiles?|specimens?|jerseys?|structures?)\b",
            sentence_lower
        ):
            has_number_before_group2_check = True

        has_group1 = has_total_of or has_number_before_group2_check

        # CHECK GROUP 2: BOTH SINGULAR AND PLURAL - track which terms found
        group2_words = [
            "fabric",
            "fabrics",
            "material",
            "materials",
            "variant",
            "variants",
            "garment",
            "garments",
            "sample",
            "samples",
            "textile",
            "textiles",
            "specimen",
            "specimens",
            "jersey",
            "jerseys",
            "structure",
            "structures",
        ]
        group2_found = [word for word in group2_words if word in sentence_lower]
        has_group2 = len(group2_found) > 0

        # CHECK GROUP 3: action words - track which terms found
        group3_words = [
            "tested",
            "produced",
            "used",
            "analyzed",
            "evaluated",
            "studied",
            "prepared",
            "examined",
            "knit",
            "knitted",
            "woven",
        ]
        # Use word boundaries to avoid false matches (e.g., "used" in "focused")
        group3_found = [word for word in group3_words if re.search(rf'\b{word}\b', sentence_lower)]
        has_group3 = len(group3_found) > 0
        
        # VALID COMBINATION CHECK
        # Group 1 (number immediately before Group 2) is REQUIRED for ALL combinations
        # Group 1 is satisfied by EITHER 1a (total of) OR 1b (number before Group 2)
        valid_combination = False
        combination_type = ""

        if has_group1 and has_group2 and has_group3:
            valid_combination = True
            combination_type = "Group 1 + 2 + 3 (BEST)"
            found_best_combination = True  # Mark that we found the best combination
        elif has_group1 and has_group2:
            valid_combination = True
            combination_type = "Group 1 + 2 (COMMON)"
        elif has_group1 and has_group3:
            valid_combination = True
            combination_type = "Group 1 + 3 (RARE)"
        # REMOVED: Group 2 + 3 without Group 1 - user requires number IMMEDIATELY before Group 2
        
        if not valid_combination:
            continue
        
        print(f"\n     Sentence {i}: VALID COMBINATION - {combination_type}")
        print(f"     Text: '{sentence.strip()[:250]}'")

        # Show what triggered Group 1
        # Group 1 = (1a: "total of") OR (1b: number immediately before Group 2)
        print(f"     GROUP 1 SATISFIED: {has_group1}")
        if has_total_of:
            total_of_pos = sentence_lower.find("total of")
            context = sentence_lower[max(0, total_of_pos-20):min(len(sentence_lower), total_of_pos+80)]
            print(f"       → Via 'total of': '...{context}...'")
        if has_number_before_group2_check:
            print(f"       → Via number IMMEDIATELY before Group 2")
        if not has_group1:
            print(f"       → NOT SATISFIED (need 'total of' OR number before Group 2)")

        group2_terms = ", ".join(group2_found) if group2_found else "none"
        print(f"     GROUP 2 (fabric terms): {has_group2}. Found: {group2_terms}")
        group3_terms = ", ".join(group3_found) if group3_found else "none"
        print(f"     GROUP 3 (action words): {has_group3}. Found: {group3_terms}")
        
        # ===== PRIORITY 1: ARABIC NUMERALS (most common) =====
        # Must be WHOLE, POSITIVE numbers only
        # Number must be IMMEDIATELY before Group 2 term (no words in between)
        # IMPORTANT: Find ALL numbers in sentence, take the highest
        # IMPORTANT: Exclude sample IDs like "S18", "F3", etc. (must be standalone)
        digit_patterns = [
            r"total\s+of\s+(?<![a-zA-Z0-9.])(\d+)(?![a-zA-Z0-9.])\s+(?:fabrics?|materials?|samples?|variants?|garments?|textiles?|specimens?|jerseys?|structures?)",
            r"(?<![a-zA-Z0-9.])(\d+)(?![a-zA-Z0-9.])\s+(?:different\s+)?types?\s+of\s+(?:[\w-]+\s+){0,6}(?:fabrics?|materials?|samples?|variants?|garments?|textiles?|specimens?|jerseys?|structures?)",
            r"(?<![a-zA-Z0-9.])(\d+)(?![a-zA-Z0-9.])\s+(?:fabrics?|materials?|samples?|variants?|garments?|textiles?|specimens?|jerseys?|structures?)",
            r"(?:tested|produced|used|analyzed|evaluated|studied|prepared|examined|knit|knitted|woven)\s+(?<![a-zA-Z0-9.])(\d+)(?![a-zA-Z0-9.])\s+(?:fabrics?|materials?|samples?|variants?|garments?|textiles?|specimens?|jerseys?|structures?)",
        ]

        # Find ALL matching numbers in this sentence (don't break early)
        # IMPORTANT: Only accept if Group 3 term is NEARBY (not at end of sentence)
        for pattern in digit_patterns:
            for match in re.finditer(pattern, sentence_lower):
                num = int(match.group(1))
                if 1 <= num <= 100:
                    # SKIP common section headers (appear in all studies)
                    match_context = sentence_lower[max(0, match.start()-5):min(len(sentence_lower), match.end()+40)]
                    section_headers = [
                        r'\d+\s+materials?\s+and\s+methods?',
                        r'\d+\s+introduction',
                        r'\d+\s+results?',
                        r'\d+\s+discussion',
                        r'\d+\s+conclusion',
                        r'\d+\s+experimental',
                        r'\d+\s+methodology',
                        r'\d+\s+background',
                        r'\d+\s+literature\s+review',
                    ]
                    is_section_header = False
                    for header_pattern in section_headers:
                        if re.search(header_pattern, match_context, re.IGNORECASE):
                            print(f"    ⚠ SKIPPING {num}: Section header detected")
                            is_section_header = True
                            break
                    if is_section_header:
                        continue
                    # Get position of this number match
                    match_pos = match.start()

                    # Find closest Group 3 term position
                    closest_group3_distance = float('inf')
                    for g3_word in group3_words:
                        for g3_match in re.finditer(rf'\b{g3_word}\b', sentence_lower):
                            distance = abs(g3_match.start() - match_pos)
                            closest_group3_distance = min(closest_group3_distance, distance)

                    # Only accept if Group 3 is within 100 characters (same clause/phrase)
                    # This filters out section headers like "2 Materials and Methods"
                    if closest_group3_distance > 100:
                        print(f"    ⚠ SKIPPING {num}: Group 3 term too far away ({closest_group3_distance} chars)")
                        continue

                    if not explicit_count or num > explicit_count:
                        explicit_count = num
                        print("\n    ***************************************************************************")
                        print("    ***************************************************************************")
                        print(f"    ✓ FOUND arabic numeral: {num} (Group 3 within {closest_group3_distance} chars)")
                        print("    ***************************************************************************")
                        print("    ***************************************************************************")
                        print()
                        print(f"     Sentence: '{sentence.strip()[:200]}...'")
        
        # ===== PRIORITY 2: WORD NUMBERS (one, two, three, etc.) - CASE-INSENSITIVE =====
        # Number word must be IMMEDIATELY before Group 2 term (no words in between)
        # ONLY exception: "X different types of [Group2]" or "X types of [Group2]"
        # IMPORTANT: Find ALL numbers in sentence, take the highest
        word_patterns = [
            rf"\b({word_pattern_check})\b\s+(?:different\s+)?types?\s+of\s+(?:[\w-]+\s+){{0,6}}(?:fabrics?|materials?|samples?|variants?|garments?|textiles?|specimens?|jerseys?|structures?)",
            rf"\b({word_pattern_check})\b\s+(?:fabrics?|materials?|samples?|variants?|garments?|textiles?|specimens?|jerseys?|structures?)",
            rf"\b({word_pattern_check})\b\s+(?:fabrics?|materials?|samples?|variants?|garments?|textiles?|specimens?|jerseys?|structures?)\s+of",
        ]

        # Find ALL matching numbers in this sentence (don't break early)
        # IMPORTANT: Only accept if Group 3 term is NEARBY (not at end of sentence)
        for pattern in word_patterns:
            for match in re.finditer(pattern, sentence_lower):
                word_num = match.group(1)
                if word_num in word_to_num:
                    found_count = word_to_num[word_num]

                    # DEBUG: Show exact match context
                    match_start = match.start()
                    match_end = match.end()
                    context_start = max(0, match_start - 20)
                    context_end = min(len(sentence_lower), match_end + 20)
                    match_context_full = sentence_lower[context_start:context_end]
                    print(f"    [DEBUG] Matched '{word_num}' at pos {match_start}: '...{match_context_full}...'")

                    # Get position of this number match
                    match_pos = match.start()

                    # Find closest Group 3 term position
                    closest_group3_distance = float('inf')
                    for g3_word in group3_words:
                        for g3_match in re.finditer(rf'\b{g3_word}\b', sentence_lower):
                            distance = abs(g3_match.start() - match_pos)
                            closest_group3_distance = min(closest_group3_distance, distance)

                    # Only accept if Group 3 is within 100 characters (same clause/phrase)
                    if closest_group3_distance > 100:
                        print(f"    ⚠ SKIPPING '{word_num}' ({found_count}): Group 3 term too far away ({closest_group3_distance} chars)")
                        continue

                    if not explicit_count or found_count > explicit_count:
                        explicit_count = found_count
                        print("\n    ***************************************************************************")
                        print("    ***************************************************************************")
                        print(f"    ✓ FOUND word number: '{word_num}' → {found_count} (Group 3 within {closest_group3_distance} chars)")
                        print("    ***************************************************************************")
                        print("    ***************************************************************************")
                        print()
                        print(f"     Sentence: '{sentence.strip()[:200]}...'")
        
        # ===== PRIORITY 3: ROMAN NUMERALS (rare but possible) =====
        # IMPORTANT: Find ALL numbers in sentence, take the highest
        roman_pattern = "|".join(roman_to_num.keys())
        roman_patterns = [
            rf"(?:^|\s)({roman_pattern})(?:\s+)(?:fabrics?|materials?|samples?|variants?|garments?|textiles?|specimens?|jerseys?|structures?)\b",
        ]

        # Find ALL matching numbers in this sentence (don't break early)
        # IMPORTANT: Only accept if Group 3 term is NEARBY (not at end of sentence)
        for pattern in roman_patterns:
            for match in re.finditer(pattern, sentence_lower):
                roman_num = match.group(1)
                if roman_num in roman_to_num:
                    found_count = roman_to_num[roman_num]

                    # Get position of this number match
                    match_pos = match.start()

                    # Find closest Group 3 term position
                    closest_group3_distance = float('inf')
                    for g3_word in group3_words:
                        for g3_match in re.finditer(rf'\b{g3_word}\b', sentence_lower):
                            distance = abs(g3_match.start() - match_pos)
                            closest_group3_distance = min(closest_group3_distance, distance)

                    # Only accept if Group 3 is within 100 characters (same clause/phrase)
                    if closest_group3_distance > 100:
                        print(f"    ⚠ SKIPPING '{roman_num.upper()}' ({found_count}): Group 3 term too far away ({closest_group3_distance} chars)")
                        continue

                    if not explicit_count or found_count > explicit_count:
                        explicit_count = found_count
                        print("\n    ***************************************************************************")
                        print("    ***************************************************************************")
                        print(f"    ✓ FOUND roman numeral: '{roman_num.upper()}' → {found_count} (Group 3 within {closest_group3_distance} chars)")
                        print("    ***************************************************************************")
                        print("    ***************************************************************************")
                        print()
                        print(f"     Sentence: '{sentence.strip()[:200]}...'")

        # After checking ALL patterns in this sentence, continue to next sentence
        # Don't return early - we want to find the HIGHEST number across ALL sentences
        if found_best_combination and explicit_count:
            print(f"\n    → Found Groups 1+2+3 in this sentence with count = {explicit_count}")
            print(f"    → Continuing to check remaining sentences for higher numbers...")
    
    if explicit_count:
        print("\n    ═══════════════════════════════════════════════")
        print(f"    DECISION: EXPLICIT COUNT = {explicit_count}")
        print("    HIGHEST NUMBER FOUND ACROSS ALL SENTENCES")
        print("    ═══════════════════════════════════════════════\n")
        return explicit_count

    print("\n    Result: NO explicit count found using group logic")

    # ===== TABLE FALLBACK: Group 2 term + "table" =====
    # ONLY run table extraction if we did NOT find the BEST combination (Groups 1+2+3)
    # If we found Groups 1+2+3 together, that's conclusive - don't waste time on tables
    if not found_best_combination:
        print("\n    TABLE FALLBACK: No Groups 1+2+3 found, checking for Group 2 term + 'table'...")

        group2_words = [
            "fabric", "fabrics",
            "garment", "garments",
            "sample", "samples",
        ]

        table_fallback_count = None
        extracted_tables = set()  # Track which tables we've already extracted

        for i, sentence in enumerate(sentences):
            sentence_lower = sentence.lower()

            # OPTIMIZATION: Only process sentences that contain numbers
            has_any_number = (
                re.search(r'\d', sentence_lower) or  # Arabic numerals
                re.search(rf'\b({word_pattern_check})\b', sentence_lower) or  # Word numbers
                re.search(rf'\b({"|".join(roman_to_num.keys())})\b', sentence_lower)  # Roman numerals
            )
            if not has_any_number:
                continue

            # Check if sentence has a Group 2 term AND "table"
            group2_found_here = [word for word in group2_words if word in sentence_lower]
            has_group2 = len(group2_found_here) > 0
            table_match = re.search(r"table\s+(\d+)", sentence_lower, re.IGNORECASE)

            if has_group2 and table_match:
                table_num = int(table_match.group(1))

                # Skip if we've already extracted this table
                if table_num in extracted_tables:
                    print(f"\n    ⊗ Skipping Table {table_num} (already extracted)")
                    continue

                print(f"\n    ✓ Found Group 2 term + Table {table_num}")
                print(f"    Sentence {i}: '{sentence.strip()[:300]}...'")
                group2_terms_list = ", ".join(group2_found_here)
                print(f"    Group 2 Terms Present: {group2_terms_list}")
                print(f"    Attempting to parse Table {table_num}...")

                # Try to extract and count rows from the specified table
                table_count = extract_table_row_count(pdf_path, table_num)
                extracted_tables.add(table_num)  # Mark this table as extracted

                if table_count and not table_fallback_count:
                    table_fallback_count = table_count
                    print(f"    ✓ Extracted {table_count} rows from Table {table_num} (stored as fallback)")
                elif not table_count:
                    print(f"    ✗ Could not parse Table {table_num}")

        if table_fallback_count:
            print(f"\n    Stored table fallback count: {table_fallback_count}")
        else:
            print("    Result: NO valid table reference found")
    else:
        print("\n    TABLE FALLBACK: Skipped (already found Groups 1+2+3 together)")
        table_fallback_count = None

    # If we found a table fallback count, use it
    if table_fallback_count:
        print("\n    ═══════════════════════════════════════════════")
        print(f"    DECISION: TABLE FALLBACK COUNT = {table_fallback_count}")
        print("    Using table row count from earlier")
        print("    ═══════════════════════════════════════════════\n")
        return table_fallback_count
    
    # ==================================================================
    # PRIORITY 2: SINGLE SAMPLE
    # ==================================================================
    print("\n    PRIORITY 2: Checking for single-sample study...")
    single_sample_patterns = [
        r"\ba\s+(?:smart|novel|new|single)\s+fabrics?\b",
        r"\bone\s+fabrics?\b",
        r"\bsingle\s+(?:fabrics?|materials?|samples?)\b",
    ]
    
    for pattern in single_sample_patterns:
        if re.search(pattern, search_text, re.IGNORECASE):
            match = re.search(pattern, search_text, re.IGNORECASE)
            context_start = max(0, match.start() - 50)
            context_end = min(len(search_text), match.end() + 50)
            context = search_text[context_start:context_end]
            print(f"    ✓ Found: '{context}'")
            print("\n    ═══════════════════════════════════════════════")
            print("    DECISION: SINGLE SAMPLE STUDY = 1")
            print("    ═══════════════════════════════════════════════\n")
            return 1
    
    print("    Result: NOT a single sample study")
    
    # ==================================================================
    # PRIORITY 3: TABLE REFERENCE
    # ==================================================================
    print("\n    PRIORITY 3: Looking for table references...")
    print(
        "    Required in same sentence: (fabric/fabrics OR material/materials OR sample/samples OR variant/variants OR garment/garments OR textile/textiles OR specimen/specimens) AND 'table'"
    )
    
    table_numbers = set()
    
    for i, sentence in enumerate(sentences):
        sentence_lower = sentence.lower()
        has_sample_word = any(
            word in sentence_lower
            for word in [
                "fabric",
                "fabrics",
                "material",
                "materials",
                "sample",
                "samples",
                "garment",
                "garments",
                "textile",
                "textiles",
                "specimen",
                "specimens",
                "variant",
                "variants",
            ]
        )
        has_table_ref = re.search(r"table\s+(\d+)", sentence_lower, re.IGNORECASE)
        
        if has_sample_word and has_table_ref:
            table_num = has_table_ref.group(1)
            table_numbers.add(table_num)
            print(f"    ✓ Sentence {i}: Has sample word + Table {table_num}")
            print(f"     Text: '{sentence.strip()[:200]}...'")
    
    if not table_numbers:
        print("    Result: NO table reference found")
        print("\n    ═══════════════════════════════════════════════")
        print("    DECISION: No indicators found = 1")
        print("    ═══════════════════════════════════════════════\n")
        return 1
    
    print(f"\n    ✓ Found table references: {sorted(table_numbers)}")
    
    # ==================================================================
    # PRIORITY 4: EXTRACT TABLES
    # ==================================================================
    print("\n    PRIORITY 4: Extracting tables to count rows...")
    best_count = None
    best_method = None
    
    # CAMELOT
    print("\n    --- Method A: CAMELOT ---")
    try:
        tables_lattice = camelot.read_pdf(pdf_path, pages="all", flavor="lattice")
        tables_stream = camelot.read_pdf(pdf_path, pages="all", flavor="stream")
        all_camelot_tables = list(tables_lattice) + list(tables_stream)
        print(f"    Found {len(all_camelot_tables)} total tables")
        
        for idx, table in enumerate(all_camelot_tables):
            df = table.df
            rows = len(df) - 1
            print(f"\n    Table {idx + 1}:")
            print(f"     Dimensions: {len(df)} rows × {len(df.columns)} cols")
            print(f"     HEADER ROW: {str(df.iloc[0].tolist())[:150]}...")
            if len(df) > 1:
                print(f"     DATA ROW 1: {str(df.iloc[1].tolist())[:150]}...")
            
            header_text = " ".join(df.iloc[0].astype(str)).lower()
            fabric_keywords = [
                "fabric",
                "fabrics",
                "material",
                "materials",
                "sample",
                "samples",
                "fiber",
                "fibers",
                "composition",
                "compositions",
                "thickness",
                "density",
                "densities",
                "weight",
                "weights",
                "gsm",
                "structure",
                "structures",
                "weave",
                "weaves",
                "knit",
                "knits",
                "variant",
                "variants",
                "garment",
                "garments",
                "textile",
                "textiles",
                "specimen",
                "specimens",
            ]
            score = sum([kw in header_text for kw in fabric_keywords])
            print(f"     Fabric keyword score: {score}/{len(fabric_keywords)}")
            
            if score >= 2 and rows > 0:
                print("     ✓✓✓ THIS LOOKS LIKE THE SAMPLE TABLE ✓✓✓")
                if not best_count:
                    best_count = rows
                    best_method = f"Camelot (table {idx + 1})"
            else:
                print(f"     ✗ REJECTED (score {score} < 2 or rows={rows})")
    except Exception as e:
        print(f"    CameLot ERROR: {e}")
    
    # TABULA
    if not best_count:
        print("\n    --- Method B: TABULA ---")
        try:
            dfs = tabula.read_pdf(pdf_path, pages="all", multiple_tables=True, silent=True)
            print(f"    Found {len(dfs)} tables")
            
            for idx, df in enumerate(dfs):
                rows = len(df) - 1 if len(df) > 1 else len(df)
                print(f"\n    Table {idx + 1}:")
                print(f"     Dimensions: {len(df)} rows × {len(df.columns)} cols")
                print(f"     COLUMNS: {str(df.columns.tolist())[:150]}...")
                if len(df) > 0:
                    print(f"     DATA ROW 1: {str(df.iloc[0].tolist())[:150]}...")
                
                header_text = " ".join([str(col).lower() for col in df.columns])
                fabric_keywords = [
                    "fabric",
                    "fabrics",
                    "material",
                    "materials",
                    "sample",
                    "samples",
                    "fiber",
                    "fibers",
                    "composition",
                    "compositions",
                    "thickness",
                    "density",
                    "densities",
                    "weight",
                    "weights",
                    "gsm",
                    "structure",
                    "structures",
                    "weave",
                    "weaves",
                    "knit",
                    "knits",
                    "variant",
                    "variants",
                    "garment",
                    "garments",
                    "textile",
                    "textiles",
                    "specimen",
                    "specimens",
                ]
                score = sum([kw in header_text for kw in fabric_keywords])
                print(f"     Fabric keyword score: {score}/{len(fabric_keywords)}")
                
                if score >= 2 and rows > 0:
                    print("     ✓✓✓ THIS LOOKS LIKE THE SAMPLE TABLE ✓✓✓")
                    if not best_count:
                        best_count = rows
                        best_method = f"Tabula (table {idx + 1})"
                else:
                    print(f"     ✗ REJECTED (score {score} < 2 or rows={rows})")
        except Exception as e:
            print(f"    Tabula ERROR: {e}")
    
    # PDFPLUMBER
    if not best_count:
        print("\n    --- Method C: PDFPLUMBER ---")
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    tables = page.extract_tables()
                    if not tables:
                        continue
                    print(f"    Page {page_num + 1}: Found {len(tables)} table(s)")
                    
                    for t_idx, table in enumerate(tables):
                        if not table or len(table) <= 1:
                            continue
                        rows = len(table) - 1
                        print(f"\n    Page {page_num + 1}, Table {t_idx + 1}:")
                        print(f"     Dimensions: {len(table)} rows × {len(table[0])} cols")
                        print(f"     HEADER: {str(table[0])[:150]}...")
                        if len(table) > 1:
                            print(f"     DATA ROW 1: {str(table[1])[:150]}...")
                        
                        header_text = " ".join(
                            [str(cell).lower() if cell else "" for cell in table[0]]
                        )
                        fabric_keywords = [
                            "fabric",
                            "fabrics",
                            "material",
                            "materials",
                            "sample",
                            "samples",
                            "fiber",
                            "fibers",
                            "composition",
                            "compositions",
                            "thickness",
                            "density",
                            "densities",
                            "weight",
                            "weights",
                            "gsm",
                            "structure",
                            "structures",
                            "weave",
                            "weaves",
                            "knit",
                            "knits",
                            "variant",
                            "variants",
                            "garment",
                            "garments",
                            "textile",
                            "textiles",
                            "specimen",
                            "specimens",
                        ]
                        score = sum([kw in header_text for kw in fabric_keywords])
                        print(f"     Fabric keyword score: {score}/{len(fabric_keywords)}")
                        
                        if score >= 2:
                            print("     ✓✓✓ THIS LOOKS LIKE THE SAMPLE TABLE ✓✓✓")
                            if not best_count:
                                best_count = rows
                                best_method = f"pdfplumber (page {page_num + 1}, table {t_idx + 1})"
                        else:
                            print(f"     ✗ REJECTED (score {score} < 2)")
        except Exception as e:
            print(f"    pdfplumber ERROR: {e}")
    
    # FINAL DECISION
    if best_count:
        print("\n    ═══════════════════════════════════════════════")
        print(f"    DECISION: Extracted {best_count} rows")
        print(f"    Method: {best_method}")
        print("    ═══════════════════════════════════════════════\n")
        return best_count
    else:
        print("\n    ═══════════════════════════════════════════════")
        print("    DECISION: All extraction methods failed = 1")
        print("    ═══════════════════════════════════════════════\n")
        return 1

# ================== TEST FUNCTIONS ==================
def test_process_pdfs():
    """Test function to process all PDFs and extract metadata."""
    print(f"Loading input Excel: {INPUT_EXCEL}")
    
    if not INPUT_EXCEL.exists():
        print(f"ERROR: Input file not found -> {INPUT_EXCEL}")
        pytest.fail(f"Input file not found: {INPUT_EXCEL}")
    
    wb_input = openpyxl.load_workbook(INPUT_EXCEL)
    ws_input = wb_input.active
    
    study_ids = []
    for row in ws_input.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            study_ids.append(str(row[0]).strip())
    
    print(f"Found {len(study_ids)} study IDs in input Excel")
    
    # ================== PROCESS EACH PDF ==================
    output_columns = [
        "Study Number",
        "Study title",
        "Year of Publish",
        "Name of first-listed author",
        "Number of Sample Fabrics",
    ]
    output_rows = []
    
    for study_id in study_ids:
        print(f"\n{'=' * 60}")
        print(f"Processing Study {study_id}")
        print(f"{'=' * 60}")
        
        # Find PDF
        pdf_path = PDF_FOLDER / f"{study_id}.pdf"
        if not os.path.exists(pdf_path):
            print(f"  PDF not found -> {pdf_path}")
            output_rows.append(
                {
                    "Study Number": study_id,
                    "Study title": "PDF not found",
                    "Year of Publish": "N/A",
                    "Name of first-listed author": "N/A",
                    "Number of Sample Fabrics": 0,
                }
            )
            continue
        
        # Extract full text
        doc = fitz.open(pdf_path)
        full_text = ""
        for page in doc:
            full_text += page.get_text()
        doc.close()
        print(f"  Extracted {len(full_text)} characters of text")
        
        # --------- EXTRACT TITLE ---------
        meta_title = extract_title_with_formatting(str(pdf_path))
        print(f"  Title: {meta_title}")
        
        # --------- EXTRACT YEAR ---------
        meta_year = extract_year_from_text(full_text)
        print(f"  Year: {meta_year}")
        
        # --------- EXTRACT AUTHOR ---------
        meta_first_last = extract_first_author_with_formatting(str(pdf_path))
        print(f"  Author: {meta_first_last}")
        
        # --------- EXTRACT NUMBER OF SAMPLES (REGEX) ---------
        sample_count = extract_sample_count_from_table(str(pdf_path), full_text)
        print(f"  Samples: {sample_count}")
        
        # --------- TRY LLM FOR MISSING DATA ---------
        missing = []
        if not meta_title:
            missing.append("title")
        if not meta_year:
            missing.append("year")
        if not meta_first_last:
            missing.append("author")
        
        if missing:
            print(
                f"  Missing: {', '.join(missing) if missing else 'none, but trying LLM for samples'}"
            )
            print("  Trying LLM...")
            short_text = full_text[:4000]
            
            # LLM for title
            if "title" in missing:
                try:
                    title_prompt = f"What is the main article title? Return only the title.\n\nTEXT:\n{short_text[:2000]}"
                    llm_title = call_local_llm(title_prompt).strip()
                    if llm_title and len(llm_title) > 10 and len(llm_title) < 300:
                        meta_title = llm_title
                        print(f"    LLM found title: {meta_title}")
                except Exception as e:
                    print(f"    LLM title error: {e}")
            
            # LLM for year
            if "year" in missing:
                try:
                    year_prompt = f"What year was this published? Return only a 4-digit year.\n\nTEXT:\n{short_text[:2000]}"
                    llm_year = call_local_llm(year_prompt).strip()
                    year_match = re.search(r"\b(19[9]\d|20[0-2]\d)\b", llm_year)
                    if year_match:
                        meta_year = int(year_match.group(1))
                        print(f"    LLM found year: {meta_year}")
                except Exception as e:
                    print(f"    LLM year error: {e}")
            
            # LLM for author
            if "author" in missing:
                try:
                    author_prompt = f"Who is the first author? Return only the last name (family name).\n\nTEXT:\n{short_text}"
                    llm_author = call_local_llm(author_prompt).strip()
                    llm_author = llm_author.replace('"', "").replace("'", "").strip()
                    author_match = re.search(r"\b([A-Z][a-z]{2,})\b", llm_author)
                    if author_match:
                        meta_first_last = author_match.group(1)
                        print(f"    LLM found author: {meta_first_last}")
                except Exception as e:
                    print(f"    LLM author error: {e}")
        
        # --------- SET DEFAULTS FOR MISSING VALUES ---------
        if not meta_title:
            meta_title = "Title not extracted"
        if not meta_year:
            meta_year = "Year not found"
        if not meta_first_last:
            meta_first_last = "Author not found"
        
        # --------- APPLY 100 CHARACTER LIMIT ---------
        if meta_title and len(str(meta_title)) > 100:
            meta_title = str(meta_title)[:100]
        if meta_first_last and len(str(meta_first_last)) > 100:
            meta_first_last = str(meta_first_last)[:100]
        
        print("\n  FINAL METADATA:")
        print(f"    Title: {meta_title}")
        print(f"    Year: {meta_year}")
        print(f"    Author: {meta_first_last}")
        print(f"    Samples: {sample_count}")
        
        # --------- ADD TO OUTPUT ---------
        output_rows.append(
            {
                "Study Number": study_id,
                "Study title": meta_title,
                "Year of Publish": meta_year,
                "Name of first-listed author": meta_first_last,
                "Number of Sample Fabrics": sample_count,
            }
        )
    
    # ================== WRITE OUTPUT EXCEL ==================
    print(f"\n{'=' * 60}")
    print(f"Writing output to: {OUTPUT_EXCEL}")
    print(f"{'=' * 60}")
    
    wb_output = Workbook()
    ws_output = wb_output.active
    ws_output.title = "Metadata"
    
    # Write header
    ws_output.append(output_columns)
    
    # Write data rows
    for row_dict in output_rows:
        row_values = [row_dict.get(col, "") for col in output_columns]
        ws_output.append(row_values)
    
    wb_output.save(OUTPUT_EXCEL)
    print(f"✓ Output saved: {OUTPUT_EXCEL}")
    print(f"✓ Processed {len(output_rows)} studies")

if __name__ == "__main__":
    # Run the test function directly when script is executed
    test_process_pdfs()
