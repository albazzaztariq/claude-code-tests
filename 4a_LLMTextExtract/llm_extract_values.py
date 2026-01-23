"""
LLM Text Extract - Value Extraction with Qwen 2.5 7B

Extracts metric values from filtered text using LLM via Ollama.

PROCESS:
1. Create copy of metadata Excel from Pipeline 3
2. Read all _filtered.txt files in Extracted Text folder
3. For each filtered text file:
   - Parse metric sentences
   - Send to Qwen 2.5 7B LLM (via Ollama) for value extraction
   - If NO values extracted -> mark metadata as "No Metrics" (red row)
4. Create final dataset Excel with extracted values
5. Save metadata copy and final dataset

INPUTS:
- Metadata Excel file (from Pipeline 3)
- Filtered text files (*_filtered.txt from Pipeline 2)
- LLMTextExtractPrompt.txt (system prompt)
- MetricsSearchable.txt (schema columns)

OUTPUTS:
- 4a_Metadata.xlsx (metadata copy with "No Metrics" breakpoints)
- 4a_FinalDataset.xlsx (extracted values in schema format)
"""

import os
import sys
from pathlib import Path
from typing import Dict, List, Tuple
import time
import shutil

# Import timing and corpus utilities
sys.path.insert(0, str(Path(__file__).parent.parent))
from Timers.timing_utils import save_timer
from Utils.corpus_utils import load_corpus_json

# Load environment variables
from dotenv import load_dotenv
load_dotenv()

# Check for required libraries
try:
    import pandas as pd
except ImportError:
    print("ERROR: Missing required library 'pandas'")
    print("Install it with: pip install pandas openpyxl")
    sys.exit(1)

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
except ImportError:
    print("ERROR: Missing required library 'openpyxl'")
    print("Install it with: pip install openpyxl")
    sys.exit(1)

try:
    from rapidfuzz import fuzz, process
except ImportError:
    print("ERROR: Missing required library 'rapidfuzz'")
    print("Install it with: pip install rapidfuzz")
    sys.exit(1)

import subprocess
import re


def load_extract_prompt() -> str:
    """Load the extraction prompt from Config/LLMTextExtractPrompt.txt."""
    prompt_file = Path(__file__).parent.parent / "Config" / "LLMTextExtractPrompt.txt"

    if not prompt_file.exists():
        print(f"ERROR: Extract prompt file not found: {prompt_file.name}")
        sys.exit(1)

    with open(prompt_file, 'r', encoding='utf-8') as f:
        return f.read()


# Load extraction prompt from config
EXTRACT_PROMPT = load_extract_prompt()


def load_schema_columns() -> List[str]:
    """Load schema columns from MetricsSearchable.txt."""
    schema_file = Path(__file__).parent.parent / "Config" / "MetricsSearchable.txt"

    with open(schema_file, 'r', encoding='utf-8') as f:
        columns = [line.strip() for line in f if line.strip()]

    return columns


def fuzzy_match_metric(metric_name: str, schema_cols: List[str], threshold: int = 60) -> Tuple[str, float]:
    """
    Fuzzy match metric name to schema columns.

    Returns:
        (matched_column, confidence) or (metric_name, 0.0) if no match
    """
    metric_lower = metric_name.lower()

    # 1. Exact match
    for schema_col in schema_cols:
        if metric_lower == schema_col.lower():
            return schema_col, 1.0

    # 2. Substring match
    substring_matches = []
    for schema_col in schema_cols:
        schema_lower = schema_col.lower()
        if metric_lower in schema_lower or schema_lower in metric_lower:
            substring_matches.append(schema_col)

    if substring_matches:
        result = process.extractOne(metric_name, substring_matches, scorer=fuzz.ratio)
        if result:
            return result[0], 0.9

    # 3. Fuzzy match
    result = process.extractOne(
        metric_name,
        schema_cols,
        scorer=fuzz.token_set_ratio,
        score_cutoff=threshold
    )

    if result:
        matched_col, score, _ = result
        return matched_col, score / 100.0
    else:
        # No match - return original name
        return metric_name, 0.0


# Load schema once at startup
SCHEMA_COLUMNS = load_schema_columns()


def call_qwen(metric_sentence: str, max_retries: int = 3) -> str:
    """
    Call Qwen 2.5 7B via Ollama for metric value extraction.

    Args:
        metric_sentence: Sentence containing metric and value
        max_retries: Number of retry attempts if call fails

    Returns:
        Extracted value or "No value found"
    """
    # Build prompt with metric sentence
    prompt = EXTRACT_PROMPT.replace("{metric_sentence}", metric_sentence)

    for attempt in range(max_retries):
        try:
            # Call Ollama
            cmd = [
                "ollama", "run", "qwen2.5:7b-instruct-q5_K_M",
                prompt
            ]

            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                encoding='utf-8',
                timeout=20  # 20 second timeout per call
            )

            if result.returncode != 0:
                error_msg = result.stderr.strip() if result.stderr else "Unknown error"
                if attempt < max_retries - 1:
                    print(f"      [RETRY {attempt + 1}/{max_retries}] Ollama error: {error_msg}")
                    time.sleep(1)
                    continue
                return f"Error: Ollama failed - {error_msg}"

            # Clean response (remove ANSI codes and whitespace)
            response = result.stdout.strip()
            response = re.sub(r'\x1b\[[0-9;]*[a-zA-Z]', '', response)  # Remove ANSI codes
            response = re.sub(r'\[?[0-9]+[hlmKG]', '', response)  # Remove spinner codes
            response = response.strip()

            # Return cleaned response
            return response if response else "No value found"

        except subprocess.TimeoutExpired:
            if attempt < max_retries - 1:
                print(f"      [RETRY {attempt + 1}/{max_retries}] LLM call timed out")
                time.sleep(1)
                continue
            # After all retries exhausted, return error instead of exiting
            print(f"      [SKIP] LLM call timed out after {max_retries} retries")
            return "Error: LLM timeout"

        except Exception as e:
            if attempt < max_retries - 1:
                print(f"      [RETRY {attempt + 1}/{max_retries}] Exception: {str(e)}")
                time.sleep(1)
                continue
            return f"Error: {str(e)}"

    return "Error: Max retries reached"


def parse_filtered_text(text: str) -> List[Tuple[str, str]]:
    """
    Parse filtered text file into metric-sentence pairs.

    Args:
        text: Content of filtered text file

    Returns:
        List of (metric_label, sentence) tuples
    """
    pairs = []
    lines = text.split('\n')

    i = 0
    while i < len(lines):
        line = lines[i].strip()

        # Look for "Metric Match:" lines
        if line.startswith("Metric Match:"):
            metric_label = line.replace("Metric Match:", "").strip()

            # Get the next non-empty line(s) as the sentence
            i += 1
            sentence_parts = []
            while i < len(lines) and lines[i].strip() and not lines[i].startswith("Metric Match:"):
                sentence_parts.append(lines[i].strip())
                i += 1

            if sentence_parts:
                sentence = " ".join(sentence_parts)
                # Combine metric label and sentence for LLM
                combined = f"Metric Match: {metric_label}\n{sentence}"
                pairs.append((metric_label, combined))
        else:
            i += 1

    return pairs


def process_filtered_file(filtered_path: Path, track_first_call: bool = False) -> tuple[List[Tuple[str, str]], float]:
    """
    Process a single filtered text file.

    Args:
        filtered_path: Path to _filtered.txt file
        track_first_call: Unused parameter (kept for compatibility)

    Returns:
        Tuple of (list of (metric_name, value) tuples, total_llm_time for this file)
    """
    # Read filtered text
    with open(filtered_path, 'r', encoding='utf-8') as f:
        text = f.read()

    # Parse into metric-sentence pairs
    pairs = parse_filtered_text(text)

    if len(pairs) == 0:
        return [], 0.0

    results = []  # Changed from dict to list to allow duplicate metrics
    total_llm_time = 0.0

    for idx, (metric_label, combined_text) in enumerate(pairs, 1):
        # Call LLM to extract value (track all calls)
        call_start = time.time()
        extracted = call_qwen(combined_text)
        call_time = time.time() - call_start
        total_llm_time += call_time

        # If error or no value found, skip this metric
        extracted_lower = extracted.lower()
        if ("error" in extracted_lower or
            "no value" in extracted_lower or
            "no metrics" in extracted_lower or
            extracted.strip() == ""):
            continue

        # Fuzzy match metric name to schema
        normalized_metric, confidence = fuzzy_match_metric(metric_label, SCHEMA_COLUMNS)

        # Store normalized metric name and value
        results.append((normalized_metric, extracted))

        # Small delay between extractions
        time.sleep(0.5)

    return results, total_llm_time


def extract_all_values(extracted_text_folder: Path) -> Tuple[Path, Path]:
    """
    Extract values from all filtered text files and create outputs.

    Args:
        extracted_text_folder: Folder containing _filtered.txt files

    Returns:
        (metadata_excel_path, final_dataset_path)
    """
    query_folder = extracted_text_folder.parent

    # Load corpus metadata from JSON
    all_papers = load_corpus_json(query_folder)

    # Create paper lookup dict by study_number for fast access (only papers with study_number)
    papers_by_num = {p["study_number"]: p for p in all_papers if "study_number" in p}

    # Find all filtered text files
    filtered_files = sorted(extracted_text_folder.glob("*_filtered.txt"))

    # Process each file
    all_extractions = {}
    no_metrics_count = 0
    total_llm_time = 0.0

    total_start = time.time()

    for idx, filtered_path in enumerate(filtered_files, 1):
        pdf_num = int(filtered_path.stem.replace("_filtered", ""))

        # Track LLM time (Timer 8)
        results, file_llm_time = process_filtered_file(filtered_path)

        # Accumulate LLM time
        total_llm_time += file_llm_time

        if len(results) == 0:
            # No metrics extracted - mark in metadata
            # Update paper entry in JSON
            if pdf_num in papers_by_num:
                papers_by_num[pdf_num]["breakpoint"] = "No Metrics"
                no_metrics_count += 1
        else:
            # Metrics extracted - store for final dataset (list of tuples)
            if pdf_num not in all_extractions:
                all_extractions[pdf_num] = []
            all_extractions[pdf_num].extend(results)

    total_time = time.time() - total_start

    # Save Timer 8: Total LLM extraction time
    if total_llm_time > 0:
        save_timer(query_folder, "7_llm_extraction_total", total_llm_time)

    # Create 4a_Metadata.xlsx from updated JSON data
    from Utils.corpus_utils import convert_corpus_json_to_excel_data
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    metadata_rows = convert_corpus_json_to_excel_data(all_papers)
    metadata_excel_path = query_folder / "4a_Metadata.xlsx"

    # Write to Excel with colors
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    ws = wb.create_sheet("Corpus")

    # Write header row
    for col_idx, value in enumerate(metadata_rows[0], 1):
        ws.cell(row=1, column=col_idx, value=value).font = Font(bold=True)

    # Write data rows with colors

    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

    for row_idx, row_data in enumerate(metadata_rows[1:], 2):
        color_code = row_data[-1]  # Last element is color code
        actual_data = row_data[:-1]  # All except color code

        # Write data
        for col_idx, value in enumerate(actual_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

        # Apply color
        if color_code == "green":
            fill = green_fill
        elif color_code == "orange":
            fill = orange_fill
        elif color_code == "red":
            fill = red_fill
        else:
            fill = None

        if fill:
            for col_idx in range(1, len(actual_data) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # Column widths
    ws.column_dimensions['A'].width = 10  # PDF Number
    ws.column_dimensions['B'].width = 16  # FileType
    ws.column_dimensions['C'].width = 15  # Breakpoint
    ws.column_dimensions['D'].width = 10  # Access

    wb.save(metadata_excel_path)

    # Create final dataset Excel
    final_dataset_path = extracted_text_folder.parent / "4a_FinalDataset.xlsx"

    # Create DataFrame with schema columns
    columns = ["PDF Number"] + SCHEMA_COLUMNS

    # Build rows - ONE ROW PER EXTRACTED VALUE (not one row per PDF)
    rows = []
    for pdf_num, extractions in sorted(all_extractions.items()):
        # extractions is now a list of (metric, value) tuples
        for metric, value in extractions:
            # Create one row per value
            row = {"PDF Number": pdf_num, metric: value}
            rows.append(row)

    df = pd.DataFrame(rows, columns=columns)

    # Save to Excel
    df.to_excel(final_dataset_path, index=False, engine='openpyxl')

    # Summary - count total values extracted (not just PDFs)
    total_values = sum(len(extractions) for extractions in all_extractions.values())
    print(f"{total_values} Metrics extracted from text")

    return metadata_excel_path, final_dataset_path


def find_latest_query_folder():
    """Find the latest Query folder in Output directory."""
    output_base = Path(__file__).parent.parent / "Output"
    query_folders = [d for d in output_base.iterdir() if d.is_dir() and "Query" in d.name]
    if not query_folders:
        return None
    return max(query_folders, key=lambda d: d.stat().st_mtime)


def main():
    """CLI entry point for LLM value extraction."""
    import argparse

    parser = argparse.ArgumentParser(description="Extract metric values using Qwen 2.5 7B LLM")
    parser.add_argument("extracted_text_folder", nargs='?', type=Path, help="Folder containing _filtered.txt files")

    args = parser.parse_args()

    # Auto-detect if no arguments provided
    if args.extracted_text_folder is None:
        query_folder = find_latest_query_folder()
        if query_folder is None:
            print("ERROR: No Query folder found in Output directory")
            sys.exit(1)

        args.extracted_text_folder = query_folder / "Extracted Text"

        print(f"Auto-detected Query folder: {query_folder.name}")

    if not args.extracted_text_folder.exists():
        print(f"ERROR: Extracted text folder not found: {args.extracted_text_folder.name}")
        sys.exit(1)

    extract_all_values(args.extracted_text_folder)


if __name__ == "__main__":
    main()
