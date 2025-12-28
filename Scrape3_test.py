import os
import re
import sys
from pathlib import Path
import fitz  # PyMuPDF
import requests
import openpyxl
from openpyxl import Workbook
import pdfplumber
import camelot
import tabula
import pytest

# ================== CONFIGURATION ==================
OLLAMA_URL = "http://localhost:11434/api/generate"
OLLAMA_MODEL = "gemma2:2b"
BASE_DIR = Path(
    r"C:\Users\azt12\OneDrive\Documents\Wrestling Robe\Materials Science - Wickability\Studies for Analysis by LLM AI\Datafiles & Python Scripts"
)
INPUT_EXCEL = BASE_DIR / "Simplified Table Format.xlsx"
OUTPUT_EXCEL = BASE_DIR / "Sample-Level-Metadata.xlsx"
PDF_FOLDER = Path(
    r"C:\Users\azt12\OneDrive\Documents\Wrestling Robe\Materials Science - Wickability\Studies for Analysis by LLM AI"
)

# ================== HELPER FUNCTIONS (not tests) ==================
def call_local_llm(prompt: str) -> str:
    """Call local Ollama LLM."""
    payload = {
        "model": OLLAMA_MODEL,
        "prompt": prompt,
        "stream": False,
    }
    resp = requests.post(OLLAMA_URL, json=payload, timeout=240)
    resp.raise_for_status()
    data = resp.json()
    return data.get("response", "")

def extract_title_with_formatting(pdf_path: str) -> str:
    """
    Extract title by finding the largest font size in first page.
    """
    doc = fitz.open(pdf_path)
    page = doc[0]
    blocks = page.get_text("dict")["blocks"]
    text_items = []
    for block in blocks:
        if "lines" in block:
            for line in block["lines"]:
                for span in line["spans"]:
                    text = span["text"].strip()
                    size = span["size"]
                    if text and len(text) > 3:
                        text_items.append(
                            {
                                "text": text,
                                "size": size,
                                "y": span["bbox"][1],
                                "x": span["bbox"][0],
                            }
                        )
    if not text_items:
        doc.close()
        return None
    text_items.sort(key=lambda x: (x["y"], x["x"]))
    skip_phrases = [
        "preprint",
        "not peer reviewed",
        "copyright",
        "license",
        "doi:",
        "arxiv",
        "biorxiv",
        "medrxiv",
        "journal of",
        "page ",
        "volume ",
        "issue ",
    ]
    candidates = []
    for item in text_items[:30]:
        if any(phrase in item["text"].lower() for phrase in skip_phrases):
            continue
        if len(item["text"]) < 10 and item["size"] < 15:
            continue
        candidates.append(item)
    if not candidates:
        doc.close()
        return None
    max_size = max(c["size"] for c in candidates)
    title_parts = []
    last_y = -999
    for item in candidates:
        if item["size"] >= max_size - 2:
            if abs(item["y"] - last_y) > 5:
                if title_parts and abs(item["y"] - last_y) > 20:
                    break
            title_parts.append(item["text"])
            last_y = item["y"]
        elif title_parts:
            break
    doc.close()
    if title_parts:
        full_title = " ".join(title_parts)
        full_title = re.sub(r"\s+", " ", full_title)
        return full_title.strip()
    return None

def extract_first_author_with_formatting(pdf_path: str) -> str:
    """
    Extract first author by finding text after the title.
    """
    doc = fitz.open(pdf_path)
    page = doc[0]
    blocks = page.get_text("dict")["blocks"]
    text_items = []
    for block in blocks:
        if "lines" in block:
            for line in block["lines"]:
                for span in line["spans"]:
                    text = span["text"].strip()
                    if text:
                        text_items.append(
                            {
                                "text": text,
                                "size": span["size"],
                                "y": span["bbox"][1],
                            }
                        )
    if not text_items:
        doc.close()
        return None
    text_items.sort(key=lambda x: x["y"])
    # Find max font size (title)
    max_size = max(item["size"] for item in text_items[:30])
    # Find where title ends
    title_end_idx = 0
    for i, item in enumerate(text_items[:30]):
        if item["size"] >= max_size - 2:
            title_end_idx = i + 1
    # Look for author in next 10 items
    for i in range(title_end_idx, min(title_end_idx + 10, len(text_items))):
        text = text_items[i]["text"]
        # Skip institutional lines
        if any(
            word in text.lower()
            for word in ["university", "school", "department", "@", "http"]
        ):
            continue
        # Look for any capitalized word that's 3+ letters (last name)
        words = text.split()
        for j, word in enumerate(words):
            # Clean the word
            clean_word = re.sub(r"[^A-Za-z]", "", word)
            # If it's capitalized and 3+ letters, it's probably a name
            if clean_word and clean_word[0].isupper() and len(clean_word) >= 3:
                # If next word is also capitalized, this is first name, take next word (last name)
                if j + 1 < len(words):
                    next_word = re.sub(r"[^A-Za-z]", "", words[j + 1])
                    if next_word and next_word[0].isupper() and len(next_word) >= 3:
                        doc.close()
                        return next_word  # Return last name
                # Otherwise just return this word
                doc.close()
                return clean_word
    doc.close()
    return None

def extract_year_from_text(text: str) -> int:
    """
    Extract publication year.
    """
    year_pattern = r"\b(19[9]\d|20[0-2]\d)\b"
    first_part = text[:6000]
    keywords_patterns = [
        (r"published[:\s]+.*?(\d{4})", 1),
        (r"received[:\s]+.*?(\d{4})", 1),
        (r"accepted[:\s]+.*?(\d{4})", 1),
        (r"copyright[:\s]+.*?(\d{4})", 1),
        (r"©[:\s]*(\d{4})", 1),
        (r"(\d{4})\s+published", 1),
        (r"online:?\s+.*?(\d{4})", 1),
        (r"\b(\d{4})\s+elsevier", 1),
        (r"\b(\d{4})\s+wiley", 1),
        (r"\b(\d{4})\s+springer", 1),
    ]
    for pattern, group in keywords_patterns:
        matches = re.finditer(pattern, first_part, re.IGNORECASE)
        for match in matches:
            year_str = match.group(group)
            year_int = int(year_str)
            if 1990 <= year_int <= 2025:
                return year_int
    lines = first_part.split("\n")
    for i, line in enumerate(lines[:50]):
        if re.match(r"^\s*\d{4}\s*$", line):
            year_int = int(line.strip())
            if 2000 <= year_int <= 2025:
                return year_int
    all_years = re.findall(year_pattern, first_part)
    if all_years:
        valid_years = [int(y) for y in all_years if 1990 <= int(y) <= 2025]
        if valid_years:
            from collections import Counter
            year_counts = Counter(valid_years)
            multiple_years = [y for y, count in year_counts.items() if count > 1]
            if multiple_years:
                return max(multiple_years)
            return max(valid_years)
    return None

def extract_table_row_count(pdf_path: str, table_number: int) -> int:
    """
    Extract row count from a specific table number in the PDF.
    Returns the number of data rows (excluding header) if found, None otherwise.
    Tries all three methods and shows what each finds for comparison.
    """
    print(f"\n    --- Extracting Table {table_number} ---")

    best_count = None
    best_method = None

    # Try CAMELOT
    print(f"\n    Trying CAMELOT...")
    try:
        tables_lattice = camelot.read_pdf(pdf_path, pages="all", flavor="lattice")
        tables_stream = camelot.read_pdf(pdf_path, pages="all", flavor="stream")
        all_camelot_tables = list(tables_lattice) + list(tables_stream)

        if table_number <= len(all_camelot_tables):
            table = all_camelot_tables[table_number - 1]
            df = table.df
            rows = len(df) - 1  # Exclude header
            print(f"    Camelot: Found Table {table_number} with {rows} data rows")
            print(f"    First column values:")
            for idx, val in enumerate(df.iloc[:, 0]):
                print(f"      Row {idx}: {val}")
            if rows > 0 and not best_count:
                best_count = rows
                best_method = "Camelot"
        else:
            print(f"    Camelot: Table {table_number} not found (only {len(all_camelot_tables)} tables)")
    except Exception as e:
        print(f"    Camelot ERROR: {e}")

    # Try TABULA
    print(f"\n    Trying TABULA...")
    try:
        dfs = tabula.read_pdf(pdf_path, pages="all", multiple_tables=True, silent=True)
        if table_number <= len(dfs):
            df = dfs[table_number - 1]
            rows = len(df) - 1 if len(df) > 1 else len(df)
            print(f"    Tabula: Found Table {table_number} with {rows} data rows")
            print(f"    First column values:")
            for idx, val in enumerate(df.iloc[:, 0]):
                print(f"      Row {idx}: {val}")
            if rows > 0 and not best_count:
                best_count = rows
                best_method = "Tabula"
        else:
            print(f"    Tabula: Table {table_number} not found (only {len(dfs)} tables)")
    except Exception as e:
        print(f"    Tabula ERROR: {e}")

    # Try PDFPLUMBER
    print(f"\n    Trying PDFPLUMBER...")
    try:
        with pdfplumber.open(pdf_path) as pdf:
            table_count = 0
            found = False
            for page_num, page in enumerate(pdf.pages):
                tables = page.extract_tables()
                for table in tables:
                    table_count += 1
                    if table_count == table_number:
                        rows = len(table) - 1  # Exclude header
                        print(f"    PDFPlumber: Found Table {table_number} on page {page_num + 1} with {rows} data rows")
                        print(f"    First column values:")
                        for idx, row in enumerate(table):
                            first_cell = row[0] if row else ""
                            print(f"      Row {idx}: {first_cell}")
                        if rows > 0 and not best_count:
                            best_count = rows
                            best_method = "PDFPlumber"
                        found = True
                        break
                if found:
                    break
            if not found:
                print(f"    PDFPlumber: Table {table_number} not found")
    except Exception as e:
        print(f"    PDFPlumber ERROR: {e}")

    if best_count:
        print(f"\n    ✓ Best result: {best_count} rows from {best_method}")
        return best_count
    else:
        print(f"\n    ✗ Could not extract Table {table_number} from any method")
        return None

def extract_sample_count_from_table(pdf_path: str, full_text: str) -> int:
    """
    Extract number of samples with SIMPLE, CLEAR, VISIBLE logic.
    """
    print("\n    === SAMPLE DETECTION DIAGNOSTICS ===")
    search_text = full_text
    print(f"    Searching through {len(search_text)} characters of text")
    
    # ==================================================================
    # PRIORITY 1: EXPLICIT COUNT - GROUP LOGIC
    # ==================================================================
    print("\n    PRIORITY 1: Looking for EXPLICIT COUNT statements...")
    print("    GROUP 1a: 'total of' + number immediately following")
    print("    GROUP 1b: number that immediately precedes a Group 2 term (with optional words)")
    print(
        "    GROUP 2: fabrics/materials/variants/garments/samples/textiles/specimens (singular + plural)"
    )
    print(
        "    GROUP 3: tested/produced/used/analyzed/evaluated/studied/prepared/examined"
    )
    print("    Numbers: Arabic (1-100), Roman (I-L), Words (one-fifty)")
    print("    ")
    print("    Valid combinations:")
    print("    - Group 1a OR 1b + Group 2 + Group 3 (BEST)")
    print("    - Group 2 + Group 3 (COMMON)")
    print("    - Group 1a OR 1b + Group 2 (COMMON)")
    print("    - Group 1a OR 1b + Group 3 (RARE)")
    
    explicit_count = None
    # Split on period/question/exclamation followed by space and capital letter
    # This handles both "word. Next" and "word.\nNext" patterns
    sentences = re.split(r'[.!?](?=\s+[A-Z])', search_text)
    print(f"    Split text into {len(sentences)} sentences")
    
    # ==================================================================
    # PRIORITY 0: SAMPLE NUMBER COLUMN - SOURCE OF TRUTH
    # ==================================================================
    print("\n    PRIORITY 0: Looking for tables with Sample Number columns...")
    print("    Searching for column headers: 'Sample No.', 'S. No', 'S. Number', 'Sample Number'")

    sample_column_headers = ["sample no.", "s. no", "s. number", "sample number", "sample no", "s.no", "s.no."]

    # Try to find a table with sample number column header
    try:
        # Try CAMELOT
        tables_lattice = camelot.read_pdf(pdf_path, pages="all", flavor="lattice")
        tables_stream = camelot.read_pdf(pdf_path, pages="all", flavor="stream")
        all_tables = list(tables_lattice) + list(tables_stream)

        for idx, table in enumerate(all_tables):
            df = table.df
            if len(df) == 0:
                continue

            # Check if any column header matches our sample number patterns
            header_row = df.iloc[0].astype(str).str.lower()
            for col_idx, header in enumerate(header_row):
                if any(sample_header in header for sample_header in sample_column_headers):
                    rows = len(df) - 1  # Exclude header
                    print(f"\n    ✓✓✓ FOUND SAMPLE NUMBER COLUMN ✓✓✓")
                    print(f"    Table {idx + 1}, Column '{df.iloc[0, col_idx]}' matches sample number pattern")
                    print(f"    Rows in this table: {rows}")
                    print(f"    First column values:")
                    for row_idx, val in enumerate(df.iloc[:, col_idx]):
                        print(f"      Row {row_idx}: {val}")

                    if rows > 0:
                        print("\n    ═══════════════════════════════════════════════")
                        print(f"    DECISION: SAMPLE NUMBER COLUMN = {rows}")
                        print("    THIS IS THE SOURCE OF TRUTH - RETURNING NOW")
                        print("    ═══════════════════════════════════════════════\n")
                        return rows
    except Exception as e:
        print(f"    ERROR searching for sample number columns: {e}")

    print("    Result: NO sample number column found")

    # Number word to digit mapping - EVERY NUMBER 1-50 (CASE-INSENSITIVE)
    # Include both lowercase and capitalized versions
    word_to_num = {
        "one": 1, "One": 1,
        "two": 2, "Two": 2,
        "three": 3, "Three": 3,
        "four": 4, "Four": 4,
        "five": 5, "Five": 5,
        "six": 6, "Six": 6,
        "seven": 7, "Seven": 7,
        "eight": 8, "Eight": 8,
        "nine": 9, "Nine": 9,
        "ten": 10, "Ten": 10,
        "eleven": 11, "Eleven": 11,
        "twelve": 12, "Twelve": 12,
        "thirteen": 13, "Thirteen": 13,
        "fourteen": 14, "Fourteen": 14,
        "fifteen": 15, "Fifteen": 15,
        "sixteen": 16, "Sixteen": 16,
        "seventeen": 17, "Seventeen": 17,
        "eighteen": 18, "Eighteen": 18,
        "nineteen": 19, "Nineteen": 19,
        "twenty": 20, "Twenty": 20,
        "twenty-one": 21, "Twenty-one": 21,
        "twenty-two": 22, "Twenty-two": 22,
        "twenty-three": 23, "Twenty-three": 23,
        "twenty-four": 24, "Twenty-four": 24,
        "twenty-five": 25, "Twenty-five": 25,
        "twenty-six": 26, "Twenty-six": 26,
        "twenty-seven": 27, "Twenty-seven": 27,
        "twenty-eight": 28, "Twenty-eight": 28,
        "twenty-nine": 29, "Twenty-nine": 29,
        "thirty": 30, "Thirty": 30,
        "thirty-one": 31, "Thirty-one": 31,
        "thirty-two": 32, "Thirty-two": 32,
        "thirty-three": 33, "Thirty-three": 33,
        "thirty-four": 34, "Thirty-four": 34,
        "thirty-five": 35, "Thirty-five": 35,
        "thirty-six": 36, "Thirty-six": 36,
        "thirty-seven": 37, "Thirty-seven": 37,
        "thirty-eight": 38, "Thirty-eight": 38,
        "thirty-nine": 39, "Thirty-nine": 39,
        "forty": 40, "Forty": 40,
        "forty-one": 41, "Forty-one": 41,
        "forty-two": 42, "Forty-two": 42,
        "forty-three": 43, "Forty-three": 43,
        "forty-four": 44, "Forty-four": 44,
        "forty-five": 45, "Forty-five": 45,
        "forty-six": 46, "Forty-six": 46,
        "forty-seven": 47, "Forty-seven": 47,
        "forty-eight": 48, "Forty-eight": 48,
        "forty-nine": 49, "Forty-nine": 49,
        "fifty": 50, "Fifty": 50,
    }
    
    # Roman numerals - EVERY NUMBER 1-50
    roman_to_num = {
        "i": 1,
        "ii": 2,
        "iii": 3,
        "iv": 4,
        "v": 5,
        "vi": 6,
        "vii": 7,
        "viii": 8,
        "ix": 9,
        "x": 10,
        "xi": 11,
        "xii": 12,
        "xiii": 13,
        "xiv": 14,
        "xv": 15,
        "xvi": 16,
        "xvii": 17,
        "xviii": 18,
        "xix": 19,
        "xx": 20,
        "xxi": 21,
        "xxii": 22,
        "xxiii": 23,
        "xxiv": 24,
        "xxv": 25,
        "xxvi": 26,
        "xxvii": 27,
        "xxviii": 28,
        "xxix": 29,
        "xxx": 30,
        "xxxi": 31,
        "xxxii": 32,
        "xxxiii": 33,
        "xxxiv": 34,
        "xxxv": 35,
        "xxxvi": 36,
        "xxxvii": 37,
        "xxxviii": 38,
        "xxxix": 39,
        "xl": 40,
        "xli": 41,
        "xlii": 42,
        "xliii": 43,
        "xliv": 44,
        "xlv": 45,
        "xlvi": 46,
        "xlvii": 47,
        "xlviii": 48,
        "xlix": 49,
        "l": 50,
    }
    
    # ===== TABLE FALLBACK: Group 2 term + "table" =====
    # If a sentence has a Group 2 term + "table", parse the table number and store as fallback
    # Don't return immediately - keep looking for explicit counts
    print("\n    TABLE FALLBACK: Checking for Group 2 term + 'table'...")

    group2_words = [
        "fabric", "fabrics", "material", "materials", "variant", "variants",
        "garment", "garments", "sample", "samples", "textile", "textiles",
        "specimen", "specimens",
    ]

    table_fallback_count = None

    for i, sentence in enumerate(sentences):
        sentence_lower = sentence.lower()

        # Check if sentence has a Group 2 term AND "table"
        group2_found_here = [word for word in group2_words if word in sentence_lower]
        has_group2 = len(group2_found_here) > 0
        table_match = re.search(r"table\s+(\d+)", sentence_lower, re.IGNORECASE)

        if has_group2 and table_match:
            table_num = int(table_match.group(1))
            print(f"\n    ✓ Found Group 2 term + Table {table_num}")
            print(f"    Sentence {i}: '{sentence.strip()[:300]}...'")
            group2_terms_list = ", ".join(group2_found_here)
            print(f"    Group 2 Terms Present: {group2_terms_list}")
            print(f"    Attempting to parse Table {table_num}...")

            # Try to extract and count rows from the specified table
            table_count = extract_table_row_count(pdf_path, table_num)
            if table_count and not table_fallback_count:
                table_fallback_count = table_count
                print(f"    ✓ Extracted {table_count} rows from Table {table_num} (stored as fallback)")
            elif not table_count:
                print(f"    ✗ Could not parse Table {table_num}")

    if table_fallback_count:
        print(f"\n    Stored table fallback count: {table_fallback_count}")
    else:
        print("    Result: NO valid table reference found")
    
    # ===== SUPERSEDING CASE 2: "total of" + NUMBER immediately followed by Group 2 =====
    # This takes absolute priority - if found, use it and stop searching
    print("\n    SUPERSEDING CASE 2: Checking for 'total of' + number + Group 2 term...")
    
    word_pattern_check = "|".join(word_to_num.keys())
    
    for i, sentence in enumerate(sentences):
        sentence_lower = sentence.lower()
        
        # Check for "total of" + arabic number + group2
        total_of_arabic = re.search(
            r"total\s+of\s+(\d+)\s+(?:different\s+)?(?:types?\s+of\s+)?(?:fabrics?|materials?|samples?|variants?|garments?|textiles?|specimens?)\b",
            sentence_lower
        )
        if total_of_arabic:
            num = int(total_of_arabic.group(1))
            if 1 <= num <= 100:
                print(f"\n    ✓✓✓ SUPERSEDING CASE FOUND - 'total of' + {num} ✓✓✓")
                print(f"    Sentence {i}: '{sentence.strip()[:200]}...'")
                print("\n    ═══════════════════════════════════════════════")
                print(f"    DECISION: EXPLICIT COUNT = {num}")
                print("    SUPERSEDES ALL OTHER METHODS - RETURNING NOW")
                print("    ═══════════════════════════════════════════════\n")
                return num
        
        # Check for "total of" + word number + group2 (CASE-INSENSITIVE)
        total_of_word = re.search(
            rf"total\s+of\s+({word_pattern_check})\s+(?:different\s+)?(?:types?\s+of\s+)?(?:fabrics?|materials?|samples?|variants?|garments?|textiles?|specimens?)\b",
            sentence_lower
        )
        if total_of_word:
            word_num = total_of_word.group(1)
            if word_num in word_to_num:
                num = word_to_num[word_num]
                print(f"\n    ✓✓✓ SUPERSEDING CASE FOUND - 'total of' + '{word_num}' → {num} ✓✓✓")
                print(f"    Sentence {i}: '{sentence.strip()[:200]}...'")
                print("\n    ═══════════════════════════════════════════════")
                print(f"    DECISION: EXPLICIT COUNT = {num}")
                print("    SUPERSEDES ALL OTHER METHODS - RETURNING NOW")
                print("    ═══════════════════════════════════════════════\n")
                return num
    
    print("    Result: NO 'total of' + number + Group 2 found")
    
    # ===== MAIN GROUP LOGIC =====
    print("\n    MAIN GROUP LOGIC: Checking all sentences for valid combinations...")
    
    for i, sentence in enumerate(sentences):
        sentence_lower = sentence.lower()

        # CHECK GROUP 1: "total of" (1a) OR number immediately before group 2 (1b)
        # These are TWO SEPARATE cases
        has_total_of = "total of" in sentence_lower

        # Check if WHOLE, POSITIVE number IMMEDIATELY precedes group 2 words
        # IMMEDIATELY means the very next word, with NO words in between
        # Pattern: number + optional whitespace + Group 2 word (NO words in between)
        has_number_before_group2_check = False

        # STRICT pattern - number must be followed by Group 2 term with NO intervening words
        # Only whitespace allowed between number and Group 2 term
        if re.search(
            r"(?<![0-9.])(\d+)(?![0-9.])\s+(?:fabrics?|materials?|samples?|variants?|garments?|textiles?|specimens?)\b",
            sentence_lower
        ):
            has_number_before_group2_check = True

        # Check for word numbers IMMEDIATELY before Group 2 (CASE-INSENSITIVE)
        if re.search(
            rf"\b({word_pattern_check})\s+(?:fabrics?|materials?|samples?|variants?|garments?|textiles?|specimens?)\b",
            sentence_lower
        ):
            has_number_before_group2_check = True

        has_group1 = has_total_of or has_number_before_group2_check

        # CHECK GROUP 2: BOTH SINGULAR AND PLURAL - track which terms found
        group2_words = [
            "fabric",
            "fabrics",
            "material",
            "materials",
            "variant",
            "variants",
            "garment",
            "garments",
            "sample",
            "samples",
            "textile",
            "textiles",
            "specimen",
            "specimens",
        ]
        group2_found = [word for word in group2_words if word in sentence_lower]
        has_group2 = len(group2_found) > 0

        # CHECK GROUP 3: action words - track which terms found
        group3_words = [
            "tested",
            "produced",
            "used",
            "analyzed",
            "evaluated",
            "studied",
            "prepared",
            "examined",
        ]
        group3_found = [word for word in group3_words if word in sentence_lower]
        has_group3 = len(group3_found) > 0
        
        # VALID COMBINATION CHECK
        # Group 1 is satisfied by EITHER 1a (total of) OR 1b (number before Group 2)
        valid_combination = False
        combination_type = ""
        
        if has_group1 and has_group2 and has_group3:
            valid_combination = True
            combination_type = "Group 1 + 2 + 3 (BEST)"
        elif has_group2 and has_group3:
            valid_combination = True
            combination_type = "Group 2 + 3 (COMMON)"
        elif has_group1 and has_group2:
            valid_combination = True
            combination_type = "Group 1 + 2 (COMMON)"
        elif has_group1 and has_group3:
            valid_combination = True
            combination_type = "Group 1 + 3 (RARE)"
        
        if not valid_combination:
            continue
        
        print(f"\n     Sentence {i}: VALID COMBINATION - {combination_type}")
        print(f"     Text: '{sentence.strip()[:250]}'")

        # Show what triggered Group 1
        if has_total_of:
            # Find where "total of" appears
            total_of_pos = sentence_lower.find("total of")
            context = sentence_lower[max(0, total_of_pos-20):min(len(sentence_lower), total_of_pos+80)]
            print(f"     Has Group 1 (total of): True - Found at: '...{context}...'")
        else:
            print(f"     Has Group 1 (total of): False")

        group2_terms = ", ".join(group2_found) if group2_found else "none"
        print(f"     Has Group 2: {has_group2}. Terms present: {group2_terms}")
        group3_terms = ", ".join(group3_found) if group3_found else "none"
        print(f"     Has Group 3: {has_group3}. Terms present: {group3_terms}")
        print(f"     Has number IMMEDIATELY before group 2: {has_number_before_group2_check}")
        
        # ===== PRIORITY 1: ARABIC NUMERALS (most common) =====
        # Must be WHOLE, POSITIVE numbers only
        # Number must be IMMEDIATELY before Group 2 term (no words in between)
        digit_patterns = [
            r"total\s+of\s+(?<![0-9.])(\d+)(?![0-9.])\s+(?:fabrics?|materials?|samples?|variants?|garments?|textiles?|specimens?)",
            r"(?<![0-9.])(\d+)(?![0-9.])\s+(?:fabrics?|materials?|samples?|variants?|garments?|textiles?|specimens?)",
            r"(?:tested|produced|used|analyzed|evaluated|studied|prepared|examined)\s+(?<![0-9.])(\d+)(?![0-9.])\s+(?:fabrics?|materials?|samples?|variants?|garments?|textiles?|specimens?)",
        ]
        
        for pattern in digit_patterns:
            match = re.search(pattern, sentence_lower)
            if match:
                num = int(match.group(1))
                if 1 <= num <= 100:
                    if not explicit_count or num > explicit_count:
                        explicit_count = num
                        print(f"\n    ✓ FOUND arabic numeral: {num}")
                        print(f"     Sentence: '{sentence.strip()[:200]}...'")
                    break
        
        # ===== PRIORITY 2: WORD NUMBERS (one, two, three, etc.) - CASE-INSENSITIVE =====
        # Number word must be IMMEDIATELY before Group 2 term (no words in between)
        word_patterns = [
            rf"\b({word_pattern_check})\b\s+(?:fabrics?|materials?|samples?|variants?|garments?|textiles?|specimens?)",
            rf"\b({word_pattern_check})\b\s+(?:fabrics?|materials?|samples?|variants?|garments?|textiles?|specimens?)\s+of",
        ]
        
        for pattern in word_patterns:
            match = re.search(pattern, sentence_lower)
            if match:
                word_num = match.group(1)
                if word_num in word_to_num:
                    found_count = word_to_num[word_num]
                    if not explicit_count or found_count > explicit_count:
                        explicit_count = found_count
                        print(f"\n    ✓ FOUND word number: '{word_num}' → {found_count}")
                        print(f"     Sentence: '{sentence.strip()[:200]}...'")
                    break
        
        # ===== PRIORITY 3: ROMAN NUMERALS (rare but possible) =====
        roman_pattern = "|".join(roman_to_num.keys())
        roman_patterns = [
            rf"(?:^|\s)({roman_pattern})(?:\s+)(?:fabrics?|materials?|samples?|variants?|garments?|textiles?|specimens?)\b",
        ]
        
        for pattern in roman_patterns:
            match = re.search(pattern, sentence_lower)
            if match:
                roman_num = match.group(1)
                if roman_num in roman_to_num:
                    found_count = roman_to_num[roman_num]
                    if not explicit_count or found_count > explicit_count:
                        explicit_count = found_count
                        print(f"\n    ✓ FOUND roman numeral: '{roman_num.upper()}' → {found_count}")
                        print(f"     Sentence: '{sentence.strip()[:200]}...'")
                    break
    
    if explicit_count:
        print("\n    ═══════════════════════════════════════════════")
        print(f"    DECISION: EXPLICIT COUNT = {explicit_count}")
        print("    SUPERSEDES ALL OTHER METHODS - RETURNING NOW")
        print("    ═══════════════════════════════════════════════\n")
        return explicit_count

    print("\n    Result: NO explicit count found using group logic")

    # If we found a table fallback count earlier, use it
    if table_fallback_count:
        print("\n    ═══════════════════════════════════════════════")
        print(f"    DECISION: TABLE FALLBACK COUNT = {table_fallback_count}")
        print("    Using table row count from earlier")
        print("    ═══════════════════════════════════════════════\n")
        return table_fallback_count
    
    # ==================================================================
    # PRIORITY 2: SINGLE SAMPLE
    # ==================================================================
    print("\n    PRIORITY 2: Checking for single-sample study...")
    single_sample_patterns = [
        r"\ba\s+(?:smart|novel|new|single)\s+fabrics?\b",
        r"\bone\s+fabrics?\b",
        r"\bsingle\s+(?:fabrics?|materials?|samples?)\b",
    ]
    
    for pattern in single_sample_patterns:
        if re.search(pattern, search_text, re.IGNORECASE):
            match = re.search(pattern, search_text, re.IGNORECASE)
            context_start = max(0, match.start() - 50)
            context_end = min(len(search_text), match.end() + 50)
            context = search_text[context_start:context_end]
            print(f"    ✓ Found: '{context}'")
            print("\n    ═══════════════════════════════════════════════")
            print("    DECISION: SINGLE SAMPLE STUDY = 1")
            print("    ═══════════════════════════════════════════════\n")
            return 1
    
    print("    Result: NOT a single sample study")
    
    # ==================================================================
    # PRIORITY 3: TABLE REFERENCE
    # ==================================================================
    print("\n    PRIORITY 3: Looking for table references...")
    print(
        "    Required in same sentence: (fabric/fabrics OR material/materials OR sample/samples OR variant/variants OR garment/garments OR textile/textiles OR specimen/specimens) AND 'table'"
    )
    
    table_numbers = set()
    
    for i, sentence in enumerate(sentences):
        sentence_lower = sentence.lower()
        has_sample_word = any(
            word in sentence_lower
            for word in [
                "fabric",
                "fabrics",
                "material",
                "materials",
                "sample",
                "samples",
                "garment",
                "garments",
                "textile",
                "textiles",
                "specimen",
                "specimens",
                "variant",
                "variants",
            ]
        )
        has_table_ref = re.search(r"table\s+(\d+)", sentence_lower, re.IGNORECASE)
        
        if has_sample_word and has_table_ref:
            table_num = has_table_ref.group(1)
            table_numbers.add(table_num)
            print(f"    ✓ Sentence {i}: Has sample word + Table {table_num}")
            print(f"     Text: '{sentence.strip()[:200]}...'")
    
    if not table_numbers:
        print("    Result: NO table reference found")
        print("\n    ═══════════════════════════════════════════════")
        print("    DECISION: No indicators found = 1")
        print("    ═══════════════════════════════════════════════\n")
        return 1
    
    print(f"\n    ✓ Found table references: {sorted(table_numbers)}")
    
    # ==================================================================
    # PRIORITY 4: EXTRACT TABLES
    # ==================================================================
    print("\n    PRIORITY 4: Extracting tables to count rows...")
    best_count = None
    best_method = None
    
    # CAMELOT
    print("\n    --- Method A: CAMELOT ---")
    try:
        tables_lattice = camelot.read_pdf(pdf_path, pages="all", flavor="lattice")
        tables_stream = camelot.read_pdf(pdf_path, pages="all", flavor="stream")
        all_camelot_tables = list(tables_lattice) + list(tables_stream)
        print(f"    Found {len(all_camelot_tables)} total tables")
        
        for idx, table in enumerate(all_camelot_tables):
            df = table.df
            rows = len(df) - 1
            print(f"\n    Table {idx + 1}:")
            print(f"     Dimensions: {len(df)} rows × {len(df.columns)} cols")
            print(f"     HEADER ROW: {str(df.iloc[0].tolist())[:150]}...")
            if len(df) > 1:
                print(f"     DATA ROW 1: {str(df.iloc[1].tolist())[:150]}...")
            
            header_text = " ".join(df.iloc[0].astype(str)).lower()
            fabric_keywords = [
                "fabric",
                "fabrics",
                "material",
                "materials",
                "sample",
                "samples",
                "fiber",
                "fibers",
                "composition",
                "compositions",
                "thickness",
                "density",
                "densities",
                "weight",
                "weights",
                "gsm",
                "structure",
                "structures",
                "weave",
                "weaves",
                "knit",
                "knits",
                "variant",
                "variants",
                "garment",
                "garments",
                "textile",
                "textiles",
                "specimen",
                "specimens",
            ]
            score = sum([kw in header_text for kw in fabric_keywords])
            print(f"     Fabric keyword score: {score}/{len(fabric_keywords)}")
            
            if score >= 2 and rows > 0:
                print("     ✓✓✓ THIS LOOKS LIKE THE SAMPLE TABLE ✓✓✓")
                if not best_count:
                    best_count = rows
                    best_method = f"Camelot (table {idx + 1})"
            else:
                print(f"     ✗ REJECTED (score {score} < 2 or rows={rows})")
    except Exception as e:
        print(f"    CameLot ERROR: {e}")
    
    # TABULA
    if not best_count:
        print("\n    --- Method B: TABULA ---")
        try:
            dfs = tabula.read_pdf(pdf_path, pages="all", multiple_tables=True, silent=True)
            print(f"    Found {len(dfs)} tables")
            
            for idx, df in enumerate(dfs):
                rows = len(df) - 1 if len(df) > 1 else len(df)
                print(f"\n    Table {idx + 1}:")
                print(f"     Dimensions: {len(df)} rows × {len(df.columns)} cols")
                print(f"     COLUMNS: {str(df.columns.tolist())[:150]}...")
                if len(df) > 0:
                    print(f"     DATA ROW 1: {str(df.iloc[0].tolist())[:150]}...")
                
                header_text = " ".join([str(col).lower() for col in df.columns])
                fabric_keywords = [
                    "fabric",
                    "fabrics",
                    "material",
                    "materials",
                    "sample",
                    "samples",
                    "fiber",
                    "fibers",
                    "composition",
                    "compositions",
                    "thickness",
                    "density",
                    "densities",
                    "weight",
                    "weights",
                    "gsm",
                    "structure",
                    "structures",
                    "weave",
                    "weaves",
                    "knit",
                    "knits",
                    "variant",
                    "variants",
                    "garment",
                    "garments",
                    "textile",
                    "textiles",
                    "specimen",
                    "specimens",
                ]
                score = sum([kw in header_text for kw in fabric_keywords])
                print(f"     Fabric keyword score: {score}/{len(fabric_keywords)}")
                
                if score >= 2 and rows > 0:
                    print("     ✓✓✓ THIS LOOKS LIKE THE SAMPLE TABLE ✓✓✓")
                    if not best_count:
                        best_count = rows
                        best_method = f"Tabula (table {idx + 1})"
                else:
                    print(f"     ✗ REJECTED (score {score} < 2 or rows={rows})")
        except Exception as e:
            print(f"    Tabula ERROR: {e}")
    
    # PDFPLUMBER
    if not best_count:
        print("\n    --- Method C: PDFPLUMBER ---")
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    tables = page.extract_tables()
                    if not tables:
                        continue
                    print(f"    Page {page_num + 1}: Found {len(tables)} table(s)")
                    
                    for t_idx, table in enumerate(tables):
                        if not table or len(table) <= 1:
                            continue
                        rows = len(table) - 1
                        print(f"\n    Page {page_num + 1}, Table {t_idx + 1}:")
                        print(f"     Dimensions: {len(table)} rows × {len(table[0])} cols")
                        print(f"     HEADER: {str(table[0])[:150]}...")
                        if len(table) > 1:
                            print(f"     DATA ROW 1: {str(table[1])[:150]}...")
                        
                        header_text = " ".join(
                            [str(cell).lower() if cell else "" for cell in table[0]]
                        )
                        fabric_keywords = [
                            "fabric",
                            "fabrics",
                            "material",
                            "materials",
                            "sample",
                            "samples",
                            "fiber",
                            "fibers",
                            "composition",
                            "compositions",
                            "thickness",
                            "density",
                            "densities",
                            "weight",
                            "weights",
                            "gsm",
                            "structure",
                            "structures",
                            "weave",
                            "weaves",
                            "knit",
                            "knits",
                            "variant",
                            "variants",
                            "garment",
                            "garments",
                            "textile",
                            "textiles",
                            "specimen",
                            "specimens",
                        ]
                        score = sum([kw in header_text for kw in fabric_keywords])
                        print(f"     Fabric keyword score: {score}/{len(fabric_keywords)}")
                        
                        if score >= 2:
                            print("     ✓✓✓ THIS LOOKS LIKE THE SAMPLE TABLE ✓✓✓")
                            if not best_count:
                                best_count = rows
                                best_method = f"pdfplumber (page {page_num + 1}, table {t_idx + 1})"
                        else:
                            print(f"     ✗ REJECTED (score {score} < 2)")
        except Exception as e:
            print(f"    pdfplumber ERROR: {e}")
    
    # FINAL DECISION
    if best_count:
        print("\n    ═══════════════════════════════════════════════")
        print(f"    DECISION: Extracted {best_count} rows")
        print(f"    Method: {best_method}")
        print("    ═══════════════════════════════════════════════\n")
        return best_count
    else:
        print("\n    ═══════════════════════════════════════════════")
        print("    DECISION: All extraction methods failed = 1")
        print("    ═══════════════════════════════════════════════\n")
        return 1

# ================== TEST FUNCTIONS ==================
def test_process_pdfs():
    """Test function to process all PDFs and extract metadata."""
    print(f"Loading input Excel: {INPUT_EXCEL}")
    
    if not INPUT_EXCEL.exists():
        print(f"ERROR: Input file not found -> {INPUT_EXCEL}")
        pytest.fail(f"Input file not found: {INPUT_EXCEL}")
    
    wb_input = openpyxl.load_workbook(INPUT_EXCEL)
    ws_input = wb_input.active
    
    study_ids = []
    for row in ws_input.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            study_ids.append(str(row[0]).strip())
    
    print(f"Found {len(study_ids)} study IDs in input Excel")
    
    # ================== PROCESS EACH PDF ==================
    output_columns = [
        "Study Number",
        "Study title",
        "Year of Publish",
        "Name of first-listed author",
        "Number of Sample Fabrics",
    ]
    output_rows = []
    
    for study_id in study_ids:
        print(f"\n{'=' * 60}")
        print(f"Processing Study {study_id}")
        print(f"{'=' * 60}")
        
        # Find PDF
        pdf_path = PDF_FOLDER / f"{study_id}.pdf"
        if not os.path.exists(pdf_path):
            print(f"  PDF not found -> {pdf_path}")
            output_rows.append(
                {
                    "Study Number": study_id,
                    "Study title": "PDF not found",
                    "Year of Publish": "N/A",
                    "Name of first-listed author": "N/A",
                    "Number of Sample Fabrics": 0,
                }
            )
            continue
        
        # Extract full text
        doc = fitz.open(pdf_path)
        full_text = ""
        for page in doc:
            full_text += page.get_text()
        doc.close()
        print(f"  Extracted {len(full_text)} characters of text")
        
        # --------- EXTRACT TITLE ---------
        meta_title = extract_title_with_formatting(str(pdf_path))
        print(f"  Title: {meta_title}")
        
        # --------- EXTRACT YEAR ---------
        meta_year = extract_year_from_text(full_text)
        print(f"  Year: {meta_year}")
        
        # --------- EXTRACT AUTHOR ---------
        meta_first_last = extract_first_author_with_formatting(str(pdf_path))
        print(f"  Author: {meta_first_last}")
        
        # --------- EXTRACT NUMBER OF SAMPLES (REGEX) ---------
        sample_count = extract_sample_count_from_table(str(pdf_path), full_text)
        print(f"  Samples: {sample_count}")
        
        # --------- TRY LLM FOR MISSING DATA ---------
        missing = []
        if not meta_title:
            missing.append("title")
        if not meta_year:
            missing.append("year")
        if not meta_first_last:
            missing.append("author")
        
        if missing:
            print(
                f"  Missing: {', '.join(missing) if missing else 'none, but trying LLM for samples'}"
            )
            print("  Trying LLM...")
            short_text = full_text[:4000]
            
            # LLM for title
            if "title" in missing:
                try:
                    title_prompt = f"What is the main article title? Return only the title.\n\nTEXT:\n{short_text[:2000]}"
                    llm_title = call_local_llm(title_prompt).strip()
                    if llm_title and len(llm_title) > 10 and len(llm_title) < 300:
                        meta_title = llm_title
                        print(f"    LLM found title: {meta_title}")
                except Exception as e:
                    print(f"    LLM title error: {e}")
            
            # LLM for year
            if "year" in missing:
                try:
                    year_prompt = f"What year was this published? Return only a 4-digit year.\n\nTEXT:\n{short_text[:2000]}"
                    llm_year = call_local_llm(year_prompt).strip()
                    year_match = re.search(r"\b(19[9]\d|20[0-2]\d)\b", llm_year)
                    if year_match:
                        meta_year = int(year_match.group(1))
                        print(f"    LLM found year: {meta_year}")
                except Exception as e:
                    print(f"    LLM year error: {e}")
            
            # LLM for author
            if "author" in missing:
                try:
                    author_prompt = f"Who is the first author? Return only the last name (family name).\n\nTEXT:\n{short_text}"
                    llm_author = call_local_llm(author_prompt).strip()
                    llm_author = llm_author.replace('"', "").replace("'", "").strip()
                    author_match = re.search(r"\b([A-Z][a-z]{2,})\b", llm_author)
                    if author_match:
                        meta_first_last = author_match.group(1)
                        print(f"    LLM found author: {meta_first_last}")
                except Exception as e:
                    print(f"    LLM author error: {e}")
        
        # --------- SET DEFAULTS FOR MISSING VALUES ---------
        if not meta_title:
            meta_title = "Title not extracted"
        if not meta_year:
            meta_year = "Year not found"
        if not meta_first_last:
            meta_first_last = "Author not found"
        
        # --------- APPLY 100 CHARACTER LIMIT ---------
        if meta_title and len(str(meta_title)) > 100:
            meta_title = str(meta_title)[:100]
        if meta_first_last and len(str(meta_first_last)) > 100:
            meta_first_last = str(meta_first_last)[:100]
        
        print("\n  FINAL METADATA:")
        print(f"    Title: {meta_title}")
        print(f"    Year: {meta_year}")
        print(f"    Author: {meta_first_last}")
        print(f"    Samples: {sample_count}")
        
        # --------- ADD TO OUTPUT ---------
        output_rows.append(
            {
                "Study Number": study_id,
                "Study title": meta_title,
                "Year of Publish": meta_year,
                "Name of first-listed author": meta_first_last,
                "Number of Sample Fabrics": sample_count,
            }
        )
    
    # ================== WRITE OUTPUT EXCEL ==================
    print(f"\n{'=' * 60}")
    print(f"Writing output to: {OUTPUT_EXCEL}")
    print(f"{'=' * 60}")
    
    wb_output = Workbook()
    ws_output = wb_output.active
    ws_output.title = "Metadata"
    
    # Write header
    ws_output.append(output_columns)
    
    # Write data rows
    for row_dict in output_rows:
        row_values = [row_dict.get(col, "") for col in output_columns]
        ws_output.append(row_values)
    
    wb_output.save(OUTPUT_EXCEL)
    print(f"✓ Output saved: {OUTPUT_EXCEL}")
    print(f"✓ Processed {len(output_rows)} studies")

if __name__ == "__main__":
    # Run the test function directly when script is executed
    test_process_pdfs()
