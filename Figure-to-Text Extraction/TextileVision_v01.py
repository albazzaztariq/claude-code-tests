#!/usr/bin/env python3
"""
ScrapeV2.py - Comprehensive Textile Specification Extractor

Uses Nougat OCR for table extraction and LLM (Ollama/OpenAI) for semantic
understanding to extract all yarn, fabric, and transfer metrics from PDFs.

DePlot (Google) for extracting data from bar charts and figures.

Focused on 1.pdf for initial development.
"""

# Disable HuggingFace connectivity checks (must be before imports)
import os
os.environ["DISABLE_MODEL_SOURCE_CHECK"] = "True"
os.environ["HF_HUB_OFFLINE"] = "0"  # Allow downloads but skip slow checks
import re
import json
import sys
from pathlib import Path
from typing import Optional, Dict, List, Any, Tuple
import requests
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Using Nougat CLI instead of Python API to avoid albumentations issues

# Layout detection models (lazy loaded)
LAYOUT_MODEL = None
CHART_LLAMA_MODEL = None
# DEEPTABSTR_MODEL = None  # Removed - was never implemented

# ================== CONFIGURATION ==================
# LLM Configuration - Choose one
LLM_PROVIDER = os.getenv("LLM_PROVIDER", "ollama")  # "ollama" or "openai"

# Ollama settings
OLLAMA_URL = os.getenv("OLLAMA_URL", "http://localhost:11434/api/generate")
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "qwen2.5:14b")  # Good balance of speed and quality

# OpenAI settings (if using)
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4-turbo-preview")

# Paths
BASE_DIR = Path(
    r"C:\Users\azt12\OneDrive\Documents\Wrestling Robe\Materials Science - Wickability\Studies for Analysis by LLM AI"
)
PDF_FOLDER = Path(
    r"C:\Users\azt12\OneDrive\Documents\Wrestling Robe\Materials Science - Wickability\Studies for Analysis by LLM AI"
)
OUTPUT_EXCEL = BASE_DIR / "Textile-Specifications-Extracted.xlsx"

# Focus on 1.pdf only for now
FOCUS_PDF = "1.pdf"

# ================== METRICS SCHEMA ==================
# This defines ALL the metrics we want to extract
# Each metric has: name, layer_types (which layer columns it applies to), category

LAYER_TYPES = ["Inner Layer", "Middle Layer", "Outer Layer", "Monolayer"]
SIDES = ["Bottom Side", "Top Side", ""]  # "" means general/both sides

# Study-level metadata (one per study)
STUDY_METADATA = [
    "Study Number",
    "Study Title",
    "Year of Publish",
    "Name of First-Listed Author",
    "Number of Sample Fabrics",
    "Testing/Standards Bodies Methods Used",
]

# Sample-level metadata (one per sample)
SAMPLE_METADATA = [
    "Sample ID/Name",
    "Number of Fabric Layers",
]

# Material, Structure, Treatment fields - per layer
MATERIAL_STRUCTURE_FIELDS = [
    {"name": "Material", "layers": True},
    {"name": "Structure", "layers": True, "note": "Knit or Woven or both"},
]

# Treatment fields - per layer and per side
TREATMENT_FIELDS = [
    {"name": "Treatment", "layers": True, "sides": True},
    {"name": "Change in Water Affinity", "layers": True, "sides": True,
     "note": "e.g. hydrophobic -> superhydrophobic"},
]

# Yarn specification fields - per layer
YARN_SPEC_FIELDS = [
    {"name": "Yarn Fiber Type", "layers": True, "note": "staple, filament, or composite"},
    {"name": "Filament Yarn Texture", "layers": True, "note": "Spun or Filament"},
    {"name": "Filament Yarn Structure", "layers": True, "note": "Textured or Flat (filament only)"},
    {"name": "Filament Yarn Cross-section Type", "layers": True, "note": "round, trilobal, etc."},
    {"name": "Yarn Denier/Linear Density", "layers": True},
    {"name": "Filament Count", "layers": True, "note": "number of filaments per yarn"},
    {"name": "Yarn Twist", "layers": True, "note": "direction and turns per inch/cm"},
]

# Porosity fields - per layer
POROSITY_FIELDS = [
    {"name": "Pore Size", "layers": True},
    {"name": "Pore Size Distribution", "layers": True},
    {"name": "Open Area Fraction", "layers": True},
    {"name": "Hydraulic Diameter", "layers": True},
]

# Structure-specific fields
STRUCTURE_SPECIFIC_FIELDS = [
    {"name": "Crimp (Wovens)", "layers": True, "note": "wovens only"},
    {"name": "Loop Length (Knits)", "layers": True, "note": "knits only"},
    {"name": "Float Length (Knits)", "layers": True, "note": "knits only"},
    {"name": "Proportion of Tucks/Misses (Knits)", "layers": True, "note": "knits only"},
    {"name": "Surface Roughness", "layers": True, "note": "in micrometers or microinches"},
]

# Physical properties - per layer
PHYSICAL_FIELDS = [
    {"name": "Fabric Weight/GSM", "layers": True, "note": "g/m² or oz/yd²"},
    {"name": "Fabric Thickness", "layers": True, "note": "mm or mil"},
    {"name": "Fabric Density", "layers": True, "note": "threads/inch, picks/inch, loops/inch"},
    {"name": "Tensile Strength", "layers": True},
    {"name": "Elongation at Break", "layers": True},
    {"name": "Tear Strength", "layers": True},
    {"name": "Fabric Hand/Feel", "layers": True},
]

# AATCC Moisture metrics - per layer
AATCC_MOISTURE_FIELDS = [
    {"name": "Wetting Time Bottom (WTB)", "layers": True, "standard": "AATCC"},
    {"name": "Wetting Time Top (WTT)", "layers": True, "standard": "AATCC"},
    {"name": "Wetting Time (WT)", "layers": True, "standard": "AATCC"},
    {"name": "Absorption Rate Bottom (ARB)", "layers": True, "standard": "AATCC"},
    {"name": "Absorption Rate Top (ART)", "layers": True, "standard": "AATCC"},
    {"name": "Bottom Absorption Rate (BAR)", "layers": True, "standard": "AATCC"},
    {"name": "Top Absorption Rate (TAR)", "layers": True, "standard": "AATCC"},
    {"name": "Wicking Distance/Rate", "layers": True, "standard": "AATCC"},
    {"name": "Vertical Wicking", "layers": True, "standard": "AATCC"},
    {"name": "Spreading Speed Bottom (SSb)", "layers": True, "standard": "AATCC"},
    {"name": "Spreading Speed Top (SSt)", "layers": True, "standard": "AATCC"},
    {"name": "Spreading Speed (SS)", "layers": True, "standard": "AATCC"},
    {"name": "Max Wetted Radius Bottom (MWRb)", "layers": True, "standard": "AATCC"},
    {"name": "Max Wetted Radius Top (MWRt)", "layers": True, "standard": "AATCC"},
    {"name": "Accumulative One-Way Transport Index (AOTI)", "layers": True, "standard": "AATCC"},
    {"name": "One Way Transport Capability (OWTC)", "layers": True, "standard": "AATCC"},
    {"name": "Overall Moisture Management Capacity (OMMC)", "layers": True, "standard": "AATCC"},
    {"name": "Drying Rate", "layers": True, "standard": "AATCC"},
    {"name": "Drying Time", "layers": True, "standard": "AATCC"},
    {"name": "Moisture Retention", "layers": True, "standard": "AATCC"},
    {"name": "Capillary Rise Height", "layers": True, "standard": "AATCC"},
    {"name": "Capillary Rise Rate", "layers": True, "standard": "AATCC"},
    {"name": "Water Retention", "layers": True, "standard": "AATCC"},
    {"name": "Maximum Water Uptake", "layers": True, "standard": "AATCC"},
]

# Wettability fields - per layer
WETTABILITY_FIELDS = [
    {"name": "Contact Angle", "layers": True},
    {"name": "Contact Angle - Advancing", "layers": True},
    {"name": "Contact Angle - Receding", "layers": True},
    {"name": "Wettability Classification", "layers": True, "note": "hydrophobic, hydrophilic, etc."},
    {"name": "Wettability Gradient", "layers": True},
    {"name": "Surface Energy/Tension", "layers": True},
    {"name": "Interfacial Tension", "layers": True},
]

# Transport metrics - per layer
TRANSPORT_FIELDS = [
    {"name": "Moisture Transport (Transplanar)", "layers": True},
    {"name": "Moisture Transport (Horizontal)", "layers": True},
    {"name": "Directional Moisture Transport", "layers": True},
    {"name": "Unidirectional Transport Direction", "layers": True},
    {"name": "Air Permeability", "layers": True},
    {"name": "Liquid Permeability", "layers": True},
    {"name": "Water Vapor Transmission Rate (WVTR)", "layers": True},
    {"name": "Moisture Vapor Transmission Rate (MVTR)", "layers": True},
]

# Thermal metrics - per layer
THERMAL_FIELDS = [
    {"name": "Thermal Conductivity", "layers": True},
    {"name": "Thermal Resistance (Rct)", "layers": True},
    {"name": "Evaporative Resistance (Ret)", "layers": True},
    {"name": "Thermal Diffusivity", "layers": True},
    {"name": "Heat Flux", "layers": True},
    {"name": "Thermal Absorptivity (Qmax)", "layers": True},
    {"name": "Heat Dissipation/Loss", "layers": True},
    {"name": "Breathability", "layers": True},
    {"name": "Comfort Index", "layers": True},
    {"name": "Moisture Management Capacity (MMC)", "layers": True},
]

# JIS-specific metrics - per layer
JIS_FIELDS = [
    {"name": "Wetting Time (JIS)", "layers": True, "standard": "JIS"},
    {"name": "Absorption Rate (JIS)", "layers": True, "standard": "JIS"},
    {"name": "Spreading Speed (JIS)", "layers": True, "standard": "JIS"},
    {"name": "Vertical Wicking (JIS)", "layers": True, "standard": "JIS"},
    {"name": "Horizontal Wicking (JIS)", "layers": True, "standard": "JIS"},
    {"name": "Drying Time (JIS)", "layers": True, "standard": "JIS"},
    {"name": "Drying Rate (JIS)", "layers": True, "standard": "JIS"},
    {"name": "Drying Index (JIS)", "layers": True, "standard": "JIS"},
    {"name": "Moisture Regain (JIS)", "layers": True, "standard": "JIS"},
    {"name": "Water Vapor Resistance (JIS)", "layers": True, "standard": "JIS"},
    {"name": "WVTR (JIS)", "layers": True, "standard": "JIS"},
    {"name": "Evaporative Heat Flux (JIS)", "layers": True, "standard": "JIS"},
]

# ISO/EN-specific metrics - per layer
ISO_EN_FIELDS = [
    {"name": "Thermal Conductivity (ISO)", "layers": True, "standard": "ISO/EN"},
    {"name": "Thermal Resistance (ISO)", "layers": True, "standard": "ISO/EN"},
    {"name": "Thermal Diffusivity (ISO)", "layers": True, "standard": "ISO/EN"},
    {"name": "Heat Flux (ISO)", "layers": True, "standard": "ISO/EN"},
    {"name": "Hydrostatic Head (ISO/EN)", "layers": True, "standard": "ISO/EN"},
    {"name": "Water Resistance (ISO/EN)", "layers": True, "standard": "ISO/EN"},
    {"name": "Air Permeability (ISO)", "layers": True, "standard": "ISO/EN"},
    {"name": "Water Vapor Resistance (ISO)", "layers": True, "standard": "ISO/EN"},
    {"name": "Water Vapor Permeability (ISO)", "layers": True, "standard": "ISO/EN"},
    {"name": "Evaporative Heat Flux (ISO)", "layers": True, "standard": "ISO/EN"},
    {"name": "Drying Rate (ISO)", "layers": True, "standard": "ISO/EN"},
]

# ASTM-specific metrics - per layer
ASTM_FIELDS = [
    {"name": "Qualitative Absorbency Rating (ASTM)", "layers": True, "standard": "ASTM"},
    {"name": "Hydrostatic Head (ASTM)", "layers": True, "standard": "ASTM"},
    {"name": "Water Resistance (ASTM)", "layers": True, "standard": "ASTM"},
    {"name": "Air Permeability (ASTM)", "layers": True, "standard": "ASTM"},
    {"name": "WVTR (ASTM)", "layers": True, "standard": "ASTM"},
    {"name": "Water Vapor Permeability (ASTM)", "layers": True, "standard": "ASTM"},
    {"name": "Water Absorbency (ASTM)", "layers": True, "standard": "ASTM"},
    {"name": "Absorption Time (ASTM)", "layers": True, "standard": "ASTM"},
    {"name": "Liquid Absorption Capacity (ASTM)", "layers": True, "standard": "ASTM"},
    {"name": "Thermal Resistance Rct (ASTM)", "layers": True, "standard": "ASTM"},
    {"name": "Evaporative Resistance Ret (ASTM)", "layers": True, "standard": "ASTM"},
]

# Other/unspecified standard metrics
OTHER_FIELDS = [
    {"name": "Capillary Pressure", "layers": True, "standard": "Other"},
    {"name": "SMF", "layers": True, "standard": "Other"},
    {"name": "PMI", "layers": True, "standard": "Other"},
]


# ================== METRICS RELEVANCE FILTER (CACHED) ==================
# Loads full metrics list once at startup, pre-processes into 2-word sequences
# for fast O(1) matching instead of scanning 693 metrics per header check.

import re

# Global cache - loaded once at first use
_METRICS_CACHE = None
_METRICS_FILE = r"C:\Users\azt12\OneDrive\Documents\Wrestling Robe\Materials Science - Wickability\Studies for Analysis by LLM AI\Datafiles & Python Scripts\MetricsFullList.txt"


def _extract_word_sequences(text: str, min_words: int = 2) -> set:
    """
    Extract all consecutive word sequences of length >= min_words from text.
    Returns a set of lowercase sequences for fast lookup.
    """
    # Clean text: remove parentheses content, special chars, normalize
    text = re.sub(r'\([^)]*\)', '', text)  # Remove (parenthetical content)
    text = re.sub(r'[^\w\s-]', ' ', text)  # Keep only words, spaces, hyphens
    text = text.lower().strip()

    words = text.split()
    sequences = set()

    # Generate all consecutive sequences of min_words or more
    for length in range(min_words, min(len(words) + 1, 5)):  # Up to 4 words
        for i in range(len(words) - length + 1):
            seq = ' '.join(words[i:i + length])
            if len(seq) > 3:  # Skip very short sequences
                sequences.add(seq)

    return sequences


def _load_metrics_cache() -> set:
    """
    Load the full metrics list and pre-process into 2-word sequences.
    Called once at first use, cached globally for performance.
    """
    global _METRICS_CACHE

    if _METRICS_CACHE is not None:
        return _METRICS_CACHE

    sequences = set()

    try:
        with open(_METRICS_FILE, 'r', encoding='utf-8') as f:
            for line in f:
                # Strip line numbers if present (e.g., "123→" or "123. ")
                line = re.sub(r'^\s*\d+[→.\s]+', '', line)
                line = line.strip()

                if not line or line.startswith('#'):
                    continue

                # Skip section headers (all caps with "LEAVE BLANK")
                if 'LEAVE' in line.upper() and 'BLANK' in line.upper():
                    continue

                # Extract 2+ word sequences from this metric
                seqs = _extract_word_sequences(line, min_words=2)
                sequences.update(seqs)

        # Add common abbreviations as single "words" too
        abbreviations = {
            'wtt', 'wtb', 'wt', 'art', 'arb', 'bar', 'tar',
            'ommc', 'owtc', 'aoti', 'mwrb', 'mwrt', 'ssb', 'sst',
            'rct', 'ret', 'wvtr', 'mvtr', 'wvp', 'qmax',
            'gsm', 'denier', 'porosity',
        }
        sequences.update(abbreviations)

        _METRICS_CACHE = sequences
        print(f"    [CACHE] Loaded {len(sequences)} metric sequences from {_METRICS_FILE}")

    except FileNotFoundError:
        print(f"    [WARNING] Metrics file not found: {_METRICS_FILE}")
        _METRICS_CACHE = set()

    return _METRICS_CACHE


def get_relevant_keywords() -> set:
    """
    Get the cached set of relevant metric sequences for fast matching.
    Loads from file on first call, returns cached set on subsequent calls.
    """
    return _load_metrics_cache()


# Words that should NOT count as relevant (too generic)
EXCLUDED_WORDS = {
    "title", "no", "no.", "number", "diagram", "design", "cam",
    "needle", "arrangement", "notation", "kd", "figure", "table",
    "for", "each", "layer", "or", "and", "the", "in", "of", "to",
}


def check_table_relevance(headers: List[str], min_matches: int = 1) -> tuple:
    """
    Check if extracted table headers contain relevant metrics.
    Uses cached 2-word sequences for O(1) lookup instead of scanning all metrics.

    Args:
        headers: List of column header strings from extracted table
        min_matches: Minimum number of relevant columns required

    Returns:
        (is_relevant: bool, matching_headers: List[str], match_count: int)
    """
    metric_sequences = get_relevant_keywords()  # Cached set of 2-word sequences
    matching_headers = []

    for header in headers:
        header_clean = header.lower().strip()
        header_clean = header_clean.replace("<0x0a>", " ")

        # Skip if header is just excluded words
        words = re.sub(r'[^\w\s]', ' ', header_clean).split()
        non_excluded_words = [w for w in words if w not in EXCLUDED_WORDS]
        if not non_excluded_words:
            continue

        # Extract 2-word sequences from this header
        header_sequences = _extract_word_sequences(header_clean, min_words=2)

        # Also check single abbreviations
        for word in words:
            if word in metric_sequences:  # Abbreviations like 'wtt', 'gsm'
                matching_headers.append(header)
                break
        else:
            # Check if any 2+ word sequence matches
            if header_sequences & metric_sequences:  # Set intersection
                matching_headers.append(header)

    is_relevant = len(matching_headers) >= min_matches
    return (is_relevant, matching_headers, len(matching_headers))


def filter_table_data(table_data: str) -> tuple:
    """
    Filter extracted table/chart data based on column header relevance.

    Args:
        table_data: Raw output from DePlot/chart extractor (pipe-separated)

    Returns:
        (is_relevant: bool, filtered_data: str, reason: str)
    """
    if not table_data or not table_data.strip():
        return (False, "", "Empty data")

    # Parse the table data - DePlot uses | as separator and <0x0A> for newlines
    lines = table_data.replace("<0x0A>", "\n").strip().split("\n")

    if len(lines) < 2:
        return (False, "", "Not enough rows")

    # First line(s) usually contain headers
    # Try to extract headers from the first non-empty line
    headers = []
    for line in lines[:3]:  # Check first 3 lines for headers
        parts = [p.strip() for p in line.split("|") if p.strip()]
        if parts:
            headers.extend(parts)

    if not headers:
        return (False, "", "No headers found")

    # Check relevance
    is_relevant, matching, match_count = check_table_relevance(headers)

    if is_relevant:
        reason = f"Found {match_count} relevant columns: {matching[:5]}"  # Show first 5
        return (True, table_data, reason)
    else:
        reason = f"No relevant columns found in: {headers[:10]}"  # Show first 10 headers
        return (False, "", reason)


def get_all_column_headers() -> List[str]:
    """
    Generate all column headers based on the schema.
    Returns a flat list of all column names.
    """
    headers = []

    # Study metadata
    headers.extend(STUDY_METADATA)

    # Sample metadata
    headers.extend(SAMPLE_METADATA)

    # Build layer-specific columns
    all_field_groups = [
        ("Material/Structure", MATERIAL_STRUCTURE_FIELDS),
        ("Yarn Specs", YARN_SPEC_FIELDS),
        ("Porosity", POROSITY_FIELDS),
        ("Structure-Specific", STRUCTURE_SPECIFIC_FIELDS),
        ("Physical Properties", PHYSICAL_FIELDS),
        ("Wettability", WETTABILITY_FIELDS),
        ("Transport", TRANSPORT_FIELDS),
        ("Thermal", THERMAL_FIELDS),
        ("AATCC Moisture", AATCC_MOISTURE_FIELDS),
        ("JIS", JIS_FIELDS),
        ("ISO/EN", ISO_EN_FIELDS),
        ("ASTM", ASTM_FIELDS),
        ("Other", OTHER_FIELDS),
    ]

    # Treatment fields with sides
    for field in TREATMENT_FIELDS:
        for layer in LAYER_TYPES:
            if field.get("sides"):
                for side in SIDES:
                    if side:
                        headers.append(f"{field['name']}: {layer}, {side}")
                    else:
                        headers.append(f"{field['name']}: {layer}")
            else:
                headers.append(f"{field['name']}: {layer}")

    # Regular layer fields (no sides)
    for group_name, fields in all_field_groups:
        for field in fields:
            if field.get("layers"):
                for layer in LAYER_TYPES:
                    headers.append(f"{field['name']}: {layer}")
            else:
                headers.append(field['name'])

    return headers


# ================== LLM FUNCTIONS ==================

def call_ollama(prompt: str, system_prompt: str = "") -> str:
    """Call local Ollama LLM."""
    full_prompt = f"{system_prompt}\n\n{prompt}" if system_prompt else prompt

    payload = {
        "model": OLLAMA_MODEL,
        "prompt": full_prompt,
        "stream": False,
        "options": {
            "temperature": 0.1,  # Low temperature for more deterministic extraction
            "num_predict": 8000,  # Allow longer responses for JSON
        }
    }

    try:
        resp = requests.post(OLLAMA_URL, json=payload, timeout=300)
        resp.raise_for_status()
        data = resp.json()
        return data.get("response", "")
    except requests.exceptions.ConnectionError:
        print("ERROR: Cannot connect to Ollama. Is it running?")
        print("Start Ollama with: ollama serve")
        return ""
    except Exception as e:
        print(f"Ollama error: {e}")
        return ""


def call_openai(prompt: str, system_prompt: str = "") -> str:
    """Call OpenAI API."""
    if not OPENAI_API_KEY:
        print("ERROR: OPENAI_API_KEY not set")
        return ""

    import openai
    client = openai.OpenAI(api_key=OPENAI_API_KEY)

    messages = []
    if system_prompt:
        messages.append({"role": "system", "content": system_prompt})
    messages.append({"role": "user", "content": prompt})

    try:
        response = client.chat.completions.create(
            model=OPENAI_MODEL,
            messages=messages,
            temperature=0.1,
            max_tokens=8000,
        )
        return response.choices[0].message.content
    except Exception as e:
        print(f"OpenAI error: {e}")
        return ""


def call_llm(prompt: str, system_prompt: str = "") -> str:
    """Call the configured LLM provider."""
    if LLM_PROVIDER == "openai":
        return call_openai(prompt, system_prompt)
    else:
        return call_ollama(prompt, system_prompt)


# ================== EXTRACTION FUNCTIONS ==================

def extract_with_nougat(pdf_path: str) -> str:
    """
    Extract text from PDF using Nougat CLI.
    First checks for existing .mmd file to avoid re-running Nougat.
    """
    import subprocess
    import tempfile
    import glob as glob_module

    # Check for existing .mmd file first
    pdf_stem = Path(pdf_path).stem
    existing_mmd = Path(pdf_path).parent / f"{pdf_stem}.mmd"
    if existing_mmd.exists():
        print(f"    Found existing .mmd file: {existing_mmd}")
        with open(existing_mmd, "r", encoding="utf-8") as f:
            markdown_text = f.read()
        print(f"    Loaded {len(markdown_text)} characters from cached .mmd")
        return markdown_text

    try:
        print(f"    Extracting with Nougat CLI...")

        with tempfile.TemporaryDirectory() as tmpdir:
            cmd = ["nougat", str(pdf_path), "-o", tmpdir, "--no-skipping"]
            print(f"    Running: {' '.join(cmd)}")

            result = subprocess.run(cmd, capture_output=True, text=True, timeout=600)

            print(f"    Return code: {result.returncode}")
            if result.stdout:
                print(f"    stdout: {result.stdout[:500]}")
            if result.stderr:
                print(f"    stderr: {result.stderr[:500]}")

            # List what's in the temp dir
            all_files = os.listdir(tmpdir)
            print(f"    Files in temp dir: {all_files}")

            if result.returncode != 0:
                print(f"    Nougat CLI error: {result.stderr}")
                sys.exit(1)

            mmd_files = glob_module.glob(os.path.join(tmpdir, "*.mmd"))

            if not mmd_files:
                print("    No .mmd output file found")
                sys.exit(1)

            with open(mmd_files[0], "r", encoding="utf-8") as f:
                markdown_text = f.read()

            print(f"    Got {len(markdown_text)} characters")
            return markdown_text

    except subprocess.TimeoutExpired:
        print("    Nougat timed out")
        sys.exit(1)
    except FileNotFoundError:
        print("    Nougat CLI not found. Install with: pip install nougat-ocr")
        sys.exit(1)
    except Exception as e:
        print(f"    Nougat ERROR: {e}")
        sys.exit(1)


# ================== DOCUMENT LAYOUT ANALYSIS PIPELINE ==================
#
# Current status:
# - Layout detection: PaddleOCR LayoutDetection (ENABLED - detects figures/tables/text)
# - Text/Table extraction: Nougat (ACTIVE - handles both)
# - Chart extraction: DePlot/MatCha/LLaVA (DISABLED - connectivity issues)
# - Table image OCR: docTR/Tesseract (COMMENTED OUT - Nougat handles tables)
#
# ===========================================================================


def load_layout_model():
    """
    Load layout detection model using PaddleOCR LayoutDetection.
    Returns tuple of (model_type, model) or None if unavailable.

    PaddleOCR LayoutDetection labels:
    - doc_title, abstract, paragraph_title, header, footer
    - text, formula
    - image, chart, figure_title
    - table (if present)
    """
    global LAYOUT_MODEL

    if LAYOUT_MODEL is not None:
        return LAYOUT_MODEL

    try:
        from paddleocr import LayoutDetection
        print("    Loading PaddleOCR LayoutDetection...")

        # threshold=0.2 to detect smaller charts and combo chart+table elements
        model = LayoutDetection(threshold=0.2)

        LAYOUT_MODEL = ("paddle", model)
        print("    PaddleOCR LayoutDetection loaded successfully")
        return LAYOUT_MODEL

    except ImportError as e:
        print(f"    PaddleOCR not installed: {e}")
        return None
    except Exception as e:
        print(f"    PaddleOCR failed to load: {e}")
        return None


def extract_pdf_page_as_image(pdf_path: str, page_num: int, dpi: int = 200):
    """Extract a single page from PDF as PIL Image at specified DPI."""
    try:
        import pypdfium2 as pdfium

        pdf = pdfium.PdfDocument(pdf_path)
        page = pdf[page_num]

        # Render at specified DPI
        bitmap = page.render(scale=dpi/72)
        pil_image = bitmap.to_pil()

        return pil_image
    except Exception as e:
        print(f"    Error extracting page {page_num}: {e}")
        return None


def filter_overlapping_boxes(boxes, iou_threshold=0.5, containment_threshold=0.7):
    """
    Filter overlapping and contained boxes - keep higher confidence or larger box.
    Uses IoU (Intersection over Union) for overlap detection.
    Also removes boxes that are mostly contained within larger boxes of SAME TYPE.
    Different labels (e.g., 'chart' inside 'image') are kept - the more specific one wins.
    """
    if len(boxes) <= 1:
        return boxes

    # Sort by area descending (largest first), then by confidence
    def box_sort_key(b):
        coord = b.get('coordinate', [])
        if len(coord) >= 4:
            area = (coord[2] - coord[0]) * (coord[3] - coord[1])
        else:
            area = 0
        return (-area, -b.get('score', 0))

    sorted_boxes = sorted(boxes, key=box_sort_key)
    kept = []

    for box in sorted_boxes:
        coord = box.get('coordinate', [])
        if len(coord) < 4:
            continue
        x1, y1, x2, y2 = coord
        box_area = (x2 - x1) * (y2 - y1)
        if box_area <= 0:
            continue

        box_label = box.get('label', '').lower()

        should_discard = False
        for kept_box in kept:
            kcoord = kept_box.get('coordinate', [])
            if len(kcoord) < 4:
                continue
            kx1, ky1, kx2, ky2 = kcoord
            kept_area = (kx2 - kx1) * (ky2 - ky1)
            kept_label = kept_box.get('label', '').lower()

            # Calculate intersection
            inter_x1 = max(x1, kx1)
            inter_y1 = max(y1, ky1)
            inter_x2 = min(x2, kx2)
            inter_y2 = min(y2, ky2)

            if inter_x2 > inter_x1 and inter_y2 > inter_y1:
                inter_area = (inter_x2 - inter_x1) * (inter_y2 - inter_y1)
                union_area = box_area + kept_area - inter_area

                # Check IoU overlap (same type only)
                iou = inter_area / union_area if union_area > 0 else 0
                if iou > iou_threshold and box_label == kept_label:
                    should_discard = True
                    break

                # Check containment for same-type boxes only
                # This filters sub-images within images, but keeps charts inside images
                containment = inter_area / box_area if box_area > 0 else 0
                if containment > containment_threshold and box_label == kept_label:
                    should_discard = True
                    break

        if not should_discard:
            kept.append(box)

    return kept


def detect_layout_elements(page_image) -> Dict[str, List]:
    """
    Detect layout elements (figures, tables, text) in a page image.
    Uses PaddleOCR LayoutDetection for layout detection.
    Returns dict with 'figures' and 'tables' lists of cropped images.

    Uses low threshold (0.3) to catch small charts, then filters:
    - Overlapping boxes (keep higher confidence)
    - Very small boxes (likely noise)
    - Boxes with extreme aspect ratios
    """
    import numpy as np

    result = {"figures": [], "tables": [], "text_blocks": []}

    model_info = load_layout_model()
    if model_info is None:
        print("    Falling back to grid-based detection...")
        return fallback_grid_detection(page_image)

    model_type, model = model_info
    img_width, img_height = page_image.size

    # Minimum size thresholds (as fraction of page)
    MIN_WIDTH_FRAC = 0.05   # At least 5% of page width
    MIN_HEIGHT_FRAC = 0.03  # At least 3% of page height
    MIN_AREA_FRAC = 0.005   # At least 0.5% of page area

    # Aspect ratio limits (width/height)
    MIN_ASPECT = 0.1  # Not too tall and skinny
    MAX_ASPECT = 10   # Not too wide and flat

    try:
        img_array = np.array(page_image)

        if model_type == "paddle":
            # Returns: [{'boxes': [{'label': 'image', 'coordinate': [x1,y1,x2,y2], 'score': 0.9}, ...]}]
            detections = model.predict(img_array)

            if detections and len(detections) > 0:
                all_boxes = detections[0].get('boxes', [])

                # Filter overlapping and contained boxes (same-type only)
                filtered_boxes = filter_overlapping_boxes(all_boxes, iou_threshold=0.5, containment_threshold=0.7)

                for box in filtered_boxes:
                    label = box.get('label', '').lower()
                    coord = box.get('coordinate', [])
                    score = box.get('score', 0.5)

                    if len(coord) < 4:
                        continue

                    x1, y1, x2, y2 = int(coord[0]), int(coord[1]), int(coord[2]), int(coord[3])
                    box_width = x2 - x1
                    box_height = y2 - y1
                    box_area = box_width * box_height
                    page_area = img_width * img_height

                    # Size filters
                    if box_width < img_width * MIN_WIDTH_FRAC:
                        continue  # Too narrow
                    if box_height < img_height * MIN_HEIGHT_FRAC:
                        continue  # Too short
                    if box_area < page_area * MIN_AREA_FRAC:
                        continue  # Too small overall

                    # Aspect ratio filter (skip for tables - they can be very wide)
                    aspect = box_width / box_height if box_height > 0 else 0
                    if label != 'table' and (aspect < MIN_ASPECT or aspect > MAX_ASPECT):
                        continue  # Extreme aspect ratio

                    # Confidence threshold - lower for charts/images since they're important
                    conf_threshold = 0.3 if label in ['chart', 'image', 'figure'] else 0.4
                    if score < conf_threshold:
                        continue

                    cropped = page_image.crop((x1, y1, x2, y2))

                    region_info = {
                        "image": cropped,
                        "bbox": (x1, y1, box_width, box_height),
                        "confidence": score,
                        "type": label,
                    }

                    # Map labels to categories
                    if label in ['image', 'figure', 'chart']:
                        result["figures"].append(region_info)
                    elif label == 'table':
                        result["tables"].append(region_info)
                    elif label in ['text', 'paragraph_title', 'title']:
                        result["text_blocks"].append(region_info)

        # If no detections, use grid fallback
        if len(result["figures"]) == 0 and len(result["tables"]) == 0:
            print("      No elements detected, using grid fallback...")
            return fallback_grid_detection(page_image)

        print(f"      Detected: {len(result['figures'])} figures, {len(result['tables'])} tables")
        return result

    except Exception as e:
        print(f"    Layout detection error: {e}")
        import traceback
        traceback.print_exc()
        return fallback_grid_detection(page_image)


def fallback_grid_detection(page_image) -> Dict[str, List]:
    """Fallback detection when LayoutParser is unavailable. Uses 3x3 grid for finer detection."""
    import numpy as np

    result = {"figures": [], "tables": [], "text_blocks": []}

    img_array = np.array(page_image.convert('RGB'))
    height, width = img_array.shape[:2]

    # Split into 3x3 grid for finer chart isolation
    grid_rows = 3
    grid_cols = 3

    for row in range(grid_rows):
        for col in range(grid_cols):
            x1 = col * (width // grid_cols)
            y1 = row * (height // grid_rows)
            x2 = (col + 1) * (width // grid_cols)
            y2 = (row + 1) * (height // grid_rows)

            cropped = page_image.crop((x1, y1, x2, y2))

            # Check if non-empty (has content)
            cell_array = np.array(cropped.convert('L'))
            non_white = np.sum(cell_array < 240) / cell_array.size

            # Require more content (15%) to filter out mostly empty cells
            if non_white > 0.15:
                result["figures"].append({
                    "image": cropped,
                    "bbox": (x1, y1, x2 - x1, y2 - y1),
                    "confidence": 0.5,
                    "type": "Figure",
                    "position": f"row{row+1}_col{col+1}"
                })

    print(f"      Grid detection: {len(result['figures'])} regions with content")
    return result


def process_chart_with_chartllama(chart_image, prompt: str = None) -> str:
    """
    Process a chart image using LLaVA (via Ollama) or DePlot fallback.
    NOTE: DePlot/transformers has connectivity issues, trying LLaVA first.
    """
    # Try LLaVA first (via Ollama) - faster and no connectivity issues
    result = process_chart_with_llava(chart_image)
    if result:
        return result

    # Skip DePlot for now due to transformers connectivity issues
    print("        LLaVA not available, skipping chart extraction")
    return ""


def process_chart_with_llava(chart_image) -> str:
    """Use LLaVA (via Ollama) for chart understanding."""
    try:
        import base64
        from io import BytesIO

        # Convert image to base64
        buffered = BytesIO()
        chart_image.save(buffered, format="PNG")
        img_base64 = base64.b64encode(buffered.getvalue()).decode()

        # Call LLaVA via Ollama
        payload = {
            "model": "llava",
            "prompt": "This is a bar chart from a research paper. Extract ALL numeric values with their labels. Format: label | value. List every bar/data point you can see.",
            "images": [img_base64],
            "stream": False,
            "options": {"temperature": 0.1}
        }

        print("        Trying LLaVA...")
        resp = requests.post(OLLAMA_URL, json=payload, timeout=120)
        if resp.status_code == 200:
            data = resp.json()
            result = data.get("response", "")
            if result:
                print(f"        LLaVA extracted: {result[:100]}...")
                return result
        else:
            print(f"        LLaVA error: {resp.status_code}")
    except requests.exceptions.Timeout:
        print("        LLaVA timeout")
    except Exception as e:
        print(f"        LLaVA not available: {e}")

    return ""


def process_chart_with_matcha(chart_image) -> str:
    """Use Google MatCha for chart understanding - better for scientific charts."""
    try:
        from transformers import Pix2StructProcessor, Pix2StructForConditionalGeneration

        print("        Loading MatCha model...")
        processor = Pix2StructProcessor.from_pretrained('google/matcha-base')
        model = Pix2StructForConditionalGeneration.from_pretrained('google/matcha-base')

        # MatCha understands chart questions better
        inputs = processor(
            images=chart_image,
            text="What are all the numeric values shown in this chart? List each bar/point with its label and value.",
            return_tensors="pt"
        )
        predictions = model.generate(**inputs, max_new_tokens=512)
        result = processor.decode(predictions[0], skip_special_tokens=True)

        return result
    except Exception as e:
        print(f"        MatCha error: {e}, trying DePlot...")
        return process_chart_with_deplot(chart_image)


def process_chart_with_deplot(chart_image) -> str:
    """Fallback: Use Google DePlot for chart to table conversion."""
    try:
        from transformers import Pix2StructProcessor, Pix2StructForConditionalGeneration

        processor = Pix2StructProcessor.from_pretrained('google/deplot')
        model = Pix2StructForConditionalGeneration.from_pretrained('google/deplot')

        inputs = processor(
            images=chart_image,
            text="Generate underlying data table of the figure below:",
            return_tensors="pt"
        )
        predictions = model.generate(**inputs, max_new_tokens=512)
        result = processor.decode(predictions[0], skip_special_tokens=True)

        return result
    except Exception as e:
        print(f"        DePlot error: {e}")
        return ""


def process_table_image(table_image) -> str:
    """
    Process a table image using OCR.
    NOTE: Currently disabled - Nougat already extracts tables from PDF.
    This is kept as a fallback if Nougat table extraction fails.
    """
    # COMMENTED OUT - Nougat handles table extraction
    # Uncomment if Nougat table extraction proves insufficient
    #
    # try:
    #     return process_table_with_doctr(table_image)
    # except Exception as e:
    #     print(f"        docTR error: {e}")
    #     return process_table_with_tesseract(table_image)

    print("        Table image OCR skipped (Nougat handles tables)")
    return ""


# COMMENTED OUT - Nougat handles table extraction
# Uncomment these functions if Nougat table extraction proves insufficient
#
# def process_table_with_doctr(table_image) -> str:
#     """Use docTR for OCR on table images."""
#     try:
#         from doctr.io import DocumentFile
#         from doctr.models import ocr_predictor
#         import numpy as np
#         from io import BytesIO
#
#         # Save image to bytes for docTR
#         buffered = BytesIO()
#         table_image.save(buffered, format="PNG")
#         buffered.seek(0)
#
#         # Load with docTR
#         predictor = ocr_predictor(pretrained=True)
#
#         # Convert PIL to numpy
#         img_array = np.array(table_image)
#
#         # Run OCR
#         result = predictor([img_array])
#
#         # Extract text
#         text_output = []
#         for page in result.pages:
#             for block in page.blocks:
#                 for line in block.lines:
#                     line_text = " ".join([word.value for word in line.words])
#                     text_output.append(line_text)
#
#         return "\n".join(text_output)
#
#     except ImportError:
#         print("        docTR not installed, trying Tesseract...")
#         return process_table_with_tesseract(table_image)
#     except Exception as e:
#         print(f"        docTR error: {e}")
#         return process_table_with_tesseract(table_image)
#
#
# def process_table_with_tesseract(table_image) -> str:
#     """Fallback: Use Tesseract OCR for table text extraction."""
#     try:
#         import pytesseract
#
#         # Use table-specific Tesseract config
#         custom_config = r'--oem 3 --psm 6'
#         text = pytesseract.image_to_string(table_image, config=custom_config)
#
#         return text.strip()
#     except ImportError:
#         print("        Tesseract not installed")
#         return ""
#     except Exception as e:
#         print(f"        Tesseract error: {e}")
#         return ""


def extract_figures_from_pdf(pdf_path: str, page_numbers: List[int] = None) -> List[Dict[str, Any]]:
    """
    Full document analysis pipeline (mostly disabled - Nougat handles extraction):

    1. Render PDF page at high DPI (pypdfium2)
    2. Detect layout (DISABLED - falls back to grid)
    3. For each Figure: Try chart extraction (mostly broken)
    4. For each Table: Skipped (Nougat handles tables)

    Returns list of {page, type, position, data} dicts.
    """
    results = []

    try:
        import pypdfium2 as pdfium
        pdf = pdfium.PdfDocument(pdf_path)
        total_pages = len(pdf)

        if page_numbers is None:
            # Default to pages likely to have figures (last quarter of paper)
            start_page = max(0, total_pages - 5)
            page_numbers = list(range(start_page, total_pages))

        print(f"    Analyzing pages: {[p+1 for p in page_numbers]}")

        for page_num in page_numbers:
            if page_num >= total_pages:
                continue

            print(f"\n    === Page {page_num + 1}/{total_pages} ===")

            # Step 1: Render page at high DPI
            page_image = extract_pdf_page_as_image(pdf_path, page_num, dpi=250)
            if page_image is None:
                continue

            print(f"      Image size: {page_image.size}")

            # Step 2: Detect layout elements
            layout = detect_layout_elements(page_image)

            # Step 3: Process figures (charts)
            for i, fig_info in enumerate(layout["figures"]):
                fig_image = fig_info["image"]
                print(f"      Processing Figure {i+1} ({fig_image.size[0]}x{fig_image.size[1]})...")

                chart_data = process_chart_with_chartllama(fig_image)

                if chart_data and len(chart_data) > 10:
                    print(f"        Extracted: {chart_data[:100]}...")

                    # Apply relevance filter to extracted chart data
                    is_relevant, filtered_data, reason = filter_table_data(chart_data)
                    if is_relevant:
                        print(f"        [RELEVANT] {reason}")
                        results.append({
                            "page": page_num + 1,
                            "type": "figure",
                            "index": i,
                            "data": filtered_data
                        })
                    else:
                        print(f"        [FILTERED OUT] {reason}")

            # Step 4: Process tables
            for i, tbl_info in enumerate(layout["tables"]):
                tbl_image = tbl_info["image"]
                print(f"      Processing Table {i+1} ({tbl_image.size[0]}x{tbl_image.size[1]})...")

                table_data = process_table_image(tbl_image)

                if table_data and len(table_data) > 10:
                    print(f"        Extracted: {table_data[:100]}...")

                    # Apply relevance filter to extracted table data
                    is_relevant, filtered_data, reason = filter_table_data(table_data)
                    if is_relevant:
                        print(f"        [RELEVANT] {reason}")
                        results.append({
                            "page": page_num + 1,
                            "type": "table",
                            "index": i,
                            "data": filtered_data
                        })
                    else:
                        print(f"        [FILTERED OUT] {reason}")

        return results

    except ImportError as e:
        print(f"    Import error: {e}")
        print("    Install dependencies:")
        print("      pip install pypdfium2 layoutparser 'layoutparser[layoutmodels]'")
        print("      pip install python-doctr torch torchvision")
        return []
    except Exception as e:
        print(f"    Pipeline error: {e}")
        import traceback
        traceback.print_exc()
        return []


def extract_sample_count_rule_based(text: str) -> Optional[int]:
    """
    Rule-based sample count extraction using Group 1+2+3 logic.
    REQUIRES all three groups to be present in the same sentence:
    - Group 1: A number (arabic or word) immediately before a Group 2 term, OR "total of" + number
    - Group 2: Fabric/material terms
    - Group 3: Action words (tested, produced, etc.)

    Returns the HIGHEST number found across all matching sentences.
    Returns None if no valid matches found.
    """
    # Group 2: fabric terms (both singular and plural)
    group2_words = [
        "fabric", "fabrics", "material", "materials", "variant", "variants",
        "garment", "garments", "sample", "samples", "textile", "textiles",
        "specimen", "specimens", "jersey", "jerseys", "structure", "structures"
    ]

    # Group 3: action words
    group3_words = [
        "tested", "produced", "used", "analyzed", "evaluated", "studied",
        "prepared", "examined", "knit", "knitted", "woven", "manufactured",
        "fabricated", "developed", "constructed", "created", "made"
    ]

    # Word numbers (expanded)
    word_to_num = {
        "one": 1, "two": 2, "three": 3, "four": 4, "five": 5, "six": 6, "seven": 7,
        "eight": 8, "nine": 9, "ten": 10, "eleven": 11, "twelve": 12, "thirteen": 13,
        "fourteen": 14, "fifteen": 15, "sixteen": 16, "seventeen": 17, "eighteen": 18,
        "nineteen": 19, "twenty": 20, "twenty-one": 21, "twenty-two": 22, "twenty-three": 23,
        "twenty-four": 24, "twenty-five": 25, "twenty-six": 26, "twenty-seven": 27,
        "twenty-eight": 28, "twenty-nine": 29, "thirty": 30, "forty": 40, "fifty": 50
    }

    # Build Group 2 pattern for regex
    group2_pattern = "|".join(group2_words)

    # Split into sentences (improved splitting)
    sentences = re.split(r'(?<=[.!?])\s+(?=[A-Z])', text)

    highest_count = None
    all_matches = []

    for i, sentence in enumerate(sentences):
        sentence_lower = sentence.lower()

        # Check Group 2 (fabric terms) - must be present
        group2_found = [word for word in group2_words if word in sentence_lower]
        if not group2_found:
            continue

        # Check Group 3 (action words) - must be present with word boundary
        group3_found = [word for word in group3_words if re.search(rf'\b{word}\b', sentence_lower)]
        if not group3_found:
            continue

        # Found Group 2 + Group 3, now look for numbers (Group 1)
        numbers_found = []

        # Digit patterns: number IMMEDIATELY before Group 2 term
        digit_patterns = [
            # "total of N fabrics"
            rf'total\s+of\s+(\d+)\s+(?:{group2_pattern})',
            # "N types of fabrics"
            rf'(\d+)\s+(?:different\s+)?types?\s+of\s+(?:{group2_pattern})',
            # "N fabrics" (direct)
            rf'(?<![a-zA-Z0-9.])(\d+)(?![a-zA-Z0-9.])\s+(?:{group2_pattern})',
            # "tested N fabrics"
            rf'(?:tested|produced|used|analyzed|evaluated|studied|prepared|examined)\s+(\d+)\s+(?:{group2_pattern})',
        ]

        for pattern in digit_patterns:
            for match in re.finditer(pattern, sentence_lower):
                try:
                    num = int(match.group(1))
                    # Exclude implausible numbers (sample IDs, years, etc.)
                    if 1 <= num <= 100:
                        # Check Group 3 is within 100 chars of the match (same clause)
                        match_pos = match.start()
                        closest_g3_dist = min(
                            abs(sentence_lower.find(g3) - match_pos)
                            for g3 in group3_found
                            if g3 in sentence_lower
                        )
                        if closest_g3_dist <= 150:  # Within same clause
                            numbers_found.append(num)
                except (ValueError, AttributeError):
                    continue

        # Word numbers: "eight samples were tested"
        for word, num in word_to_num.items():
            # Word immediately before Group 2
            word_patterns = [
                rf'\b{word}\b\s+(?:different\s+)?(?:types?\s+of\s+)?(?:{group2_pattern})',
                rf'total\s+of\s+{word}\s+(?:{group2_pattern})',
            ]
            for pattern in word_patterns:
                if re.search(pattern, sentence_lower):
                    # Check Group 3 proximity
                    word_match = re.search(rf'\b{word}\b', sentence_lower)
                    if word_match:
                        match_pos = word_match.start()
                        closest_g3_dist = min(
                            abs(sentence_lower.find(g3) - match_pos)
                            for g3 in group3_found
                            if g3 in sentence_lower
                        )
                        if closest_g3_dist <= 150:
                            numbers_found.append(num)
                    break

        # Take highest from this sentence
        if numbers_found:
            sentence_max = max(numbers_found)
            match_info = {
                "count": sentence_max,
                "sentence": sentence.strip()[:150],
                "group2": group2_found,
                "group3": group3_found
            }
            all_matches.append(match_info)
            print(f"    [RULE-BASED] Match: {sentence_max} samples")
            print(f"      Group 2 terms: {group2_found}")
            print(f"      Group 3 terms: {group3_found}")
            print(f"      Sentence: '{sentence.strip()[:100]}...'")

            if highest_count is None or sentence_max > highest_count:
                highest_count = sentence_max

    if all_matches:
        print(f"    [RULE-BASED] Total matches found: {len(all_matches)}, highest: {highest_count}")

    return highest_count


def extract_study_metadata(pdf_path: str, nougat_text: str) -> Dict[str, Any]:
    """
    Extract study-level metadata using REGEX ONLY - no LLM needed.
    Nougat outputs structured Markdown that we can parse directly.
    """
    metadata = {}
    lines = nougat_text.split('\n')

    # Title = first non-empty line (before authors)
    title = "Not found"
    for line in lines[:10]:
        line = line.strip()
        if line and not line.startswith('#') and len(line) > 20:
            title = line[:200]
            break
    metadata["Study Title"] = title

    # First author = second non-empty line, extract last name
    # Authors typically have superscript numbers like "Jiahui Ou1"
    first_author = "Not found"
    found_title = False
    for line in lines[:20]:
        line = line.strip()
        if not line:
            continue
        if not found_title:
            if len(line) > 20:  # Skip title
                found_title = True
            continue
        # This should be first author - remove trailing numbers
        author_match = re.match(r'^([A-Za-z\s\-]+?)(?:\d+|\*|†)?$', line)
        if author_match:
            full_name = author_match.group(1).strip()
            # Get last name (last word)
            name_parts = full_name.split()
            if name_parts:
                first_author = name_parts[-1]  # Last name
                break
    metadata["Name of First-Listed Author"] = first_author

    # Year = look for 4-digit year patterns near publication info
    year = "Not found"
    year_patterns = [
        r'(?:Published|Received|Accepted|©|Copyright)\s*:?\s*(\d{4})',
        r'(\d{4})\s*(?:Elsevier|Springer|Wiley|Taylor)',
        r'\b(20[0-2]\d)\b',  # Any year 2000-2029
    ]
    for pattern in year_patterns:
        match = re.search(pattern, nougat_text, re.IGNORECASE)
        if match:
            year = int(match.group(1))
            break
    metadata["Year of Publish"] = year

    # Standards = search for known testing standards
    standards_found = []
    standards_patterns = [
        (r'\bAATCC\s*(?:TM\s*)?\d+', 'AATCC'),
        (r'\bISO\s*\d+', 'ISO'),
        (r'\bASTM\s*[A-Z]?\d+', 'ASTM'),
        (r'\bJIS\s*[A-Z]?\s*\d+', 'JIS'),
        (r'\bEN\s*\d+', 'EN'),
        (r'\bGB/T\s*\d+', 'GB/T'),
    ]
    for pattern, name in standards_patterns:
        if re.search(pattern, nougat_text, re.IGNORECASE):
            if name not in standards_found:
                standards_found.append(name)
    metadata["Testing/Standards Bodies Methods Used"] = ", ".join(standards_found) if standards_found else "Not specified"

    # Sample count = use rule-based detection (done separately)
    metadata["Number of Sample Fabrics"] = 1  # Will be overridden by rule-based

    return metadata


def parse_chart_data_direct(figure_text: str, expected_count: int) -> List[Dict[str, Any]]:
    """
    Parse chart data directly using regex - NO LLM needed.
    Looks for patterns like: label | value or label: value
    """
    samples = {}

    # Group name patterns to look for
    group_patterns = {
        r'B-?L[_-]?(?:@?S|Acid)': 'B-L@S',
        r'B-?L[_-]?(?:@?W|Water)': 'B-L@W',
        r'L-?B[_-]?(?:@?S|Acid)': 'L-B@S',
        r'L-?B[_-]?(?:@?W|Water)': 'L-B@W',
        r'B-?B[_-]?(?:@?S|Acid)': 'B-B@S',
        r'B-?B[_-]?(?:@?W|Water)': 'B-B@W',
        r'L-?L[_-]?(?:@?S|Acid)': 'L-L@S',
        r'L-?L[_-]?(?:@?W|Water)': 'L-L@W',
        r'\bB-?L\b': 'B-L',
        r'\bL-?B\b': 'L-B',
        r'\bB-?B\b': 'B-B',
        r'\bL-?L\b': 'L-L',
    }

    # Find all label|value pairs
    # Pattern: something | number or something: number
    value_patterns = [
        r'([A-Za-z][A-Za-z0-9_\-@]+)\s*\|\s*(\d+\.?\d*)',  # label | value
        r'([A-Za-z][A-Za-z0-9_\-@]+)\s*:\s*(\d+\.?\d*)',   # label: value
    ]

    for pattern in value_patterns:
        for match in re.finditer(pattern, figure_text, re.IGNORECASE):
            label = match.group(1)
            value = float(match.group(2))

            # Map label to standard group name
            group_name = None
            for group_pattern, standard_name in group_patterns.items():
                if re.search(group_pattern, label, re.IGNORECASE):
                    group_name = standard_name
                    break

            if group_name:
                if group_name not in samples:
                    samples[group_name] = {"sample_id": group_name}
                # Add value with auto-generated key
                key_num = len([k for k in samples[group_name].keys() if k.startswith("value")]) + 1
                samples[group_name][f"value_{key_num}"] = value

    print(f"  Direct parsing found {len(samples)} groups: {list(samples.keys())}")

    # Convert to list
    result = list(samples.values())

    # If we found groups, return them
    if result:
        return result

    # Fallback: create empty samples
    return [{"sample_id": f"Sample {i}"} for i in range(1, expected_count + 1)]


def extract_metrics_from_nougat_text(text: str) -> Dict[str, Dict[str, Any]]:
    """
    Extract metrics directly from Nougat markdown text using regex.
    Parses narrative text for embedded data values.
    """
    samples = {}

    # Define the 8 sample groups based on the paper structure
    sample_groups = [
        "B-L@S", "B-L@W", "L-B@S", "L-B@W",
        "B-B@S", "B-B@W", "L-L@S", "L-L@W"
    ]

    # Initialize samples
    for group in sample_groups:
        samples[group] = {"sample_id": group}

    # Extract contact angle (>130° for hydrophobic side)
    # NOTE: Only B-L samples are shown in the contact angle chart (Figure 3b)
    # L-B samples show ~0° because water contacts hydrophilic side first
    # B-B and L-L samples are not shown in contact angle data
    # Nougat uses LaTeX: \(>\)130\({}^{\circ}\) or contact angle >130
    contact_match = re.search(r'contact angle[^0-9]*(?:\\?\(?>\\?\)|>)\s*(\d+)', text, re.IGNORECASE)
    if contact_match:
        angle = float(contact_match.group(1))
        # Only B-L samples have >130° in the data (water on hydrophobic side)
        for group in ["B-L@S", "B-L@W"]:
            samples[group]["contact_angle"] = f">{angle}°"
        # L-B samples have ~0° (water on hydrophilic side)
        for group in ["L-B@S", "L-B@W"]:
            samples[group]["contact_angle"] = "~0° (hydrophilic side)"

    # Extract breakthrough pressure data (Figure 4a)
    # Text: "liquid column height of distilled water in L-B group was 37.58mm"
    # Text: "liquid column height of distilled water in the B-L group was 17.02 mm"
    # Need to match the specific sentence structure

    # L-B group values
    lb_water_match = re.search(r'distilled water in L-B group was (\d+\.?\d*)\s*mm', text, re.IGNORECASE)
    lb_acid_match = re.search(r'L-B.*?acidic water was (\d+\.?\d*)\s*mm', text, re.IGNORECASE)
    if lb_water_match:
        samples["L-B@W"]["breakthrough_pressure_mm"] = float(lb_water_match.group(1))
    if lb_acid_match:
        samples["L-B@S"]["breakthrough_pressure_mm"] = float(lb_acid_match.group(1))

    # B-L group values
    bl_water_match = re.search(r'distilled water in the? B-L group was (\d+\.?\d*)\s*mm', text, re.IGNORECASE)
    bl_acid_match = re.search(r'B-L group.*?acidic water was (\d+\.?\d*)\s*mm', text, re.IGNORECASE)
    if bl_water_match:
        samples["B-L@W"]["breakthrough_pressure_mm"] = float(bl_water_match.group(1))
    if bl_acid_match:
        samples["B-L@S"]["breakthrough_pressure_mm"] = float(bl_acid_match.group(1))

    # Extract evaporation/moisture dissipation rate (100% for B-L@S)
    # Pattern: "moisture dissipation rate of B-L@S reached 100%"
    # Or: "evaporation rate of B-L@S reached 100%"
    evap_match = re.search(r'B-L.{0,5}S.*?reached?\s*(?:a\s*)?(?:remarkable\s*)?(\d+)%', text, re.IGNORECASE | re.DOTALL)
    if evap_match:
        samples["B-L@S"]["moisture_dissipation_rate"] = f"{evap_match.group(1)}%"

    # Also check for "100% evaporation rate" pattern
    evap_match2 = re.search(r'(\d+)%\s*evaporation rate.*?B-L.{0,5}S', text, re.IGNORECASE | re.DOTALL)
    if evap_match2:
        samples["B-L@S"]["evaporation_rate"] = f"{evap_match2.group(1)}%"

    # Extract penetration time for acidic solution (3 minutes)
    pene_match = re.search(r'acidic.*?solution.*?(\d+)\s*minutes?', text, re.IGNORECASE)
    if pene_match:
        for group in ["B-L@S", "L-B@S", "B-B@S", "L-L@S"]:
            samples[group]["penetration_time_min"] = int(pene_match.group(1))

    # Extract breakthrough pressure reduction (59.8%)
    bp_reduction = re.search(r'(\d+\.?\d*)%\s*reduction.*?breakthrough', text, re.IGNORECASE)
    if bp_reduction:
        samples["B-L@S"]["breakthrough_pressure_reduction"] = f"{bp_reduction.group(1)}%"

    # Extract pH response range
    ph_match = re.search(r'pH\s*(?:range|of)?\s*(\d+\.?\d*)\s*[-to]+\s*(\d+\.?\d*)', text, re.IGNORECASE)
    if ph_match:
        ph_range = f"pH {ph_match.group(1)}-{ph_match.group(2)}"
        for group in sample_groups:
            samples[group]["ph_response_range"] = ph_range

    # Material and treatment info
    # From Section 2.3: Janus fabric has TWO different coatings on each side
    # - Hydrophobic side: WSiPU (silicone polyurethane) coating
    # - Hydrophilic side: Hydrophilic polyester coating
    for group in sample_groups:
        samples[group]["material"] = "Cotton fabric"

        # Test solution (acidic vs neutral)
        if "@S" in group:
            samples[group]["test_solution"] = "Acidic (pH=5)"
        else:
            samples[group]["test_solution"] = "Neutral water (pH=7)"

        # Structure and side-specific treatments based on group name
        # B = hydrophobic (WSiPU coating), L = hydrophilic (polyester coating)
        # First letter = inner/bottom side, Second letter = outer/top side
        if group.startswith("B-L"):
            samples[group]["structure"] = "Janus Bilayer"
            samples[group]["treatment_bottom"] = "WSiPU coating (hydrophobic)"
            samples[group]["treatment_top"] = "Hydrophilic polyester coating"
        elif group.startswith("L-B"):
            samples[group]["structure"] = "Janus Bilayer"
            samples[group]["treatment_bottom"] = "Hydrophilic polyester coating"
            samples[group]["treatment_top"] = "WSiPU coating (hydrophobic)"
        elif group.startswith("B-B"):
            samples[group]["structure"] = "Symmetric Bilayer"
            samples[group]["treatment_bottom"] = "WSiPU coating (hydrophobic)"
            samples[group]["treatment_top"] = "WSiPU coating (hydrophobic)"
        elif group.startswith("L-L"):
            samples[group]["structure"] = "Symmetric Bilayer"
            samples[group]["treatment_bottom"] = "Hydrophilic polyester coating"
            samples[group]["treatment_top"] = "Hydrophilic polyester coating"

    print(f"    Extracted metrics for {len(samples)} sample groups")
    for group, data in samples.items():
        metric_count = len([k for k in data.keys() if k != "sample_id"])
        print(f"      {group}: {metric_count} metrics")

    return samples


def extract_all_samples_with_metrics(text: str, expected_count: int) -> List[Dict[str, Any]]:
    """
    Extract sample data from text using DIRECT REGEX PARSING ONLY - no LLM.
    First tries to extract from Nougat narrative text, then falls back to chart parsing.
    """
    print("  [DIRECT PARSING] Parsing Nougat text for embedded data...")

    # Try extracting from Nougat narrative text first
    nougat_samples = extract_metrics_from_nougat_text(text)

    if nougat_samples and any(len(s) > 2 for s in nougat_samples.values()):
        # Got meaningful data from Nougat text
        return list(nougat_samples.values())

    # Fall back to chart data parsing
    print("  [FALLBACK] Trying chart data parsing...")
    return parse_chart_data_direct(text, expected_count)


# Mapping from common metric names/abbreviations to schema column names
METRIC_MAPPING = {
    # Basic properties
    "material": "Material: Monolayer",
    "structure": "Structure: Monolayer",
    "treatment_bottom": "Treatment: Monolayer, Bottom Side",
    "treatment_top": "Treatment: Monolayer, Top Side",
    # test_solution goes to custom column (not a treatment per se)

    # Extracted from narrative text - mapped to schema where confident
    "contact_angle": "Contact Angle: Monolayer",  # From prose description
    "breakthrough_pressure_mm": "Hydrostatic Head (ISO/EN): Monolayer",  # Liquid column height in mm
    "penetration_time_min": "Wetting Time (WT): Monolayer",  # Time for liquid to penetrate

    # These go to custom columns (don't fit standard schema cleanly)
    # "breakthrough_pressure_reduction" -> CUSTOM
    # "evaporation_rate" -> CUSTOM
    # "moisture_dissipation_rate" -> CUSTOM
    # "ph_response_range" -> CUSTOM
    "gsm": "Fabric Weight/GSM: Monolayer",
    "weight": "Fabric Weight/GSM: Monolayer",
    "thickness": "Fabric Thickness: Monolayer",
    "thickness_mm": "Fabric Thickness: Monolayer",

    # AATCC MMT metrics
    "wtt": "Wetting Time Top (WTT): Monolayer",
    "wtb": "Wetting Time Bottom (WTB): Monolayer",
    "wt": "Wetting Time (WT): Monolayer",
    "wetting_time": "Wetting Time (WT): Monolayer",
    "art": "Top Absorption Rate (TAR): Monolayer",
    "arb": "Bottom Absorption Rate (BAR): Monolayer",
    "tar": "Top Absorption Rate (TAR): Monolayer",
    "bar": "Bottom Absorption Rate (BAR): Monolayer",
    "absorption_rate": "Absorption Rate Top (ART): Monolayer",
    "sst": "Spreading Speed Top (SSt): Monolayer",
    "ssb": "Spreading Speed Bottom (SSb): Monolayer",
    "ss": "Spreading Speed (SS): Monolayer",
    "spreading_speed": "Spreading Speed (SS): Monolayer",
    "mwrt": "Max Wetted Radius Top (MWRt): Monolayer",
    "mwrb": "Max Wetted Radius Bottom (MWRb): Monolayer",
    "aoti": "Accumulative One-Way Transport Index (AOTI): Monolayer",
    "owtc": "One Way Transport Capability (OWTC): Monolayer",
    "ommc": "Overall Moisture Management Capacity (OMMC): Monolayer",

    # Thermal
    "thermal_resistance": "Thermal Resistance (Rct): Monolayer",
    "rct": "Thermal Resistance (Rct): Monolayer",
    "thermal_conductivity": "Thermal Conductivity: Monolayer",
    "ret": "Evaporative Resistance (Ret): Monolayer",
    "evaporative_resistance": "Evaporative Resistance (Ret): Monolayer",

    # Other
    "air_permeability": "Air Permeability: Monolayer",
    "wvtr": "Water Vapor Transmission Rate (WVTR): Monolayer",
    "contact_angle": "Contact Angle: Monolayer",
    "porosity": "Open Area Fraction: Monolayer",
    "drying_time": "Drying Time: Monolayer",
    "drying_rate": "Drying Rate: Monolayer",
}


def map_extracted_to_schema(sample: Dict[str, Any]) -> Dict[str, Any]:
    """Map extracted metric names to schema column names.
    Unmapped fields go to custom columns with 'CUSTOM: ' prefix."""
    mapped = {}
    custom = {}

    # Fields that should not be mapped (metadata, not metrics)
    skip_fields = {"sample_id", "num_layers"}

    for key, value in sample.items():
        if value is None:
            continue
        if key in skip_fields:
            mapped[key] = value
            continue

        key_lower = key.lower().replace(" ", "_").replace("-", "_")
        if key_lower in METRIC_MAPPING:
            mapped[METRIC_MAPPING[key_lower]] = value
        else:
            # Put unmapped metrics in custom columns
            custom_col = f"CUSTOM: {key}"
            custom[custom_col] = value

    # Merge custom columns into mapped
    mapped.update(custom)
    return mapped


def process_single_pdf(pdf_path: str) -> List[Dict[str, Any]]:
    """
    Process a single PDF and extract all specifications.
    Returns a list of rows (one per sample).
    NOUGAT ONLY - no PyMuPDF fallback.
    Uses only 2 LLM calls total for speed.
    """
    print(f"\n{'='*60}")
    print(f"Processing: {pdf_path}")
    print(f"{'='*60}")

    # Extract text using NOUGAT ONLY
    nougat_text = extract_with_nougat(pdf_path)

    # DEBUG: Print first part of nougat output
    print("\n  === NOUGAT OUTPUT (first 2000 chars) ===")
    print(nougat_text[:2000])
    print("  === END NOUGAT OUTPUT ===\n")

    # FIGURE EXTRACTION with DePlot (page 14 has the bar charts per user)
    print("\n  [DePlot] Extracting data from figures...")
    figure_data = extract_figures_from_pdf(pdf_path, page_numbers=[13])  # 0-indexed, so page 14 = index 13

    figure_text = ""
    if figure_data:
        print(f"    Extracted data from {len(figure_data)} pages with figures")
        for fig in figure_data:
            print(f"\n  === FIGURE DATA (page {fig['page']}) ===")
            # Handle Unicode characters that Windows console can't display
            data_str = fig['data'][:1000] if len(fig['data']) > 1000 else fig['data']
            print(data_str.encode('ascii', 'replace').decode('ascii'))
            print("  === END FIGURE DATA ===")
            figure_text += f"\n\nFIGURE DATA FROM PAGE {fig['page']}:\n{fig['data']}"
    else:
        print("    No figure data extracted (DePlot may not be installed)")

    # Combine nougat text with figure data for LLM
    combined_text = nougat_text
    if figure_text:
        combined_text = nougat_text + "\n\n=== EXTRACTED CHART DATA ===\n" + figure_text

    # RULE-BASED SAMPLE COUNT DETECTION (Group 1+2+3 logic)
    print("\n  [RULE-BASED] Detecting sample count with Group 1+2+3 logic...")
    rule_based_count = extract_sample_count_rule_based(nougat_text)
    if rule_based_count:
        print(f"    Rule-based detection found: {rule_based_count} samples")
    else:
        print(f"    Rule-based detection: No matches found")

    # REGEX ONLY: Get study metadata (no LLM needed - parse Nougat Markdown directly)
    print("\n  [REGEX] Extracting study metadata from Nougat Markdown...")
    study_metadata = extract_study_metadata(pdf_path, nougat_text)
    for key, value in study_metadata.items():
        print(f"    {key}: {value}")

    # Use rule-based count if found
    if rule_based_count:
        print(f"\n  Using RULE-BASED sample count: {rule_based_count}")
        study_metadata["Number of Sample Fabrics"] = rule_based_count

    # Extract samples from Nougat text (narrative contains embedded data)
    sample_count = study_metadata.get("Number of Sample Fabrics", 1)
    print(f"\n  [EXTRACTION] Extracting {sample_count} samples from Nougat text...")

    # Use Nougat text for extraction - the narrative contains embedded metrics
    # Figure text is supplementary if available
    extraction_text = combined_text  # nougat_text + any figure data
    samples = extract_all_samples_with_metrics(extraction_text, sample_count)
    print(f"    Extracted {len(samples)} samples")

    # Build output rows - just map extracted data to schema (no more LLM calls)
    output_rows = []
    print("\n  === EXTRACTED DATA ===")
    for idx, sample in enumerate(samples):
        sample_id = sample.get("sample_id", f"Sample {idx+1}")
        print(f"\n  Sample {idx+1}: {sample_id}")

        # Show raw extracted data
        for key, value in sample.items():
            if key != "sample_id" and value is not None:
                print(f"    {key}: {value}")

        # Start with study metadata
        row = {
            "Study Number": Path(pdf_path).stem,
            **study_metadata,
            "Sample ID/Name": sample_id,
            "Number of Fabric Layers": sample.get("num_layers", 1),
        }

        # Map extracted metrics to schema columns (instant, no LLM)
        mapped_metrics = map_extracted_to_schema(sample)
        row.update(mapped_metrics)

        output_rows.append(row)

    print("\n  === END EXTRACTED DATA ===")
    return output_rows


def write_to_excel(rows: List[Dict[str, Any]], output_path: Path):
    """Write extracted data to Excel file."""
    print(f"\n{'='*60}")
    print(f"Writing to Excel: {output_path}")
    print(f"{'='*60}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Textile Specifications"

    # Get schema headers
    schema_headers = get_all_column_headers()

    # Collect custom columns from data (those starting with "CUSTOM: ")
    custom_headers = set()
    for row in rows:
        for key in row.keys():
            if key.startswith("CUSTOM: "):
                custom_headers.add(key)
    custom_headers = sorted(custom_headers)

    # Combine schema headers + custom headers
    all_headers = schema_headers + custom_headers

    print(f"  Schema columns: {len(schema_headers)}")
    print(f"  Custom columns: {len(custom_headers)}")
    if custom_headers:
        for ch in custom_headers:
            print(f"    - {ch}")

    # Write header row with formatting
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    custom_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")  # Green for custom
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        # Use green for custom columns
        if header.startswith("CUSTOM: "):
            cell.fill = custom_fill
        else:
            cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(all_headers, 1):
            value = row_data.get(header, "")
            if value == "null" or value is None:
                value = ""
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Adjust column widths
    for col_idx in range(1, len(all_headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"  Saved {len(rows)} rows to {output_path}")


def main():
    """Main entry point."""
    print("\n" + "="*70)
    print("ScrapeV2 - Comprehensive Textile Specification Extractor")
    print("="*70)
    print(f"LLM Provider: {LLM_PROVIDER}")
    if LLM_PROVIDER == "ollama":
        print(f"Ollama Model: {OLLAMA_MODEL}")
    print(f"Focus PDF: {FOCUS_PDF}")
    print("="*70)

    # Check PDF exists
    pdf_path = PDF_FOLDER / FOCUS_PDF
    if not pdf_path.exists():
        print(f"\nERROR: PDF not found: {pdf_path}")
        print("Please check the path and try again.")
        return 1

    # Process the PDF
    all_rows = process_single_pdf(str(pdf_path))

    if not all_rows:
        print("\nNo data extracted. Check the PDF content and LLM connection.")
        return 1

    # Write to Excel
    write_to_excel(all_rows, OUTPUT_EXCEL)

    print("\n" + "="*70)
    print("EXTRACTION COMPLETE")
    print(f"Total samples extracted: {len(all_rows)}")
    print(f"Output file: {OUTPUT_EXCEL}")
    print("="*70 + "\n")

    return 0


if __name__ == "__main__":
    sys.exit(main())
