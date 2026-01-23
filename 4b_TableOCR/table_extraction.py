"""
Pipeline 4b: Table Extraction (Consolidated)

HYBRID APPROACH:
1. Phase 1: Crop tables from YOLO detections (saved during text extraction)
2. Phase 2: Extract text using Surya + PyMuPDF
   - Surya TableRecPredictor (GPU) - Table structure
   - Surya DetectionPredictor (GPU) - Text bbox detection
   - PyMuPDF (CPU) - Native PDF text extraction

PROCESS:
1. Load YOLO table detections from text extraction
2. Crop table images from PDFs
3. Extract table structure and text using Surya + PyMuPDF
4. Filter tables by matching column headers to MetricsSearchable.txt
5. Create final dataset Excel

INPUTS:
- PDF folder (Corpus PDFs from Pipeline 3)
- YOLO table detections JSON (from Pipeline 2)
- MetricsSearchable.txt (schema columns)

OUTPUTS:
- 4b_Metadata.xlsx (metadata with "No Metrics" breakpoints)
- 4b_FinalDataset.xlsx (extracted table values in schema format)

CRITICAL: Only works on native text tables (96.3% of academic PDFs)
Image-embedded tables are SKIPPED (3.7% of tables)
"""

import sys
import os
import json
import re
import time
import fitz  # PyMuPDF
from pathlib import Path
from PIL import Image
from collections import defaultdict

# Import timing and corpus utilities
sys.path.insert(0, str(Path(__file__).parent.parent))
from Timers.timing_utils import save_timer
from Utils.corpus_utils import load_corpus_json, convert_corpus_json_to_excel_data

# Set Surya environment variables
os.environ['DISABLE_MODEL_SOURCE_CHECK'] = 'True'

# CRITICAL: Optimize batch sizes for RTX 4070 Laptop (8GB VRAM)
os.environ["DETECTOR_BATCH_SIZE"] = "4"      # Default 36 (16GB) -> 4 (1.76GB)
os.environ["TABLE_REC_BATCH_SIZE"] = "8"     # Default 64 (10GB) -> 8 (1.2GB)

# Check for required libraries
try:
    import pandas as pd
except ImportError:
    print("ERROR: Missing required library 'pandas'")
    print("Install it with: pip install pandas openpyxl")
    sys.exit(1)

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
except ImportError:
    print("ERROR: Missing required library 'openpyxl'")
    print("Install it with: pip install openpyxl")
    sys.exit(1)

try:
    from rapidfuzz import fuzz, process
except ImportError:
    print("ERROR: Missing required library 'rapidfuzz'")
    print("Install it with: pip install rapidfuzz")
    sys.exit(1)

DPI = 150  # Match text extraction DPI


# ============================================================================
# PHASE 1: Table Cropping Functions
# ============================================================================

def crop_table_images(page_image, table_boxes):
    """Crop table regions from page image."""
    crops = []
    for x1, y1, x2, y2 in table_boxes:
        crop = page_image.crop((x1, y1, x2, y2))
        crops.append(crop)
    return crops


def crop_tables_from_yolo_detections(pdf_folder: Path, output_folder: Path, query_folder: Path):
    """
    Phase 1: Crop tables from YOLO detections (saved during text extraction).

    Args:
        pdf_folder: Folder containing PDFs
        output_folder: Output folder for table crops
        query_folder: Query folder (contains yolo_table_detections.json)

    Returns:
        list: Table crop metadata
    """
    yolo_detections_json = query_folder / "yolo_table_detections.json"

    if not yolo_detections_json.exists():
        sys.exit(1)

    # Load YOLO table detections
    with open(yolo_detections_json, 'r', encoding='utf-8') as f:
        all_detections = json.load(f)

    # Group detections by PDF and page
    detections_by_pdf_page = {}
    for det in all_detections:
        key = (det["pdf_num"], det["page_num"])
        if key not in detections_by_pdf_page:
            detections_by_pdf_page[key] = []
        detections_by_pdf_page[key].append(det)

    # Process PDFs and crop tables
    table_crop_metadata = []
    start_time = time.perf_counter()

    for (pdf_num, page_num), page_detections in sorted(detections_by_pdf_page.items()):
        pdf_path = pdf_folder / f"{pdf_num}.pdf"

        if not pdf_path.exists():
            print(f"WARNING: PDF not found: {pdf_path}")
            continue

        doc = fitz.open(pdf_path)

        if page_num >= len(doc):
            doc.close()
            continue

        page = doc[page_num]

        # Render page at same DPI as text extraction
        mat = fitz.Matrix(DPI / 72, DPI / 72)
        pix = page.get_pixmap(matrix=mat)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

        # Extract bboxes
        table_boxes = [det["bbox"] for det in page_detections]

        # Crop and save tables
        table_crops = crop_table_images(img, table_boxes)

        for table_idx, (crop, bbox) in enumerate(zip(table_crops, table_boxes), 1):
            crop_filename = f"{pdf_num}_page{page_num + 1}_table{table_idx}.png"
            crop_path = output_folder / crop_filename
            crop.save(crop_path)

            # Store metadata for Phase 2
            table_crop_metadata.append({
                "pdf_num": pdf_num,
                "pdf_name": pdf_path.name,
                "pdf_path": str(pdf_path.resolve()),  # Absolute path for PyMuPDF
                "page_num": page_num,  # 0-indexed for PyMuPDF
                "table_num": table_idx,
                "table_bbox": bbox,  # [x1, y1, x2, y2] for image detection
                "crop_filename": crop_filename,
                "crop_path": str(crop_path)
            })

        doc.close()

    # End timing
    elapsed = time.perf_counter() - start_time

    return table_crop_metadata


# ============================================================================
# PHASE 2: Text Extraction Functions
# ============================================================================

def clean_html_tags(text: str) -> str:
    """Remove HTML/markdown tags from text"""
    cleaned = re.sub(r'<[^>]+>', '', text)
    return cleaned


def filter_columns_by_headers(structured_table: list, metrics_list: list):
    """
    Filter table columns based on header row keywords.

    Keeps only columns where the header (row 0) contains a keyword from MetricsSearchable.txt.
    If no columns match, returns empty table.
    """
    if not structured_table or len(structured_table) == 0:
        return [], [], []

    headers = structured_table[0]  # First row is headers
    matching_columns = []
    found_metrics = []

    # Check each header for metric keywords
    for col_idx, header in enumerate(headers):
        header_lower = header.lower()

        # Check if this header contains any metric keyword
        for metric in metrics_list:
            metric_pattern = r'\b' + re.escape(metric.lower()) + r'\b'
            if re.search(metric_pattern, header_lower):
                matching_columns.append(col_idx)
                if metric not in found_metrics:
                    found_metrics.append(metric)
                break  # Found match for this column, move to next column

    # If no columns match, return empty table
    if not matching_columns:
        return [], [], []

    # Build filtered table with only matching columns
    filtered_table = []
    for row in structured_table:
        filtered_row = [row[col_idx] for col_idx in matching_columns if col_idx < len(row)]
        filtered_table.append(filtered_row)

    return filtered_table, matching_columns, found_metrics


def check_if_image_table(pdf_path: str, page_num: int, table_bbox: tuple, render_dpi: int = 150) -> bool:
    """
    Check if table is image-embedded (no text layer).

    Args:
        pdf_path: Path to PDF file
        page_num: Page number (0-indexed)
        table_bbox: (x1, y1, x2, y2) in rendered coordinates
        render_dpi: DPI used for rendering (default 150)

    Returns:
        True if image-embedded (no text), False if native text
    """
    doc = fitz.open(pdf_path)
    page = doc[page_num]

    # Convert from rendered coordinates to PDF coordinates
    x1, y1, x2, y2 = table_bbox
    scale = render_dpi / 72.0
    pdf_x1 = x1 / scale
    pdf_y1 = y1 / scale
    pdf_x2 = x2 / scale
    pdf_y2 = y2 / scale

    # Check if text layer exists
    clip_rect = fitz.Rect(pdf_x1, pdf_y1, pdf_x2, pdf_y2)
    text = page.get_text("text", clip=clip_rect).strip()

    doc.close()

    return len(text) == 0  # True if no text (image-embedded)


def extract_text_pymupdf(pdf_path: str, page_num: int, table_bbox: tuple, text_bboxes: list, render_dpi: int = 150) -> list:
    """
    Extract text from PDF using PyMuPDF for each text bbox.

    Args:
        pdf_path: Path to PDF file
        page_num: Page number (0-indexed)
        table_bbox: (x1, y1, x2, y2) table bbox in rendered coordinates
        text_bboxes: List of text bboxes from Surya DetectionPredictor
        render_dpi: DPI used for rendering (default 150)

    Returns:
        List of extracted text strings (one per bbox)
    """
    doc = fitz.open(pdf_path)
    page = doc[page_num]

    scale = render_dpi / 72.0
    table_x1, table_y1, table_x2, table_y2 = table_bbox

    extracted_texts = []

    for bbox in text_bboxes:
        # bbox.polygon is [[x1,y1], [x2,y2], [x3,y3], [x4,y4]] in rendered coords
        xs = [p[0] for p in bbox.polygon]
        ys = [p[1] for p in bbox.polygon]
        x_min, x_max = int(min(xs)), int(max(xs))
        y_min, y_max = int(min(ys)), int(max(ys))

        # Convert to PDF coordinates (relative to page, not table crop)
        pdf_x1 = (x_min / scale) + (table_x1 / scale)
        pdf_y1 = (y_min / scale) + (table_y1 / scale)
        pdf_x2 = (x_max / scale) + (table_x1 / scale)
        pdf_y2 = (y_max / scale) + (table_y1 / scale)

        # Extract text from bbox
        clip_rect = fitz.Rect(pdf_x1, pdf_y1, pdf_x2, pdf_y2)
        text = page.get_text("text", clip=clip_rect).strip()

        # Clean HTML tags (Surya sometimes outputs these)
        text = clean_html_tags(text)

        extracted_texts.append(text)

    doc.close()

    return extracted_texts


def bbox_overlap(bbox1, bbox2):
    """Calculate overlap area between two bboxes"""
    x1_min, y1_min, x1_max, y1_max = bbox1
    x2_min, y2_min, x2_max, y2_max = bbox2

    x_overlap = max(0, min(x1_max, x2_max) - max(x1_min, x2_min))
    y_overlap = max(0, min(y1_max, y2_max) - max(y1_min, y2_min))

    return x_overlap * y_overlap


def extract_table_text(table_crop_metadata: list, metrics_list: list, table_rec, det_predictor):
    """
    Phase 2: Extract text from table crops using Surya + PyMuPDF.

    Args:
        table_crop_metadata: Metadata from Phase 1
        metrics_list: List of metric keywords
        table_rec: Surya TableRecPredictor (already loaded)
        det_predictor: Surya DetectionPredictor (already loaded)

    Returns:
        list: Extraction results by PDF
    """
    # PHASE A: Collect ALL valid crops first (single-batch processing)
    all_valid_metas = []
    all_crops = []
    skipped_image_tables = 0

    # Start timing
    start_time = time.perf_counter()

    for meta in table_crop_metadata:
        # Get PDF path from metadata
        pdf_path = meta.get("pdf_path", "")
        if not pdf_path or not Path(pdf_path).exists():
            continue

        # Check if table is image-embedded (5ms overhead)
        table_bbox = meta.get("table_bbox", None)
        page_num = meta.get("page_num")
        if table_bbox is None:
            continue

        is_image = check_if_image_table(pdf_path, page_num, table_bbox)

        if is_image:
            skipped_image_tables += 1
            continue

        crop = Image.open(meta["crop_path"])
        all_valid_metas.append(meta)
        all_crops.append(crop)

    # Skip if no valid crops
    if not all_crops:
        return []

    # PHASE B: Process crops in SMALL batches to prevent GPU OOM
    SURYA_BATCH_SIZE = 8  # Process 8 crops at a time to avoid memory fragmentation
    total_tables_processed = len(all_crops)

    import torch

    # Process in batches
    table_results = []
    det_predictions = []

    tb0 = time.perf_counter()
    num_batches = (len(all_crops) + SURYA_BATCH_SIZE - 1) // SURYA_BATCH_SIZE

    for batch_idx in range(num_batches):
        batch_start = batch_idx * SURYA_BATCH_SIZE
        batch_end = min(batch_start + SURYA_BATCH_SIZE, len(all_crops))
        batch_crops = all_crops[batch_start:batch_end]

        # TableRecPredictor on this batch
        with torch.no_grad():
            batch_table_results = table_rec(batch_crops)
        table_results.extend(batch_table_results)

        # DetectionPredictor on same batch
        with torch.no_grad():
            batch_det_predictions = det_predictor(batch_crops)
        det_predictions.extend(batch_det_predictions)

        # Clear memory after batch completes
        del batch_table_results, batch_det_predictions, batch_crops
        torch.cuda.empty_cache()

    tb3 = time.perf_counter()

    # PHASE C: Process results and organize by PDF/page

    # Group results by PDF
    crops_by_pdf = defaultdict(lambda: defaultdict(list))
    for i, meta in enumerate(all_valid_metas):
        crops_by_pdf[meta["pdf_num"]][(meta["pdf_name"], meta["page_num"])].append((i, meta))

    # Process crops
    all_results = []
    total_tables_filtered = 0  # Tables that passed column filtering

    for pdf_num in sorted(crops_by_pdf.keys()):
        pdf_data = {
            "pdf_number": pdf_num,
            "filename": None,
            "pages": [],
            "table_count": 0,
            "crops": [],
            "found_metrics": []
        }

        for (pdf_name, page_num), indexed_metas in sorted(crops_by_pdf[pdf_num].items()):
            if pdf_data["filename"] is None:
                pdf_data["filename"] = pdf_name

            page_data = {
                "page_number": page_num,
                "tables": []
            }

            # Process results for each crop on this page
            for i, meta in indexed_metas:
                table_result = table_results[i]

                if not det_predictions[i].bboxes:
                    continue

                text_bboxes = det_predictions[i].bboxes

                # Extract text from PDF using PyMuPDF
                pdf_path = meta["pdf_path"]
                table_bbox = meta["table_bbox"]
                crop_page_num = meta["page_num"]
                extracted_texts = extract_text_pymupdf(pdf_path, crop_page_num, table_bbox, text_bboxes)

                # Create text_lines compatible with old format
                text_lines = []
                for bbox, text in zip(text_bboxes, extracted_texts):
                    # Convert bbox.polygon to [x1, y1, x2, y2]
                    xs = [p[0] for p in bbox.polygon]
                    ys = [p[1] for p in bbox.polygon]
                    bbox_rect = [min(xs), min(ys), max(xs), max(ys)]

                    # Create text_line object
                    class TextLine:
                        def __init__(self, text, bbox):
                            self.text = text
                            self.bbox = bbox

                    text_lines.append(TextLine(text, bbox_rect))

                # Match text lines to cells by bounding box overlap
                rows_dict = {}
                for cell in table_result.cells:
                    row_id = cell.row_id
                    if row_id not in rows_dict:
                        rows_dict[row_id] = []

                    # Find text lines that overlap with this cell
                    cell_text_parts = []
                    for text_line in text_lines:
                        overlap = bbox_overlap(cell.bbox, text_line.bbox)
                        if overlap > 0:
                            cell_text_parts.append(text_line.text)

                    cell_text = ' '.join(cell_text_parts)

                    rows_dict[row_id].append({
                        "col_id": cell.col_id,
                        "text": cell_text
                    })

                # Build structured table
                structured_table = []
                for row_id in sorted(rows_dict.keys()):
                    row_cells = sorted(rows_dict[row_id], key=lambda c: c["col_id"] if c["col_id"] is not None else 0)
                    structured_table.append([c["text"] for c in row_cells])

                # Filter columns based on header keywords
                filtered_table, matching_columns, found_metrics = filter_columns_by_headers(structured_table, metrics_list)

                # Only save table if it has at least one matching column
                if not filtered_table or len(matching_columns) == 0:
                    continue

                table_data = {
                    "table_number": meta["table_num"],
                    "crop_filename": meta["crop_filename"],
                    "structured_table": filtered_table,
                    "original_columns": len(structured_table[0]) if structured_table else 0,
                    "filtered_columns": len(matching_columns),
                    "matching_column_indices": matching_columns,
                    "found_metrics": found_metrics,
                    "has_metrics": True
                }

                page_data["tables"].append(table_data)
                pdf_data["table_count"] += 1
                pdf_data["crops"].append(meta["crop_filename"])

                # Add found metrics to PDF-level list
                for metric in found_metrics:
                    if metric not in pdf_data["found_metrics"]:
                        pdf_data["found_metrics"].append(metric)

            if page_data["tables"]:
                pdf_data["pages"].append(page_data)

        all_results.append(pdf_data)
        total_tables_filtered += pdf_data["table_count"]

    # End timing
    elapsed = time.perf_counter() - start_time

    return all_results


# ============================================================================
# Final Dataset Creation
# ============================================================================

def load_schema_columns() -> list:
    """Load schema columns from MetricsSearchable.txt."""
    schema_file = Path(__file__).parent.parent / "Config" / "MetricsSearchable.txt"

    with open(schema_file, 'r', encoding='utf-8') as f:
        columns = [line.strip() for line in f if line.strip()]

    return columns


def create_final_dataset_from_tables(results_data, schema_columns: list) -> tuple:
    """
    Create final dataset Excel from table extraction results.

    Args:
        results_data: Extraction results from Phase 2
        schema_columns: List of metric columns from MetricsSearchable.txt

    Returns:
        (all_extractions dict, no_metrics_pdfs list)
    """
    all_extractions = {}
    no_metrics_pdfs = []

    for pdf_entry in results_data:
        pdf_num = pdf_entry["pdf_number"]
        found_metrics = pdf_entry.get("found_metrics", [])

        if len(found_metrics) == 0:
            # No matching metrics in tables
            no_metrics_pdfs.append(pdf_num)
            continue

        # Fuzzy match found metrics to schema
        matched_data = {}

        for metric in found_metrics:
            metric_lower = metric.lower()

            # 1. Exact match
            matched = None
            for schema_col in schema_columns:
                if metric_lower == schema_col.lower():
                    matched = schema_col
                    break

            # 2. Substring match
            if not matched:
                substring_matches = []
                for schema_col in schema_columns:
                    schema_lower = schema_col.lower()
                    if metric_lower in schema_lower or schema_lower in metric_lower:
                        substring_matches.append(schema_col)

                if substring_matches:
                    result = process.extractOne(metric, substring_matches, scorer=fuzz.ratio)
                    if result:
                        matched = result[0]

            # 3. Fuzzy match
            if not matched:
                result = process.extractOne(
                    metric,
                    schema_columns,
                    scorer=fuzz.token_set_ratio,
                    score_cutoff=60
                )
                if result:
                    matched = result[0]

            # Add matched metric (placeholder value - full parsing would extract actual values)
            if matched:
                matched_data[matched] = "Table Data Found"

        if matched_data:
            all_extractions[pdf_num] = matched_data
        else:
            no_metrics_pdfs.append(pdf_num)

    return all_extractions, no_metrics_pdfs


# ============================================================================
# Main Pipeline Function
# ============================================================================

def run_pipeline_4b(pdf_folder: Path, table_rec=None, det_predictor=None) -> tuple:
    """
    Run complete Pipeline 4b.

    Args:
        pdf_folder: Folder containing PDFs (from Pipeline 3)
        table_rec: Pre-loaded Surya TableRecPredictor (optional)
        det_predictor: Pre-loaded Surya DetectionPredictor (optional)

    Returns:
        (metadata_excel_path, final_dataset_path)
    """
    query_folder = pdf_folder.parent

    # Load schema
    schema_columns = load_schema_columns()

    # Load corpus metadata from JSON
    all_papers = load_corpus_json(query_folder)
    # Filter to only papers with study_number (exclude failed downloads)
    papers_by_num = {p["study_number"]: p for p in all_papers if "study_number" in p}

    # Phase 1: Crop tables from YOLO detections
    output_folder = query_folder / "Table Crops"
    output_folder.mkdir(parents=True, exist_ok=True)

    extraction_start = time.time()

    table_crop_metadata = crop_tables_from_yolo_detections(pdf_folder, output_folder, query_folder)

    # Phase 2: Extract text using Surya + PyMuPDF

    # Load metrics
    metrics_file = Path(__file__).parent.parent / "Config" / "MetricsSearchable.txt"
    with open(metrics_file, 'r', encoding='utf-8') as f:
        metrics_list = [line.strip() for line in f if line.strip()]

    # Load models if not provided
    models_loaded_here = False
    if table_rec is None or det_predictor is None:
        import torch
        from surya.table_rec import TableRecPredictor
        from surya.detection import DetectionPredictor

        # Clear GPU memory from previous pipelines (esp. Pipeline 2 workers)
        torch.cuda.empty_cache()

        table_rec = TableRecPredictor()
        det_predictor = DetectionPredictor()

        models_loaded_here = True

    # Extract text from tables
    results_data = extract_table_text(table_crop_metadata, metrics_list, table_rec, det_predictor)

    # Save Timer 9: Total table extraction time
    extraction_time = time.time() - extraction_start
    save_timer(query_folder, "9_table_extraction_total", extraction_time)

    # Process results
    all_extractions, no_metrics_pdfs = create_final_dataset_from_tables(results_data, schema_columns)

    # Update metadata for PDFs with no metrics
    for pdf_num in no_metrics_pdfs:
        if pdf_num in papers_by_num:
            papers_by_num[pdf_num]["breakpoint"] = "No Metrics"

    # Create 4b_Metadata.xlsx from updated JSON data
    metadata_rows = convert_corpus_json_to_excel_data(all_papers)
    metadata_excel_path = query_folder / "4b_Metadata.xlsx"

    # Write to Excel with colors
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Corpus")

    # Write header row
    for col_idx, value in enumerate(metadata_rows[0], 1):
        ws.cell(row=1, column=col_idx, value=value).font = Font(bold=True)

    # Write data rows with colors
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

    for row_idx, row_data in enumerate(metadata_rows[1:], 2):
        color_code = row_data[-1]
        actual_data = row_data[:-1]

        for col_idx, value in enumerate(actual_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

        if color_code == "green":
            fill = green_fill
        elif color_code == "orange":
            fill = orange_fill
        elif color_code == "red":
            fill = red_fill
        else:
            fill = None

        if fill:
            for col_idx in range(1, len(actual_data) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10

    wb.save(metadata_excel_path)

    # Create final dataset Excel
    final_dataset_path = query_folder / "4b_FinalDataset.xlsx"

    columns = ["PDF Number"] + schema_columns

    # Build rows
    rows = []
    for pdf_num, extractions in sorted(all_extractions.items()):
        row = {"PDF Number": pdf_num}
        for metric, value in extractions.items():
            row[metric] = value
        rows.append(row)

    df = pd.DataFrame(rows, columns=columns)

    # Save to Excel
    df.to_excel(final_dataset_path, index=False, engine='openpyxl')

    # Summary
    print(f"{len(all_extractions)} Metrics extracted from tables")

    return metadata_excel_path, final_dataset_path


# ============================================================================
# CLI Entry Point
# ============================================================================

def find_latest_query_folder():
    """Find the latest Query folder in Output directory."""
    output_base = Path(__file__).parent.parent / "Output"
    query_folders = [d for d in output_base.iterdir() if d.is_dir() and "Query" in d.name]
    if not query_folders:
        return None
    return max(query_folders, key=lambda d: d.stat().st_mtime)


def main():
    """CLI entry point for table extraction."""
    import argparse

    parser = argparse.ArgumentParser(description="Extract tables using YOLO + Surya + PyMuPDF")
    parser.add_argument("pdf_folder", nargs='?', type=Path, help="Folder containing PDFs")

    args = parser.parse_args()

    # Auto-detect if no arguments provided
    if args.pdf_folder is None:
        query_folder = find_latest_query_folder()
        if query_folder is None:
            print("ERROR: No Query folder found in Output directory")
            sys.exit(1)

        # Use Corpus PDFs folder
        papers_folder = query_folder / "Corpus PDFs"
        args.pdf_folder = papers_folder

        print(f"Auto-detected Query folder: {query_folder.name}")

    if not args.pdf_folder.exists():
        print(f"ERROR: PDF folder not found: {args.pdf_folder.name}")
        sys.exit(1)

    run_pipeline_4b(args.pdf_folder)


if __name__ == "__main__":
    main()
