"""
Pipeline 4b: Table Extraction with PyMuPDF Hybrid Method

HYBRID APPROACH:
1. PaddleOCR PPStructure - Table detection (finds tables in PDFs)
2. Surya TableRecPredictor - Table structure (rows/columns/cells)
3. Surya DetectionPredictor - Text bbox detection
4. PyMuPDF - Native PDF text extraction (33x faster than Surya OCR)

PROCESS:
1. Create copy of metadata Excel from Pipeline 3
2. Extract tables from PDFs using hybrid approach
3. Match table columns to MetricsSearchable.txt
4. For PDFs with NO matching metrics -> mark as "No Metrics" (red row)
5. Create final dataset Excel with extracted table values
6. Save metadata copy and final dataset

INPUTS:
- PDF folder (OA Papers / Non-OA Papers from Pipeline 3)
- Metadata Excel file (from Pipeline 3)
- MetricsSearchable.txt (schema columns)

OUTPUTS:
- 4b_Metadata.xlsx (metadata copy with "No Metrics" breakpoints)
- 4b_FinalDataset.xlsx (extracted table values in schema format)

LIMITATIONS:
- Only works on native text tables (96.3% of academic PDFs)
- Image-embedded tables are SKIPPED (3.7% of tables)
- See: Docs-ProjectScripts/PyMuPDF_Hybrid_Method.md
"""

import sys
import os
from pathlib import Path
import shutil
import subprocess
import json
import time

# Import timing and corpus utilities
sys.path.insert(0, str(Path(__file__).parent.parent))
from Timers.timing_utils import save_timer
from Utils.corpus_utils import load_corpus_json, convert_corpus_json_to_excel_data

print("[DEBUG] Script starting...")

# Check for required libraries
try:
    import pandas as pd
except ImportError:
    print("ERROR: Missing required library 'pandas'")
    print("Install it with: pip install pandas openpyxl")
    sys.exit(1)

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
except ImportError:
    print("ERROR: Missing required library 'openpyxl'")
    print("Install it with: pip install openpyxl")
    sys.exit(1)

try:
    from rapidfuzz import fuzz, process
except ImportError:
    print("ERROR: Missing required library 'rapidfuzz'")
    print("Install it with: pip install rapidfuzz")
    sys.exit(1)


def load_schema_columns() -> list:
    """Load schema columns from MetricsSearchable.txt."""
    schema_file = Path(__file__).parent.parent / "Config" / "MetricsSearchable.txt"

    with open(schema_file, 'r', encoding='utf-8') as f:
        columns = [line.strip() for line in f if line.strip()]

    return columns


def extract_tables(pdf_folder: Path, output_folder: Path, metrics_file: Path, query_folder: Path = None):
    """
    Run table extraction pipeline (Phase 1 + Phase 2).

    Args:
        pdf_folder: Folder containing PDFs
        output_folder: Output folder for crops (Table Crops folder)
        metrics_file: Path to MetricsSearchable.txt
        query_folder: Query folder for timer saving (optional)

    Returns:
        Path to phase2_results.json
    """
    # Timer 9: Total table extraction (track Phase 1 + Phase 2 for all PDFs)
    extraction_start = time.time()

    print("[1/5] Running table extraction (YOLO + Surya)...\n")

    # Create output folder (crops go directly here, no Crops subfolder)
    output_folder.mkdir(parents=True, exist_ok=True)

    metadata_json = output_folder / "phase1_metadata.json"
    results_json = output_folder / "phase2_results.json"

    # PHASE 1: YOLO table cropping (subprocess - reuses detections from text extraction)
    print("  Phase 1: Cropping tables from YOLO detections...")
    yolo_script = Path(__file__).parent / "yolo_phase1.py"

    subprocess.run(
        [
            sys.executable,
            str(yolo_script),
            str(pdf_folder),
            str(output_folder),
            str(metadata_json)
        ],
        check=True
    )

    # Load phase 1 results
    with open(metadata_json, 'r', encoding='utf-8') as f:
        table_crops = json.load(f)

    print(f"  Phase 1 complete: {len(table_crops)} table crops detected\n")

    # PHASE 2: PyMuPDF Hybrid text extraction (direct call - no subprocess overhead)
    print("  Phase 2: Extracting text with PyMuPDF Hybrid...")

    # Import and call directly instead of subprocess
    from surya_phase2_pymupdf import process_tables

    try:
        process_tables(metadata_json, metrics_file, results_json)
        print(f"  Phase 2 complete: Results saved to {results_json.name}\n")
    except Exception as e:
        print(f"  WARNING: Phase 2 failed with error: {e}")
        print(f"  Continuing without table results\n")

    # Save Timer 9: Total table extraction (Phase 1 + Phase 2 time for all PDFs)
    extraction_time = time.time() - extraction_start
    if query_folder:
        save_timer(query_folder, "9_table_extraction_total", extraction_time)

    # Clean up metadata JSON files (keep only table crop images)
    print("  Cleaning up temporary metadata files...")
    if metadata_json.exists():
        metadata_json.unlink()
        print(f"    Deleted: {metadata_json.name}")
    if results_json.exists():
        # Need to save results_json data before deleting
        with open(results_json, 'r', encoding='utf-8') as f:
            results_data = json.load(f)
        results_json.unlink()
        print(f"    Deleted: {results_json.name}\n")
        # Return the data instead of the file path
        return results_data

    return results_json


def create_final_dataset_from_tables(results_data, schema_columns: list) -> tuple:
    """
    Create final dataset Excel from table extraction results.

    Args:
        results_data: Either Path to phase2_results.json or the data itself (list/dict)
        schema_columns: List of metric columns from MetricsSearchable.txt

    Returns:
        (all_extractions dict, no_metrics_pdfs list)
    """
    print("[2/5] Processing table extraction results...\n")

    # Load results (support both Path and data)
    if isinstance(results_data, (Path, str)):
        with open(results_data, 'r', encoding='utf-8') as f:
            results = json.load(f)
    else:
        results = results_data

    print(f"  Loaded results for {len(results)} PDFs")

    all_extractions = {}
    no_metrics_pdfs = []

    for pdf_entry in results:
        pdf_num = pdf_entry["pdf_number"]
        found_metrics = pdf_entry.get("found_metrics", [])

        print(f"  PDF #{pdf_num}: {len(found_metrics)} metric columns found")

        if len(found_metrics) == 0:
            # No matching metrics in tables
            no_metrics_pdfs.append(pdf_num)
            print(f"    [NO METRICS] Will mark as 'No Metrics'")
            continue

        # Fuzzy match found metrics to schema
        matched_data = {}

        for metric in found_metrics:
            metric_lower = metric.lower()

            # 1. Exact match
            matched = None
            for schema_col in schema_columns:
                if metric_lower == schema_col.lower():
                    matched = schema_col
                    break

            # 2. Substring match
            if not matched:
                substring_matches = []
                for schema_col in schema_columns:
                    schema_lower = schema_col.lower()
                    if metric_lower in schema_lower or schema_lower in metric_lower:
                        substring_matches.append(schema_col)

                if substring_matches:
                    result = process.extractOne(metric, substring_matches, scorer=fuzz.ratio)
                    if result:
                        matched = result[0]

            # 3. Fuzzy match
            if not matched:
                result = process.extractOne(
                    metric,
                    schema_columns,
                    scorer=fuzz.token_set_ratio,
                    score_cutoff=60
                )
                if result:
                    matched = result[0]

            # Add matched metric (placeholder value - full parsing would extract actual values)
            if matched:
                matched_data[matched] = "Table Data Found"
                print(f"    Matched: '{metric}' -> '{matched}'")

        if matched_data:
            all_extractions[pdf_num] = matched_data
        else:
            no_metrics_pdfs.append(pdf_num)
            print(f"    [NO MATCHES] No columns matched schema")

    print(f"\n  Summary:")
    print(f"    PDFs with extracted data: {len(all_extractions)}")
    print(f"    PDFs with no metrics: {len(no_metrics_pdfs)}\n")

    return all_extractions, no_metrics_pdfs


def run_pipeline_4b(pdf_folder: Path) -> tuple:
    """
    Run complete Pipeline 4b.

    Args:
        pdf_folder: Folder containing PDFs (from Pipeline 3)

    Returns:
        (metadata_excel_path, final_dataset_path)
    """
    query_folder = pdf_folder.parent

    print("=" * 80)
    print("PIPELINE 4b: TABLE EXTRACTION")
    print("=" * 80)
    print(f"PDF Folder: {pdf_folder}")
    print("=" * 80)
    print()

    # Load schema
    schema_columns = load_schema_columns()
    print(f"Loaded schema: {len(schema_columns)} metrics\n")

    # Load corpus metadata from JSON
    print(f"Loading corpus metadata...")
    all_papers = load_corpus_json(query_folder)
    # Filter to only papers with study_number (exclude failed downloads)
    papers_by_num = {p["study_number"]: p for p in all_papers if "study_number" in p}
    print(f"  Loaded {len(all_papers)} papers\n")

    # Run table extraction (save to Table Crops in Query folder)
    output_folder = query_folder / "Table Crops"
    metrics_file = Path(__file__).parent.parent / "Config" / "MetricsSearchable.txt"

    results_json = extract_tables(pdf_folder, output_folder, metrics_file, query_folder=query_folder)

    # Process results
    all_extractions, no_metrics_pdfs = create_final_dataset_from_tables(
        results_json, schema_columns
    )

    # Update metadata for PDFs with no metrics
    print("[3/5] Updating metadata for PDFs with no table metrics...\n")

    for pdf_num in no_metrics_pdfs:
        if pdf_num in papers_by_num:
            papers_by_num[pdf_num]["breakpoint"] = "No Metrics"
            print(f"  PDF #{pdf_num}: Marked as 'No Metrics'")

    # Create 4b_Metadata.xlsx from updated JSON data
    print(f"\n[4/5] Creating metadata Excel (4b_Metadata.xlsx)...")

    metadata_rows = convert_corpus_json_to_excel_data(all_papers)
    metadata_excel_path = query_folder / "4b_Metadata.xlsx"

    # Write to Excel with colors
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Corpus")

    # Write header row
    for col_idx, value in enumerate(metadata_rows[0], 1):
        ws.cell(row=1, column=col_idx, value=value).font = Font(bold=True)

    # Write data rows with colors
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

    for row_idx, row_data in enumerate(metadata_rows[1:], 2):
        color_code = row_data[-1]
        actual_data = row_data[:-1]

        for col_idx, value in enumerate(actual_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

        if color_code == "green":
            fill = green_fill
        elif color_code == "orange":
            fill = orange_fill
        elif color_code == "red":
            fill = red_fill
        else:
            fill = None

        if fill:
            for col_idx in range(1, len(actual_data) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10

    wb.save(metadata_excel_path)
    print(f"      Saved: {metadata_excel_path.name}")
    print(f"      Updated {len(no_metrics_pdfs)} papers with 'No Metrics'\n")

    # Create final dataset Excel
    print(f"[5/5] Creating final dataset Excel...")
    final_dataset_path = query_folder / "4b_FinalDataset.xlsx"

    columns = ["PDF Number"] + schema_columns

    # Build rows
    rows = []
    for pdf_num, extractions in sorted(all_extractions.items()):
        row = {"PDF Number": pdf_num}
        for metric, value in extractions.items():
            row[metric] = value
        rows.append(row)

    df = pd.DataFrame(rows, columns=columns)

    # Save to Excel
    df.to_excel(final_dataset_path, index=False, engine='openpyxl')
    print(f"      Saved: {final_dataset_path.name}")
    print(f"      Rows: {len(rows)}")
    print(f"      Columns: {len(columns)}\n")

    # Summary
    print("=" * 80)
    print("PIPELINE 4b COMPLETE")
    print("=" * 80)
    print(f"PDFs with extracted table data: {len(all_extractions)}")
    print(f"PDFs with no metrics: {len(no_metrics_pdfs)}")
    print(f"\nOutputs:")
    print(f"  Metadata: {metadata_excel_path.name}")
    print(f"  Dataset:  {final_dataset_path.name}")
    print("=" * 80)

    return metadata_excel_path, final_dataset_path


def find_latest_query_folder():
    """Find the latest Query folder in Output directory."""
    output_base = Path(__file__).parent.parent / "Output"
    query_folders = [d for d in output_base.iterdir() if d.is_dir() and "Query" in d.name]
    if not query_folders:
        return None
    return max(query_folders, key=lambda d: d.stat().st_mtime)


def main():
    """CLI entry point for table extraction."""
    import argparse

    parser = argparse.ArgumentParser(description="Extract tables using PaddleOCR + Surya")
    parser.add_argument("pdf_folder", nargs='?', type=Path, help="Folder containing PDFs")

    args = parser.parse_args()

    # Auto-detect if no arguments provided
    if args.pdf_folder is None:
        query_folder = find_latest_query_folder()
        if query_folder is None:
            print("ERROR: No Query folder found in Output directory")
            sys.exit(1)

        # Use single Corpus PDFs folder (no OA/Non-OA separation)
        papers_folder = query_folder / "Corpus PDFs"
        args.pdf_folder = papers_folder

        print(f"Auto-detected Query folder: {query_folder.name}")

    if not args.pdf_folder.exists():
        print(f"ERROR: PDF folder not found: {args.pdf_folder}")
        sys.exit(1)

    run_pipeline_4b(args.pdf_folder)


if __name__ == "__main__":
    main()
