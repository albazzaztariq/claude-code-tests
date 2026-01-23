"""
ScholarSweep Final Output Combiner

Combines outputs from Pipelines 1-4 into a single final workbook.

PROCESS:
1. Add "Extractor" column to 4a_Metadata.xlsx (mark all rows as "LLM")
2. Add "Extractor" column to 4b_Metadata.xlsx (mark all rows as "TableExtraction")
3. Combine 4a and 4b metadata into single "Extraction Results" sheet (each study has 2 rows)
4. Add "Extractor" column to 4a_FinalDataset.xlsx (mark as "LLM")
5. Add "Extractor" column to 4b_FinalDataset.xlsx (mark as "TableExtraction")
6. Combine both datasets into single "Metrics" sheet
7. Create final 3-sheet workbook:
   - Sheet 1: "Search Results" (all papers from search with upfront numbering)
   - Sheet 2: "Extraction Results" (combined metadata from 4a+4b)
   - Sheet 3: "Metrics" (combined dataset from 4a+4b)
8. Delete intermediate files

NOTE: Pipeline 1 now only downloads actual PDFs. HTML/XML/JSON files are rejected
during download, so no file extension renaming is needed.

INPUTS:
- Corpus Metadata.xlsx (from Pipeline 1+3)
- 4a_Metadata.xlsx
- 4a_FinalDataset.xlsx
- 4b_Metadata.xlsx
- 4b_FinalDataset.xlsx

OUTPUTS:
- Query Output [mmddyy-hh]H[mm]M.xlsx (3 sheets)
"""

import sys
from pathlib import Path
import time
import shutil

# Import timing and corpus utilities
sys.path.insert(0, str(Path(__file__).parent.parent))
from Timers.timing_utils import save_timer, display_timers
from Utils.corpus_utils import load_corpus_json, convert_corpus_json_to_excel_data

# Check for required libraries
try:
    import pandas as pd
except ImportError:
    print("ERROR: Missing required library 'pandas'")
    print("Install it with: pip install pandas openpyxl")
    sys.exit(1)

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    print("ERROR: Missing required library 'openpyxl'")
    print("Install it with: pip install openpyxl")
    sys.exit(1)


def combine_metadata(metadata_4a_path: Path, metadata_4b_path: Path) -> pd.DataFrame:
    """
    Combine 4a and 4b metadata into single DataFrame with "Extractor" column.
    Each study will have 2 rows - one from 4a (LLM), one from 4b (TableExtraction).

    Args:
        metadata_4a_path: Path to 4a_Metadata.xlsx
        metadata_4b_path: Path to 4b_Metadata.xlsx

    Returns:
        Combined DataFrame with "Extractor" column at beginning
    """
    # Load metadata files
    df_4a = pd.read_excel(metadata_4a_path, engine='openpyxl')
    df_4b = pd.read_excel(metadata_4b_path, engine='openpyxl')

    # Add "Extractor" column at the beginning
    df_4a.insert(0, "Extractor", "LLM")
    df_4b.insert(0, "Extractor", "TableExtraction")

    # Combine (each study now has 2 rows)
    combined = pd.concat([df_4a, df_4b], ignore_index=True)

    # Sort by PDF Number so both rows for each study are together
    if "PDF Number" in combined.columns:
        combined = combined.sort_values("PDF Number")

    return combined


def combine_datasets(dataset_4a_path: Path, dataset_4b_path: Path) -> pd.DataFrame:
    """
    Combine 4a and 4b final datasets into single DataFrame with "Extractor" column.

    Args:
        dataset_4a_path: Path to 4a_FinalDataset.xlsx
        dataset_4b_path: Path to 4b_FinalDataset.xlsx

    Returns:
        Combined DataFrame with "Extractor" column at beginning
    """
    # Load datasets
    df_4a = pd.read_excel(dataset_4a_path, engine='openpyxl')
    df_4b = pd.read_excel(dataset_4b_path, engine='openpyxl')

    # Add "Extractor" column
    df_4a.insert(0, "Extractor", "LLM")
    df_4b.insert(0, "Extractor", "TableExtraction")

    # Combine
    combined = pd.concat([df_4a, df_4b], ignore_index=True)

    return combined


def create_final_workbook(
    query_folder: Path,
    combined_metadata: pd.DataFrame,
    combined_dataset: pd.DataFrame
) -> Path:
    """
    Create final 3-sheet workbook.

    Args:
        query_folder: Query folder path
        combined_metadata: Combined metadata from 4a+4b
        combined_dataset: Combined metrics dataset from 4a+4b

    Returns:
        Path to final workbook
    """
    # Extract timestamp from query folder name (format: "mm-dd-yy-hhmm Query")
    # Convert to output format: "Query Output mm-dd-yy-hhHmmM.xlsx"
    folder_name = query_folder.name
    if " Query" in folder_name:
        timestamp_part = folder_name.replace(" Query", "")
        parts = timestamp_part.split("-")
        if len(parts) == 4:
            mm, dd, yy, hhmm = parts
            hh = hhmm[:2]
            min_part = hhmm[2:]
            output_filename = f"Query Output {mm}-{dd}-{yy}-{hh}H{min_part}M.xlsx"
        else:
            output_filename = "Query Output.xlsx"  # Fallback
    else:
        output_filename = "Query Output.xlsx"  # Fallback

    final_workbook_path = query_folder / output_filename

    # Create new workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Sheet 1: "Search Results" (from corpus_metadata.json)
    # Load corpus metadata from JSON
    all_papers = load_corpus_json(query_folder)
    corpus_rows = convert_corpus_json_to_excel_data(all_papers)

    ws1 = wb.create_sheet("Search Results")

    # Write header row
    from openpyxl.styles import Font, PatternFill
    for col_idx, value in enumerate(corpus_rows[0], 1):
        ws1.cell(row=1, column=col_idx, value=value).font = Font(bold=True)

    # Define colors
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

    # Write data rows with NEW color logic
    for row_idx, row_data in enumerate(corpus_rows[1:], 2):
        actual_data = row_data[:-1]  # All except color code
        breakpoint_value = actual_data[2] if len(actual_data) > 2 else ""  # Breakpoint column
        filetype_value = actual_data[1] if len(actual_data) > 1 else ""  # FileType column

        # Write data
        for col_idx, value in enumerate(actual_data, 1):
            ws1.cell(row=row_idx, column=col_idx, value=value)

        # NEW COLOR LOGIC (per user requirements):
        # 1. Red: Anything with content in Breakpoint column
        # 2. Orange: "Unknown" FileType + empty Breakpoint (overrides green)
        # 3. Green: Everything else (empty Breakpoint)
        if breakpoint_value and str(breakpoint_value).strip() != "":
            # Red: Any value in Breakpoint
            fill = red_fill
        elif filetype_value == "Unknown":
            # Orange: Unknown FileType (only when Breakpoint is empty)
            fill = orange_fill
        else:
            # Green: Empty Breakpoint
            fill = green_fill

        # Apply color to entire row
        for col_idx in range(1, len(actual_data) + 1):
            ws1.cell(row=row_idx, column=col_idx).fill = fill

    # Column widths
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 16
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 10

    # Sheet 2: "Extraction Results" (summary of successful extractions)
    ws2 = wb.create_sheet("Extraction Results")

    # Create summary: only PDFs where metrics were found
    # Columns: Extractor, PDF Number, Metrics Found, Data
    ws2.append(["Extractor", "PDF Number", "Metrics Found", "Data"])

    # Count metrics for each PDF/Extractor combination
    for _, row in combined_dataset.iterrows():
        extractor = row.get("Extractor", "")
        pdf_number = row.get("PDF Number", "")

        # Count non-null values (excluding Extractor and PDF Number columns)
        metric_cols = [col for col in combined_dataset.columns if col not in ["Extractor", "PDF Number"]]
        metrics_found = row[metric_cols].notna().sum()

        # Get list of metric names that have non-null values
        found_metric_names = [col for col in metric_cols if pd.notna(row[col])]
        data_list = ", ".join(found_metric_names)

        # Only add row if at least 1 metric was found
        if metrics_found > 0:
            ws2.append([extractor, pdf_number, metrics_found, data_list])

    # Format header row
    from openpyxl.styles import Font
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    # Column widths
    ws2.column_dimensions['A'].width = 18  # Extractor
    ws2.column_dimensions['B'].width = 12  # PDF Number
    ws2.column_dimensions['C'].width = 15  # Metrics Found
    ws2.column_dimensions['D'].width = 60  # Data

    # Sheet 3: "Metrics" (combined dataset from 4a+4b)
    ws3 = wb.create_sheet("Metrics")

    for row in dataframe_to_rows(combined_dataset, index=False, header=True):
        ws3.append(row)

    # Save workbook
    wb.save(final_workbook_path)

    return final_workbook_path


def cleanup_intermediate_files(query_folder: Path, metadata_4a: Path, metadata_4b: Path, dataset_4a: Path, dataset_4b: Path):
    """
    Delete intermediate files and folders.

    Args:
        query_folder: Path to Query folder
        metadata_4a: Path to 4a_Metadata.xlsx
        metadata_4b: Path to 4b_Metadata.xlsx
        dataset_4a: Path to 4a_FinalDataset.xlsx
        dataset_4b: Path to 4b_FinalDataset.xlsx
    """
    # Delete all intermediate files
    corpus_json = query_folder / "corpus_metadata.json"
    yolo_detections = query_folder / "yolo_table_detections.json"
    timers_temp = query_folder / ".timers_temp.json"

    files_to_delete = [
        metadata_4a, metadata_4b, dataset_4a, dataset_4b,
        corpus_json, yolo_detections, timers_temp
    ]

    for file_path in files_to_delete:
        if file_path.exists():
            # Retry deletion with increasing delays (Windows file lock handling)
            for attempt in range(3):
                try:
                    file_path.unlink()
                    break
                except (PermissionError, OSError):
                    if attempt < 2:
                        time.sleep(0.5 * (attempt + 1))  # 0.5s, 1.0s, then give up

    # NOTE: Staging Downloads folder is cleaned up by Pipeline 1
    # No need to clean it up here (avoids 27s of retry delays)


def run_combiner(query_folder: Path, total_runtime: float = None):
    """
    Run final output combiner.

    Args:
        query_folder: Path to Query folder containing all outputs
        total_runtime: Total runtime of entire program in seconds (optional)
    """
    # Timer 11: Start timing final output creation
    combiner_start = time.time()

    print("Creating final dataset...")

    # Find input files
    metadata_4a = query_folder / "4a_Metadata.xlsx"
    metadata_4b = query_folder / "4b_Metadata.xlsx"
    dataset_4a = query_folder / "4a_FinalDataset.xlsx"
    dataset_4b = query_folder / "4b_FinalDataset.xlsx"
    corpus_json = query_folder / "corpus_metadata.json"

    # Verify all files exist
    missing_files = []
    for file_path in [metadata_4a, metadata_4b, dataset_4a, dataset_4b, corpus_json]:
        if not file_path.exists():
            missing_files.append(file_path.name)

    if missing_files:
        sys.exit(1)

    # Combine metadata (4a + 4b -> single Corpus with 2 rows per study)
    combined_metadata = combine_metadata(metadata_4a, metadata_4b)

    # Combine datasets (4a + 4b -> single Metrics sheet)
    combined_dataset = combine_datasets(dataset_4a, dataset_4b)

    # Create final workbook (3 sheets)
    final_workbook = create_final_workbook(
        query_folder,
        combined_metadata,
        combined_dataset
    )

    # Timer 9: Save final output creation time
    combiner_time = time.time() - combiner_start
    save_timer(query_folder, "9_final_output", combiner_time)

    print(f"Operations Successful. Files located at {query_folder.name}")

    # Calculate total runtime from all timer values (if not provided)
    # Pipelines 4a (7_llm_extraction_total) and 4b (8_table_extraction_total) run in PARALLEL
    # so we take MAX instead of SUM for those two
    if total_runtime is None:
        from Timers.timing_utils import load_timers
        timers = load_timers(query_folder)

        # Separate parallel timers (4a and 4b)
        parallel_timers = [
            timers.get("7_llm_extraction_total"),
            timers.get("8_table_extraction_total")
        ]
        parallel_time = max(t for t in parallel_timers if t is not None) if any(t is not None for t in parallel_timers) else 0

        # Sum sequential timers (1-6 and 9)
        sequential_keys = ["1_openalex_total", "2_unpaywall_total", "3_corpus_download_total",
                          "4_text_extraction_total", "5_metrics_match_total", "6_llm_filter_total",
                          "9_final_output"]
        sequential_time = sum(timers.get(k, 0) or 0 for k in sequential_keys)

        total_runtime = sequential_time + parallel_time

    # Display all timers (must happen BEFORE cleanup deletes .timers_temp.json)
    display_timers(query_folder, total_runtime)

    # Cleanup (after display_timers so we can read .timers_temp.json first)
    cleanup_intermediate_files(query_folder, metadata_4a, metadata_4b, dataset_4a, dataset_4b)


def find_latest_query_folder():
    """Find the latest Query folder in Output directory."""
    output_base = Path(__file__).parent.parent / "Output"
    query_folders = [d for d in output_base.iterdir() if d.is_dir() and "Query" in d.name]
    if not query_folders:
        return None
    return max(query_folders, key=lambda d: d.stat().st_mtime)


def main():
    """CLI entry point."""
    import argparse

    parser = argparse.ArgumentParser(description="Combine ScholarSweep pipeline outputs")
    parser.add_argument("query_folder", nargs='?', type=Path, help="Query folder path")

    args = parser.parse_args()

    # Auto-detect if no arguments provided
    if args.query_folder is None:
        args.query_folder = find_latest_query_folder()
        if args.query_folder is None:
            print("ERROR: No Query folder found in Output directory")
            sys.exit(1)

        print(f"Auto-detected Query folder: {args.query_folder.name}")

    if not args.query_folder.exists():
        print(f"ERROR: Query folder not found: {args.query_folder.name}")
        sys.exit(1)

    run_combiner(args.query_folder)


if __name__ == "__main__":
    main()
