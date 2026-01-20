"""
ScholarSweep v01-1 - Clean Rewrite
API search → Deduplicate → OA check → Excel → Download

FLOW:
1. Read config file (terms, journals, fields)
2. Build OpenAlex query with AND/OR/NOT logic
3. Call API and paginate through results
4. Deduplicate by DOI
5. Check OA status via Unpaywall
6. Save to Excel (2 sheets: OA / No OA)
7. Download PDFs to OA Papers/ and Non-OA Papers/ folders

Usage:
    python ScholarSweep.py [config_file.txt]

TEMPORARY: Limited to first 250 PDFs for testing
"""

import sys
import re
import time
import json

# Check for required third-party libraries
try:
    import requests
except ImportError:
    print("ERROR: Missing required library 'requests'")
    print("Install it with: pip install requests")
    sys.exit(1)

from pathlib import Path
from datetime import datetime
from urllib.parse import urlencode
from urllib.request import urlretrieve
import os

# =============================================================================
# CONFIGURATION
# =============================================================================

OPENALEX_API_URL = "https://api.openalex.org/works"
OPENALEX_EMAIL = "albazzaz.tariq@gmail.com"  # Polite pool access (10 req/sec)
UNPAYWALL_EMAIL = "albazzaz.tariq@gmail.com"

# Default config file location
BASE_DIR = Path(__file__).parent
DEFAULT_CONFIG = BASE_DIR / "query.txt"
OUTPUT_DIR = Path(__file__).parent.parent.parent.parent / "Output"  # PROD/Output/

# Journal list import (optional)
try:
    sys.path.insert(0, str(BASE_DIR.parent.parent / "v02" / "v02-2" / "src"))
    from JournalList import TEXTILE_JOURNALS
except ImportError:
    TEXTILE_JOURNALS = {}

# =============================================================================
# CONFIG FILE PARSER
# =============================================================================

def parse_config(filepath: Path) -> dict:
    """Parse config file and return search parameters.

    Returns:
        dict with keys: terms, journals, fields, max_results
    """
    if not filepath.exists():
        raise FileNotFoundError(f"Config file not found: {filepath}")

    with open(filepath, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    # Parse sections
    config = {
        "terms": "",
        "journals": [],
        "fields": ["title", "abstract"],
        "max_results": None
    }

    current_section = None

    for line in lines:
        line = line.strip()

        # Skip comments and blanks
        if not line or line.startswith("#"):
            continue

        # Section headers
        if line.startswith("[") and line.endswith("]"):
            current_section = line[1:-1].lower()
            continue

        # Parse based on section
        if current_section == "openalex":
            if line.startswith("field:"):
                field_str = line.split(":", 1)[1].strip()
                config["fields"] = [f.strip() for f in field_str.split(",")]
            elif line.startswith("terms:"):
                config["terms"] = line.split(":", 1)[1].strip()

        elif current_section == "journals":
            # Expand acronyms to full names
            journal = line.strip()
            if journal in TEXTILE_JOURNALS.values():
                # Find full name from acronym
                for full_name, acronym in TEXTILE_JOURNALS.items():
                    if acronym == journal:
                        config["journals"].append(full_name)
                        break
            else:
                config["journals"].append(journal)

        elif current_section == "max results":
            try:
                config["max_results"] = int(line.replace(",", ""))
            except ValueError:
                pass

    if not config["terms"]:
        raise ValueError("No search terms found in config file")

    return config

# =============================================================================
# OPENALEX QUERY BUILDER
# =============================================================================

def build_openalex_query(config: dict) -> str:
    """Build OpenAlex API URL from config.

    Handles AND/OR/NOT logic in search terms.

    CRITICAL: OpenAlex requires comma-separated filters for AND logic.
    Example: (woven OR knit) AND moisture-wicking becomes:
      filter=fulltext.search:woven OR knit,fulltext.search:moisture-wicking

    Args:
        config: Dict with terms, journals, fields, max_results

    Returns:
        Full API URL with query parameters
    """
    terms = config["terms"]
    fields = config["fields"]

    # Parse terms: /AND/ splits into separate filter parts (comma-separated)
    # Within each AND group, /// = OR
    and_groups = [g.strip() for g in terms.split("/AND/") if g.strip()]

    filter_parts = []

    for and_group in and_groups:
        # Split by /// for OR terms within this AND group
        or_terms = [t.strip() for t in and_group.split("///") if t.strip()]

        # Format each term (quote multi-word phrases)
        formatted_terms = []
        for term in or_terms:
            term = term.strip()
            if ' ' in term:
                # Multi-word phrase: needs quotes
                formatted_terms.append(f'"{term}"')
            else:
                # Single word or hyphenated: no quotes
                formatted_terms.append(term)

        # Join OR terms with " OR "
        or_query = " OR ".join(formatted_terms)

        # Create filter part for this AND group
        if "fulltext" in fields:
            filter_parts.append(f"fulltext.search:{or_query}")
        elif "title" in fields and "abstract" in fields:
            filter_parts.append(f"title_and_abstract.search:{or_query}")
        elif "title" in fields:
            filter_parts.append(f"title.search:{or_query}")
        else:
            # Fallback to search param (not recommended)
            filter_parts.append(f"search:{or_query}")

    # Add journal filter if specified
    if config["journals"]:
        journal_filter = "|".join(config["journals"])
        filter_parts.append(f"primary_location.source.display_name.search:{journal_filter}")

    # Build params
    params = {
        "per_page": 200,
        "cursor": "*",
        "mailto": OPENALEX_EMAIL
    }

    if filter_parts:
        params["filter"] = ",".join(filter_parts)

    return f"{OPENALEX_API_URL}?{urlencode(params, safe=':,|')}"

# =============================================================================
# OPENALEX API CALLER
# =============================================================================

def call_openalex(url: str, max_results: int = None) -> list[dict]:
    """Call OpenAlex API and paginate through all results.

    Args:
        url: OpenAlex API URL with cursor=*
        max_results: Maximum papers to retrieve (None = unlimited)

    Returns:
        List of paper metadata dicts
    """
    papers = []
    current_url = url

    print(f"\n{'='*60}")
    print("CALLING OPENALEX API")
    print(f"{'='*60}")

    while current_url:
        print(f"  Fetching batch {len(papers)//200 + 1}...", end='', flush=True)

        try:
            response = requests.get(current_url, timeout=30)
            response.raise_for_status()
            data = response.json()
        except Exception as e:
            print(f"\n  ERROR: {e}")
            break

        results = data.get("results", [])
        print(f" {len(results)} papers")

        # Parse results
        for item in results:
            paper = {
                "title": item.get("title", ""),
                "author": (item.get("authorships", [{}])[0].get("author", {}).get("display_name", "") if item.get("authorships") else ""),
                "year": item.get("publication_year", ""),
                "doi": item.get("doi", "").replace("https://doi.org/", "") if item.get("doi") else "",
                "abstract": (item.get("abstract_inverted_index", None) and "Abstract available" or ""),
                "pdf_url": "",
                "sources": "OpenAlex"
            }

            # Get DOI URL
            if paper["doi"]:
                paper["doi_url"] = f"https://doi.org/{paper['doi']}"
            else:
                paper["doi_url"] = ""

            # Get PDF URL if available
            if item.get("open_access", {}).get("oa_url"):
                paper["pdf_url"] = item["open_access"]["oa_url"]

            papers.append(paper)

            # Check max results
            if max_results and len(papers) >= max_results:
                print(f"  Reached max results limit ({max_results})")
                return papers

        # Get next cursor
        meta = data.get("meta", {})
        next_cursor = meta.get("next_cursor")

        if next_cursor:
            current_url = url.replace("cursor=*", f"cursor={next_cursor}")
            time.sleep(0.1)  # Rate limiting
        else:
            current_url = None

    print(f"\n  Total papers fetched: {len(papers):,}")
    return papers

# =============================================================================
# DEDUPLICATION
# =============================================================================

def deduplicate(papers: list[dict]) -> list[dict]:
    """Deduplicate papers by DOI and title.

    Args:
        papers: List of paper metadata dicts

    Returns:
        Deduplicated list
    """
    print(f"\n{'='*60}")
    print("DEDUPLICATING")
    print(f"{'='*60}")
    print(f"  Before: {len(papers):,}")

    seen_dois = set()
    seen_titles = set()
    unique = []

    for paper in papers:
        doi = paper.get("doi", "")
        title = paper.get("title", None)

        # Handle None titles
        if title:
            title = title.lower().strip()
        else:
            title = ""

        # Skip if DOI or title already seen
        if doi and doi in seen_dois:
            continue
        if title and title in seen_titles:
            continue

        # Add to unique list
        unique.append(paper)

        if doi:
            seen_dois.add(doi)
        if title:
            seen_titles.add(title)

    print(f"  After: {len(unique):,}")

    # Assign study numbers
    for i, paper in enumerate(unique, 1):
        paper["study_number"] = i
        paper["status"] = "pending"

    return unique

# =============================================================================
# UNPAYWALL OA CHECKER
# =============================================================================

def check_oa_status(papers: list[dict]) -> tuple[list[dict], list[dict]]:
    """Check OA status via Unpaywall and split into OA/non-OA.

    Args:
        papers: List of paper metadata dicts

    Returns:
        Tuple of (oa_papers, non_oa_papers)
    """
    print(f"\n{'='*60}")
    print("CHECKING OA STATUS (Unpaywall)")
    print(f"{'='*60}")

    oa_papers = []
    non_oa_papers = []

    for i, paper in enumerate(papers, 1):
        if i % 50 == 0:
            print(f"  Checked {i}/{len(papers)}...", flush=True)

        doi = paper.get("doi", "")

        if not doi:
            non_oa_papers.append(paper)
            continue

        # Check Unpaywall
        try:
            url = f"https://api.unpaywall.org/v2/{doi}?email={UNPAYWALL_EMAIL}"
            response = requests.get(url, timeout=5)

            if response.status_code == 200:
                data = response.json()

                # Get best OA location
                best_oa = data.get("best_oa_location", {})
                if best_oa and best_oa.get("url_for_pdf"):
                    paper["pdf_url"] = best_oa["url_for_pdf"]
                    oa_papers.append(paper)
                else:
                    non_oa_papers.append(paper)
            else:
                non_oa_papers.append(paper)

        except Exception:
            non_oa_papers.append(paper)

        time.sleep(0.02)  # Unpaywall rate limit: 100k/day

    print(f"\n  OA papers: {len(oa_papers):,}")
    print(f"  Non-OA papers: {len(non_oa_papers):,}")

    return oa_papers, non_oa_papers

# =============================================================================
# EXCEL SAVER
# =============================================================================

def save_to_excel(oa_papers: list[dict], non_oa_papers: list[dict], filepath: Path):
    """Save papers to Excel with OA and No OA worksheets.

    Args:
        oa_papers: List of OA paper metadata
        non_oa_papers: List of non-OA paper metadata
        filepath: Path to save Excel file
    """
    print(f"\n{'='*60}")
    print("SAVING TO EXCEL")
    print(f"{'='*60}")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        headers = ["PDF Number", "Breakpoint", "Sources", "Author", "Title", "Year", "DOI", "DOI URL", "PDF URL", "Abstract"]

        def write_sheet(ws, papers, sheet_name):
            ws.title = sheet_name

            # Headers
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header).font = Font(bold=True)

            # Data
            for row, paper in enumerate(papers, 2):
                status = paper.get("status", "pending")

                ws.cell(row=row, column=1, value=paper.get("study_number", ""))
                # Breakpoint: "ScholarSweep" if failed, blank if downloaded/pending
                ws.cell(row=row, column=2, value="ScholarSweep" if status == "failed" else "")
                ws.cell(row=row, column=3, value=paper.get("sources", ""))
                ws.cell(row=row, column=4, value=paper.get("author", ""))
                ws.cell(row=row, column=5, value=paper.get("title", ""))
                ws.cell(row=row, column=6, value=paper.get("year", ""))
                ws.cell(row=row, column=7, value=paper.get("doi", ""))
                ws.cell(row=row, column=8, value=paper.get("doi_url", ""))
                ws.cell(row=row, column=9, value=paper.get("pdf_url", ""))
                ws.cell(row=row, column=10, value=paper.get("abstract", "")[:32000])

                # Apply color based on status
                if status == "downloaded":
                    fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Green
                    for c in range(1, 11):
                        ws.cell(row=row, column=c).fill = fill
                elif status == "failed":
                    fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Red
                    for c in range(1, 11):
                        ws.cell(row=row, column=c).fill = fill

            # Column widths
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 50
            ws.column_dimensions['F'].width = 8
            ws.column_dimensions['G'].width = 25
            ws.column_dimensions['H'].width = 35
            ws.column_dimensions['I'].width = 40
            ws.column_dimensions['J'].width = 80

        # Write sheets
        ws1 = wb.active
        write_sheet(ws1, oa_papers, "OA")

        ws2 = wb.create_sheet()
        write_sheet(ws2, non_oa_papers, "No OA")

        # Save
        wb.save(filepath)

        print(f"  Saved: {filepath}")
        print(f"    OA sheet: {len(oa_papers):,} papers")
        print(f"    No OA sheet: {len(non_oa_papers):,} papers")

    except ImportError:
        print("  ERROR: openpyxl not installed")
        print("  Install with: pip install openpyxl")

# =============================================================================
# PDF DOWNLOADER WITH PROGRESS BAR
# =============================================================================

def format_time(seconds):
    """Format seconds into HH:MM:SS."""
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    secs = int(seconds % 60)
    if hours > 0:
        return f"{hours:02d}:{minutes:02d}:{secs:02d}"
    else:
        return f"{minutes:02d}:{secs:02d}"

def format_size(bytes_val):
    """Format bytes into MB."""
    return f"{bytes_val / 1024 / 1024:.2f} MB"

def download_with_progress(url, filepath, study_num, paper_type, max_retries=3):
    """Download file with progress tracking and retry logic.

    Args:
        url: URL to download from
        filepath: Path to save file
        study_num: Study number for tracking
        paper_type: "OA" or "Non-OA"
        max_retries: Maximum number of download attempts (default: 3)

    Returns:
        tuple: (success: bool, size: int)
    """
    for attempt in range(max_retries):
        try:
            response = requests.get(url, stream=True, timeout=30)
            if response.status_code != 200:
                continue  # Try again

            total_size = int(response.headers.get('content-length', 0))
            downloaded = 0

            with open(filepath, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
                        downloaded += len(chunk)

            return True, downloaded
        except Exception as e:
            if attempt == max_retries - 1:
                return False, 0
            time.sleep(1)  # Wait 1 second before retry

    return False, 0

def download_pdfs(oa_papers: list[dict], non_oa_papers: list[dict],
                  oa_folder: Path, non_oa_folder: Path, excel_path: Path,
                  api_start_time: float):
    """Download PDFs to OA and Non-OA folders with progress bar.

    Args:
        oa_papers: List of OA papers with pdf_url
        non_oa_papers: List of non-OA papers with DOI
        oa_folder: Directory for OA PDFs
        non_oa_folder: Directory for non-OA PDFs
        excel_path: Path to Excel file for status updates
        api_start_time: Time when API calls started (for runtime tracking)
    """
    print(f"\n{'='*60}")
    print("DOWNLOADING PDFs")
    print(f"{'='*60}\n")

    # Count total downloads needed
    oa_downloadable = [p for p in oa_papers if p.get("pdf_url") and p.get("status") == "pending"]
    non_oa_downloadable = [p for p in non_oa_papers if p.get("doi") and p.get("status") == "pending"]

    total_downloads = len(oa_downloadable) + len(non_oa_downloadable)

    if total_downloads == 0:
        print("  No papers to download.")
        return

    print(f"Total to download: {total_downloads}")
    print(f"  OA: {len(oa_downloadable)}")
    print(f"  Non-OA: {len(non_oa_downloadable)}")
    print()

    # Download tracking
    downloaded_count = 0
    failed_count = 0
    total_bytes = 0
    download_start_time = time.time()

    def update_progress():
        """Update progress bar display."""
        elapsed = time.time() - download_start_time
        api_elapsed = time.time() - api_start_time

        # Calculate progress
        progress = downloaded_count / total_downloads if total_downloads > 0 else 0
        percent = progress * 100

        # Calculate speed
        if elapsed > 0:
            mbps = (total_bytes * 8 / 1024 / 1024) / elapsed  # Megabits per second
            avg_time_per_file = elapsed / downloaded_count if downloaded_count > 0 else 0
            remaining_files = total_downloads - downloaded_count
            eta = avg_time_per_file * remaining_files if downloaded_count > 0 else 0
        else:
            mbps = 0
            eta = 0

        # Progress bar (50 chars) - ASCII for Windows compatibility
        bar_length = 50
        filled = int(bar_length * progress)
        bar = '=' * filled + '-' * (bar_length - filled)

        # Print on same line
        runtime_str = format_time(api_elapsed)
        eta_str = format_time(eta) if eta > 0 else "--:--"

        print(f"\r[{bar}] {percent:5.1f}% | {downloaded_count}/{total_downloads} | "
              f"{mbps:6.2f} Mbps | ETA: {eta_str} | Runtime: {runtime_str}", end='', flush=True)

    # Download OA papers
    for paper in oa_downloadable:
        pdf_url = paper.get("pdf_url", "")
        study_num = paper.get("study_number", "unknown")
        filename = f"{study_num}.pdf"
        filepath = oa_folder / filename

        success, size = download_with_progress(pdf_url, filepath, study_num, "OA")

        if success:
            paper["status"] = "downloaded"
            downloaded_count += 1
            total_bytes += size
        else:
            paper["status"] = "failed"  # Mark as failed after 3 retries
            failed_count += 1

        update_progress()

    # Download non-OA papers (SciHub)
    for paper in non_oa_downloadable:
        doi = paper.get("doi", "")
        study_num = paper.get("study_number", "unknown")
        filename = f"{study_num}.pdf"
        filepath = non_oa_folder / filename

        # Try SciHub mirrors
        scihub_urls = [
            f"https://sci-hub.se/{doi}",
            f"https://sci-hub.st/{doi}",
            f"https://sci-hub.ru/{doi}"
        ]

        success = False
        for scihub_url in scihub_urls:
            result, size = download_with_progress(scihub_url, filepath, study_num, "Non-OA")
            if result:
                paper["status"] = "downloaded"
                downloaded_count += 1
                total_bytes += size
                success = True
                break

        if not success:
            paper["status"] = "failed"  # Mark as failed after trying all mirrors
            failed_count += 1

        update_progress()

    # Final progress update
    print()  # New line after progress bar

    # Summary
    elapsed = time.time() - download_start_time
    api_elapsed = time.time() - api_start_time

    print(f"\n{'='*60}")
    print("DOWNLOAD SUMMARY")
    print(f"{'='*60}")
    print(f"  Downloaded: {downloaded_count}/{total_downloads}")
    print(f"  Failed: {failed_count}")
    print(f"  Total size: {format_size(total_bytes)}")
    print(f"  Download time: {format_time(elapsed)}")
    print(f"  Total runtime: {format_time(api_elapsed)}")
    if elapsed > 0:
        avg_mbps = (total_bytes * 8 / 1024 / 1024) / elapsed
        print(f"  Average speed: {avg_mbps:.2f} Mbps")
    print(f"{'='*60}")

    # Update Excel with status changes (green for success, red for failed)
    if downloaded_count > 0 or failed_count > 0:
        save_to_excel(oa_papers, non_oa_papers, excel_path)

# =============================================================================
# TEXT EXTRACTION FROM PDFS (TEXTILE VISION)
# =============================================================================

def extract_text_from_pdfs(oa_papers: list[dict], non_oa_papers: list[dict],
                           oa_folder: Path, non_oa_folder: Path):
    """Extract text from all downloaded PDFs using TextileVision pipeline.

    Calls the PaddleOCR + Tesseract text extraction script.

    Args:
        oa_papers: List of OA papers
        non_oa_papers: List of non-OA papers
        oa_folder: Directory with OA PDFs
        non_oa_folder: Directory with non-OA PDFs
    """
    print(f"\n{'='*60}")
    print("EXTRACTING TEXT FROM PDFs")
    print(f"{'='*60}\n")

    # Import TextileVision extraction module
    import sys
    textilevision_path = Path(r"C:\Users\azt12\OneDrive\Documents\Business\Textile\WrestlingRobe\Programming\FabricETL\DEV\Code_Root\TextileVision\Doc-Img-Analysis_PipelineDev\Text\TEST_PIPELINE\v01\v01-1")
    sys.path.insert(0, str(textilevision_path))

    try:
        from paddleocr_layout_masking_tesseract import extract_text_from_pdf
    except ImportError as e:
        print(f"  ERROR: Could not import TextileVision extraction module")
        print(f"  {e}")
        print("  Skipping text extraction.")
        return

    # Collect all PDFs that were successfully downloaded
    all_papers = oa_papers + non_oa_papers
    downloaded_pdfs = []

    for paper in all_papers:
        if paper.get("status") == "downloaded":
            study_num = paper.get("study_number")
            # Check which folder it's in
            pdf_path_oa = oa_folder / f"{study_num}.pdf"
            pdf_path_non_oa = non_oa_folder / f"{study_num}.pdf"

            if pdf_path_oa.exists():
                downloaded_pdfs.append((study_num, pdf_path_oa))
            elif pdf_path_non_oa.exists():
                downloaded_pdfs.append((study_num, pdf_path_non_oa))

    if not downloaded_pdfs:
        print("  No PDFs to extract text from.")
        return

    print(f"  Extracting text from {len(downloaded_pdfs)} PDFs...")
    print()

    # Process each PDF
    for study_num, pdf_path in downloaded_pdfs:
        print(f"  Processing PDF {study_num}... ", end='', flush=True)

        try:
            # Extract text
            extracted_text = extract_text_from_pdf(pdf_path)

            # Save to .txt file next to PDF
            txt_path = pdf_path.parent / f"{study_num}.txt"
            with open(txt_path, 'w', encoding='utf-8') as f:
                f.write(extracted_text)

            print(f"✓ ({len(extracted_text):,} chars, {len(extracted_text.split()):,} words)")
        except Exception as e:
            print(f"✗ Error: {e}")

    print(f"\n{'='*60}")
    print("TEXT EXTRACTION COMPLETE")
    print(f"{'='*60}")

# =============================================================================
# MAIN ORCHESTRATOR
# =============================================================================

def main():
    """Main orchestrator."""
    print("="*60)
    print("SCHOLARSWEEP v01-1")
    print("="*60)

    # Track API start time
    api_start_time = time.time()

    # Get config file path
    if len(sys.argv) > 1:
        config_path = Path(sys.argv[1])
    else:
        config_path = DEFAULT_CONFIG

    # Parse config
    print(f"\nConfig file: {config_path}")
    config = parse_config(config_path)

    print(f"  Terms: {config['terms'][:80]}...")
    print(f"  Fields: {', '.join(config['fields'])}")
    print(f"  Journals: {len(config['journals'])} specified")
    print(f"  Max results: {config['max_results'] or 'Unlimited'}")

    # TEMPORARY: Limit to 250 for testing
    print(f"\n  *** TEST MODE: Limiting to first 250 PDFs ***")

    # Create output folders
    timestamp = datetime.now().strftime("%m-%d-%y-%H%M")
    query_folder = OUTPUT_DIR / f"{timestamp} Query"
    oa_folder = query_folder / "Downloaded Papers" / "OA Papers"
    non_oa_folder = query_folder / "Downloaded Papers" / "Non-OA Papers"

    oa_folder.mkdir(parents=True, exist_ok=True)
    non_oa_folder.mkdir(parents=True, exist_ok=True)

    excel_path = query_folder / f"{timestamp} Papers.xlsx"

    print(f"\n{'='*60}")
    print("API CALLS BEING BUILT...")
    print(f"{'='*60}")

    # Build query
    url = build_openalex_query(config)
    print("  Query built successfully.\n")

    print(f"{'='*60}")
    print("API CALLS BEING MADE...")
    print(f"{'='*60}")

    # Call API (limit to 250)
    papers = call_openalex(url, max_results=250)

    if not papers:
        print("\nResults: 0")
        print("\nNo papers found.")
        return

    print(f"\nResults: {len(papers):,} papers retrieved")

    # Deduplicate
    papers = deduplicate(papers)

    # Limit to max_results from config (for testing)
    max_pdfs = config.get('max_results', 250)
    if len(papers) > max_pdfs:
        print(f"\n  Trimming from {len(papers)} to {max_pdfs} papers (from config)...")
        papers = papers[:max_pdfs]
        # Re-assign study numbers
        for i, paper in enumerate(papers, 1):
            paper["study_number"] = i

    # Check OA status
    oa_papers, non_oa_papers = check_oa_status(papers)

    # Save to Excel
    save_to_excel(oa_papers, non_oa_papers, excel_path)

    # Download PDFs (with progress tracking)
    download_pdfs(oa_papers, non_oa_papers, oa_folder, non_oa_folder, excel_path, api_start_time)

    # Extract text from downloaded PDFs
    extract_text_from_pdfs(oa_papers, non_oa_papers, oa_folder, non_oa_folder)

    print(f"\n{'='*60}")
    print("COMPLETE")
    print(f"{'='*60}")
    print(f"  Results: {excel_path}")
    print(f"  PDFs: {query_folder / 'Downloaded Papers'}")

if __name__ == "__main__":
    main()
