#!/usr/bin/env python3
"""
ScrapeV2.py - Comprehensive Textile Specification Extractor

Uses Nougat OCR for table extraction and LLM (Ollama/OpenAI) for semantic
understanding to extract all yarn, fabric, and transfer metrics from PDFs.

Focused on 1.pdf for initial development.
"""

import os
import re
import json
import sys
from pathlib import Path
from typing import Optional, Dict, List, Any, Tuple
import requests
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Using Nougat CLI instead of Python API to avoid albumentations issues

# ================== CONFIGURATION ==================
# LLM Configuration - Choose one
LLM_PROVIDER = os.getenv("LLM_PROVIDER", "ollama")  # "ollama" or "openai"

# Ollama settings
OLLAMA_URL = os.getenv("OLLAMA_URL", "http://localhost:11434/api/generate")
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "qwen2.5:14b")  # Good balance of speed and quality

# OpenAI settings (if using)
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4-turbo-preview")

# Paths
BASE_DIR = Path(
    r"C:\Users\azt12\OneDrive\Documents\Wrestling Robe\Materials Science - Wickability\Studies for Analysis by LLM AI\Datafiles & Python Scripts"
)
PDF_FOLDER = Path(
    r"C:\Users\azt12\OneDrive\Documents\Wrestling Robe\Materials Science - Wickability\Studies for Analysis by LLM AI"
)
OUTPUT_EXCEL = BASE_DIR / "Textile-Specifications-Extracted.xlsx"

# Focus on 1.pdf only for now
FOCUS_PDF = "1.pdf"

# ================== METRICS SCHEMA ==================
# This defines ALL the metrics we want to extract
# Each metric has: name, layer_types (which layer columns it applies to), category

LAYER_TYPES = ["Inner Layer", "Middle Layer", "Outer Layer", "Monolayer"]
SIDES = ["Bottom Side", "Top Side", ""]  # "" means general/both sides

# Study-level metadata (one per study)
STUDY_METADATA = [
    "Study Number",
    "Study Title",
    "Year of Publish",
    "Name of First-Listed Author",
    "Number of Sample Fabrics",
    "Testing/Standards Bodies Methods Used",
]

# Sample-level metadata (one per sample)
SAMPLE_METADATA = [
    "Sample ID/Name",
    "Number of Fabric Layers",
]

# Material, Structure, Treatment fields - per layer
MATERIAL_STRUCTURE_FIELDS = [
    {"name": "Material", "layers": True},
    {"name": "Structure", "layers": True, "note": "Knit or Woven or both"},
]

# Treatment fields - per layer and per side
TREATMENT_FIELDS = [
    {"name": "Treatment", "layers": True, "sides": True},
    {"name": "Change in Water Affinity", "layers": True, "sides": True,
     "note": "e.g. hydrophobic -> superhydrophobic"},
]

# Yarn specification fields - per layer
YARN_SPEC_FIELDS = [
    {"name": "Yarn Fiber Type", "layers": True, "note": "staple, filament, or composite"},
    {"name": "Filament Yarn Texture", "layers": True, "note": "Spun or Filament"},
    {"name": "Filament Yarn Structure", "layers": True, "note": "Textured or Flat (filament only)"},
    {"name": "Filament Yarn Cross-section Type", "layers": True, "note": "round, trilobal, etc."},
    {"name": "Yarn Denier/Linear Density", "layers": True},
    {"name": "Filament Count", "layers": True, "note": "number of filaments per yarn"},
    {"name": "Yarn Twist", "layers": True, "note": "direction and turns per inch/cm"},
]

# Porosity fields - per layer
POROSITY_FIELDS = [
    {"name": "Pore Size", "layers": True},
    {"name": "Pore Size Distribution", "layers": True},
    {"name": "Open Area Fraction", "layers": True},
    {"name": "Hydraulic Diameter", "layers": True},
]

# Structure-specific fields
STRUCTURE_SPECIFIC_FIELDS = [
    {"name": "Crimp (Wovens)", "layers": True, "note": "wovens only"},
    {"name": "Loop Length (Knits)", "layers": True, "note": "knits only"},
    {"name": "Float Length (Knits)", "layers": True, "note": "knits only"},
    {"name": "Proportion of Tucks/Misses (Knits)", "layers": True, "note": "knits only"},
    {"name": "Surface Roughness", "layers": True, "note": "in micrometers or microinches"},
]

# Physical properties - per layer
PHYSICAL_FIELDS = [
    {"name": "Fabric Weight/GSM", "layers": True, "note": "g/m² or oz/yd²"},
    {"name": "Fabric Thickness", "layers": True, "note": "mm or mil"},
    {"name": "Fabric Density", "layers": True, "note": "threads/inch, picks/inch, loops/inch"},
    {"name": "Tensile Strength", "layers": True},
    {"name": "Elongation at Break", "layers": True},
    {"name": "Tear Strength", "layers": True},
    {"name": "Fabric Hand/Feel", "layers": True},
]

# AATCC Moisture metrics - per layer
AATCC_MOISTURE_FIELDS = [
    {"name": "Wetting Time Bottom (WTB)", "layers": True, "standard": "AATCC"},
    {"name": "Wetting Time Top (WTT)", "layers": True, "standard": "AATCC"},
    {"name": "Wetting Time (WT)", "layers": True, "standard": "AATCC"},
    {"name": "Absorption Rate Bottom (ARB)", "layers": True, "standard": "AATCC"},
    {"name": "Absorption Rate Top (ART)", "layers": True, "standard": "AATCC"},
    {"name": "Bottom Absorption Rate (BAR)", "layers": True, "standard": "AATCC"},
    {"name": "Top Absorption Rate (TAR)", "layers": True, "standard": "AATCC"},
    {"name": "Wicking Distance/Rate", "layers": True, "standard": "AATCC"},
    {"name": "Vertical Wicking", "layers": True, "standard": "AATCC"},
    {"name": "Spreading Speed Bottom (SSb)", "layers": True, "standard": "AATCC"},
    {"name": "Spreading Speed Top (SSt)", "layers": True, "standard": "AATCC"},
    {"name": "Spreading Speed (SS)", "layers": True, "standard": "AATCC"},
    {"name": "Max Wetted Radius Bottom (MWRb)", "layers": True, "standard": "AATCC"},
    {"name": "Max Wetted Radius Top (MWRt)", "layers": True, "standard": "AATCC"},
    {"name": "Accumulative One-Way Transport Index (AOTI)", "layers": True, "standard": "AATCC"},
    {"name": "One Way Transport Capability (OWTC)", "layers": True, "standard": "AATCC"},
    {"name": "Overall Moisture Management Capacity (OMMC)", "layers": True, "standard": "AATCC"},
    {"name": "Drying Rate", "layers": True, "standard": "AATCC"},
    {"name": "Drying Time", "layers": True, "standard": "AATCC"},
    {"name": "Moisture Retention", "layers": True, "standard": "AATCC"},
    {"name": "Capillary Rise Height", "layers": True, "standard": "AATCC"},
    {"name": "Capillary Rise Rate", "layers": True, "standard": "AATCC"},
    {"name": "Water Retention", "layers": True, "standard": "AATCC"},
    {"name": "Maximum Water Uptake", "layers": True, "standard": "AATCC"},
]

# Wettability fields - per layer
WETTABILITY_FIELDS = [
    {"name": "Contact Angle", "layers": True},
    {"name": "Contact Angle - Advancing", "layers": True},
    {"name": "Contact Angle - Receding", "layers": True},
    {"name": "Wettability Classification", "layers": True, "note": "hydrophobic, hydrophilic, etc."},
    {"name": "Wettability Gradient", "layers": True},
    {"name": "Surface Energy/Tension", "layers": True},
    {"name": "Interfacial Tension", "layers": True},
]

# Transport metrics - per layer
TRANSPORT_FIELDS = [
    {"name": "Moisture Transport (Transplanar)", "layers": True},
    {"name": "Moisture Transport (Horizontal)", "layers": True},
    {"name": "Directional Moisture Transport", "layers": True},
    {"name": "Unidirectional Transport Direction", "layers": True},
    {"name": "Air Permeability", "layers": True},
    {"name": "Liquid Permeability", "layers": True},
    {"name": "Water Vapor Transmission Rate (WVTR)", "layers": True},
    {"name": "Moisture Vapor Transmission Rate (MVTR)", "layers": True},
]

# Thermal metrics - per layer
THERMAL_FIELDS = [
    {"name": "Thermal Conductivity", "layers": True},
    {"name": "Thermal Resistance (Rct)", "layers": True},
    {"name": "Evaporative Resistance (Ret)", "layers": True},
    {"name": "Thermal Diffusivity", "layers": True},
    {"name": "Heat Flux", "layers": True},
    {"name": "Thermal Absorptivity (Qmax)", "layers": True},
    {"name": "Heat Dissipation/Loss", "layers": True},
    {"name": "Breathability", "layers": True},
    {"name": "Comfort Index", "layers": True},
    {"name": "Moisture Management Capacity (MMC)", "layers": True},
]

# JIS-specific metrics - per layer
JIS_FIELDS = [
    {"name": "Wetting Time (JIS)", "layers": True, "standard": "JIS"},
    {"name": "Absorption Rate (JIS)", "layers": True, "standard": "JIS"},
    {"name": "Spreading Speed (JIS)", "layers": True, "standard": "JIS"},
    {"name": "Vertical Wicking (JIS)", "layers": True, "standard": "JIS"},
    {"name": "Horizontal Wicking (JIS)", "layers": True, "standard": "JIS"},
    {"name": "Drying Time (JIS)", "layers": True, "standard": "JIS"},
    {"name": "Drying Rate (JIS)", "layers": True, "standard": "JIS"},
    {"name": "Drying Index (JIS)", "layers": True, "standard": "JIS"},
    {"name": "Moisture Regain (JIS)", "layers": True, "standard": "JIS"},
    {"name": "Water Vapor Resistance (JIS)", "layers": True, "standard": "JIS"},
    {"name": "WVTR (JIS)", "layers": True, "standard": "JIS"},
    {"name": "Evaporative Heat Flux (JIS)", "layers": True, "standard": "JIS"},
]

# ISO/EN-specific metrics - per layer
ISO_EN_FIELDS = [
    {"name": "Thermal Conductivity (ISO)", "layers": True, "standard": "ISO/EN"},
    {"name": "Thermal Resistance (ISO)", "layers": True, "standard": "ISO/EN"},
    {"name": "Thermal Diffusivity (ISO)", "layers": True, "standard": "ISO/EN"},
    {"name": "Heat Flux (ISO)", "layers": True, "standard": "ISO/EN"},
    {"name": "Hydrostatic Head (ISO/EN)", "layers": True, "standard": "ISO/EN"},
    {"name": "Water Resistance (ISO/EN)", "layers": True, "standard": "ISO/EN"},
    {"name": "Air Permeability (ISO)", "layers": True, "standard": "ISO/EN"},
    {"name": "Water Vapor Resistance (ISO)", "layers": True, "standard": "ISO/EN"},
    {"name": "Water Vapor Permeability (ISO)", "layers": True, "standard": "ISO/EN"},
    {"name": "Evaporative Heat Flux (ISO)", "layers": True, "standard": "ISO/EN"},
    {"name": "Drying Rate (ISO)", "layers": True, "standard": "ISO/EN"},
]

# ASTM-specific metrics - per layer
ASTM_FIELDS = [
    {"name": "Qualitative Absorbency Rating (ASTM)", "layers": True, "standard": "ASTM"},
    {"name": "Hydrostatic Head (ASTM)", "layers": True, "standard": "ASTM"},
    {"name": "Water Resistance (ASTM)", "layers": True, "standard": "ASTM"},
    {"name": "Air Permeability (ASTM)", "layers": True, "standard": "ASTM"},
    {"name": "WVTR (ASTM)", "layers": True, "standard": "ASTM"},
    {"name": "Water Vapor Permeability (ASTM)", "layers": True, "standard": "ASTM"},
    {"name": "Water Absorbency (ASTM)", "layers": True, "standard": "ASTM"},
    {"name": "Absorption Time (ASTM)", "layers": True, "standard": "ASTM"},
    {"name": "Liquid Absorption Capacity (ASTM)", "layers": True, "standard": "ASTM"},
    {"name": "Thermal Resistance Rct (ASTM)", "layers": True, "standard": "ASTM"},
    {"name": "Evaporative Resistance Ret (ASTM)", "layers": True, "standard": "ASTM"},
]

# Other/unspecified standard metrics
OTHER_FIELDS = [
    {"name": "Capillary Pressure", "layers": True, "standard": "Other"},
    {"name": "SMF", "layers": True, "standard": "Other"},
    {"name": "PMI", "layers": True, "standard": "Other"},
]


def get_all_column_headers() -> List[str]:
    """
    Generate all column headers based on the schema.
    Returns a flat list of all column names.
    """
    headers = []

    # Study metadata
    headers.extend(STUDY_METADATA)

    # Sample metadata
    headers.extend(SAMPLE_METADATA)

    # Build layer-specific columns
    all_field_groups = [
        ("Material/Structure", MATERIAL_STRUCTURE_FIELDS),
        ("Yarn Specs", YARN_SPEC_FIELDS),
        ("Porosity", POROSITY_FIELDS),
        ("Structure-Specific", STRUCTURE_SPECIFIC_FIELDS),
        ("Physical Properties", PHYSICAL_FIELDS),
        ("Wettability", WETTABILITY_FIELDS),
        ("Transport", TRANSPORT_FIELDS),
        ("Thermal", THERMAL_FIELDS),
        ("AATCC Moisture", AATCC_MOISTURE_FIELDS),
        ("JIS", JIS_FIELDS),
        ("ISO/EN", ISO_EN_FIELDS),
        ("ASTM", ASTM_FIELDS),
        ("Other", OTHER_FIELDS),
    ]

    # Treatment fields with sides
    for field in TREATMENT_FIELDS:
        for layer in LAYER_TYPES:
            if field.get("sides"):
                for side in SIDES:
                    if side:
                        headers.append(f"{field['name']}: {layer}, {side}")
                    else:
                        headers.append(f"{field['name']}: {layer}")
            else:
                headers.append(f"{field['name']}: {layer}")

    # Regular layer fields (no sides)
    for group_name, fields in all_field_groups:
        for field in fields:
            if field.get("layers"):
                for layer in LAYER_TYPES:
                    headers.append(f"{field['name']}: {layer}")
            else:
                headers.append(field['name'])

    return headers


# ================== LLM FUNCTIONS ==================

def call_ollama(prompt: str, system_prompt: str = "") -> str:
    """Call local Ollama LLM."""
    full_prompt = f"{system_prompt}\n\n{prompt}" if system_prompt else prompt

    payload = {
        "model": OLLAMA_MODEL,
        "prompt": full_prompt,
        "stream": False,
        "options": {
            "temperature": 0.1,  # Low temperature for more deterministic extraction
            "num_predict": 8000,  # Allow longer responses for JSON
        }
    }

    try:
        resp = requests.post(OLLAMA_URL, json=payload, timeout=300)
        resp.raise_for_status()
        data = resp.json()
        return data.get("response", "")
    except requests.exceptions.ConnectionError:
        print("ERROR: Cannot connect to Ollama. Is it running?")
        print("Start Ollama with: ollama serve")
        return ""
    except Exception as e:
        print(f"Ollama error: {e}")
        return ""


def call_openai(prompt: str, system_prompt: str = "") -> str:
    """Call OpenAI API."""
    if not OPENAI_API_KEY:
        print("ERROR: OPENAI_API_KEY not set")
        return ""

    import openai
    client = openai.OpenAI(api_key=OPENAI_API_KEY)

    messages = []
    if system_prompt:
        messages.append({"role": "system", "content": system_prompt})
    messages.append({"role": "user", "content": prompt})

    try:
        response = client.chat.completions.create(
            model=OPENAI_MODEL,
            messages=messages,
            temperature=0.1,
            max_tokens=8000,
        )
        return response.choices[0].message.content
    except Exception as e:
        print(f"OpenAI error: {e}")
        return ""


def call_llm(prompt: str, system_prompt: str = "") -> str:
    """Call the configured LLM provider."""
    if LLM_PROVIDER == "openai":
        return call_openai(prompt, system_prompt)
    else:
        return call_ollama(prompt, system_prompt)


# ================== EXTRACTION FUNCTIONS ==================

def extract_with_nougat(pdf_path: str) -> str:
    """
    Extract text from PDF using Nougat CLI.
    """
    import subprocess
    import tempfile
    import glob

    try:
        print(f"    Extracting with Nougat CLI...")

        with tempfile.TemporaryDirectory() as tmpdir:
            cmd = ["nougat", str(pdf_path), "-o", tmpdir, "--no-skipping"]
            print(f"    Running: {' '.join(cmd)}")

            result = subprocess.run(cmd, capture_output=True, text=True, timeout=600)

            if result.returncode != 0:
                print(f"    Nougat CLI error: {result.stderr}")
                sys.exit(1)

            mmd_files = glob.glob(os.path.join(tmpdir, "*.mmd"))

            if not mmd_files:
                print("    No .mmd output file found")
                sys.exit(1)

            with open(mmd_files[0], "r", encoding="utf-8") as f:
                markdown_text = f.read()

            print(f"    Got {len(markdown_text)} characters")
            return markdown_text

    except subprocess.TimeoutExpired:
        print("    Nougat timed out")
        sys.exit(1)
    except FileNotFoundError:
        print("    Nougat CLI not found. Install with: pip install nougat-ocr")
        sys.exit(1)
    except Exception as e:
        print(f"    Nougat ERROR: {e}")
        sys.exit(1)


def extract_study_metadata(pdf_path: str, nougat_text: str) -> Dict[str, Any]:
    """
    Extract study-level metadata using ONE consolidated LLM call.
    NOUGAT TEXT ONLY - no PyMuPDF.
    """
    metadata = {}

    # ONE consolidated LLM call for all metadata
    consolidated_prompt = f"""
Analyze this textile research paper and extract the following information.
Return ONLY a JSON object with these exact keys:

{{
  "title": "the main paper title",
  "year": 2021,
  "first_author_lastname": "Smith",
  "sample_count": 8,
  "standards": "AATCC, ISO 9237"
}}

RULES:
- title: The main paper title (first 200 chars max)
- year: Publication year as integer (look for "Published", "Received", "Accepted", copyright)
- first_author_lastname: Last name of first author only
- sample_count: EXACT number of fabric samples tested (look for "eight samples", "10 fabrics", sample IDs like S1-S8, table rows)
- standards: Comma-separated list of testing standards (AATCC, ISO, JIS, ASTM, etc.) or "Not specified"

PAPER TEXT:
{nougat_text[:15000]}

Return ONLY the JSON object:"""

    response = call_llm(consolidated_prompt)

    # Parse JSON response
    try:
        json_match = re.search(r'\{[\s\S]*\}', response)
        if json_match:
            data = json.loads(json_match.group())
            metadata["Study Title"] = str(data.get("title", "Not found"))[:200]
            metadata["Year of Publish"] = data.get("year", "Not found")
            metadata["Name of First-Listed Author"] = str(data.get("first_author_lastname", "Not found"))[:50]
            metadata["Number of Sample Fabrics"] = int(data.get("sample_count", 1))
            metadata["Testing/Standards Bodies Methods Used"] = data.get("standards", "Not specified")
        else:
            print(f"  Warning: Could not parse metadata JSON")
            metadata["Study Title"] = "Not found"
            metadata["Year of Publish"] = "Not found"
            metadata["Name of First-Listed Author"] = "Not found"
            metadata["Number of Sample Fabrics"] = 1
            metadata["Testing/Standards Bodies Methods Used"] = "Not specified"
    except Exception as e:
        print(f"  Warning: Metadata extraction error: {e}")
        metadata["Study Title"] = "Not found"
        metadata["Year of Publish"] = "Not found"
        metadata["Name of First-Listed Author"] = "Not found"
        metadata["Number of Sample Fabrics"] = 1
        metadata["Testing/Standards Bodies Methods Used"] = "Not specified"

    return metadata


def extract_all_samples_with_metrics(nougat_text: str, expected_count: int) -> List[Dict[str, Any]]:
    """
    ONE LLM call to extract ALL samples with ALL metrics found in the paper.
    LLM finds what's there, then we match to schema columns.
    """

    extraction_prompt = f"""
Analyze this textile research paper. Extract ALL data for ALL {expected_count} fabric samples.

For EACH sample, extract:
1. sample_id: The ID/name from the paper
2. ALL properties and metrics you can find (material, structure, thickness, GSM, any test results)

Return as JSON array. Include ANY metric you find - use the exact name from the paper.
Example format:
[
  {{
    "sample_id": "S1",
    "material": "100% polyester",
    "structure": "single jersey",
    "gsm": 150,
    "thickness_mm": 0.45,
    "WTT": 3.2,
    "WTB": 2.8,
    "ARB": 45.5,
    "ART": 42.1,
    "SSb": 2.1,
    "SSt": 1.9,
    "MWRb": 25,
    "MWRt": 22,
    "AOTI": 380,
    "OMMC": 0.65,
    ... any other metrics found ...
  }}
]

Extract ALL numeric values from tables. Include units where shown.

PAPER TEXT:
{nougat_text[:30000]}

Return ONLY the JSON array:"""

    response = call_llm(extraction_prompt)

    # Parse JSON response
    try:
        json_match = re.search(r'\[[\s\S]*\]', response)
        if json_match:
            samples = json.loads(json_match.group())
            # Pad if needed
            while len(samples) < expected_count:
                samples.append({"sample_id": f"Sample {len(samples)+1}"})
            return samples[:expected_count]
    except json.JSONDecodeError as e:
        print(f"  Warning: Failed to parse samples JSON: {e}")

    # Fallback
    return [{"sample_id": f"Sample {i}"} for i in range(1, expected_count + 1)]


# Mapping from common metric names/abbreviations to schema column names
METRIC_MAPPING = {
    # Basic properties
    "material": "Material: Monolayer",
    "structure": "Structure: Monolayer",
    "gsm": "Fabric Weight/GSM: Monolayer",
    "weight": "Fabric Weight/GSM: Monolayer",
    "thickness": "Fabric Thickness: Monolayer",
    "thickness_mm": "Fabric Thickness: Monolayer",

    # AATCC MMT metrics
    "wtt": "Wetting Time Top (WTT): Monolayer",
    "wtb": "Wetting Time Bottom (WTB): Monolayer",
    "wt": "Wetting Time (WT): Monolayer",
    "wetting_time": "Wetting Time (WT): Monolayer",
    "art": "Top Absorption Rate (TAR): Monolayer",
    "arb": "Bottom Absorption Rate (BAR): Monolayer",
    "tar": "Top Absorption Rate (TAR): Monolayer",
    "bar": "Bottom Absorption Rate (BAR): Monolayer",
    "absorption_rate": "Absorption Rate Top (ART): Monolayer",
    "sst": "Spreading Speed Top (SSt): Monolayer",
    "ssb": "Spreading Speed Bottom (SSb): Monolayer",
    "ss": "Spreading Speed (SS): Monolayer",
    "spreading_speed": "Spreading Speed (SS): Monolayer",
    "mwrt": "Max Wetted Radius Top (MWRt): Monolayer",
    "mwrb": "Max Wetted Radius Bottom (MWRb): Monolayer",
    "aoti": "Accumulative One-Way Transport Index (AOTI): Monolayer",
    "owtc": "One Way Transport Capability (OWTC): Monolayer",
    "ommc": "Overall Moisture Management Capacity (OMMC): Monolayer",

    # Thermal
    "thermal_resistance": "Thermal Resistance (Rct): Monolayer",
    "rct": "Thermal Resistance (Rct): Monolayer",
    "thermal_conductivity": "Thermal Conductivity: Monolayer",
    "ret": "Evaporative Resistance (Ret): Monolayer",
    "evaporative_resistance": "Evaporative Resistance (Ret): Monolayer",

    # Other
    "air_permeability": "Air Permeability: Monolayer",
    "wvtr": "Water Vapor Transmission Rate (WVTR): Monolayer",
    "contact_angle": "Contact Angle: Monolayer",
    "porosity": "Open Area Fraction: Monolayer",
    "drying_time": "Drying Time: Monolayer",
    "drying_rate": "Drying Rate: Monolayer",
}


def map_extracted_to_schema(sample: Dict[str, Any]) -> Dict[str, Any]:
    """Map extracted metric names to schema column names."""
    mapped = {}
    for key, value in sample.items():
        if value is None:
            continue
        key_lower = key.lower().replace(" ", "_").replace("-", "_")
        if key_lower in METRIC_MAPPING:
            mapped[METRIC_MAPPING[key_lower]] = value
        else:
            # Keep original if no mapping (for sample_id, etc.)
            mapped[key] = value
    return mapped


def process_single_pdf(pdf_path: str) -> List[Dict[str, Any]]:
    """
    Process a single PDF and extract all specifications.
    Returns a list of rows (one per sample).
    NOUGAT ONLY - no PyMuPDF fallback.
    Uses only 2 LLM calls total for speed.
    """
    print(f"\n{'='*60}")
    print(f"Processing: {pdf_path}")
    print(f"{'='*60}")

    # Extract text using NOUGAT ONLY
    nougat_text = extract_with_nougat(pdf_path)

    # DEBUG: Print first part of nougat output
    print("\n  === NOUGAT OUTPUT (first 2000 chars) ===")
    print(nougat_text[:2000])
    print("  === END NOUGAT OUTPUT ===\n")

    # LLM CALL 1: Get study metadata (consolidated - title, year, author, sample count, standards)
    print("\n  [LLM Call 1/2] Extracting study metadata...")
    study_metadata = extract_study_metadata(pdf_path, nougat_text)
    for key, value in study_metadata.items():
        print(f"    {key}: {value}")

    # LLM CALL 2: Extract ALL samples with ALL metrics in ONE call
    sample_count = study_metadata.get("Number of Sample Fabrics", 1)
    print(f"\n  [LLM Call 2/2] Extracting {sample_count} samples with metrics...")
    samples = extract_all_samples_with_metrics(nougat_text, sample_count)
    print(f"    Extracted {len(samples)} samples")

    # Build output rows - just map extracted data to schema (no more LLM calls)
    output_rows = []
    print("\n  === EXTRACTED DATA ===")
    for idx, sample in enumerate(samples):
        sample_id = sample.get("sample_id", f"Sample {idx+1}")
        print(f"\n  Sample {idx+1}: {sample_id}")

        # Show raw extracted data
        for key, value in sample.items():
            if key != "sample_id" and value is not None:
                print(f"    {key}: {value}")

        # Start with study metadata
        row = {
            "Study Number": Path(pdf_path).stem,
            **study_metadata,
            "Sample ID/Name": sample_id,
            "Number of Fabric Layers": sample.get("num_layers", 1),
        }

        # Map extracted metrics to schema columns (instant, no LLM)
        mapped_metrics = map_extracted_to_schema(sample)
        row.update(mapped_metrics)

        output_rows.append(row)

    print("\n  === END EXTRACTED DATA ===")
    return output_rows


def write_to_excel(rows: List[Dict[str, Any]], output_path: Path):
    """Write extracted data to Excel file."""
    print(f"\n{'='*60}")
    print(f"Writing to Excel: {output_path}")
    print(f"{'='*60}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Textile Specifications"

    # Get all headers
    all_headers = get_all_column_headers()

    # Write header row with formatting
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(all_headers, 1):
            value = row_data.get(header, "")
            if value == "null" or value is None:
                value = ""
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Adjust column widths
    for col_idx in range(1, len(all_headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"  ✓ Saved {len(rows)} rows to {output_path}")


def main():
    """Main entry point."""
    print("\n" + "="*70)
    print("ScrapeV2 - Comprehensive Textile Specification Extractor")
    print("="*70)
    print(f"LLM Provider: {LLM_PROVIDER}")
    if LLM_PROVIDER == "ollama":
        print(f"Ollama Model: {OLLAMA_MODEL}")
    print(f"Focus PDF: {FOCUS_PDF}")
    print("="*70)

    # Check PDF exists
    pdf_path = PDF_FOLDER / FOCUS_PDF
    if not pdf_path.exists():
        print(f"\nERROR: PDF not found: {pdf_path}")
        print("Please check the path and try again.")
        return 1

    # Process the PDF
    all_rows = process_single_pdf(str(pdf_path))

    if not all_rows:
        print("\nNo data extracted. Check the PDF content and LLM connection.")
        return 1

    # Write to Excel
    write_to_excel(all_rows, OUTPUT_EXCEL)

    print("\n" + "="*70)
    print("EXTRACTION COMPLETE")
    print(f"Total samples extracted: {len(all_rows)}")
    print(f"Output file: {OUTPUT_EXCEL}")
    print("="*70 + "\n")

    return 0


if __name__ == "__main__":
    sys.exit(main())
